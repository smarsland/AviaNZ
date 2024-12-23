# Version 3.4 18/12/24
# Authors: Stephen Marsland, Nirosha Priyadarshani, Julius Juodakis, Virginia Listanti, Giotto Frean

# This is the main class for the AviaNZ interface.

#    AviaNZ bioacoustic analysis program
#    Copyright (C) 2017--2024

#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.

#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.

#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.

# TODO: 
# 1. Check Freebird import, BatSearch output
# 2. Test Harry's overlaps into excel 
# 2a -> Replace (or not...) the mergeSegments whereever that was
# 3. James' filters list
# 3a -> replace warning about deleting excels?
# 3b -> sort out logic for actually running multiple filters (e.g. for resampling)
# 3c -> size of filter box
# 3d -> WTF merge segments things
# 3e -> tidy code
# 4. Finish Neural Networks
# 5. Finish clustering
# 6. Colour maps
# 7. Check merge with main cf media playback cleanup, CompareCalls dialog bugs changes
# 8. Fix OutputBatPasses -- JSON
# 9. Mac installation
# 10. Test and test again. And again.
# 11. If a .data file is empty, delete, restart -- check
# 12. Replace call type selection with dropdown from species
# 13. Ruth things
# 14. Is librosa used?
# 15. Tidy up the flac loader (correct place, check for space) and add a converter
# 16. Check the splitting with data files properly

import sys, os, json, platform, re, shutil, csv
from shutil import copyfile

from PyQt6 import QtCore, QtGui
from PyQt6.QtGui import QIcon, QStandardItemModel, QStandardItem, QKeySequence, QPixmap, QCursor
from PyQt6.QtWidgets import QApplication, QInputDialog, QFileDialog, QMainWindow, QToolButton, QLabel, QSlider, QScrollBar, QDoubleSpinBox, QPushButton, QListWidgetItem, QMenu, QFrame, QMessageBox, QWidgetAction, QComboBox, QTreeView, QGraphicsProxyWidget, QWidget, QVBoxLayout, QGroupBox, QSizePolicy, QHBoxLayout, QSpinBox, QAbstractSpinBox, QLineEdit, QStyle, QWizard#, QActionGroup, QShortcut
from PyQt6.QtWidgets import QGraphicsBlurEffect
# The two below will move from QtWidgets
from PyQt6.QtGui import QActionGroup, QShortcut
from PyQt6.QtCore import Qt, QDir, QTimer, QPoint, QPointF, QLocale, QModelIndex, QRectF #, QThread
from PyQt6.QtMultimedia import QAudio, QAudioFormat

import wavio
import numpy as np
from scipy.ndimage.filters import median_filter

import pyqtgraph as pg
from pyqtgraph.dockarea import DockArea, Dock
import pyqtgraph.functions as fn
import pyqtgraph.exporters as pge
from pyqtgraph.parametertree import Parameter, ParameterTree

import SupportClasses, SupportClasses_GUI
import Dialogs
import DialogsTraining
import Spectrogram
import Segment
import WaveletSegment
import WaveletFunctions
import Clustering
import colourMaps
import Shapes
import SignalProc

from functools import partial

import re

import webbrowser, copy, math
import time
import openpyxl
# TODO: Check this
from lxml import etree as ET
#import xml.etree.ElementTree as ET

pg.setConfigOption('useNumba', True)
pg.setConfigOption('background','w')
pg.setConfigOption('foreground','k')
pg.setConfigOption('antialias',True)
#print("Package import complete.")

# import pdb
# from PyQt5.QtCore import pyqtRemoveInputHook
# from pdb import set_trace
#
# def debug_trace():
#     pyqtRemoveInputHook()
#     set_trace()


class AviaNZ(QMainWindow):
    """Main class for the user interface.
    Contains most of the user interface and plotting code"""

    def __init__(self,root=None,configdir=None,CLI=False,cheatsheet=False,zooniverse=False,firstFile='', imageFile='', command=''):
        """Initialisation of the class. Load main config and bird lists from configdir.
        Also initialises the data structures and loads an initial file (specified explicitly)
        and sets up the window.
        One interesting configuration point is the DOC setting, which hides the more 'research' functions."""
        print("Starting AviaNZ...")

        super(AviaNZ, self).__init__()
        self.root = root
        self.CLI = CLI
        self.cheatsheet = cheatsheet
        self.zooniverse = zooniverse

        # configdir passes the standard user app dir based on OS.
        # At this point, the main config file should already be ensured to exist.
        self.configdir = configdir
        self.configfile = os.path.join(configdir, "AviaNZconfig.txt")
        self.ConfigLoader = SupportClasses.ConfigLoader()
        self.config = self.ConfigLoader.config(self.configfile)
        self.saveConfig = True
        print("Config loaded")

        # Load filters
        self.filtersDir = os.path.join(configdir, self.config['FiltersDir'])
        self.FilterDicts = self.ConfigLoader.filters(self.filtersDir)
        print("Filters loaded")

        # Load the birdlists - both are now necessary:
        self.shortBirdList = self.ConfigLoader.shortbl(self.config['BirdListShort'], configdir)
        if self.shortBirdList is None:
            raise OSError("Short bird list missing, cannot continue")
        self.longBirdList = self.ConfigLoader.longbl(self.config['BirdListLong'], configdir)
        if self.longBirdList is None:
            raise OSError("Long bird list missing, cannot continue")
        self.batList = self.ConfigLoader.batl(self.config['BatList'], configdir)
        if self.batList is None:
            raise OSError("Bat list missing, cannot continue")
        
        # Load the known calls so far
        self.knownCalls = {}
        for filt in self.FilterDicts.values():
            if not filt["species"] in self.knownCalls:
                self.knownCalls[filt["species"]]=[]
            
            for subf in filt["Filters"]:
                if not subf["calltype"] in self.knownCalls[filt["species"]] and not subf["calltype"]=="Any" and not subf["calltype"]=="Other":
                    self.knownCalls[filt["species"]].append(subf["calltype"])

        # avoid comma/point problem in number parsing
        QLocale.setDefault(QLocale(QLocale.Language.English,QLocale.Country.NewZealand))
        #QLocale.setDefault(QLocale(QLocale.English, QLocale.NewZealand))
        print('Locale is set to ' + QLocale().name())

        # The data structures for the segments
        self.listLabels = []
        self.listRectanglesa1 = []
        self.listRectanglesa2 = []
        self.SegmentRects = []
        self.segmentPlots = []
        self.shapePlots = []
        self.box1id = -1

        self.started = False
        self.startedInAmpl = False
        self.startTime = 0
        self.segmentsToSave = False
        self.viewCallType = False
        self.CallTypeMenu = False
        self.batmode = False

        self.lastSpecies = [{"species": "Don't Know", "certainty": 0, "filter": "M"}]
        self.DOC = self.config['DOC']
        self.extra = "none"
        self.playSpeed = 1.0
        self.playingVisible = False 

        self.noisefloor = 0

        # Whether or not the context menu allows multiple birds.
        self.multipleBirds = self.config['MultipleSpecies']

        if len(self.config['RecentFiles']) > 0:
            self.SoundFileDir = os.path.dirname(self.config['RecentFiles'][-1])
            if not os.path.isdir(self.SoundFileDir):
                self.SoundFileDir = self.config['SoundFileDir']
        else:
            self.SoundFileDir = self.config['SoundFileDir']
        self.filename = None
        self.focusRegion = None
        self.operator = self.config['operator']
        self.reviewer = self.config['reviewer']
        self.filters = []

        # For preventing callbacks involving overview panel
        self.updateRequestedByOverview = False

        # working directory
        if not os.path.isdir(self.SoundFileDir):
            print("Directory doesn't exist: making it")
            os.makedirs(self.SoundFileDir)

        #self.backupDatafiles()

        # INPUT FILE LOADING
        # search order: infile -> firstFile -> dialog
        # Make life easier for now: preload a birdsong
        if not os.path.isfile(firstFile) and not cheatsheet and not zooniverse:
            # For distribution:
            firstFile = self.SoundFileDir
            # Can also use:
            # firstFile = self.SoundFileDir + '/' + 'kiwi_1min.wav'

        if not os.path.isfile(firstFile) and not cheatsheet and not zooniverse:
            if self.CLI:
                print("file %s not found, exiting" % firstFile)
                raise OSError("No input file, cannot continue")
            else:
                # pop up a dialog to select file
                firstFile, drop = QFileDialog.getOpenFileName(self, 'Choose File', self.SoundFileDir, "WAV or BMP files (*.wav *.bmp);; Only WAV files (*.wav);; Only BMP files (*.bmp);; FLAC files (*.flac)")
                while firstFile == '':
                    msg = SupportClasses_GUI.MessagePopup("w", "Select Sound File", "Choose a sound file to proceed.\nDo you want to continue?")
                    msg.setStandardButtons(QMessageBox.StandardButton.No)
                    msg.addButton("Choose a file", QMessageBox.ButtonRole.YesRole)
                    msg.button(QMessageBox.StandardButton.No).setText("Exit")
                    reply = msg.exec()
                    if reply == 0:
                        firstFile, drop = QFileDialog.getOpenFileName(self, 'Choose File', self.SoundFileDir, "WAV or BMP files (*.wav *.bmp);; Only WAV files (*.wav);; Only BMP files (*.bmp);; FLAC files (*.flac)")
                    else:
                        sys.exit()

        # parse firstFile to dir and file parts
        if not cheatsheet and not zooniverse:
            self.SoundFileDir = os.path.dirname(firstFile)
            print("Working dir set to %s" % self.SoundFileDir)
            print("Opening file %s" % firstFile)

        # to keep code simpler, graphic options are created even in CLI mode
        # they're just not shown because QMainWindow.__init__ is skipped
        if not self.CLI:
            QMainWindow.__init__(self, root)

        # parse mouse settings
        if self.config['drawingRightBtn']:
            self.MouseDrawingButton = Qt.MouseButton.RightButton
        else:
            self.MouseDrawingButton = Qt.MouseButton.LeftButton

        self.createMenu()
        self.createFrame()

        # Boxes with area smaller than this will be ignored -
        # to avoid accidentally creating little boxes
        self.minboxsize = 0.1

        self.resetStorageArrays()
        if self.CLI:
            if cheatsheet or zooniverse:
                # use infile as directory
                # TODO: imagefile?
                print(firstFile)
                self.SoundFileDir = firstFile
                # Read folders and sub-folders
                for root, dirs, files in os.walk(firstFile):
                    for f in files:
                        if f[-4:].lower() == '.wav':
                            print(os.path.join(root, f))
                            self.loadFile(os.path.join(root, f), cs=True)
                            self.widthWindow.setValue(60)  # self.datalengthSec)
                            print('file path: ', os.path.join(root, f[:-4]))
                            self.setColourLevels(20, 50)
                            self.saveImage(os.path.join(root, f[:-4]+'.png'))
            else:
                self.loadFile(firstFile)
                while command!=():
                    c = command[0]
                    command = command[1:]
                    print("Next command to execute is %s" % c)
                    if c=="denoise":
                        self.denoise()
                    elif c=="segment":
                        self.segment()
                    else:
                        print("ERROR: %s is not a valid command" % c)
                        raise ValueError("CLI command not recognized")
                if imageFile!='':
                    # reset images to show full width if in CLI:
                    self.widthWindow.setValue(self.datalengthSec)
                    self.saveImage(imageFile)
        else:
            # Make the window and associated widgets
            self.setWindowTitle('AviaNZ')
            self.setWindowIcon(QIcon('img/AviaNZ.ico'))
            # Show the window
            if self.config['StartMaximized']:
                self.showMaximized()
                # extra toggle because otherwise Windows starts at a non-maximized size
                self.setWindowState(self.windowState() ^ Qt.WindowState.WindowMaximized)
                self.setWindowState(self.windowState() | Qt.WindowState.WindowMaximized)
            else:
                self.show()

            # Save the segments every minute
            self.timer = QTimer()
            self.timer.timeout.connect(self.saveSegments)
            self.timer.start(self.config['secsSave']*1000)

            self.listLoadFile(os.path.basename(firstFile))

        if self.DOC and not cheatsheet and not zooniverse:
            self.setOperatorReviewerDialog()

    def createMenu_qt5(self):
        """ Create the menu entries at the top of the screen and link them as appropriate.
        Some of them are initialised according to the data in the configuration file."""
        # Legacy code, tiny change in Qt6

        fileMenu = self.menuBar().addMenu("&File")
        openIcon = self.style().standardIcon(QStyle.SP_DialogOpenButton)
        fileMenu.addAction(openIcon, "&Open sound file", self.openFile, "Ctrl+O")
        # fileMenu.addAction("&Change Directory", self.chDir)
        fileMenu.addAction("Set Operator/Reviewer (Current File)", self.setOperatorReviewerDialog)
        fileMenu.addSeparator()
        for recentfile in self.config['RecentFiles']:
            fileMenu.addAction(recentfile, lambda arg=recentfile: self.openFile(arg))
        fileMenu.addSeparator()
        fileMenu.addAction("Restart Program",self.restart,"Ctrl+R")
        fileMenu.addAction(QIcon(QPixmap('img/exit.png')), "&Quit",QApplication.quit,"Ctrl+Q")

        # This is a very bad way to do this, but I haven't worked anything else out (setMenuRole() didn't work)
        # Add it a second time, then it appears!
        if platform.system() == 'Darwin':
            fileMenu.addAction("&Quit",QApplication.quit,"Ctrl+Q")

        specMenu = self.menuBar().addMenu("&Appearance")

        self.useAmplitudeTick = specMenu.addAction("Show amplitude plot", self.useAmplitudeCheck)
        self.useAmplitudeTick.setCheckable(True)
        self.useAmplitudeTick.setChecked(self.config['showAmplitudePlot'])
        self.useAmplitude = True

        self.useFilesTick = specMenu.addAction("Show file list", self.useFilesCheck)
        self.useFilesTick.setCheckable(True)
        self.useFilesTick.setChecked(self.config['showListofFiles'])

        # this can go under "Change interface settings"
        self.showOverviewSegsTick = specMenu.addAction("Show annotation overview", self.showOverviewSegsCheck)
        self.showOverviewSegsTick.setCheckable(True)
        self.showOverviewSegsTick.setChecked(self.config['showAnnotationOverview'])

        self.showPointerDetails = specMenu.addAction("Show pointer details in spectrogram", self.showPointerDetailsCheck)
        self.showPointerDetails.setCheckable(True)
        self.showPointerDetails.setChecked(self.config['showPointerDetails'])

        specMenu.addSeparator()

        colMenu = specMenu.addMenu("Choose colour map")
        colGroup = QActionGroup(self)
        for colour in self.config['ColourList']:
            cm = colMenu.addAction(colour)
            cm.setCheckable(True)
            if colour==self.config['cmap']:
                cm.setChecked(True)
            receiver = lambda checked, cmap=colour: self.setColourMap(cmap)
            cm.triggered.connect(receiver)
            colGroup.addAction(cm)
        self.invertcm = specMenu.addAction("Invert colour map",self.invertColourMap)
        self.invertcm.setCheckable(True)
        self.invertcm.setChecked(self.config['invertColourMap'])

        # specMenu.addSeparator()
        specMenu.addAction("&Change spectrogram parameters",self.showSpectrogramDialog, "Ctrl+C")

        if not self.DOC:
            specMenu.addSeparator()
            self.showDiagnosticTick = specMenu.addAction("Show training diagnostics",self.showDiagnosticDialog)
            self.showDiagnosticCNN = specMenu.addAction("Show CNN training diagnostics", self.showDiagnosticDialogCNN)
            self.extraMenu = specMenu.addMenu("Diagnostic plots")
            extraGroup = QActionGroup(self)
            for ename in ["none", "Wavelet scalogram", "Wavelet correlations", "Wind energy", "Rain", "Wind adjustment", "Filtered spectrogram, new + AA", "Filtered spectrogram, new", "Filtered spectrogram, old"]:
                em = self.extraMenu.addAction(ename)
                em.setCheckable(True)
                if ename == self.extra:
                    em.setChecked(True)
                receiver = lambda checked, ename=ename: self.setExtraPlot(ename)
                em.triggered.connect(receiver)
                extraGroup.addAction(em)

        specMenu.addSeparator()
        markMenu = specMenu.addMenu("Mark on spectrogram")
        self.showFundamental = markMenu.addAction("Fundamental frequency", self.showFundamentalFreq,"Ctrl+F")
        self.showFundamental.setCheckable(True)
        self.showFundamental.setChecked(True)
        self.showSpectral = markMenu.addAction("Spectral derivative", self.showSpectralDeriv)
        self.showSpectral.setCheckable(True)
        self.showSpectral.setChecked(False)
        if not self.DOC:
            self.showFormant = markMenu.addAction("Formants", self.showFormants)
            self.showFormant.setCheckable(True)
            self.showFormant.setChecked(False)
        self.showEnergies = markMenu.addAction("Maximum energies", self.showMaxEnergy)
        self.showEnergies.setCheckable(True)
        self.showEnergies.setChecked(False)

        # if not self.DOC:
        #     cqt = specMenu.addAction("Show CQT", self.showCQT)

        specMenu.addSeparator()

        self.readonly = specMenu.addAction("Make read only",self.makeReadOnly)
        self.readonly.setCheckable(True)
        self.readonly.setChecked(self.config['readOnly'])

        specMenu.addSeparator()
        specMenu.addAction("Interface settings", self.changeSettings)
        specMenu.addAction("Put docks back",self.dockReplace)

        actionMenu = self.menuBar().addMenu("&Actions")
        actionMenu.addAction("Delete all segments", self.deleteAll, "Ctrl+D")
        self.addRegularAction = actionMenu.addAction("Mark regular segments", self.addRegularSegments, "Ctrl+M")

        actionMenu.addSeparator()
        self.denoiseAction = actionMenu.addAction("Denoise",self.showDenoiseDialog)
        actionMenu.addAction("Add metadata about noise", self.addNoiseData, "Ctrl+N")
        #actionMenu.addAction("Find matches",self.findMatches)

        #if not self.DOC:
            #actionMenu.addAction("Filter spectrogram",self.medianFilterSpec)
            #actionMenu.addAction("Denoise spectrogram",self.denoiseImage)

        actionMenu.addSeparator()
        self.segmentAction = actionMenu.addAction("Segment",self.segmentationDialog,"Ctrl+S")

        if not self.DOC:
            actionMenu.addAction("Calculate segment statistics", self.calculateStats)
            actionMenu.addAction("Analyse shapes", self.showShapesDialog)
            actionMenu.addAction("Cluster segments", self.classifySegments)

        #actionMenu.addSeparator()
        #self.showInvSpec = actionMenu.addAction("Save sound file", self.invertSpectrogram)

        actionMenu.addSeparator()

        if not self.DOC:
            actionMenu.addAction("Export spectrogram image", self.saveImageRaw)
        actionMenu.addAction("&Export current view as image",self.saveImage,"Ctrl+I")

        # "Recognisers" menu
        recMenu = self.menuBar().addMenu("&Recognisers")
        extrarecMenu = recMenu.addMenu("Train an automated recogniser")
        extrarecMenu.addAction("Train a changepoint recogniser", self.buildChpRecogniser)
        if not self.DOC:
            extrarecMenu.addAction("Train a wavelet recogniser", self.buildRecogniser)

        extrarecMenu.addAction("Extend a recogniser with CNN", self.buildCNN)
        recMenu.addAction("Test a recogniser", self.testRecogniser)
        recMenu.addAction("Manage recognisers", self.manageFilters)
        recMenu.addAction("Customise a recogniser (use existing ROC)", self.customiseFiltersROC)

        # "Utilities" menu
        utilMenu = self.menuBar().addMenu("&Utilities")
        utilMenu.addAction("Import from Excel", self.excel2Annotation)
        utilMenu.addAction("Import from Freebird", self.tag2Annotation)
        utilMenu.addAction("Backup annotations", self.backupAnnotations)
        utilMenu.addAction("&Split WAV/DATA files", self.launchSplitter)

        helpMenu = self.menuBar().addMenu("&Help")
        helpMenu.addAction("Help", self.showHelp, "Ctrl+H")
        helpMenu.addAction("&Cheat Sheet", self.showCheatSheet)
        helpMenu.addSeparator()
        helpMenu.addAction("About", self.showAbout, "Ctrl+A")
        if platform.system() == 'Darwin':
            helpMenu.addAction("About", self.showAbout, "Ctrl+A")

    def createMenu(self):
        # In Qt6 the order of addAction changes
        """ Create the menu entries at the top of the screen and link them as appropriate.
        Some of them are initialised according to the data in the configuration file."""

        fileMenu = self.menuBar().addMenu("&File")
        openIcon = self.style().standardIcon(QStyle.StandardPixmap.SP_DialogOpenButton)
        fileMenu.addAction(openIcon, "&Open sound file", "Ctrl+O", self.openFile)
        #fileMenu.addAction(openIcon, "&Open sound file", self.openFile, "Ctrl+O")
        # fileMenu.addAction("&Change Directory", self.chDir)
        fileMenu.addAction("Set Operator/Reviewer (Current File)", self.setOperatorReviewerDialog)
        fileMenu.addSeparator()
        for recentfile in self.config['RecentFiles']:
            fileMenu.addAction(recentfile, lambda arg=recentfile: self.openFile(arg))
        fileMenu.addSeparator()
        fileMenu.addAction("Restart Program","Ctrl+R",self.restart)
        #fileMenu.addAction("Restart Program",self.restart,"Ctrl+R")
        #fileMenu.addAction(QIcon(QPixmap('img/exit.png')), "&Quit",QApplication.quit,"Ctrl+Q")
        fileMenu.addAction(QIcon(QPixmap('img/exit.png')), "&Quit","Ctrl+Q",QApplication.quit)

        # This is a very bad way to do this, but I haven't worked anything else out (setMenuRole() didn't work)
        # Add it a second time, then it appears!
        if platform.system() == 'Darwin':
            fileMenu.addAction("&Quit","Ctrl+Q",QApplication.quit)
            #fileMenu.addAction("&Quit",QApplication.quit,"Ctrl+Q")

        specMenu = self.menuBar().addMenu("&Appearance")

        self.useAmplitudeTick = specMenu.addAction("Show amplitude plot", self.useAmplitudeCheck)
        self.useAmplitudeTick.setCheckable(True)
        self.useAmplitudeTick.setChecked(self.config['showAmplitudePlot'])
        self.useAmplitude = True

        self.useFilesTick = specMenu.addAction("Show file list", self.useFilesCheck)
        self.useFilesTick.setCheckable(True)
        self.useFilesTick.setChecked(self.config['showListofFiles'])

        # this can go under "Change interface settings"
        self.showOverviewSegsTick = specMenu.addAction("Show annotation overview", self.showOverviewSegsCheck)
        self.showOverviewSegsTick.setCheckable(True)
        self.showOverviewSegsTick.setChecked(self.config['showAnnotationOverview'])

        self.showPointerDetails = specMenu.addAction("Show pointer details in spectrogram", self.showPointerDetailsCheck)
        self.showPointerDetails.setCheckable(True)
        self.showPointerDetails.setChecked(self.config['showPointerDetails'])

        specMenu.addSeparator()

        colMenu = specMenu.addMenu("Choose colour map")
        colGroup = QActionGroup(self)
        for colour in self.config['ColourList']:
            cm = colMenu.addAction(colour)
            cm.setCheckable(True)
            if colour==self.config['cmap']:
                cm.setChecked(True)
            receiver = lambda checked, cmap=colour: self.setColourMap(cmap)
            cm.triggered.connect(receiver)
            colGroup.addAction(cm)
        self.invertcm = specMenu.addAction("Invert colour map",self.invertColourMap)
        self.invertcm.setCheckable(True)
        self.invertcm.setChecked(self.config['invertColourMap'])

        # specMenu.addSeparator()
        specMenu.addAction("&Change spectrogram parameters","Ctrl+C",self.showSpectrogramDialog)
        #specMenu.addAction("&Change spectrogram parameters",self.showSpectrogramDialog, "Ctrl+C")

        if not self.DOC:
            specMenu.addSeparator()
            self.showDiagnosticTick = specMenu.addAction("Show training diagnostics",self.showDiagnosticDialog)
            self.showDiagnosticCNN = specMenu.addAction("Show CNN training diagnostics", self.showDiagnosticDialogCNN)
            self.extraMenu = specMenu.addMenu("Diagnostic plots")
            extraGroup = QActionGroup(self)
            for ename in ["none", "Wavelet scalogram", "Wavelet correlations", "Wind energy", "Rain", "Wind adjustment", "Filtered spectrogram, new + AA", "Filtered spectrogram, new", "Filtered spectrogram, old"]:
                em = self.extraMenu.addAction(ename)
                em.setCheckable(True)
                if ename == self.extra:
                    em.setChecked(True)
                receiver = lambda checked, ename=ename: self.setExtraPlot(ename)
                em.triggered.connect(receiver)
                extraGroup.addAction(em)

        specMenu.addSeparator()
        markMenu = specMenu.addMenu("Mark on spectrogram")
        self.showFundamental = markMenu.addAction("Fundamental frequency","Ctrl+F", self.showFundamentalFreq)
        self.showFundamental.setCheckable(True)
        self.showFundamental.setChecked(True)
        if not self.DOC:
            self.showSpectral = markMenu.addAction("Spectral derivative", self.showSpectralDeriv)
            self.showSpectral.setCheckable(True)
            self.showSpectral.setChecked(False)
            self.showFormant = markMenu.addAction("Formants", self.showFormants)
            self.showFormant.setCheckable(True)
            self.showFormant.setChecked(False)
        self.showEnergies = markMenu.addAction("Maximum energies", self.showMaxEnergy)
        self.showEnergies.setCheckable(True)
        self.showEnergies.setChecked(False)

        specMenu.addSeparator()

        self.readonly = specMenu.addAction("Make read only",self.makeReadOnly)
        self.readonly.setCheckable(True)
        self.readonly.setChecked(self.config['readOnly'])

        specMenu.addSeparator()
        specMenu.addAction("Interface settings", self.changeSettings)
        specMenu.addAction("Put docks back",self.dockReplace)

        actionMenu = self.menuBar().addMenu("&Actions")
        actionMenu.addAction("Delete all segments","Ctrl+D",self.deleteAll)
        self.addRegularAction = actionMenu.addAction("Mark regular segments","Ctrl+M", self.addRegularSegments)

        actionMenu.addSeparator()
        self.denoiseAction = actionMenu.addAction("Denoise",self.showDenoiseDialog)
        actionMenu.addAction("Add metadata about noise","Ctrl+N",self.addNoiseData)

        #if not self.DOC:
            #actionMenu.addAction("Filter spectrogram",self.medianFilterSpec)
            #actionMenu.addAction("Denoise spectrogram",self.denoiseImage)

        actionMenu.addSeparator()
        self.segmentAction = actionMenu.addAction("Segment","Ctrl+S",self.segmentationDialog)

        if not self.DOC:
            actionMenu.addAction("Calculate segment statistics", self.calculateStats)
            actionMenu.addAction("Analyse shapes", self.showShapesDialog)
            actionMenu.addAction("Cluster segments", self.classifySegments)

        #actionMenu.addSeparator()
        # TODO: Bug
        #self.showInvSpec = actionMenu.addAction("Save sound file", self.invertSpectrogram)

        actionMenu.addSeparator()

        if not self.DOC:
            actionMenu.addAction("Export spectrogram image", self.saveImageRaw)
        actionMenu.addAction("&Export current view as image","Ctrl+I",self.saveImage)

        # "Recognisers" menu
        recMenu = self.menuBar().addMenu("&Recognisers")
        extrarecMenu = recMenu.addMenu("Train an automated recogniser")
        extrarecMenu.addAction("Train a changepoint recogniser", self.buildChpRecogniser)
        if not self.DOC:
            extrarecMenu.addAction("Train a wavelet recogniser", self.buildRecogniser)

        extrarecMenu.addAction("Extend a recogniser with CNN", self.buildCNN)
        recMenu.addAction("Test a recogniser", self.testRecogniser)
        recMenu.addAction("Manage recognisers", self.manageFilters)
        recMenu.addAction("Customise a recogniser (use existing ROC)", self.customiseFiltersROC)

        # "Utilities" menu
        utilMenu = self.menuBar().addMenu("&Utilities")
        utilMenu.addAction("Import from Excel", self.excel2Annotation)
        utilMenu.addAction("Import from Freebird", self.tag2Annotation)
        utilMenu.addAction("Backup annotations", self.backupAnnotations)
        utilMenu.addAction("&Split WAV/DATA files", self.launchSplitter)

        helpMenu = self.menuBar().addMenu("&Help")
        helpMenu.addAction("Help","Ctrl+H", self.showHelp)
        helpMenu.addAction("&Cheat Sheet", self.showCheatSheet)
        helpMenu.addSeparator()
        helpMenu.addAction("About","Ctrl+A", self.showAbout)
        if platform.system() == 'Darwin':
            helpMenu.addAction("About","Ctrl+A", self.showAbout)

    def showAbout(self):
        """ Create the About Message Box"""
        msg = SupportClasses_GUI.MessagePopup("a", "About", ".")
        msg.exec()
        return

    def showHelp(self):
        """ Show the user manual (a pdf file), make it offline for easy access"""
        webbrowser.open_new(r'http://avianz.net/docs/AviaNZManual.pdf')

    def showCheatSheet(self):
        """ Show the cheatsheet of sample spectrograms"""
        webbrowser.open_new(r'http://www.avianz.net/index.php/resources/cheat-sheet/about-cheat-sheet')

    def launchSplitter(self):
        """ Close the main window, start splitter QMainWindow """
        print("Switching to AviaNZ Splitter")
        QApplication.exit(2)

    def createFrame(self):
        """ Creates the main window.
        This consists of a set of pyqtgraph docks with widgets in.
         d_ for docks, w_ for widgets, p_ for plots"""

        # Make the window and set its size
        self.area = DockArea()
        self.setCentralWidget(self.area)
        self.resize(1240,600)
        self.move(100,50)

        # Make the colours that are used in the interface
        # The dark ones are to draw lines instead of boxes
        self.ColourSelected = QtGui.QColor(self.config['ColourSelected'][0], self.config['ColourSelected'][1], self.config['ColourSelected'][2], self.config['ColourSelected'][3])
        self.ColourNamed = QtGui.QColor(self.config['ColourNamed'][0], self.config['ColourNamed'][1], self.config['ColourNamed'][2], self.config['ColourNamed'][3])
        self.ColourNone = QtGui.QColor(self.config['ColourNone'][0], self.config['ColourNone'][1], self.config['ColourNone'][2], self.config['ColourNone'][3])
        self.ColourPossible = QtGui.QColor(self.config['ColourPossible'][0], self.config['ColourPossible'][1], self.config['ColourPossible'][2], self.config['ColourPossible'][3])

        self.ColourSelectedDark = QtGui.QColor(self.config['ColourSelected'][0], self.config['ColourSelected'][1], self.config['ColourSelected'][2], 255)
        self.ColourNamedDark = QtGui.QColor(self.config['ColourNamed'][0], self.config['ColourNamed'][1], self.config['ColourNamed'][2], 255)
        self.ColourNoneDark = QtGui.QColor(self.config['ColourNone'][0], self.config['ColourNone'][1], self.config['ColourNone'][2], 255)
        self.ColourPossibleDark = QtGui.QColor(self.config['ColourPossible'][0], self.config['ColourPossible'][1], self.config['ColourPossible'][2], 255)

        # Make the docks and lay them out
        self.d_overview = Dock("Overview",size=(1200,150))
        self.d_ampl = Dock("Amplitude",size=(1200,150))
        self.d_spec = Dock("Spectrogram",size=(1200,300))
        self.d_controls = Dock("Controls",size=(40,90))
        self.d_files = Dock("Files",size=(40,200))
        self.d_plot = Dock("Plots",size=(1200,150))
        self.d_controls.setSizePolicy(QSizePolicy.Policy.Minimum,QSizePolicy.Policy.Minimum)
        #self.d_controls.setSizePolicy(1,1)

        self.area.addDock(self.d_files,'left')
        self.area.addDock(self.d_overview,'right',self.d_files)
        self.area.addDock(self.d_ampl,'bottom',self.d_overview)
        self.area.addDock(self.d_spec,'bottom',self.d_ampl)
        self.area.addDock(self.d_controls,'bottom',self.d_files)
        self.area.addDock(self.d_plot,'bottom',self.d_spec)

        # Store the state of the docks in case the user wants to reset it
        self.state = self.area.saveState()
        containers, docks = self.area.findAll()
        self.state_cont = [cont.sizes() for cont in containers]

        # Put content widgets in the docks:
        # OVERVIEW dock
        self.w_overview = pg.LayoutWidget()
        self.w_overview.layout.setColumnStretch(1, 10)
        self.w_overview.layout.setColumnStretch(0, 0)
        self.w_overview.layout.setColumnStretch(2, 0)
        self.d_overview.addWidget(self.w_overview)
        # this will hold both overview image and segment boxes
        self.w_overview1 = pg.GraphicsLayoutWidget()
        self.w_overview1.ci.layout.setContentsMargins(0.5, 1, 0.5, 1)
        self.w_overview1.ci.layout.setRowSpacing(0, 0)
        self.w_overview1.ci.layout.setRowStretchFactor(0, 7)
        self.w_overview1.ci.layout.setRowStretchFactor(1, 1)

        fileInfo = QHBoxLayout()
        self.fileInfoSR = QLabel()
        self.fileInfoSR.setStyleSheet("QLabel {color: #505050}")
        self.fileInfoNCh = QLabel()
        self.fileInfoNCh.setStyleSheet("QLabel {color: #505050}")
        self.fileInfoSS = QLabel()
        self.fileInfoSS.setStyleSheet("QLabel {color: #505050}")
        self.fileInfoDur = QLabel()
        self.fileInfoDur.setStyleSheet("QLabel {color: #505050}")
        fileInfo.addWidget(self.fileInfoSR)
        fileInfo.addSpacing(20)
        fileInfo.addWidget(self.fileInfoNCh)
        fileInfo.addSpacing(20)
        fileInfo.addWidget(self.fileInfoSS)
        fileInfo.addSpacing(20)
        fileInfo.addWidget(self.fileInfoDur)
        fileInfo.addStretch(5)

        self.p_overview = SupportClasses_GUI.DemousedViewBox()
        self.w_overview1.addItem(self.p_overview,row=0,col=0)
        self.p_overview2 = SupportClasses_GUI.ChildInfoViewBox(enableMouse=False, enableMenu=False)
        self.w_overview1.addItem(self.p_overview2,row=1,col=0)
        self.p_overview2.setXLink(self.p_overview)
        self.p_overview2.setPreferredHeight(25)
        self.p_overview2.setCursor(Qt.CursorShape.PointingHandCursor)

        # The buttons to move through the overview
        self.leftBtn = QPushButton()
        self.leftBtn.setIcon(QIcon("img/overview-back.png"))
        self.leftBtn.setIconSize(QtCore.QSize(7, 28))
        self.leftBtn.setMinimumWidth(16)
        self.leftBtn.clicked.connect(self.moveLeft)
        self.leftBtn.setToolTip("Move view back")
        self.rightBtn = QPushButton()
        self.rightBtn.setIcon(QIcon("img/overview-next.png"))
        self.rightBtn.setIconSize(QtCore.QSize(7, 28))
        self.rightBtn.setMinimumWidth(16)
        self.rightBtn.clicked.connect(self.moveRight)
        self.rightBtn.setToolTip("Move view forward")
        self.leftBtn.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.MinimumExpanding)
        self.rightBtn.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.MinimumExpanding)

        # Buttons to move to next/previous five minutes
        self.prev5mins=QToolButton()
        self.prev5mins.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_MediaSeekBackward))
        self.prev5mins.setMinimumSize(35, 30)
        self.prev5mins.setToolTip("Previous page")
        self.prev5mins.clicked.connect(self.movePrev5mins)
        self.next5mins=QToolButton()
        self.next5mins.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_MediaSeekForward))
        self.next5mins.setMinimumSize(35, 30)
        self.next5mins.setToolTip("Next page")
        self.next5mins.clicked.connect(self.moveNext5mins)
        self.placeInFileLabel2 = QLabel('Page')
        self.placeInFileLabel = QLabel('')
        self.placeInFileLabel.setSizePolicy(QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Minimum)
        self.placeInFileSelector = QSpinBox()
        self.placeInFileSelector.setRange(1,1)
        self.placeInFileSelector.setButtonSymbols(QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.placeInFileSelector.editingFinished.connect(self.moveTo5mins)
        self.placeInFileSelector.setMinimumHeight(25)

        # "Find next annotation" buttons
        self.annotJumpLabel = QLabel("Jump to next mark:")
        self.annotJumpG = QToolButton()
        self.annotJumpG.setIcon(QIcon('img/findnext-g.png'))
        self.annotJumpG.setToolTip("Any label")
        self.annotJumpG.setMinimumSize(35,30)
        self.annotJumpG.setIconSize(QtCore.QSize(20, 17))
        self.annotJumpG.clicked.connect(lambda: self.annotJumper(100))
        self.annotJumpY = QToolButton()
        self.annotJumpY.setIcon(QIcon('img/findnext-y.png'))
        self.annotJumpY.setToolTip("Uncertain label")
        self.annotJumpY.setMinimumSize(35,30)
        self.annotJumpY.setIconSize(QtCore.QSize(20, 17))
        self.annotJumpY.clicked.connect(lambda: self.annotJumper(99))

        # position everything in the dock
        self.w_overview.layout.addLayout(fileInfo, 0, 0, 1, 3)
        self.w_overview.addWidget(self.w_overview1, row=2, col=1)
        self.w_overview.addWidget(self.leftBtn,row=2,col=0)
        self.w_overview.addWidget(self.rightBtn,row=2,col=2)
        placeInFileBox = QHBoxLayout()
        placeInFileBox.addStretch(10)
        placeInFileBox.addWidget(self.placeInFileLabel2)
        placeInFileBox.addWidget(self.prev5mins)
        placeInFileBox.addWidget(self.placeInFileSelector)
        placeInFileBox.addWidget(self.next5mins)
        placeInFileBox.addWidget(self.placeInFileLabel)
        placeInFileBox.addStretch(4)
        placeInFileBox.addWidget(self.annotJumpLabel)
        placeInFileBox.addWidget(self.annotJumpG)
        placeInFileBox.addWidget(self.annotJumpY)
        placeInFileBox.addStretch(4)
        self.w_overview.layout.addLayout(placeInFileBox, 3, 1)

        # Corresponding keyboard shortcuts:
        self.moveLeftKey = QShortcut(QKeySequence(Qt.Key.Key_Left), self)
        self.moveLeftKey.activated.connect(self.moveLeft)
        self.moveRightKey = QShortcut(QKeySequence(Qt.Key.Key_Right), self)
        self.moveRightKey.activated.connect(self.moveRight)
        self.movePrev5minsKey = QShortcut(QKeySequence("Shift+Left"), self)
        self.movePrev5minsKey.activated.connect(self.movePrev5mins)
        self.moveNext5minsKey = QShortcut(QKeySequence("Shift+Right"), self)
        self.moveNext5minsKey.activated.connect(self.moveNext5mins)


        # AMPLITUDE dock
        self.w_ampl = pg.GraphicsLayoutWidget()
        self.p_ampl = SupportClasses_GUI.DragViewBox(self, enableMouse=False,enableMenu=False,enableDrag=False, thisIsAmpl=True)
        self.p_ampl.setAutoVisible(False, True)
        self.w_ampl.addItem(self.p_ampl,row=0,col=1)
        self.d_ampl.addWidget(self.w_ampl)

        self.w_spec = pg.GraphicsLayoutWidget()
        self.p_spec = SupportClasses_GUI.DragViewBox(self, enableMouse=False,enableMenu=False,enableDrag=self.config['specMouseAction']==3, thisIsAmpl=False)
        self.w_spec.addItem(self.p_spec,row=0,col=1)
        self.d_spec.addWidget(self.w_spec)

        self.w_plot = pg.GraphicsLayoutWidget()
        self.p_plot = self.w_plot.addViewBox(enableMouse=False,enableMenu=False)
        self.w_plot.addItem(self.p_plot,row=0,col=1)
        self.d_plot.addWidget(self.w_plot)

        # The axes
        # Time axis has to go separately in loadFile
        self.ampaxis = pg.AxisItem(orientation='left')
        self.w_ampl.addItem(self.ampaxis,row=0,col=0)
        self.ampaxis.linkToView(self.p_ampl)
        self.ampaxis.setWidth(w=65)
        self.ampaxis.setLabel('')

        self.specaxis = pg.AxisItem(orientation='left')
        if not self.zooniverse:
            self.w_spec.addItem(self.specaxis,row=0,col=0)
        self.specaxis.linkToView(self.p_spec)
        self.specaxis.setWidth(w=65)

        # Plot window also needs an axis to make them line up
        self.plotaxis = pg.AxisItem(orientation='left')
        self.w_plot.addItem(self.plotaxis,row=0,col=0)
        self.plotaxis.linkToView(self.p_plot)
        self.plotaxis.setWidth(w=65)
        self.plotaxis.setLabel('')

        # Hide diagnostic plot window until requested
        self.d_plot.hide()

        # The bar that shows playback position
        self.bar = pg.InfiniteLine(angle=90, movable=True, pen={'color': 'c', 'width': 3})
        self.bar.btn = self.MouseDrawingButton
        self.bar.sigPositionChangeFinished.connect(self.barMoved)

        # Guides that can be used in batmode
        self.guidelines = [0]*len(self.config['guidecol'])
        for gi in range(len(self.config['guidecol'])):
            self.guidelines[gi] = pg.InfiniteLine(angle=0, movable=False, pen={'color': self.config['guidecol'][gi], 'width': 2})

        # The print out at the bottom of the spectrogram with data in
        # Note: widgets cannot be directly added to GraphicsLayout, so need to convert them to proxy GraphicsWidgets using the proxy
        self.pointData = QLabel()
        self.pointData.setStyleSheet("QLabel { background-color : white; color : #CC0000; }")
        self.pointDataProxy = QGraphicsProxyWidget()
        self.pointDataProxy.setWidget(self.pointData)
        self.segInfo = QLabel()
        self.segInfo.setStyleSheet("QLabel { background-color : white; color : #CC0000; }")
        self.segInfoProxy = QGraphicsProxyWidget()
        self.segInfoProxy.setWidget(self.segInfo)

        # The various plots
        self.overviewImage = pg.ImageItem(enableMouse=False)
        self.p_overview.addItem(self.overviewImage)
        self.overviewImageRegion = SupportClasses_GUI.LinearRegionItemO(pen=pg.mkPen(120,80,200, width=2),
                hoverPen=pg.mkPen(60, 40, 230, width=3.5))
        # This is needed for compatibility with other shaded rectangles:
        self.overviewImageRegion.lines[0].btn = Qt.MouseButton.RightButton
        self.overviewImageRegion.lines[1].btn = Qt.MouseButton.RightButton
        self.p_overview.addItem(self.overviewImageRegion, ignoreBounds=True)
        self.amplPlot = pg.PlotDataItem()
        self.p_ampl.addItem(self.amplPlot)
        self.specPlot = pg.ImageItem()

        # TODO: Useful?
        self.blurEffect = QGraphicsBlurEffect(blurRadius=1.1)
        self.specPlot.setGraphicsEffect(self.blurEffect)

        self.p_spec.addItem(self.specPlot)
        if self.MouseDrawingButton==Qt.MouseButton.RightButton:
            self.p_ampl.unsetCursor()
            self.specPlot.unsetCursor()
            self.bar.setCursor(Qt.CursorShape.OpenHandCursor)
        else:
            self.p_ampl.setCursor(QtGui.QCursor(QPixmap('img/cursor.bmp'), 0, 0))
            self.specPlot.setCursor(QtGui.QCursor(QPixmap('img/cursor.bmp'), 0, 0))
            self.bar.unsetCursor()

        # Connect up the listeners
        self.p_ampl.scene().sigMouseClicked.connect(self.mouseClicked_ampl)
        self.p_spec.scene().sigMouseClicked.connect(self.mouseClicked_spec)

        # Connect up so can disconnect if not selected...
        self.p_spec.scene().sigMouseMoved.connect(self.mouseMoved)
        self.w_spec.addItem(self.segInfoProxy, row=2, col=1)
        self.w_spec.addItem(self.pointDataProxy, row=3, col=1)

        # The content of the other two docks
        self.w_controls = pg.LayoutWidget()
        self.d_controls.addWidget(self.w_controls)
        self.w_files = pg.LayoutWidget()
        self.d_files.addWidget(self.w_files)

        # Button to move to the next file in the list
        self.nextFileBtn=QToolButton()
        self.nextFileBtn.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_MediaSkipForward))
        self.nextFileBtn.clicked.connect(self.openNextFile)
        self.nextFileBtn.setToolTip("Open next file")
        self.w_files.addWidget(self.nextFileBtn,row=0,col=1)

        # The buttons inside the controls dock
        self.playButton = QToolButton()
        self.playButton.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_MediaPlay))
        self.playButton.setIconSize(QtCore.QSize(20, 20))
        self.playButton.setToolTip("Play visible")
        self.playButton.clicked.connect(self.playVisible)
        self.playKey = QShortcut(QKeySequence("Space"), self)
        self.playKey.activated.connect(self.playVisible)

        self.stopButton = QToolButton()
        self.stopButton.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_MediaStop))
        self.stopButton.setIconSize(QtCore.QSize(20, 20))
        self.stopButton.setToolTip("Stop playback")
        self.stopButton.clicked.connect(self.stopPlayback)

        self.playSegButton = QToolButton()
        self.playSegButton.setIcon(QIcon('img/playsegment.png'))
        self.playSegButton.setIconSize(QtCore.QSize(20, 20))
        self.playSegButton.setToolTip("Play selected")
        self.playSegButton.clicked.connect(self.playSelectedSegment)

        self.speedButton = QToolButton()
        self.speedButton.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)
        self.speedButton.setIcon(QIcon('img/playSlow-w.png'))
        self.speedButton.setIconSize(QtCore.QSize(20, 20))
        self.speedButton.setToolTip("Playback speed")

        # The playback speed options
        speedMenu = QMenu()
        extraGroup = QActionGroup(self)
        for ename in ["4","3","2","1",u'\u00BE',u'\u00BD',u'\u00BC']:
            em = speedMenu.addAction(ename)
            em.setCheckable(True)
            if ename == "1":
                em.setChecked(True)
            receiver = lambda checked, ename=ename: self.setSpeed(ename)
            em.triggered.connect(receiver)
            extraGroup.addAction(em)
        self.speedButton.setMenu(speedMenu)

        self.quickDenoiseButton = QToolButton()
        self.quickDenoiseButton.setIcon(QIcon('img/denoisesegment.png'))
        self.quickDenoiseButton.setIconSize(QtCore.QSize(20, 20))
        self.quickDenoiseButton.setToolTip("Denoise segment")
        self.quickDenoiseButton.clicked.connect(self.denoiseSeg)
        self.playBandLimitedSegButton = QToolButton()
        self.playBandLimitedSegButton.setIcon(QtGui.QIcon('img/playBandLimited.png'))
        self.playBandLimitedSegButton.setIconSize(QtCore.QSize(20, 20))
        self.playBandLimitedSegButton.setToolTip("Play selected-band limited")
        self.playBandLimitedSegButton.clicked.connect(self.playBandLimitedSegment)

        self.floorSlider = QSlider(Qt.Orientation.Horizontal)
        self.floorSlider.setMinimum(0)
        self.floorSlider.setMaximum(100)
        self.floorSlider.valueChanged.connect(self.floorSliderMoved)

        # Volume, brightness and contrast sliders.
        # Need to pass true (config) values to set up correct initial positions
        self.specControls = SupportClasses_GUI.BrightContrVol(self.config['brightness'], self.config['contrast'], self.config['invertColourMap'], horizontal=False)
        self.specControls.colChanged.connect(self.setColourLevels)
        self.specControls.volChanged.connect(self.volSliderMoved)

        # Confirm button - auto ups the certainty to 100
        self.confirmButton = QPushButton("   Confirm labels")
        self.confirmButton.clicked.connect(self.confirmSegment)
        self.confirmButton.setIcon(QIcon(QPixmap('img/check-mark2.png')))
        self.confirmButton.setStyleSheet("QPushButton {padding: 3px 3px 3px 3px}")
        self.confirmButton.setToolTip("Set all labels in this segment as certain")

        # Delete segment button. We have to get rid of the extra event args
        self.deleteButton = QPushButton("  Delete segment")
        self.deleteButton.clicked.connect(lambda _ : self.deleteSegment())
        self.deleteButton.setIcon(QIcon(QPixmap('img/deleteL.png')))
        self.deleteButton.setStyleSheet("QPushButton {padding: 3px 3px 3px 3px}")

        # export selected sound
        self.exportSoundBtn = QPushButton("  Save sound clip")
        self.exportSoundBtn.clicked.connect(lambda _ : self.saveSelectedSound(False))
        self.exportSoundBtn.setIcon(QIcon(QPixmap('img/storage2.png')))
        self.exportSoundBtn.setToolTip("Export the selected segment to a file")

        # export selected sound
        if not self.DOC:
            self.exportSlowSoundBtn = QPushButton("  Save slow sound clip")
            self.exportSlowSoundBtn.clicked.connect(lambda _ : self.saveSelectedSound(True))
            self.exportSlowSoundBtn.setIcon(QIcon(QPixmap('img/storage2.png')))
            self.exportSlowSoundBtn.setToolTip("Export the selected sound to a file at different speed")

        # flips buttons to Disabled state
        self.refreshSegmentControls()

        # The spinbox for changing the width shown in the controls dock
        windowLabel = QLabel('Visible window (seconds)')
        windowLabel.setAlignment(Qt.AlignmentFlag.AlignBottom)
        self.widthWindow = QDoubleSpinBox()
        self.widthWindow.setSingleStep(1.0)
        self.widthWindow.setDecimals(2)
        self.widthWindow.setValue(self.config['windowWidth'])
        self.widthWindow.valueChanged[float].connect(self.changeWidth)

        # Place all these widgets in the Controls dock
        self.w_controls.addWidget(self.playButton,row=0,col=0)
        self.w_controls.addWidget(self.playSegButton,row=0,col=1)
        self.w_controls.addWidget(self.playBandLimitedSegButton,row=0,col=2)
        self.w_controls.addWidget(self.speedButton,row=0,col=3)
        self.w_controls.addWidget(self.stopButton,row=1,col=0)
        self.w_controls.addWidget(self.quickDenoiseButton,row=1,col=2)

        self.w_controls.addWidget(QLabel('Noise floor'),row=2,col=0)
        self.w_controls.addWidget(self.floorSlider,row=2,col=1,colspan=3)
        self.w_controls.addWidget(self.specControls, row=3, col=0, rowspan=2, colspan=4)

        self.w_controls.addWidget(QLabel('Visible window'),row=8,col=0,colspan=4)
        self.w_controls.addWidget(self.widthWindow,row=9,col=0,colspan=2)
        self.w_controls.addWidget(QLabel('seconds'), row=9, col=2, colspan=2)
        # Spacer because pyqtgraph can't add spacer items
        spacer = QWidget()
        self.w_controls.addWidget(spacer, row=10, col=0, colspan=4)
        self.w_controls.layout.setRowMinimumHeight(10, 3)

        # Empty widget to add in the gridlayout
        segContrs = QGroupBox("Selected segment")
        segContrs.setStyleSheet("QGroupBox:title{color: #505050; font-weight: 50}")
        segContrsBox = QVBoxLayout()
        segContrs.setLayout(segContrsBox)
        segContrsBox.addWidget(self.confirmButton)
        segContrsBox.addWidget(self.deleteButton)
        segContrsBox.addWidget(self.exportSoundBtn)
        if not self.DOC:
            segContrsBox.addWidget(self.exportSlowSoundBtn)
        self.w_controls.addWidget(segContrs, row=12, col=0, colspan=4)

        # A slider to move through the file easily
        self.scrollSlider = QScrollBar(Qt.Orientation.Horizontal)
        self.scrollSlider.valueChanged.connect(self.scroll)
        self.d_spec.addWidget(self.scrollSlider)

        # List to hold the list of files
        self.listFiles = SupportClasses_GUI.LightedFileList(self.ColourNone, self.ColourPossibleDark, self.ColourNamed)
        self.listFiles.itemDoubleClicked.connect(self.listLoadFile)

        self.w_files.addWidget(QLabel('Double click to open'),row=0,col=0)
        self.w_files.addWidget(QLabel('Icon marks annotation certainty'),row=1,col=0)
        self.w_files.addWidget(self.listFiles,row=2,colspan=2)

        # The context menu (drops down on mouse click) to select birds
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.menuBirdList = QMenu()
        self.menuBirdOther = QMenu('Other',self.menuBirdList)
        self.menuBirdAll = QMenu('All',self.menuBirdOther)
        # New line to allow multiple selections
        self.menuBirdList.installEventFilter(self)
        self.menuBirdOther.installEventFilter(self)
        self.menuBirdAll.installEventFilter(self)
        self.fillBirdList()
        # Hack to get the type of an ROI
        p_spec_r = SupportClasses_GUI.ShadedRectROI(0, 0)
        self.ROItype = type(p_spec_r)

        # Listener for key presses
        self.w_ampl.installEventFilter(self)
        self.w_spec.installEventFilter(self)

        # Statusbar
        self.statusLeft = QLabel("Left")
        self.statusLeft.setFrameStyle(QFrame.Shape.Panel | QFrame.Shadow.Sunken)
        self.statusBM = QLabel("")
        self.statusBM.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.statusBM.setFrameStyle(QFrame.Shape.Panel | QFrame.Shadow.Sunken)
        self.statusRO = QLabel("")
        self.statusRO.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.statusRO.setFrameStyle(QFrame.Shape.Panel | QFrame.Shadow.Sunken)
        self.statusRight = QLabel("")
        self.statusRight.setAlignment(Qt.AlignmentFlag.AlignRight)
        self.statusRight.setFrameStyle(QFrame.Shape.Panel | QFrame.Shadow.Sunken)
        # Style
        # statusStyle='QLabel {border:transparent}'
        # self.statusLeft.setStyleSheet(statusStyle)
        # self.statusRO.setStyleSheet(statusStyle)
        # self.statusRight.setStyleSheet(statusStyle)
        self.statusBar().addPermanentWidget(self.statusLeft,3)
        self.statusBar().addPermanentWidget(self.statusBM,1)
        self.statusBar().addPermanentWidget(self.statusRO,1)
        self.statusBar().addPermanentWidget(self.statusRight,2)

        # Set the message in the status bar
        self.statusLeft.setText("Ready")
        self.statusRO.setText("Read-only mode" if self.config['readOnly'] else "")

        # Function calls to check if should show various parts of the interface, whether dragging boxes or not
        self.makeReadOnly()
        self.useAmplitudeCheck()
        self.useFilesCheck()
        self.showOverviewSegsCheck()
        self.dragRectsTransparent()
        self.showPointerDetailsCheck()

        # Set the focus ready for action
        self.w_spec.setFocus()

    def toggleBatMode(self):
        """ Enables/disables GUI elements when bat mode is entered/left.
            Called on every load.
        """
        if self.batmode:
            print("Bat mode: on")
        else:
            print("Bat mode: off")

        if self.batmode:
            self.useAmplitudeTick.setChecked(False)
            # otherwise leave as it was
        self.useAmplitudeTick.setEnabled(not self.batmode)
        self.useAmplitudeCheck()

        if not self.DOC:
            self.showDiagnosticTick.setEnabled(not self.batmode)
            self.extraMenu.setEnabled(not self.batmode)
            self.setExtraPlot("none")
            self.showFormant.setEnabled(not self.batmode)

        #self.showInvSpec.setVisible(self.batmode)
        self.showFundamental.setEnabled(not self.batmode)
        #self.showSpectral.setEnabled(not self.batmode)
        self.showEnergies.setEnabled(not self.batmode)

        self.addRegularAction.setEnabled(not self.batmode)
        self.denoiseAction.setEnabled(not self.batmode)
        self.segmentAction.setEnabled(not self.batmode)

        self.playButton.setEnabled(not self.batmode)
        self.stopButton.setEnabled(not self.batmode)
        self.specControls.volSlider.setEnabled(not self.batmode)
        self.specControls.volIcon.setEnabled(not self.batmode)

        text = "Bat mode" if self.batmode else ""
        self.statusBM.setText(text)

        # Also need to enter read-only mode:
        # no editing of segments, can only change labels on the premade one
        # But it is a costly operation, so check if needed:
        if self.batmode and not self.readonly.isChecked():
            self.readonly.setChecked(True)
            self.makeReadOnly()
        elif self.batmode and self.readonly.isChecked():
            pass
        elif self.readonly.isChecked() and not self.batmode:
            self.readonly.setChecked(False)
            self.makeReadOnly()
        else:  # not checked, not batmode
            pass
        self.readonly.setEnabled(not self.batmode)

    def refreshSegmentControls(self):
        """ Toggles all the segment controls on/off when a segment
            is (de)selected. Call this after changing self.box1id.
            Remember to update this when segment controls change!
        """
        # Basic buttons that toggle on any segment selection
        if self.DOC:
            btns = [self.deleteButton, self.playSegButton, self.quickDenoiseButton, self.exportSoundBtn]
        else:
            btns = [self.deleteButton, self.playSegButton, self.quickDenoiseButton, self.exportSoundBtn, self.exportSlowSoundBtn]

        # Buttons that should be allowed in batmode
        batbtns = [self.deleteButton]

        # if self.box1id is not -1, flip on, otherwise off
        if self.box1id<0:
            for btn in btns:
                btn.setEnabled(False)
            self.playBandLimitedSegButton.setEnabled(False)
            self.confirmButton.setEnabled(False)
        else:
            if self.batmode:
                for btn in batbtns:
                    btn.setEnabled(True)
            else:
                for btn in btns:
                    btn.setEnabled(True)

                # Special case for BandLimitedButton because it requires set freq bands
                if type(self.listRectanglesa2[self.box1id]) == self.ROItype:
                    # it's a rectangle box:
                    self.playBandLimitedSegButton.setEnabled(True)
                else:
                    # it's a 0 to inf segment:
                    self.playBandLimitedSegButton.setEnabled(False)

            # Special case for Confirm button because it requires yellow segment
            self.confirmButton.setEnabled(False)
            for sp in self.segments[self.box1id][4]:
                if sp["certainty"]<100 and sp["species"]!="Don't Know":
                    self.confirmButton.setEnabled(True)
                    break
    
    def parse_short_list_item(self,item,unsure=False):
        # Determine certainty
        # Add ? marks if Ctrl menu is called
        
        searchForCall = re.search(r' \[(.*?)\]$',item)
        call = None if searchForCall is None else searchForCall.group(1)
        beforeCall = item if call is None else item.split(" ["+call+"]")[0]

        searchForSpecies = re.search(r' \((.*?)\)$',beforeCall)
        if searchForSpecies is not None:
            species = searchForSpecies.group(1)
            genus = beforeCall.split(" ("+species+")")[0]
        else:
            # try > format
            beforeCallSplit = beforeCall.split('>')
            if len(beforeCallSplit)==1:
                species = None
                genus = beforeCall
            else:
                species = beforeCallSplit[1]
                genus = beforeCallSplit[0]
        
        if species is None:
            mergedName = genus
        else:
            mergedName = genus + " (" + species + ")"

        if unsure and item != "Don't Know":
            cert = 50
            mergedName = mergedName+'?'
        elif item == "Don't Know":
            cert = 0
        else:
            cert = 100

        return mergedName, cert, call
    
    def makeBatLists(self,unsure=False):
        # There aren't many options, but anyway...
        if self.config['ReorderList'] and hasattr(self,'segments') and self.box1id>-1:
            for key in self.segments[self.box1id].keys:
                # Either move the label to the top of the list, or delete the last
                if key[0] in self.batList:
                    self.batList.remove(key[0])
                else:
                    del self.batList[-1]
                self.batList.insert(0,key[0])

        # Create menu items and mark them
        # (we assume that bat list is always short enough to fit in one column)
        for item in self.batList:
            species, cert, call = self.parse_short_list_item(item,unsure)

            bat = self.menuBirdList.addAction(species)
            bat.setCheckable(True)
            if hasattr(self,'segments') and self.segments[self.box1id].hasLabel(item, cert):
                bat.setChecked(True)
            bat.triggered.connect(partial(self.batSelected, species))
            self.menuBirdList.addAction(bat)
    
    def makeShortBirdLists(self,unsure=False):        
        # reorder
        if self.config['ReorderList'] and hasattr(self,'segments') and self.box1id>-1:
            for species,certainty,calltype in self.segments[self.box1id].getKeysWithCalltypes():
                label = species if calltype is None else species+" ["+calltype+"]"
                # Either move the label to the top of the list, or delete the last
                if label in self.shortBirdList:
                    self.shortBirdList.remove(label)
                else:
                    del self.shortBirdList[-1]
                self.shortBirdList.insert(0,label)
        
        self.shortBirdList = [x for x in self.shortBirdList if x != ""] + [x for x in self.shortBirdList if x == ""] # just put the blanks at the end

        # Create menu items
        for item in self.shortBirdList[:15]:
            if not item=="":
                species, cert, calltype = self.parse_short_list_item(item,unsure)
                label = species if calltype is None else species+" ["+calltype+"]"
                bird = self.menuBirdList.addAction(label)
                bird.setCheckable(True)
                if calltype is None:
                    if hasattr(self,'segments') and self.segments[self.box1id].hasLabel(species, cert):
                        if self.segments[self.box1id].getCalltype(species, cert) is None: # we only want to mark the species in the menu if doesn't have a calltype.
                            bird.setChecked(True)
                        bird.triggered.connect(partial(self.birdAndCallSelected, species, "Any"))
                else:
                    if hasattr(self,'segments') and self.segments[self.box1id].getCalltype(species, cert)==calltype:
                        bird.setChecked(True)
                    bird.triggered.connect(partial(self.birdAndCallSelected, species, calltype))
        
        for item in self.shortBirdList[15:]:
            if not item=="":
                species, cert, calltype = self.parse_short_list_item(item,unsure)
                label = species if calltype is None else species+" ["+calltype+"]"
                bird = self.menuBirdOther.addAction(label)
                bird.setCheckable(True)
                if calltype is None:
                    if hasattr(self,'segments') and self.segments[self.box1id].hasLabel(species, cert):
                        if self.segments[self.box1id].getCalltype(species, cert) is None: # we don't want to mark the species in the menu if it has a calltype.
                            bird.setChecked(True)
                        bird.triggered.connect(partial(self.birdAndCallSelected, species, "Any"))
                else:
                    if hasattr(self,'segments') and self.segments[self.box1id].getCalltype(species, cert)==calltype:
                        bird.setChecked(True)
                    bird.triggered.connect(partial(self.birdAndCallSelected, species, calltype))

    def makeFullBirdListByLetter(self,unsure=False):
        allBirdTree = {}
        if self.longBirdList is not None:
            for longBirdEntry in self.longBirdList:
                # Add ? marks if Ctrl menu is called
                if unsure and longBirdEntry != "Don't Know" and longBirdEntry != "Other":
                    longBirdEntry = longBirdEntry+'?'

                if '>' in longBirdEntry:
                    speciesLevel1,speciesLevel2 = longBirdEntry.split('>')
                else:
                    speciesLevel1,speciesLevel2 = longBirdEntry, ""
                
                species = speciesLevel1 if speciesLevel2=="" else speciesLevel1 + " ("+speciesLevel2+")"
                calls = ["Any"]
                if species in self.knownCalls:
                    calls+=self.knownCalls[species].copy()
                calls.append("Other")

                firstLetter = speciesLevel1[0].upper()

                if not firstLetter in allBirdTree:
                    allBirdTree[firstLetter] = {}
                
                if not speciesLevel1 in allBirdTree[firstLetter]:
                    allBirdTree[firstLetter][speciesLevel1] = {}
                
                if speciesLevel2 == "":
                    allBirdTree[firstLetter][speciesLevel1] = {None: calls}
                else:
                    if not speciesLevel2 in allBirdTree[firstLetter][speciesLevel1]:
                        allBirdTree[firstLetter][speciesLevel1][speciesLevel2] = calls

        for letter in allBirdTree:
            letterMenu = QMenu(letter,self.menuBirdAll)
            for speciesLevel1 in allBirdTree[letter]:
                speciesLevel1Menu = QMenu(speciesLevel1,letterMenu)
                if None in allBirdTree[letter][speciesLevel1]: # no species, go straight to call
                    for call in allBirdTree[letter][speciesLevel1][None]:
                        callAction = speciesLevel1Menu.addAction(call)
                        callAction.triggered.connect(partial(self.birdAndCallSelected, speciesLevel1, call))
                else:
                    for speciesLevel2 in allBirdTree[letter][speciesLevel1]:
                        species = speciesLevel1 + " (" + speciesLevel2 + ")"
                        speciesLevel2Menu = QMenu(speciesLevel2,speciesLevel1Menu)
                        for call in allBirdTree[letter][speciesLevel1][speciesLevel2]:
                            callAction = speciesLevel2Menu.addAction(call)
                            callAction.triggered.connect(partial(self.birdAndCallSelected, species, call))
                        speciesLevel1Menu.addMenu(speciesLevel2Menu)
                letterMenu.addMenu(speciesLevel1Menu)
            self.menuBirdAll.addMenu(letterMenu)
        
    def fillBirdList(self,unsure=False):
        """ Sets the contents of the context menu.
        The first 20 items are in the first menu, the next in a second menu.
        Any extras go into the combobox at the end of the second list.
        This is called a lot because the order of birds in the list changes since the last choice is moved to the top of the list.
        When calltype-level display is on, fills the list with possible call types from a filter (if available)
        """
        print("running FILLBIRDLIST",self.batmode)

        self.menuBirdList.clear()
        self.menuBirdOther.clear()
        self.menuBirdAll.clear()

        if self.batmode:
            self.makeBatLists()
        else:
            self.makeShortBirdLists()
            self.menuBirdList.addMenu(self.menuBirdOther)
            self.makeFullBirdListByLetter()
            self.menuBirdOther.addMenu(self.menuBirdAll)

    def fillFileList(self,dir,fileName):
        """ Generates the list of files for the file listbox.
            dir: directory to use.
            fileName: currently opened file (marks it in the list).
        """
        if not os.path.isdir(dir):
            print("ERROR: directory %s doesn't exist" % dir)
            return

        self.listFiles.fill(dir, fileName)

    def resetStorageArrays(self):
        """ Called when new files are loaded.
        Resets the variables that hold the data to be saved and/or plotted.
        """

        # Remove the segments
        self.removeSegments()

        # Check if media is playing and stop it if so
        if hasattr(self,'media_obj'):
            self.stopPlayback()
            #print(self.media_obj.isPlayingorPaused(),self.media_obj.state())
            #if self.media_obj.isPlayingorPaused():
                #print("stopping")
                #self.stopPlayback()
                #self.media_thread.quit()
                #self.media_thread.wait()
            #del self.media_obj

        # This is a flag to say if the next thing that the user clicks on should be a start or a stop for segmentation
        if self.started:
            # This is the second click, so should pay attention and close the segment
            # Stop the mouse motion connection, remove the drawing boxes
            if self.startedInAmpl:
                try:
                    self.p_ampl.scene().sigMouseMoved.disconnect()
                except Exception:
                    pass
                self.p_ampl.removeItem(self.vLine_a)
            else:
                try:
                    self.p_spec.scene().sigMouseMoved.disconnect()
                except Exception:
                    pass
                # Add the other mouse move listener back
                if self.showPointerDetails.isChecked():
                    self.p_spec.scene().sigMouseMoved.connect(self.mouseMoved)
                self.p_spec.removeItem(self.vLine_s)
            self.p_ampl.removeItem(self.drawingBox_ampl)
            self.p_spec.removeItem(self.drawingBox_spec)
        self.started = False
        self.startedInAmpl = False
        self.segmentsToSave = False

        # Keep track of start points and selected buttons
        self.windowStart = 0
        self.prevBoxCol = self.config['ColourNone']
        self.bar.setValue(0)

        if self.multipleBirds:
            self.multipleBirds = self.config['MultipleSpecies']

        # reset buttons which require segment selection
        self.refreshSegmentControls()

        # Remove any fundamental frequencies drawn
        self.showFundamental.setChecked(False)
        for r in self.segmentPlots:
            self.p_spec.removeItem(r)
        self.segmentPlots=[]

        # Remove any shape marks
        for sh in self.shapePlots:
            try:
                self.p_spec.removeItem(sh)
            except Exception:
                pass

        # Remove spectral derivatives
        try:
            self.showSpectral.setChecked(False)
            self.p_spec.removeItem(self.derivPlot)
        except Exception:
            pass

        # Remove formants
        if not self.DOC:
            self.showFormant.setChecked(False)
            try:
                self.p_spec.removeItem(self.formantPlot)
            except Exception:
                pass

        # remove max energies
        self.showEnergies.setChecked(False)
        try:
            self.p_spec.removeItem(self.energyPlot)
        except Exception:
            pass

        # Cheatsheet: remove the freq labels
        if self.zooniverse and hasattr(self,'label1'):
            self.p_spec.removeItem(self.label1)
            self.p_spec.removeItem(self.label2)
            self.p_spec.removeItem(self.label3)
            self.p_spec.removeItem(self.label4)
            self.p_spec.removeItem(self.label5)

    def openFile(self, fileName=None):
        """ This handles the menu items for opening a file.
        Pops up a file selection dialog if no fileName provided.
        Splits the directory name and filename out, and then passes the filename to the loader."""

        if fileName is None:
            # File -> Open or splash screen:
            fileName, drop = QFileDialog.getOpenFileName(self, 'Choose File', self.SoundFileDir, "WAV or BMP files (*.wav *.bmp);; Only WAV files (*.wav);; Only BMP files (*.bmp);; FLAC files (*.flac)")
        # (it is provided when this is called by File -> [recent file clicked])
        success = 1
        SoundFileDirOld = self.SoundFileDir
        if self.filename is not None:
            fileNameOld = os.path.basename(self.filename)
        if fileName != '':
            print("Opening file %s" % fileName)
            self.SoundFileDir = os.path.dirname(fileName)
            success = self.listLoadFile(os.path.basename(fileName))
        if success==1:
            print("Warning: could not load file, reloading current file")
            self.SoundFileDir = SoundFileDirOld
            if hasattr(self, 'fileNameOld'):
                self.filename = os.path.join(self.SoundFileDir, fileNameOld)
                self.listLoadFile(fileNameOld)

    def listLoadFile(self,current):
        """ Listener for when the user clicks on a filename (also called by openFile() )
        Does the safety checks for file existence etc.
        Prepares the program for a new file.
        Saves the segments of the current file, resets flags and calls loadFile().
        """

        # Need name of file
        if type(current) is QListWidgetItem:
            current = current.text()
            current = re.sub('\/.*', '', current)

        fullcurrent = os.path.join(self.SoundFileDir, current)
        if not os.path.isdir(fullcurrent):
            if not os.path.isfile(fullcurrent):
                print("File %s does not exist!" % fullcurrent)
                return(1)
            # avoid files with no data (Tier 1 has 0Kb .wavs)
            if os.stat(fullcurrent).st_size == 0:
                print("Cannot open file %s of size 0!" % fullcurrent)
                return(1)
            if os.stat(fullcurrent).st_size < 1000:
                print("File %s appears to have only header" % fullcurrent)
                return(1)
            if fullcurrent.lower().endswith('.wav'):
                with open(fullcurrent, 'br') as f:
                    if f.read(4) != b'RIFF':
                        print("WAV file %s not formatted correctly" % fullcurrent)
                        return(1)
                self.batmode = False
            elif fullcurrent.lower().endswith('.bmp'):
                with open(fullcurrent, 'br') as f:
                    if f.read(2) != b'BM':
                        print("BMP file %s not formatted correctly" % fullcurrent)
                        return(1)
                self.batmode = True
                # loadFile will determine mode and update GUI based on self.batmode
            elif fullcurrent.lower().endswith('.flac'):
                self.batmode = False
            else:
                print("Unrecognized format of file %s " % fullcurrent)
                return(1)

            # Setting this to True forces initial segment save
            # self.segmentsToSave = True

            # Calls the noise data checks, segment saving, recent file updaters
            if self.filename is not None:
                self.closeFile()

        # Update the file list to show the right location
        i=0
        lof = self.listFiles.listOfFiles
        # This is skipped on first load, when len=0
        if len(lof)>0:
            while i<len(lof)-1 and lof[i].fileName() != current:
                i+=1
            if lof[i].isDir() or (i == len(lof)-1 and lof[i].fileName() != current):
                dir = QDir(self.SoundFileDir)
                dir.cd(lof[i].fileName())
                self.SoundFileDir=str(dir.absolutePath())

        # Now repopulate the listbox
        self.fillFileList(self.SoundFileDir, current)

        # If a file was clicked, open it
        if not os.path.isdir(fullcurrent):
            self.loadFile(fullcurrent)

        # self.listFiles.setCurrentItem(current)

        return(0)

    def loadFile(self, name=None):
        """ This does the work of loading a file.
        We are using wavio to do the reading. We turn the data into a float, but do not normalise it (/2^(15)).
        For 2 channels, just take the first one.
        Normalisation can cause problems for some segmentations, e.g. Harma.

        This method also gets the spectrogram to plot it, loads the segments from a *.data file, and
        passes the new data to any of the other classes that need it.
        Then it sets up the audio player and fills in the appropriate time data in the window, and makes
        the scroll bar and overview the appropriate lengths.

        name: full path to the file to be loaded. If None, loads the next section of the current file
        """

        # Reset interface and switch to right mode. 
        # We assume that whoever starts this, already set self.batmode correctly
        self.resetStorageArrays()
        self.toggleBatMode()

        with pg.ProgressDialog("Loading..", 0, 6) as dlg:
            dlg.setCancelButton(None)
            dlg.setWindowIcon(QIcon('img/Avianz.ico'))
            dlg.setWindowTitle('AviaNZ')
            dlg.show()
            dlg.update()
            if name is not None:
                if not os.path.exists(name):
                    print("ERROR: tried to open non-existing file %s", name)
                    return
                self.filename = name

                # Create an instance of the Signal Processing class
                if not hasattr(self, 'sp'):
                    if self.cheatsheet:
                        self.sp = Spectrogram.Spectrogram(512,256, 0, 0)
                    else:
                        self.sp = Spectrogram.Spectrogram(self.config['window_width'], self.config['incr'], self.config['minFreq'], self.config['maxFreq'])

                self.currentFileSection = 0

                if hasattr(self, 'timeaxis') and not self.zooniverse:
                    self.w_spec.removeItem(self.timeaxis)

                # TODO: Could specify others -- cf Raven
                # Check if the filename is in standard DOC format
                # Which is xxxxxx_xxxxxx.wav or ccxx_cccc_xxxxxx_xxxxxx.wav (c=char, x=0-9), could have _ afterward
                # So this checks for the 6 ints _ 6 ints part anywhere in string
                DOCRecording = re.search('(\d{6})_(\d{6})',name[-17:-4])

                if DOCRecording:
                    self.startTime = DOCRecording.group(2)

                    #if int(self.startTime[:2]) > 8 or int(self.startTime[:2]) < 8:
                    if int(self.startTime[:2]) > 17 or int(self.startTime[:2]) < 7: # 6pm to 6am
                        print("Night time DOC recording")
                    else:
                        print("Day time DOC recording")
                        # TODO: And modify the order of the bird list
                    self.startTime = int(self.startTime[:2]) * 3600 + int(self.startTime[2:4]) * 60 + int(self.startTime[4:6])
                    self.timeaxis = SupportClasses_GUI.TimeAxisHour(orientation='bottom',linkView=self.p_ampl)
                else:
                    self.startTime = 0
                    self.timeaxis = SupportClasses_GUI.TimeAxisMin(orientation='bottom',linkView=self.p_ampl)

                if not self.zooniverse:
                    self.w_spec.addItem(self.timeaxis, row=1, col=1)

                # This next line is a hack to make the axis update
                #self.changeWidth(self.widthWindow.value())

            dlg += 1
            dlg.update()

            # Read in the file and make the spectrogram
            # Determine where to start and how much to read for this page (in seconds):
            self.startRead = max(0,self.currentFileSection*self.config['maxFileShow']-self.config['fileOverlap'])
            # avoid files with no data (Tier 1 has 0Kb .wavs)
            if os.stat(self.filename).st_size == 0:
                self.statusLeft.setText("File appears empty")
                return

            # main read-in:
            if self.batmode:
                self.sp.minFreqShow = self.config['minFreqBats']
                self.sp.maxFreqShow = self.config['maxFreqBats']
                successread = self.sp.readBmp(name)
                if successread>0:
                    print("ERROR: file not loaded")
                    return
                # this assumes that the entire file is always loaded in BMP mode
                self.datalength = self.sp.fileLength
            else:
                self.sp.minFreqShow = self.config['minFreq']
                self.sp.maxFreqShow = self.config['maxFreq']
                if self.startRead == 0:
                    lenRead = self.config['maxFileShow'] + self.config['fileOverlap']
                else:
                    lenRead = self.config['maxFileShow'] + 2*self.config['fileOverlap']

                 
                if self.filename.lower().endswith('.wav'):
                    self.sp.readWav(self.filename, lenRead, self.startRead)
                elif self.filename.lower().endswith('.flac'):
                    self.sp.readFlac(self.filename)
                    self.sp.readWav('/home/marslast/output.wav',lenRead,self.startRead)

                # resample to 16K if needed (Spectrogram will determine)
                if self.cheatsheet:
                    self.sp.resample(16000)
                    self.sp.maxFreqShow = 8000

                # Parse wav format details based on file header:
                self.datalength = np.shape(self.sp.data)[0]

                # self.sp.audioFormat will be set
                # self.sp.fileLength will be determined from wav header
                # self.sp.minFreq and maxFreq will be set based on sample rate
                # self.sp.*Show will be set based on Spectrogram settings

            dlg += 1
            dlg.update()

            self.datalengthSec = self.datalength / self.sp.audioFormat.sampleRate()
            #self.datalengthSec = self.datalength / self.sp.sampleRate
            #print("Length of file is ", self.datalengthSec, " seconds (", self.datalength, " samples) loaded from ", self.sp.fileLength / self.sp.sampleRate, "seconds (", self.sp.fileLength, " samples) with sample rate ",self.sp.sampleRate, " Hz.")

            if name is not None:  # i.e. starting a new file, not next section
                if self.datalength != self.sp.fileLength:
                    self.nFileSections = int(np.ceil(self.sp.fileLength/self.datalength))
                    self.prev5mins.setEnabled(False)
                    self.next5mins.setEnabled(True)
                    self.movePrev5minsKey.setEnabled(False)
                    self.moveNext5minsKey.setEnabled(True)
                else:
                    self.nFileSections = 1
                    self.prev5mins.setEnabled(False)
                    self.next5mins.setEnabled(False)
                    self.movePrev5minsKey.setEnabled(False)
                    self.moveNext5minsKey.setEnabled(False)
                #print('number of pages: ', self.nFileSections)

            # Update overview info
            if self.nFileSections == 1:
                self.placeInFileLabel.setText("(%d s in 1 page)" % self.datalengthSec)
                self.placeInFileSelector.setVisible(False)
                self.placeInFileLabel2.setVisible(False)
            else:
                self.placeInFileLabel2.setVisible(True)
                self.placeInFileSelector.setVisible(True)
                self.placeInFileSelector.setValue(self.currentFileSection+1)
                self.placeInFileSelector.setMaximum(self.nFileSections)
                self.placeInFileLabel.setText("of %d (%d s in page)" % (self.nFileSections, self.datalengthSec))
            self.fileInfoSR.setText("<b>Sampling rate:</b> %d Hz" % self.sp.audioFormat.sampleRate())
            #self.fileInfoSR.setText("<b>Sampling rate:</b> %d Hz" % self.sp.sampleRate)
            self.fileInfoNCh.setText("<b>Channels:</b> %d" % self.sp.audioFormat.channelCount())
            self.fileInfoSS.setText("<b>Sample format:</b> %s" % str(self.sp.audioFormat.sampleFormat()).split('.')[-1])
            self.fileInfoDur.setText("<b>Duration:</b> %d min %d s" % divmod(self.sp.fileLength // self.sp.audioFormat.sampleRate(), 60))
            #self.fileInfoDur.setText("<b>Duration:</b> %d min %d s" % divmod(self.sp.fileLength // self.sp.sampleRate, 60))

            if not self.batmode:
                # Create the main spectrogram
                _ = self.sp.spectrogram(window_width=self.config['window_width'], incr=self.config['incr'],window=self.config['windowType'],sgType=self.config['sgType'],sgScale=self.config['sgScale'],nfilters=self.config['nfilters'],mean_normalise=self.config['sgMeanNormalise'],equal_loudness=self.config['sgEqualLoudness'],onesided=self.config['sgOneSided'])
                # For batmode, the spectrogram is already created.
            # Normalize the spectrogram, appropriately for the current mode and user settings
            self.setSpectrogram()

            # ANNOTATIONS: init empty list
            self.segments = Segment.SegmentList()
            # Load any previous segments stored
            if os.path.isfile(self.filename + '.data') and os.stat(self.filename+'.data').st_size > 0:
                # Populate it, add the metadata attribute
                # (note: we're overwriting the JSON duration with actual full wav size)
                hasmetadata = self.segments.parseJSON(self.filename+'.data', self.datalength / self.sp.audioFormat.sampleRate())
                #hasmetadata = self.segments.parseJSON(self.filename+'.data', self.sp.fileLength / self.sp.sampleRate)
                if not hasmetadata:
                    self.segments.metadata["Operator"] = self.operator
                    self.segments.metadata["Reviewer"] = self.reviewer
                    self.segmentsToSave = True
                self.operator = self.segments.metadata.get("Operator", self.operator)
                self.reviewer = self.segments.metadata.get("Reviewer",self.reviewer)

                #self.segmentsToSave = True

                # if there are any multi-species segments,
                # switch the option on regardless of user preference
                for s in self.segments:
                    if len(s[4])>1:
                        if not self.multipleBirds:
                            self.multipleBirds = True
                    for species,certainty,calltype in s.getKeysWithCalltypes():
                        if not species in self.knownCalls:
                            self.knownCalls[species]=[]
                        if not calltype is None:
                            if not calltype in self.knownCalls[species]:
                                self.knownCalls[species].append(calltype)
            else:
                self.segments.metadata = {"Operator": self.operator, "Reviewer": self.reviewer, "Duration": self.datalength / self.sp.audioFormat.sampleRate()}
                #self.segments.metadata = {"Operator": self.operator, "Reviewer": self.reviewer, "Duration": self.datalength / self.sp.sampleRate}

            # Bat mode: initialize with an empty segment for the entire file
            if self.batmode and len(self.segments)==0:
                species = [{"species": "Don't Know", "certainty": 0, "filter": "M"}]
                # SRM: TODO: keep this? If so, needs a parameter...
                self.useClicks = True
                if self.useClicks:
                    result = self.sp.clickSearch()
                    if result is not None:
                        start = self.convertSpectoAmpl(result[0])
                        end = self.convertSpectoAmpl(result[1])
                    else:
                        start = 0
                        end = self.datalength / self.sp.audioFormat.sampleRate()
                        #end = self.sp.fileLength / self.sp.sampleRate
                else:
                    start = 0
                    end = self.datalength / self.sp.audioFormat.sampleRate()
                    #end = self.sp.fileLength / self.sp.sampleRate
                newSegment = Segment.Segment([start, end, 0, 0, species])
                self.segments.append(newSegment)
                self.segmentsToSave = True
                self.refreshFileColor()

            self.drawProtocolMarks()

            self.statusRight.setText("Operator: " + str(self.operator) + ", Reviewer: " + str(self.reviewer))

            if hasattr(self,'seg'):
                self.seg.setNewData(self.sp)
            else:
                self.seg = Segment.Segmenter(self.sp, self.sp.audioFormat.sampleRate())
                #self.seg = Segment.Segmenter(self.sp, self.sp.sampleRate)

            # Update the Dialogs
            # Also close any ones that could get buggy when moving between bird-bat modes
            if hasattr(self,'spectrogramDialog'):
                if self.batmode:
                    self.spectrogramDialog.reject()
                self.spectrogramDialog.setValues(self.sp.minFreq,self.sp.maxFreq,self.sp.minFreqShow,self.sp.maxFreqShow)
            if hasattr(self,'denoiseDialog'):
                if self.batmode:
                    self.denoiseDialog.reject()
                self.denoiseDialog.setValues(self.sp.minFreq,self.sp.maxFreq)

            # Delete any denoising backups from the previous file
            if hasattr(self,'audiodata_backup'):
                self.audiodata_backup = None
            #self.showInvSpec.setChecked(False)

            self.timeaxis.setOffset(self.startRead+self.startTime)

            # Set the window size
            if self.batmode:
                self.windowSize = self.datalengthSec
            else:
                self.windowSize = self.config['windowWidth']
            self.timeaxis.setRange(0, self.windowSize)
            self.widthWindow.setRange(0.5, self.datalengthSec)

            # Reset if the file is shorter than the window
            if self.datalengthSec < self.windowSize or self.batmode:
                self.windowSize = self.datalengthSec
            self.widthWindow.setValue(self.windowSize)

            # Decide which axis scaling to use
            if self.windowSize<3:
                self.timeaxis.setShowMS(True)

            self.totalTime = self.convertMillisecs(1000*self.datalengthSec)

            if not self.CLI:
                if not self.batmode:
                    # Initialise the sound and bar moving timer
                    #if not hasattr(self,'media_thread'):
                        #self.media_thread = QThread()
                        #self.NotifyTimer = QTimer(self)
                        #self.NotifyTimer.timeout.connect(self.movePlaySlider)
                    self.media_obj = SupportClasses_GUI.ControllableAudio(self.sp,useBar=True)
                    self.media_obj.NotifyTimer.timeout.connect(self.movePlaySlider)

                    #self.media_obj.NotifyTimer.timeout.connect(self.movePlaySlider)
                    #self.media_obj.moveToThread(self.media_thread)
                    #self.media_thread.start()

                    #self.mediaThread = threading.Thread(target=self.media_obj.play)
                    #self.mediaThread.start()

                # Set the length of the scrollbar.
                self.scrollSlider.setRange(0,int(np.shape(self.sg)[0] - self.convertAmpltoSpec(self.widthWindow.value())))
                self.scrollSlider.setValue(0)

                self.drawOverview()
                dlg += 1
                dlg.update()
                self.drawfigMain()
                self.setWindowTitle('AviaNZ: Manual Processing ' + self.filename)
                dlg += 1
                dlg.update()
                self.w_spec.setFocus()
                self.statusLeft.setText("Ready")
            else:
                self.drawfigMain()

    def openNextFile(self):
        """ Listener for next file >> button.
        Get the next file in the list and call the loader. """

        # If the user has navigated away from the dir with currently open file, return:
        if self.listFiles.soundDir != os.path.dirname(self.filename):
            self.SoundFileDir = os.path.dirname(self.filename)
            self.fillFileList(self.SoundFileDir, os.path.basename(self.filename))

        i=self.listFiles.currentRow()
        if i+1<len(self.listFiles):
            self.listFiles.setCurrentRow(i+1)
            self.listLoadFile(self.listFiles.currentItem())
        else:
            # Tell the user they've finished
            msg = SupportClasses_GUI.MessagePopup("d", "Last file", "You've finished processing the folder")
            msg.exec()

    def showPointerDetailsCheck(self):
        """ Listener for the menuitem that sets if detailed info should be shown when hovering over spectrogram.
        Turning this off saves lots of CPU performance."""

        self.config['showPointerDetails'] = self.showPointerDetails.isChecked()
        if self.showPointerDetails.isChecked():
            self.p_spec.scene().sigMouseMoved.connect(self.mouseMoved)
            self.w_spec.addItem(self.pointDataProxy, row=3, col=1)
        else:
            self.p_spec.scene().sigMouseMoved.disconnect()
            self.w_spec.removeItem(self.pointDataProxy)

    def dragRectsTransparent(self):
        """ Listener for the check menu item that decides if the user wants the dragged rectangles to have colour or not.
        It's a switch from Brush to Pen or vice versa.
        """
        if self.config['transparentBoxes']:
            for box in self.listRectanglesa2:
                if type(box) == self.ROItype:
                    col = box.brush.color()
                    col.setAlpha(255)
                    box.transparent = True
                    box.setBrush(pg.mkBrush(None))
                    box.setPen(pg.mkPen(col,width=1))
                    box.update()
                    col.setAlpha(100)
        else:
            for box in self.listRectanglesa2:
                if type(box) == self.ROItype:
                    col = box.pen.color()
                    col.setAlpha(self.ColourNamed.alpha())
                    box.transparent = False
                    box.setBrush(pg.mkBrush(col))
                    box.setPen(pg.mkPen(None))
                    box.update()
                    col.setAlpha(100)

    def useAmplitudeCheck(self):
        """ Listener for the check menu item saying if the user wants to see the waveform.
        Does not remove the dock, just hides it. It's therefore easy to replace, but could have some performance overhead.
        """
        if self.useAmplitudeTick.isChecked():
            self.useAmplitude = True
            self.d_ampl.show()
        else:
            self.useAmplitude = False
            self.d_ampl.hide()
        self.config['showAmplitudePlot'] = self.useAmplitudeTick.isChecked()

    def useFilesCheck(self):
        """ Listener to process if the user swaps the check menu item to see the file list. """
        if self.useFilesTick.isChecked():
            self.d_files.show()
        else:
            self.d_files.hide()
        self.config['showListofFiles'] = self.useFilesTick.isChecked()

    def showOverviewSegsCheck(self):
        """ Listener to process if the user swaps the check menu item to see the overview segment boxes. """
        if self.showOverviewSegsTick.isChecked():
            self.p_overview2.show()
        else:
            self.p_overview2.hide()
        self.config['showAnnotationOverview'] = self.showOverviewSegsTick.isChecked()

    def makeReadOnly(self):
        """ Listener to process the check menu item to make the plots read only.
        Turns off the listeners for the amplitude and spectrogram plots.
        Also has to go through all of the segments, turn off the listeners, and make them unmovable.
        """
        self.config['readOnly'] = self.readonly.isChecked()
        self.statusRO.setText("Read-only mode" if self.config['readOnly'] else "")
        if self.readonly.isChecked():
            # This is for accepting drag boxes or not
            self.p_spec.enableDrag = False

            # When clicking is used to draw segments/boxes, read-only changes are implemented in the button signals.
            # Because connecting-disconnecting slots is very dirty.

            # This will re-make segment boxes with correct moving abilities:
            if hasattr(self, 'sp'):
                self.removeSegments(delete=False)
                self.drawfigMain(remaking=True)
        else:
            self.p_spec.enableDrag = self.config['specMouseAction']==3
            if hasattr(self, 'sp'):
                self.removeSegments(delete=False)
                self.drawfigMain(remaking=True)

    def dockReplace(self):
        """ Listener for if the docks should be replaced menu item.
            A rewrite of pyqtgraph.dockarea.restoreState.
            """
        containers, docks = self.area.findAll()
        # Main recursion of restoreState:
        self.area.buildFromState(self.state['main'], docks, self.area, missing='error')
        # RestoreState doesn't restore non-floating window sizes 
        self.d_plot.hide()
        containers, docks = self.area.findAll()
        # Basically say that left panel and controls should be as small as possible:
        self.d_controls.setSizePolicy(QSizePolicy.Policy.Minimum,QSizePolicy.Policy.Minimum)
        containers[1].setSizePolicy(QSizePolicy.Policy.Minimum,QSizePolicy.Policy.Minimum)
        #self.useAmplitudeTick.setChecked(True)
        #self.useAmplitude = True
        #self.config['showAmplitudePlot'] = True
        self.useFilesTick.setChecked(True)
        self.config['showListofFiles'] = True
        self.showOverviewSegsTick.setChecked(True)
        self.config['showAnnotationOverview'] = True
        self.useAmplitudeCheck()
        # for cont in range(len(containers)):
        #     containers[cont].setSizes(self.state_cont[cont])

    def showFundamentalFreq(self):
        """ Calls the Spectrogram class to compute, and then draws, the fundamental frequency"""

        with pg.BusyCursor():
            if self.showFundamental.isChecked():
                self.statusLeft.setText("Drawing fundamental frequency...")
                segs = self.sp.drawFundFreq(self.seg)

                # Get the individual pieces
                self.segmentPlots = []
                # Draw each contiguous "segment" of fund freq
                for s in segs:
                    self.segmentPlots.append(pg.PlotDataItem())
                    self.segmentPlots[-1].setData(s[0], s[1], pen=pg.mkPen('r', width=3))
                    self.p_spec.addItem(self.segmentPlots[-1])
            else:
                self.statusLeft.setText("Removing fundamental frequency...")
                for r in self.segmentPlots:
                    self.p_spec.removeItem(r)
                self.segmentPlots = []
            self.statusLeft.setText("Ready")

    def showMaxEnergy(self):
        """ Calls the Spectrogram class to compute, and then draws, the maximum energy"""

        with pg.BusyCursor():
            if self.showEnergies.isChecked():
                self.statusLeft.setText("Drawing max energies...")
                x, y = self.sp.max_energy(self.sg)
                self.energyPlot = pg.ScatterPlotItem()
                self.energyPlot.setBrush(None)
                self.energyPlot.setData(x, y, brush=pg.mkBrush((0, 255, 0, 130)), pen=pg.mkPen(None), size=5)

                self.p_spec.addItem(self.energyPlot)
            else:
                self.statusLeft.setText("Removing max energies...")
                self.p_spec.removeItem(self.energyPlot)
            self.statusLeft.setText("Ready")

    def showSpectralDeriv(self):
        """ Calls the Spectrogram class to compute, and then draws, the spectral derivatives"""

        with pg.BusyCursor():
            if self.showSpectral.isChecked():
                self.statusLeft.setText("Drawing spectral derivative...")
                x, y = self.sp.drawSpectralDeriv()
            
                if x is not None:
                    self.derivPlot = pg.ScatterPlotItem()
                    self.derivPlot.setData(x, y, pen=pg.mkPen('b', width=5))
                    self.p_spec.addItem(self.derivPlot)
            else:
                self.statusLeft.setText("Removing spectral derivative...")
                if hasattr(self, 'derivPlot'):
                    self.p_spec.removeItem(self.derivPlot)
            self.statusLeft.setText("Ready")

    def showFormants(self):
        """ Calls the Spectrogram class to compute, and then draws, the formants"""

        with pg.BusyCursor():
            if self.showFormant.isChecked():
                self.statusLeft.setText("Drawing formants...")
                x, y = self.sp.drawFormants()
                self.formantPlot = pg.ScatterPlotItem()
                self.formantPlot.setData(x, y, pen=pg.mkPen('b', width=0.05))
                self.p_spec.addItem(self.formantPlot)
            else:
                self.statusLeft.setText("Removing formants...")
                if hasattr(self, 'formantPlot'):
                    self.p_spec.removeItem(self.formantPlot)
            self.statusLeft.setText("Ready")

    def drawGuidelines(self):
        # Frequency guides for bat mode

        if self.config['guidelinesOn']=='always' or (self.config['guidelinesOn']=='bat' and self.batmode):
            for gi in range(len(self.guidelines)):
                self.guidelines[gi].setValue(self.convertFreqtoY(self.config['guidepos'][gi]))
                self.guidelines[gi].setPen(color=self.config['guidecol'][gi], width=2)
                self.p_spec.addItem(self.guidelines[gi], ignoreBounds=True)
        else:
            # easy way to hide
            for g in self.guidelines:
                g.setValue(-1000)

    # def showCQT(self):
    #     cqt = self.sp.comp_cqt()
    #     print(np.shape(cqt),np.shape(self.sg))
    #     self.specPlot.setImage(10*np.log10(np.real(cqt*np.conj(cqt))).T)
    #     self.p_spec.setXRange(0, np.shape(cqt)[1], update=True, padding=0)

    # ==============
    # Code for drawing and using the main figure

    def convertAmpltoSpec(self,x):
        """ Unit conversion """
        if self.batmode:
            incr = 512
        else:
            incr = self.config['incr']
        return x*self.sp.audioFormat.sampleRate()/incr
        #return x*self.sp.sampleRate/incr

    def convertSpectoAmpl(self,x):
        """ Unit conversion """
        if self.batmode:
            incr = 512
        else:
            incr = self.config['incr']
        return x*incr/self.sp.audioFormat.sampleRate()
        #return x*incr/self.sp.sampleRate

    def convertMillisecs(self,millisecs):
        """ Unit conversion """
        seconds = (millisecs / 1000) % 60
        minutes = (millisecs / (1000 * 60)) % 60
        return "%02d" % minutes+":"+"%02d" % seconds

    def convertYtoFreq(self,y,sgy=None):
        """ Unit conversion """
        if sgy is None:
            sgy = np.shape(self.sg)[1]
        return y * self.sp.audioFormat.sampleRate()//2 / sgy + self.sp.minFreqShow
        #return y * self.sp.sampleRate//2 / sgy + self.sp.minFreqShow

    def convertFreqtoY(self,f):
        """ Unit conversion """
        sgy = np.shape(self.sg)[1]
        return (f-self.sp.minFreqShow) * sgy / (self.sp.audioFormat.sampleRate()//2)
        #return (f-self.sp.minFreqShow) * sgy / (self.sp.sampleRate//2)

    def drawOverview(self):
        """ On loading a new file, update the overview figure to show where you are up to in the file.
        Also, compute the new segments for the overview, make sure that the listeners are connected
        for clicks on them, and disconnect old listeners. """
        self.overviewImage.setImage(self.sg)
        self.overviewImageRegion.setBounds([0, len(self.sg)])
        self.overviewImageRegion.setRegion([0, self.convertAmpltoSpec(self.widthWindow.value())])
        try:
            self.overviewImageRegion.sigRegionChangeFinished.disconnect()
        except Exception:
            pass
        self.overviewImageRegion.sigRegionChangeFinished.connect(self.updateOverview)

        # Three y values are No. not known, No. known, No. possible
        # widthOverviewSegment is in seconds
        numSegments = int(np.ceil(np.shape(self.sg)[0]/self.convertAmpltoSpec(self.config['widthOverviewSegment'])))
        self.widthOverviewSegment = np.shape(self.sg)[0]//numSegments

        self.overviewSegments = np.zeros((numSegments,3))

        # Delete the overview segments
        for r in self.SegmentRects:
            self.p_overview2.removeItem(r)
        self.SegmentRects = []

        # Add new overview segments
        for i in range(numSegments):
            r = SupportClasses_GUI.ClickableRectItem(i*self.widthOverviewSegment, 0, self.widthOverviewSegment, 1)
            r.setPen(pg.mkPen(100, 100, 100))
            r.setBrush(pg.mkBrush('w'))
            self.SegmentRects.append(r)
            self.p_overview2.addItem(r)
        try:
            self.p_overview2.sigChildMessage.disconnect()
        except Exception:
            pass
        self.p_overview2.sigChildMessage.connect(self.overviewSegmentClicked)
        self.p_overview2.setYRange(-0.2, 1, padding=0.02)

    def overviewSegmentClicked(self,x):
        """ Listener for an overview segment being clicked on.
        Work out which one, and move the region appropriately. Calls updateOverview to do the work. """
        minX, maxX = self.overviewImageRegion.getRegion()
        halfwin = (maxX-minX)/2
        self.overviewImageRegion.setRegion([x-halfwin, x+halfwin])

    def updateOverview(self):
        """ Listener for when the overview box is changed. Other functions call it indirectly by setRegion.
        Does the work of keeping all the plots in the right place as the overview moves.
        It sometimes updates a bit slowly. """

        if hasattr(self, 'media_obj'):
            if self.media_obj.isPlayingorPaused():
                self.stopPlayback()

        minX, maxX = self.overviewImageRegion.getRegion()

        # (The region bounds are checked against spec size in our subclass)

        # Temporarily block callback, and update window size (because setRegion may have changed it to fit bounds)
        self.updateRequestedByOverview = True
        self.widthWindow.setValue(self.convertSpectoAmpl(maxX-minX))
        self.p_ampl.setXRange(self.convertSpectoAmpl(minX), self.convertSpectoAmpl(maxX), update=True, padding=0)
        self.p_spec.setXRange(minX, maxX, update=True, padding=0)
        self.p_plot.setXRange(self.convertSpectoAmpl(minX), self.convertSpectoAmpl(maxX), update=True, padding=0)

        # I know the next two lines SHOULD be unnecessary. But they aren't!
        self.p_ampl.setXRange(self.convertSpectoAmpl(minX), self.convertSpectoAmpl(maxX), padding=0)
        self.p_spec.setXRange(minX, maxX, padding=0)

        if self.extra != "none" and "Filtered spectrogram" not in self.extra:
            self.p_plot.setXRange(self.convertSpectoAmpl(minX), self.convertSpectoAmpl(maxX), padding=0)
        if "Filtered spectrogram" in self.extra:
            self.p_plot.setXRange(minX, maxX, padding=0)
        elif self.extra=="Wavelet scalogram":
            self.p_plot.setXRange(self.convertSpectoAmpl(minX)*4, self.convertSpectoAmpl(maxX)*4)
        # self.setPlaySliderLimits(1000.0*self.convertSpectoAmpl(minX),1000.0*self.convertSpectoAmpl(maxX))
        self.scrollSlider.setValue(int(minX))
        self.config['windowWidth'] = self.convertSpectoAmpl(maxX-minX)
        # self.saveConfig = True
        self.timeaxis.update()
        QApplication.processEvents()
        self.updateRequestedByOverview = False

    def setfigs(self):
        height = self.sp.audioFormat.sampleRate() // 2 / np.shape(self.sg)[1]
        #height = self.sp.sampleRate // 2 / np.shape(self.sg)[1]
        pixelstart = int(self.sp.minFreqShow/height)
        pixelend = int(self.sp.maxFreqShow/height)

        self.overviewImage.setImage(self.sg[:,pixelstart:pixelend])
        self.overviewImageRegion.setBounds([0, len(self.sg)])
        self.specPlot.setImage(self.sg[:,pixelstart:pixelend])
        self.setExtraPlot(self.extra)

        self.setColourMap(self.config['cmap'])
        self.setColourLevels()

    def drawfigMain(self,remaking=False):
        """ Draws the main amplitude and spectrogram plots and any segments on them.
        Has to do some work to get the axis labels correct.
        """

        if len(self.sp.data)>0 and not self.batmode:
            self.amplPlot.setData(np.linspace(0.0,self.datalengthSec,num=self.datalength,endpoint=True),self.sp.data)

        self.timeaxis.setLabel('')

        self.setfigs()

        # Sort out the spectrogram frequency axis
        # The constants here are divided by 1000 to get kHz, and then remember the top is sampleRate/2

        # There are two options for logarithmic axis (Mel/Bark): keep the numbers equally spaced, but correct the labels, or keep the numbers but space the labels correctly.
        # I'm doing the first for now, although it isn't as good.

        FreqRange = self.sp.maxFreqShow-self.sp.minFreqShow
        height = self.sp.audioFormat.sampleRate() // 2 / np.shape(self.sg)[1]
        #height = self.sp.sampleRate // 2 / np.shape(self.sg)[1]
        SpecRange = FreqRange/height
        self.drawGuidelines()

        if self.zooniverse:

            labels = [0,int(FreqRange//4000),int(FreqRange//2000),int(3*FreqRange//4000),int(FreqRange//1000)]
            if self.config['sgScale'] == 'Mel Frequency':
                for i in range(len(labels)):
                    labels[i] = self.sp.convertHztoMel(labels[i])
            elif self.config['sgScale'] == 'Bark Frequency':
                for i in range(len(labels)):
                    labels[i] = self.sp.convertHztoBark(labels[i])
        
            offset=6
            txt='<span style="color: #0F0; font-size:20pt">%s</div>'%str(labels[0])
            #txt='<span style="color: #0F0; font-size:20pt">%s</div>'%str(0)
            self.label1 = pg.TextItem(html=txt, color='g', anchor=(0,0))
            self.p_spec.addItem(self.label1)
            self.label1.setPos(0,0+offset)

            txt='<span style="color: #0F0; font-size:20pt">%s</div>'%str(labels[1])
            #txt='<span style="color: #0F0; font-size:20pt">%s</div>'%str(int(FreqRange//4000))
            self.label2 = pg.TextItem(html=txt, color='g', anchor=(0,0))
            self.p_spec.addItem(self.label2)
            self.label2.setPos(0,SpecRange/4+offset)

            txt='<span style="color: #0F0; font-size:20pt">%s</div>'%str(labels[2])
            #txt='<span style="color: #0F0; font-size:20pt">%s</div>'%str(int(FreqRange//2000))
            self.label3 = pg.TextItem(html=txt, color='g', anchor=(0,0))
            self.p_spec.addItem(self.label3)
            self.label3.setPos(0,SpecRange/2+offset)

            txt='<span style="color: #0F0; font-size:20pt">%s</div>'%str(labels[3])
            #txt='<span style="color: #0F0; font-size:20pt">%s</div>'%str(int(3*FreqRange//4000))
            self.label4 = pg.TextItem(html=txt, color='g', anchor=(0,0))
            self.p_spec.addItem(self.label4)
            self.label4.setPos(0,3*SpecRange/4+offset)

            txt='<span style="color: #0F0; font-size:20pt">%s</div>'%str(labels[4])
            #txt='<span style="color: #0F0; font-size:20pt">%s</div>'%str(int(FreqRange//1000))
            self.label5 = pg.TextItem(html=txt, color='g', anchor=(0,0))
            self.p_spec.addItem(self.label5)
            self.label5.setPos(0,SpecRange+offset)
        else:
            labels = [self.sp.minFreqShow, self.sp.minFreqShow+FreqRange/4, self.sp.minFreqShow+FreqRange/2, self.sp.minFreqShow+3*FreqRange/4, self.sp.minFreqShow+FreqRange]

            if self.config['sgScale'] == 'Mel Frequency':
                for i in range(len(labels)):
                    labels[i] = self.sp.convertHztoMel(labels[i])
                self.specaxis.setLabel('Mels')
            elif self.config['sgScale'] == 'Bark Frequency':
                for i in range(len(labels)):
                    labels[i] = self.sp.convertHztoBark(labels[i])*1000
                self.specaxis.setLabel('Barks')
            else:
                self.specaxis.setLabel('kHz')
       
            #self.specaxis.setTicks([[(0,round(labels[0]/1000,2)),(SpecRange/4,round(labels[1]/1000,2)),(SpecRange/2,round(labels[2]/1000,2)),(3*SpecRange/4,round(labels[3]/1000,2)),(SpecRange,round(labels[4]/1000,2))]])
            ticks = [(0,round(labels[0]/1000,2)),(SpecRange/4,round(labels[1]/1000,2)),(SpecRange/2,round(labels[2]/1000,2)),(3*SpecRange/4,round(labels[3]/1000,2)),(SpecRange,round(labels[4]/1000,2))]
            ticks = [[(tick[0], "%.1f" % tick[1] ) for tick in ticks]]
            self.specaxis.setTicks(ticks)

            #self.specaxis.setTicks([[(0,round(self.sp.minFreqShow/1000, 2)),
                                 #(SpecRange/4,round(self.sp.minFreqShow/1000+FreqRange/4000, 2)),
                                 #(SpecRange/2,round(self.sp.minFreqShow/1000+FreqRange/2000, 2)),
                                 #(3*SpecRange/4,round(self.sp.minFreqShow/1000+3*FreqRange/4000, 2)),
                                 #(SpecRange,round(self.sp.minFreqShow/1000+FreqRange/1000, 2))]])

        self.updateOverview()
        self.textpos = int((self.sp.maxFreqShow-self.sp.minFreqShow)/height) #+ self.config['textoffset']

        # If there are segments, show them
        if not self.cheatsheet and not self.zooniverse:
            for count in range(len(self.segments)):
                self.addSegment(self.segments[count][0], self.segments[count][1], self.segments[count][2], self.segments[count][3], self.segments[count][4], False, count, remaking, coordsAbsolute=True)

            # This is the moving bar for the playback
            self.p_spec.addItem(self.bar, ignoreBounds=True)

        QApplication.processEvents()

    def setSpeed(self,speed):
        # TODO: when called from the playback, would ideally set the speed back to 1 -> check this
        # So how to find the right action?
        #self.speedButton.menu().setCurrentIndex(3)
        if type(speed) is str:
            # convert Unicode fractions to floats
            speedchar = ord(speed)
            if speedchar == 188:
                speed = 0.25
            elif speedchar == 189:
                speed = 0.5
            elif speedchar == 190:
                speed = 0.75
        self.playSpeed = 1/float(speed)
        self.media_obj.setSpeed(self.playSpeed)
        #print("playback speed:",self.playSpeed)

    def setExtraPlot(self, plotname):
        """ Reacts to menu clicks and updates or hides diagnostic plot window."""
        self.extra = plotname

        # Clear plot before updating/proceeding
        self.clearDiagnostic()

        if self.extra != "none":
            self.d_plot.show()
        else:
            self.d_plot.hide()

        # Plot wavelet scalogram
        if self.extra == "Wavelet scalogram":
            self.plotExtra = pg.ImageItem()
            self.p_plot.addItem(self.plotExtra)

            # Passing dummy spInfo because we only use this for a function
            ws = WaveletSegment.WaveletSegment(spInfo={}, wavelet='dmey2')
            e = ws.computeWaveletEnergy(self.sp.data, self.sp.audioFormat.sampleRate(), window=0.25, inc=0.25)
            #e = ws.computeWaveletEnergy(self.sp.data, self.sp.sampleRate, window=0.25, inc=0.25)
            # e is 2^nlevels x nseconds

            # show only leaf nodes:
            #print(np.shape(e))
            e = np.log(e[30:62,:])

            self.plotExtra.setLookupTable(colourMaps.getLookupTable("Inferno"))
            self.plotExtra.setImage(e.T)
            self.plotaxis.setLabel('Wavelet node')

        # plot wc correlations
        if self.extra == "Wavelet correlations":
            self.plotExtra = pg.ImageItem()
            self.p_plot.addItem(self.plotExtra)

            # Preprocess
            # TODO: Other samplerates?
            data = self.sp.resample(16000)
            data = SignalProc.bandpassFilter(data, self.sp.audioFormat.sampleRate(), 100, 16000)
            #data = SignalProc.bandpassFilter(data, self.sp.sampleRate, 100, 16000)

            # passing dummy spInfo because we only use this for a function
            ws = WaveletSegment.WaveletSegment(spInfo={}, wavelet='dmey2')
            e = ws.computeWaveletEnergy(self.sp.data, self.sp.audioFormat.sampleRate())
            #e = ws.computeWaveletEnergy(self.sp.data, self.sp.sampleRate)
            annotation = np.zeros(np.shape(e)[1])
            for s in self.segments:
                annotation[math.floor(s[0]):math.ceil(s[1])] = 1
            w0 = np.where(annotation == 0)[0]
            w1 = np.where(annotation == 1)[0]

            r = np.zeros((64, np.shape(e)[1]))
            for count in range(62):
                # Just compute_r from WaveletSegment
                corr = (np.mean(e[count, w1]) - np.mean(e[count,w0]))/np.std(e[count, :]) * np.sqrt(len(w0)*len(w1))/len(annotation)
                # Map a long vector of rs to different image areas
                level = int(math.log(count+2, 2))
                node = count+2 - 2**level
                node = WaveletFunctions.graycode(node)
                r[node * 2**(6-level) : (node+1) * 2**(6-level), level] = corr
            r[:, 0] = np.linspace(np.min(r), np.max(r), num=64)
            # Propagate along x
            for tmult in range(10, len(annotation)):
                r[:, tmult] = r[:, tmult-10]

            self.plotExtra.setLookupTable(colourMaps.getLookupTable("Viridis"))
            self.plotExtra.setImage(r.T)
            self.plotaxis.setLabel('Frequency bin')

        # Plot estimated wind and signal levels for a pre-set node
        if self.extra == "Wind adjustment":
            # TODO
            TGTSAMPLERATE = 16000
            minchpwin = 32/TGTSAMPLERATE
            chpwin = round(0.25/minchpwin)*minchpwin
            print("Will use window of", chpwin, "s")
            # Resample and generate WP w/ all nodes for the current page
            if self.sp.audioFormat.sampleRate() != TGTSAMPLERATE:
            #if self.sp.sampleRate != TGTSAMPLERATE:
                datatoplot = self.sp.resample(TGTSAMPLERATE)
            else:
                datatoplot = self.sp.data
            WF = WaveletFunctions.WaveletFunctions(data=datatoplot, wavelet='dmey2', maxLevel=5, samplerate=TGTSAMPLERATE)
            WF.WaveletPacket(range(31, 63))
            # list all the node frequency centers
            node_freqs = [sum(WaveletFunctions.getWCFreq(n, TGTSAMPLERATE))/2 for n in range(31, 63)]

            xs = np.arange(0, self.datalengthSec, chpwin)
            # xs = xs[:-1] # TODO TMP only for bittern
            datalen = len(xs)

            # extract energies from each leaf node
            Es = np.zeros((datalen, 32))
            i = 0
            for node in range(31, 63):
                E, _ = WF.extractE(node, chpwin)
                Es[:,i] = E[:datalen]
                i += 1
                print("node %d extracted" % node)

            # extract unadjusted signal energy
            tgt_node = 46-31
            Es = np.log(Es)
            sig_lvl = Es[:,tgt_node]

            # estimate wind level in each window
            wind_lvlR = np.zeros(len(xs))
            wind_lvlC = np.zeros(len(xs))
            wind_lvlCO = np.zeros(len(xs))

            # select 10 % frames with lowest overall energy:
            # (note that root node does not need downsampling)
            bgpow, _ = WF.extractE(0, chpwin, wpantialias=False)
            numframes = round(0.1*len(bgpow))
            quietframes = np.argpartition(bgpow, numframes)[:numframes]
            # From these, determine "background level" for the target nodes:
            bgpow = np.mean(sig_lvl[quietframes])

            oversubalpha = 1.2

            def calc_wind(energies, nodecentres, tgt):
                # fits an adjustment model based on energies, nodecentres,
                # excluding tgt node.
                regy = np.delete(energies, tgt)
                regy = np.delete(regy, 45-31)
                regy = np.delete(regy, 44-31)
                regx = np.delete(nodecentres, tgt)
                regx = np.delete(regx, 45-31)
                regx = np.delete(regx, 44-31)

                # remove two extreme nodes as they have filtering artifacts
                regx = np.delete(regx, 16)
                regx = np.delete(regx, 0)
                regy = np.delete(regy, 16)
                regy = np.delete(regy, 0)
                # Drop nodes that are too high outside wind freq range
                freqmask = np.where(regx<6000)

                regy = regy[freqmask]
                regx = np.log(regx[freqmask])

                # Robust reg
                # NOTE: this is our own QuantReg reimplementation
                # really tailored to our polynomial fit
                # TODO see if sklearn model, to be added in v1.0, is any better
                regx_poly = np.column_stack((np.ones(len(regx)), regx, regx**2, regx**3))
                pol = WaveletFunctions.QuantReg(regy, regx_poly, q=0.20, max_iter=250, p_tol=1e-3)
                predR = pol(np.log(nodecentres[tgt]))

                # Cubic fit
                pol = np.polynomial.polynomial.Polynomial.fit(regx, regy, 3)
                predC = pol(np.log(nodecentres[tgt]))

                # oversubtracted cubic
                predCO = (predC - bgpow)*oversubalpha + bgpow
                return predR, predC, predCO

            for w in range(len(xs)):
                wind_lvlR[w], wind_lvlC[w], wind_lvlCO[w] = calc_wind(Es[w,:], node_freqs, tgt_node)

            self.p_legend = pg.LegendItem()
            self.p_legend.setParentItem(self.p_plot)
            self.plotExtra = pg.PlotDataItem(xs, sig_lvl)
            self.plotExtra.setPen(fn.mkPen(color='k', width=2))
            self.p_legend.addItem(self.plotExtra, 'node')
            self.plotExtra2 = pg.PlotDataItem(xs, wind_lvlC)
            self.plotExtra2.setPen(fn.mkPen(color='r', width=2))
            self.p_legend.addItem(self.plotExtra2, 'cubic')
            self.plotExtra3 = pg.PlotDataItem(xs, wind_lvlR)
            self.plotExtra3.setPen(fn.mkPen(color=(0, 200, 0), width=2))
            self.p_legend.addItem(self.plotExtra3, 'robust')
            self.plotExtra4 = pg.PlotDataItem(xs, wind_lvlCO)
            self.plotExtra4.setPen(fn.mkPen(color=(80,30,255), width=2))
            self.p_legend.addItem(self.plotExtra4, 'cub-over')
            self.plotaxis.setLabel('Log mean power')
            self.p_plot.addItem(self.plotExtra)
            self.p_plot.addItem(self.plotExtra2)
            self.p_plot.addItem(self.plotExtra3)
            self.p_plot.addItem(self.plotExtra4)

        # plot energy in "wind" band
        if self.extra == "Wind energy":
            we_mean = np.zeros(int(np.ceil(self.datalengthSec)))
            we_std = np.zeros(int(np.ceil(self.datalengthSec)))
            for w in range(int(np.ceil(self.datalengthSec))):
                data = self.sp.data[int(w*self.sp.audioFormat.sampleRate()):int((w+1)*self.sp.audioFormat.sampleRate())]
                #data = self.sp.data[int(w*self.sp.sampleRate):int((w+1)*self.sp.sampleRate)]
                post = Segment.PostProcess(configdir=self.configdir, audioData=data, sampleRate=self.sp.audioFormat.sampleRate(), segments=[], subfilter={})
                #post = Segment.PostProcess(configdir=self.configdir, audioData=data, sampleRate=self.sp.sampleRate, segments=[], subfilter={})
                m, std, _ = post.wind_cal(data, self.sp.audioFormat.sampleRate())
                #m, std, _ = post.wind_cal(data, self.sp.sampleRate)
                we_mean[w] = m
                we_std[w] = std

            print('mean wind: ', we_mean)
            self.plotExtra = pg.PlotDataItem(np.arange(int(np.ceil(self.datalengthSec))), we_mean)
            self.plotExtra.setPen(fn.mkPen(color='k', width=2))
            self.plotExtra2 = pg.PlotDataItem(np.arange(int(np.ceil(self.datalengthSec))), we_mean+we_std)
            self.plotExtra2.setPen(fn.mkPen('c'))
            self.plotExtra3 = pg.PlotDataItem(np.arange(int(np.ceil(self.datalengthSec))), we_mean-we_std)
            self.plotExtra3.setPen(fn.mkPen('c'))
            self.plotExtra4 = pg.PlotDataItem(np.arange(int(np.ceil(self.datalengthSec))), np.ones(int(np.ceil(self.datalengthSec)))*1.5)
            self.plotExtra4.setPen(fn.mkPen('r'))
            self.plotaxis.setLabel('Mean (SD) power, V^2/Hz')
            self.p_plot.addItem(self.plotExtra)
            self.p_plot.addItem(self.plotExtra2)
            self.p_plot.addItem(self.plotExtra3)
            self.p_plot.addItem(self.plotExtra4)

        # plot energy in "rain"
        if self.extra == "Rain":
            we_mean = np.zeros(int(np.ceil(self.datalengthSec)))
            we_std = np.zeros(int(np.ceil(self.datalengthSec)))
            for w in range(int(self.datalength/self.sp.audioFormat.sampleRate())):
            #for w in range(int(self.datalength/self.sp.sampleRate)):
                data = self.sp.data[int(w*self.sp.audioFormat.sampleRate()):int((w+1)*self.sp.audioFormat.sampleRate())]
                #data = self.sp.data[int(w*self.sp.sampleRate):int((w+1)*self.sp.sampleRate)]
                tempsp = Spectrogram.Spectrogram()
                tempsp.data = data
                sgRaw = tempsp.spectrogram()
                # Normalise
                sgRaw -= np.mean(sgRaw, axis=0)
                sgRaw /= np.max(np.abs(sgRaw), axis=0)
                we_mean[w] = np.mean(np.mean(sgRaw, axis=0))
                we_std[w] = np.std(np.std(sgRaw, axis=0))
            self.plotExtra = pg.PlotDataItem(np.arange(self.datalengthSec), we_mean)
            self.plotExtra.setPen(fn.mkPen(color='k', width=2))
            self.plotExtra2 = pg.PlotDataItem(np.arange(self.datalengthSec), we_mean + we_std)
            self.plotExtra2.setPen(fn.mkPen('r'))
            self.plotExtra3 = pg.PlotDataItem(np.arange(self.datalengthSec), we_mean - we_std)
            self.plotExtra3.setPen(fn.mkPen('r'))
            self.p_plot.addItem(self.plotExtra)
            self.p_plot.addItem(self.plotExtra2)
            self.p_plot.addItem(self.plotExtra3)

        # plot spectrogram of only the filtered band:
        if self.extra == "Filtered spectrogram, new + AA" or self.extra == "Filtered spectrogram, new" or self.extra == "Filtered spectrogram, old":
            self.plotExtra = pg.ImageItem()
            self.p_plot.addItem(self.plotExtra)

            try:
                plotNodes = json.load(open(os.path.join(self.filtersDir, 'plotnodes.txt')))
            except:
                print("Couldn't load file, using default list")
                plotNodes = [48, 49, 52]

            # resample
            if self.sp.audioFormat.sampleRate() != 16000:
            #if self.sp.sampleRate != 16000:
                audiodata = self.sp.resample(16000)
            else:
                audiodata = self.sp.data

            WF = WaveletFunctions.WaveletFunctions(data=audiodata, wavelet='dmey2', maxLevel=5, samplerate=16000)

            # For now, not using antialiasFilter in the reconstructions as it's quick anyway
            if self.extra == "Filtered spectrogram, new + AA":
                WF.WaveletPacket(plotNodes, 'symmetric', True, True)
                C = WF.reconstructWP2(plotNodes[0], True, False)[:len(self.sp.data)]
                for node in plotNodes[1:]:
                    C = C + WF.reconstructWP2(node, True, False)[:len(C)]
            if self.extra == "Filtered spectrogram, new":
                WF.WaveletPacket(plotNodes, 'symmetric', False)
                C = WF.reconstructWP2(plotNodes[0], True, False)[:len(self.sp.data)]
                for node in plotNodes[1:]:
                    C = C + WF.reconstructWP2(node, True, False)[:len(C)]
            if self.extra == "Filtered spectrogram, old":
                WF.WaveletPacket(plotNodes, 'symmetric', False)
                C = WF.reconstructWP2(plotNodes[0], False)[:len(self.sp.data)]
                for node in plotNodes[1:]:
                    C = C + WF.reconstructWP2(node, False)[:len(C)]

            # Reconstructed signal was @ 16 kHz,
            # so we upsample to get equal sized spectrograms
            # TODO: Check this one
            if self.sp.audioFormat.sampleRate() != 16000:
            #if self.sp.sampleRate != 16000:
                C = self.sp.resample(self.sp.audioFormat.sampleRate(), C)
                #C = self.sp.resample(self.sp.sampleRate, C)
            tempsp = Spectrogram.Spectrogram()
            tempsp.data = C
            sgRaw = tempsp.spectrogram()
            sgHeightReduction = np.shape(sgRaw)[1]*16000//self.sp.audioFormat.sampleRate()
            #sgHeightReduction = np.shape(sgRaw)[1]*16000//self.sp.sampleRate
            sgRaw = sgRaw[:, :sgHeightReduction]
            maxsg = max(np.min(sgRaw), 1e-9)
            tempsp = np.abs(np.where(sgRaw == 0, 0.0, 10.0 * np.log10(sgRaw / maxsg)))

            self.plotExtra.setLookupTable(colourMaps.getLookupTable("Inferno"))
            self.plotExtra.setImage(tempsp)

            # set axis. Always at 0:sampleRate//2
            #minX, maxX = self.overviewImageRegion.getRegion()
            #self.p_plot.setXRange(minX, maxX, padding=0)
            MaxFreq = 8000
            height = 16000 // 2 / np.shape(tempsp)[1]
            SpecRange = MaxFreq/height
            #self.plotaxis.setTicks([[(0, 0.0), (SpecRange/4,round(MaxFreq/4000, 2)), (SpecRange/2,round(MaxFreq/2000, 2)), (3*SpecRange/4,round(3*MaxFreq/4000, 2)), (SpecRange,round(MaxFreq/1000, 2))]])
            ticks = [(0, 0.0), (SpecRange/4,round(MaxFreq/4000, 2)), (SpecRange/2,round(MaxFreq/2000, 2)), (3*SpecRange/4,round(3*MaxFreq/4000, 2)), (SpecRange,round(MaxFreq/1000, 2))]
            ticks = [[(tick[0], "%.1f" % tick[1] ) for tick in ticks]]
            self.plotaxis.setTicks(ticks)

            self.plotaxis.setLabel('kHz')

    def updateRegion_spec(self):
        """ This is the listener for when a segment box is changed in the spectrogram.
        It updates the position of the matching box, and also the text within it.
        """
        sender = self.sender()
        i = 0
        while self.listRectanglesa2[i] != sender and i<len(self.listRectanglesa2):
            i = i+1
        if i==len(self.listRectanglesa2):
            print("ERROR: segment not found!")
            return

        # update the overview boxes, step 1
        self.refreshOverviewWith(self.segments[i], delete=True)

        # fix the position of the text label
        if type(sender) == self.ROItype:
            # using box coordinates
            x1 = self.convertSpectoAmpl(sender.pos()[0])
            x2 = self.convertSpectoAmpl(sender.pos()[0]+sender.size()[0])
            self.segments[i][2] = self.convertYtoFreq(sender.pos()[1])
            self.segments[i][3] = self.convertYtoFreq(sender.pos()[1]+sender.size()[1])
            self.listLabels[i].setPos(sender.pos()[0], self.textpos)
        else:
            # using segment coordinates
            x1 = self.convertSpectoAmpl(sender.getRegion()[0])
            x2 = self.convertSpectoAmpl(sender.getRegion()[1])
            self.listLabels[i].setPos(sender.getRegion()[0], self.textpos)

        # update the amplitude segment
        self.listRectanglesa1[i].blockSignals(True)
        self.listRectanglesa1[i].setRegion([x1,x2])
        self.listRectanglesa1[i].blockSignals(False)
        self.segmentsToSave = True

        self.segments[i][0] = x1 + self.startRead
        self.segments[i][1] = x2 + self.startRead

        # update the overview boxes, step 2
        self.refreshOverviewWith(self.segments[i])

    def updateRegion_ampl(self):
        """ This is the listener for when a segment box is changed in the waveform plot.
        It updates the position of the matching box, and also the text within it.
        """
        sender = self.sender()
        i = 0
        while self.listRectanglesa1[i] != sender and i<len(self.listRectanglesa1):
            i = i+1
        if i==len(self.listRectanglesa1):
            print("Segment not found!")
        else:
            # update the overview boxes, step 1
            self.refreshOverviewWith(self.segments[i], delete=True)

            # fix the position of the text label
            x1 = self.convertAmpltoSpec(sender.getRegion()[0])
            x2 = self.convertAmpltoSpec(sender.getRegion()[1])
            self.listLabels[i].setPos(x1,self.textpos)

            # update the corresponding spectrogram segment
            self.listRectanglesa2[i].blockSignals(True)
            if type(self.listRectanglesa2[i]) == self.ROItype:
                # update the box
                y1 = self.listRectanglesa2[i].pos().y()
                y2 = self.listRectanglesa2[i].size().y()
                self.listRectanglesa2[i].setPos(pg.Point(x1,y1))
                self.listRectanglesa2[i].setSize(pg.Point(x2-x1,y2))
            else:
                # update the segment
                self.listRectanglesa2[i].setRegion([x1,x2])
            self.segmentsToSave = True
            self.listRectanglesa2[i].blockSignals(False)

            self.segments[i][0] = sender.getRegion()[0] + self.startRead
            self.segments[i][1] = sender.getRegion()[1] + self.startRead

            # update the overview boxes, step 2
            self.refreshOverviewWith(self.segments[i])

    def addRegularSegments(self):
        """ Add regular segments to the spectrogram.
        For most people drawProtocolMarks is better.
        """

        if self.box1id>-1:
            self.deselectSegment(self.box1id)
        segtimes = [(seg[0], seg[1]) for seg in self.segments]
        i = 0
        print("Adding segments (%d s every %d s)" %(self.config['protocolSize'], self.config['protocolInterval']))
        while i < self.segments.metadata["Duration"]:
            # check for segment presence in case of double click or other issues
            if len(segtimes)>0 and (i, i+self.config['protocolSize']) in segtimes:
                print("segment already exists, skipping")
            else:
                self.addSegment(i, i + self.config['protocolSize'], coordsAbsolute=True)
            i += self.config['protocolInterval']
        self.segmentsToSave = True

    def drawProtocolMarks(self):
        """ If check-ignore protocol is used, mark check-ignore limits.
        Also called when the relevant parameters are changed in interface settings.
        """

        # Clean old marks, if any
        if hasattr(self, 'protocolMarks'):
            for m in self.protocolMarks:
                self.p_spec.removeItem(m)
        self.protocolMarks = []

        if self.config['protocolOn']:
            linePen = pg.mkPen(self.config['protocolLineCol'], width=self.config['protocolLineWidth'])
            linestart = 0

            # Pages >1 start with an overlap zone, so need to offset marks:
            if self.currentFileSection > 0:
                linestart += self.config['fileOverlap']
            #while linestart < self.datalength/self.sp.sampleRate:
            while linestart < self.datalength/self.sp.audioFormat.sampleRate():
                lineend = min(self.datalength/self.sp.audioFormat.sampleRate(), linestart + self.config['protocolSize'])
                #lineend = min(self.datalength/self.sp.sampleRate, linestart + self.config['protocolSize'])
                line = pg.ROI(pos=(self.convertAmpltoSpec(linestart),0), size=(self.convertAmpltoSpec(lineend-linestart),0), movable=False, pen=linePen)
                lline = pg.InfiniteLine(pos=self.convertAmpltoSpec(linestart), angle=90, movable=False, pen=linePen)
                rline = pg.InfiniteLine(pos=self.convertAmpltoSpec(lineend), angle=90, movable=False, pen=linePen)
                self.protocolMarks.append(line)
                self.protocolMarks.append(lline)
                self.protocolMarks.append(rline)
                self.p_spec.addItem(line)
                self.p_spec.addItem(lline)
                self.p_spec.addItem(rline)
                linestart += self.config['protocolInterval']

    def refreshOverviewWith(self, segment, delete=False):
        """Recalculates the overview box colours and refreshes their display.
        To be used when segments are added, deleted or moved.
        Takes Segments as an input and either removes or adds to the box counters."""

        # Work out which overview segment this segment is in (could be more than one)
        # max/min deal with segments continuing past the edge of current page
        inds = max(0, int(self.convertAmpltoSpec(segment[0]-self.startRead) / self.widthOverviewSegment))
        inde = min(int(self.convertAmpltoSpec(segment[1]-self.startRead) / self.widthOverviewSegment), len(self.overviewSegments)-1)

        for label in segment[4]:
            if label["certainty"] == 0:
                # "red" label counter
                if delete:
                    self.overviewSegments[inds:inde+1,0] -= 1
                else:
                    self.overviewSegments[inds:inde+1,0] += 1
            elif label["certainty"] == 100:
                # "green" label counter
                if delete:
                    self.overviewSegments[inds:inde + 1, 1] -= 1
                else:
                    self.overviewSegments[inds:inde + 1, 1] += 1
            else:
                # "yellow" label counter
                if delete:
                    self.overviewSegments[inds:inde + 1, 2] -= 1
                else:
                    self.overviewSegments[inds:inde + 1, 2] += 1

        if np.any(self.overviewSegments<0):
            print("Warning: something went wrong with overview colors!")
            print(self.overviewSegments)

        # set the colour of these boxes in the overview
        for box in range(inds, inde + 1):
            if self.overviewSegments[box,0] > 0:
                self.SegmentRects[box].setBrush(self.ColourNone)
            elif self.overviewSegments[box,2] > 0:
                self.SegmentRects[box].setBrush(self.ColourPossible)
            elif self.overviewSegments[box,1] > 0:
                self.SegmentRects[box].setBrush(self.ColourNamed)
            else:
                # boxes w/o segments
                self.SegmentRects[box].setBrush(pg.mkBrush('w'))
            self.SegmentRects[box].update()
        # Deleting is almost always paired with redoing, so no need to refresh twice
        if not delete:
            self.refreshFileColor()

    def addSegment(self,startpoint,endpoint,y1=0,y2=0,species=[],saveSeg=True,index=-1,remaking=False,coordsAbsolute=False):
        """ When a new segment is created, does the work of creating it and connecting its
        listeners. Also updates the relevant overview segment.
        If a segment is too long for the current section, truncates it.
        Args:
        startpoint, endpoint - in secs, either from page start, or absolute (then set coordsAbsolute=True)
        y1, y2 should be the frequencies (between 0 and Fs//2)
        species - list of labels (including certainties, .data format)
        saveSeg - store the created segment on self.segments. Set to False when drawing the saved ones.
        remaking - can be turned to True to reuse existing graphics objects
        coordsAbsolute - set to True to accept start,end in absolute coords (from file start)
        """
        #print("Segment added at %d-%d, %d-%d" % (startpoint, endpoint, y1, y2))
        if self.box1id>-1:
            self.deselectSegment(self.box1id)

        # Make sure startpoint and endpoint are in the right order
        if startpoint > endpoint:
            temp = startpoint
            startpoint = endpoint
            endpoint = temp
        # same for freqs
        if y1 > y2:
            temp = y1
            y1 = y2
            y2 = temp
        # since we allow passing empty list here:
        if len(species) == 0:
            species = [{"species": "Don't Know", "certainty": 0, "filter": "M"}]
        else:
            species = copy.deepcopy(species)

        if coordsAbsolute:
            # convert from absolute times to relative-to-page times
            startpoint = startpoint - self.startRead
            endpoint = endpoint - self.startRead

        if not saveSeg:
            # check if this segment fits in the current spectrogram page
            if endpoint < 0 or startpoint > self.datalengthSec:
                print("Warning: a segment was not shown")
                show = False
            elif y1!=0 and y2!=0 and (y1 > self.sp.maxFreqShow or y2 < self.sp.minFreqShow):
                print("Warning: a segment was not shown")
                show = False
            else:
                show = True
        else:
            self.segmentsToSave = True
            show = True

        if saveSeg or show:
            # Create a Segment. This will check for errors and standardize the labels
            # Note: we convert time from _relative to page_ to _relative to file start_
            newSegment = Segment.Segment([startpoint+self.startRead, endpoint+self.startRead, y1, y2, species])

            # Add the segment to the data
            if saveSeg:
                self.segments.append(newSegment)

            self.refreshFileColor()

        if not show:
            # Add a None element into the array so that the correct boxids work
            if remaking:
                self.listRectanglesa1[index] = None
                self.listRectanglesa2[index] = None
                self.listLabels[index] = None
            else:
                self.listRectanglesa1.append(None)
                self.listRectanglesa2.append(None)
                self.listLabels.append(None)
            return
        # Otherwise, this is a visible segment.

        # --- rest of this function only does the graphics ---
        cert = min([lab["certainty"] for lab in species])
        if cert == 0:
            self.prevBoxCol = self.ColourNone
        elif cert == 100:
            self.prevBoxCol = self.ColourNamed
        else:
            self.prevBoxCol = self.ColourPossible

        self.refreshOverviewWith(newSegment)

        segsMovable = not (self.config['readOnly'])
        scenerect = QRectF(0, 0, np.shape(self.sg)[0], np.shape(self.sg)[1])

        # Add the segment in both plots and connect up the listeners
        p_ampl_r = SupportClasses_GUI.LinearRegionItem2(self, brush=self.prevBoxCol, movable=segsMovable, bounds=[0, self.datalengthSec])
        self.p_ampl.addItem(p_ampl_r, ignoreBounds=True)
        p_ampl_r.setRegion([startpoint, endpoint])
        p_ampl_r.sigRegionChangeFinished.connect(self.updateRegion_ampl)

        # Full-height segments:
        # SM: 30/7/22 Change here -- get rid of full-height segments for non-bats
        
        #if y1==0 and y2==0:
            # filled-in segments normally, transparent ones for bats:
            #p_spec_r = None
            #if not self.batmode: #and not self.config['transparentBoxes']:
                #p_spec_r = SupportClasses_GUI.LinearRegionItem2(self, brush=self.prevBoxCol, movable=segsMovable, bounds=[0, np.shape(self.sg)[0]])
            #else:
                #p_spec_r = SupportClasses_GUI.LinearRegionItem2(self, pen=pg.mkPen(self.prevBoxCol, width=6), movable=segsMovable, bounds=[0, np.shape(self.sg)[0]])
                #p_spec_r.setBrush(None)
            #p_spec_r.setRegion([self.convertAmpltoSpec(startpoint), self.convertAmpltoSpec(endpoint)])
        if self.batmode:
            # transparent segments for bats:
            p_spec_r = None
            p_spec_r = SupportClasses_GUI.LinearRegionItem2(self, pen=pg.mkPen(self.prevBoxCol, width=6), movable=segsMovable, bounds=[0, np.shape(self.sg)[0]])
            p_spec_r.setBrush(None)
            p_spec_r.setRegion([self.convertAmpltoSpec(startpoint), self.convertAmpltoSpec(endpoint)])
        # rectangle boxes:
        else:
            if y1==0 and y2==0:
                y2 = self.sp.audioFormat.sampleRate()//2
                #y2 = self.sp.sampleRate//2
            specy1 = self.convertFreqtoY(max(y1, self.sp.minFreqShow))
            specy2 = self.convertFreqtoY(min(y2, self.sp.maxFreqShow))
            startpointS = QPointF(self.convertAmpltoSpec(startpoint), specy1)
            endpointS = QPointF(self.convertAmpltoSpec(endpoint), specy2)
            p_spec_r = SupportClasses_GUI.ShadedRectROI(startpointS, endpointS - startpointS, movable=segsMovable, maxBounds=scenerect, parent=self)
            if self.config['transparentBoxes']:
                col = self.prevBoxCol.rgb()
                col = QtGui.QColor(col)
                col.setAlpha(255)
                p_spec_r.transparent = True
                p_spec_r.setBrush(None)
                p_spec_r.setHoverBrush(None)
                p_spec_r.setPen(pg.mkPen(col,width=1))
                col.setAlpha(100)
            else:
                p_spec_r.setBrush(pg.mkBrush(self.prevBoxCol))
                col = self.prevBoxCol
                col.setAlpha(180)
                p_spec_r.transparent = False
                p_spec_r.setHoverBrush(pg.mkBrush(col))
                p_spec_r.setPen(pg.mkPen(None))
                col.setAlpha(100)
        self.p_spec.addItem(p_spec_r, ignoreBounds=True)
        p_spec_r.sigRegionChangeFinished.connect(self.updateRegion_spec)

        # Put the text into the box
        label = pg.TextItem(text="new", color='k', anchor=(0,1))
        self.p_spec.addItem(label)
        label.setPos(self.convertAmpltoSpec(startpoint), self.textpos)

        # Add the segments to the relevent lists
        if remaking:
            self.listRectanglesa1[index] = p_ampl_r
            self.listRectanglesa2[index] = p_spec_r
            self.listLabels[index] = label
        else:
            self.listRectanglesa1.append(p_ampl_r)
            self.listRectanglesa2.append(p_spec_r)
            self.listLabels.append(label)

        # mark this as the current segment
        if index>-1:
            self.box1id = index
        else:
            self.box1id = len(self.listLabels) - 1

        # update its displayed label
        self.updateText(self.box1id)

    def selectSegment(self, boxid):
        """ Changes the segment colors and enables playback buttons."""
        # print("selected %d" % boxid)
        self.box1id = boxid
        self.refreshSegmentControls()

        # Helps dealing with edge effects for various review functions
        if boxid>len(self.listRectanglesa1) or self.listRectanglesa1[boxid] is None:
            return

        self.prevBoxCol = self.listRectanglesa1[boxid].brush.color()
        brush = fn.mkBrush(self.ColourSelected)
        if self.listRectanglesa1[boxid] is not None and self.listRectanglesa2[boxid] is not None:
            self.listRectanglesa1[boxid].setBrush(brush)
            self.listRectanglesa2[boxid].setBrush(brush)
            self.listRectanglesa1[boxid].setHoverBrush(brush)
            self.listRectanglesa2[boxid].setHoverBrush(brush)

            self.listRectanglesa1[boxid].update()
            self.listRectanglesa2[boxid].update()

        # Show details of selection
        self.segInfo.setText(self.segments[boxid].infoString())

    def deselectSegment(self, boxid):
        """ Restores the segment colors and disables playback buttons."""
        # print("deselected %d" % boxid)
        self.box1id = -1
        self.refreshSegmentControls()
        # Hide details of selection
        self.segInfo.setText("")

        # Helps dealing with edge effects for various review functions
        if boxid>len(self.listRectanglesa1) or self.listRectanglesa1[boxid] is None:
            return

        # Filled-in segments normally, transparent ones for bats:
        # (This is somewhat convoluted to keep amplitude segments updated even in bat mode,
        # as they are used for tracking prevBoxCol)
        col = self.prevBoxCol
        col.setAlpha(100)
        self.listRectanglesa1[boxid].setBrush(fn.mkBrush(col))
        if not self.batmode:
            self.listRectanglesa2[boxid].setBrush(fn.mkBrush(col))
        else:
            self.listRectanglesa2[boxid].setBrush(None)
            self.listRectanglesa2[boxid].setPen(self.prevBoxCol, width=6)

        col.setAlpha(180)
        self.listRectanglesa1[boxid].setHoverBrush(fn.mkBrush(col))
        self.listRectanglesa2[boxid].setHoverBrush(fn.mkBrush(col))
        col.setAlpha(100)

        if self.config['transparentBoxes'] and type(self.listRectanglesa2[boxid]) == self.ROItype:
            col = self.prevBoxCol.rgb()
            col = QtGui.QColor(col)
            col.setAlpha(255)
            self.listRectanglesa2[boxid].setBrush(pg.mkBrush(None))
            self.listRectanglesa2[boxid].setHoverBrush(pg.mkBrush(None))
            self.listRectanglesa2[boxid].setPen(col,width=1)
            col.setAlpha(100)

        self.listRectanglesa1[boxid].update()
        self.listRectanglesa2[boxid].update()

### mouse management

    def mouseMoved(self,evt):
        """ Listener for mouse moves.
        If the user moves the mouse in the spectrogram, print the time, frequency, power for the mouse location. """
        if not self.showPointerDetails.isChecked():
            return
        elif self.p_spec.sceneBoundingRect().contains(evt):
            mousePoint = self.p_spec.mapSceneToView(evt)
            indexx = int(mousePoint.x())
            indexy = int(mousePoint.y())
            if indexx > 0 and indexx < np.shape(self.sg)[0] and indexy > 0 and indexy < np.shape(self.sg)[1]:
                time = self.convertSpectoAmpl(mousePoint.x()) + self.currentFileSection * self.config['maxFileShow'] - (self.currentFileSection>0)*self.config['fileOverlap'] + self.startTime
                seconds = time % 60
                minutes = (time//60) % 60
                hours = (time//3600) % 24
                if hours>0:
                    self.pointData.setText('time=%.2d:%.2d:%05.2f (hh:mm:ss.ms), freq=%0.1f (Hz), power=%0.1f (dB)' % (hours,minutes,seconds, mousePoint.y() * self.sp.audioFormat.sampleRate()//2 / np.shape(self.sg)[1] + self.sp.minFreqShow, self.sg[indexx, indexy]))
                    #self.pointData.setText('time=%.2d:%.2d:%05.2f (hh:mm:ss.ms), freq=%0.1f (Hz), power=%0.1f (dB)' % (hours,minutes,seconds, mousePoint.y() * self.sp.sampleRate//2 / np.shape(self.sg)[1] + self.sp.minFreqShow, self.sg[indexx, indexy]))
                else:
                    self.pointData.setText('time=%.2d:%05.2f (mm:ss.ms), freq=%0.1f (Hz), power=%0.1f (dB)' % (minutes,seconds, mousePoint.y() * self.sp.audioFormat.sampleRate()//2 / np.shape(self.sg)[1] + self.sp.minFreqShow, self.sg[indexx, indexy]))
                    #self.pointData.setText('time=%.2d:%05.2f (mm:ss.ms), freq=%0.1f (Hz), power=%0.1f (dB)' % (minutes,seconds, mousePoint.y() * self.sp.sampleRate//2 / np.shape(self.sg)[1] + self.sp.minFreqShow, self.sg[indexx, indexy]))

    def mouseClicked_ampl(self,evt):
        """ Listener for if the user clicks on the amplitude plot.
        If there is a box selected, get its colour.
        If the user has clicked inside the scene, they could be
        (1) clicking in an already existing box -> select it
        (2) clicking anywhere else -> start a box
        (3) clicking a second time to finish a box -> create the segment
        """
        pos = evt.scenePos()

        # If any box is selected, deselect (wherever clicked)
        wasSelected = self.box1id
        if self.box1id>-1:
            self.deselectSegment(self.box1id)

        # If clicked inside scene:
        if self.p_ampl.sceneBoundingRect().contains(pos):
            mousePoint = self.p_ampl.mapSceneToView(pos)

            # If this is the second click and not a box, close the segment
            if self.started:
                # Can't finish boxes in ampl plot
                if self.config['specMouseAction']>1:
                    if self.startedInAmpl:
                        # started in ampl and finish in ampl,
                        # so continue as usual to draw a segment
                        pass
                    else:
                        # started in spec so ignore this bullshit
                        return

                # remove the drawing box:
                self.p_spec.removeItem(self.vLine_s)
                self.p_ampl.removeItem(self.vLine_a)
                self.p_ampl.removeItem(self.drawingBox_ampl)
                self.p_spec.removeItem(self.drawingBox_spec)
                # disconnect GrowBox listeners, leave the position listener
                self.p_ampl.scene().sigMouseMoved.disconnect()
                self.p_spec.scene().sigMouseMoved.disconnect()
                if self.showPointerDetails.isChecked():
                    self.p_spec.scene().sigMouseMoved.connect(self.mouseMoved)

                # If the user has pressed shift, copy the last species and don't use the context menu
                # If they pressed Control, add ? to the names
                # Possibly, if they pressed the Windows key, use call type menu
                modifiers = QApplication.keyboardModifiers()
                if modifiers == Qt.KeyboardModifier.ShiftModifier:
                    self.addSegment(self.start_ampl_loc, max(mousePoint.x(),0.0),species=self.lastSpecies)
                elif modifiers == Qt.KeyboardModifier.ControlModifier:
                    self.addSegment(self.start_ampl_loc,max(mousePoint.x(),0.0))
                    # Context menu
                    self.fillBirdList(unsure=True)
                    self.menuBirdList.popup(QPoint(int(evt.screenPos().x()), int(evt.screenPos().y())))
                elif modifiers == Qt.KeyboardModifier.MetaModifier:
                    # TODO: SRM: Check
                    # TODO: Check fillBirdList and toggleViewSp and whether they compete
                    self.addSegment(self.start_ampl_loc, max(mousePoint.x(),0.0),species=self.lastSpecies)
                    if self.viewCallType is False:
                        self.viewCallType = True
                        # Calltype context menu
                        self.fillBirdList()
                        self.menuBirdList.popup(QPoint(int(evt.screenPos().x()), int(evt.screenPos().y())))
                        self.viewCallType = False
                    else:
                        self.menuBirdList.popup(QPoint(int(evt.screenPos().x()), int(evt.screenPos().y())))
                else:
                    self.addSegment(self.start_ampl_loc,max(mousePoint.x(),0.0))
                    # Context menu
                    self.fillBirdList()
                    self.menuBirdList.popup(QPoint(int(evt.screenPos().x()), int(evt.screenPos().y())))
                self.p_ampl.setFocus()

                # The new segment is now selected and can be played
                self.selectSegment(self.box1id)
                self.started = not(self.started)
                self.startedInAmpl = False

                # Reset cursor to not drawing (or leave as drawing if LMB draws)
                if self.MouseDrawingButton==Qt.MouseButton.RightButton:
                    self.p_ampl.unsetCursor()
                    self.specPlot.unsetCursor()
            # If this is the first click:
            else:
                # If this is right click (drawing mode):
                # (or whatever you want)
                if evt.button() == self.MouseDrawingButton:
                    if self.config['readOnly']:
                        return
                    # this would prevent starting boxes in ampl plot
                    # if self.config['specMouseAction']>1:
                    #    return

                    nonebrush = self.ColourNone
                    self.start_ampl_loc = mousePoint.x()

                    # spectrogram plot bar and mouse followers:
                    self.vLine_s = pg.InfiniteLine(angle=90, movable=False,pen={'color': 'r', 'width': 3})
                    self.p_spec.addItem(self.vLine_s, ignoreBounds=True)
                    self.vLine_s.setPos(self.convertAmpltoSpec(self.start_ampl_loc))

                    self.drawingBox_spec = pg.LinearRegionItem(brush=nonebrush)
                    self.p_spec.addItem(self.drawingBox_spec, ignoreBounds=True)
                    self.drawingBox_spec.setRegion([self.convertAmpltoSpec(self.start_ampl_loc), self.convertAmpltoSpec(self.start_ampl_loc)])
                    self.p_spec.scene().sigMouseMoved.connect(self.GrowBox_spec)

                    # amplitude plot bar and mouse followers:
                    self.vLine_a = pg.InfiniteLine(angle=90, movable=False,pen={'color': 'r', 'width': 3})
                    self.p_ampl.addItem(self.vLine_a, ignoreBounds=True)
                    self.vLine_a.setPos(self.start_ampl_loc)

                    self.drawingBox_ampl = pg.LinearRegionItem(brush=nonebrush)
                    self.p_ampl.addItem(self.drawingBox_ampl, ignoreBounds=True)
                    self.drawingBox_ampl.setRegion([self.start_ampl_loc, self.start_ampl_loc])
                    self.p_ampl.scene().sigMouseMoved.connect(self.GrowBox_ampl)

                    self.started = not (self.started)
                    self.startedInAmpl = True

                    # Force cursor to drawing
                    self.p_ampl.setCursor(QtGui.QCursor(QPixmap('img/cursor.bmp'), 0, 0))
                    self.specPlot.setCursor(QtGui.QCursor(QPixmap('img/cursor.bmp'), 0, 0))
                # if this is left click (selection mode):
                else:
                    # Check if the user has clicked in a box
                    # Note: Returns the first one it finds, i.e. the newest
                    box1id = -1
                    for count in range(len(self.listRectanglesa1)):
                        if self.listRectanglesa1[count] is not None:
                            x1, x2 = self.listRectanglesa1[count].getRegion()
                            if x1 <= mousePoint.x() and x2 >= mousePoint.x():
                                box1id = count
                                break

                    # User clicked in a segment:
                    if box1id > -1:
                        # select the segment:
                        self.selectSegment(box1id)
                        # is it the first click on this segment?
                        if wasSelected==box1id:
                            # popup dialog
                            modifiers = QApplication.keyboardModifiers()
                            if modifiers == Qt.KeyboardModifier.ControlModifier:
                                self.fillBirdList(unsure=True)
                            elif modifiers == Qt.KeyboardModifier.MetaModifier:
                                # TODO: SRM: Check
                                # TODO: Check fillBirdList and toggleViewSp and whether they compete
                                self.addSegment(self.start_ampl_loc, max(mousePoint.x(),0.0),species=self.lastSpecies)
                                if self.viewCallType is False:
                                    self.viewCallType = True
                                    # Calltype context menu
                                    self.fillBirdList()
                                    self.menuBirdList.popup(QPoint(int(evt.screenPos().x()), int(evt.screenPos().y())))
                                    self.viewCallType = False
                                else:
                                    self.menuBirdList.popup(QPoint(int(evt.screenPos().x()), int(evt.screenPos().y())))
                            else:
                                self.fillBirdList()
                            self.menuBirdList.popup(QPoint(int(evt.screenPos().x()), int(evt.screenPos().y())))

    def mouseClicked_spec(self,evt):
        """ Listener for if the user clicks on the spectrogram plot.
        See the amplitude version (mouseClicked_ampl()) for details. Although much of the code is a repeat,
        it is separated for clarity.
        """
        pos = evt.scenePos()

        # if any box is selected, deselect (wherever clicked)
        wasSelected = self.box1id
        if self.box1id>-1:
            self.deselectSegment(self.box1id)

        # When drawing boxes near scene borders, it's easy to release mouse outside scene,
        # and all the dragging gets messed up then. We map such cases to closest scene positions here:
        if self.started and self.config['specMouseAction']==3:
            bounds = self.p_spec.sceneBoundingRect()
            if not bounds.contains(pos):
                newX = min(bounds.right(), max(bounds.left(), pos.x()))
                newY = min(bounds.bottom(), max(bounds.top(), pos.y()))
                pos.setX(newX)
                pos.setY(newY)

        # If clicked inside scene:
        if self.p_spec.sceneBoundingRect().contains(pos):
            mousePoint = self.p_spec.mapSceneToView(pos)

            # If this is the second click, close the segment/box
            # Note: can finish segment with either left or right click
            if self.started:
                if self.config['specMouseAction']>1 and self.startedInAmpl:
                    # Started in ampl, and spec is used for boxes, so can't continue here
                    return

                # Remove the drawing box:
                if not self.config['specMouseAction']>1:
                    self.p_spec.removeItem(self.vLine_s)
                    self.p_ampl.scene().sigMouseMoved.disconnect()
                self.p_ampl.removeItem(self.vLine_a)
                self.p_ampl.removeItem(self.drawingBox_ampl)
                self.p_spec.removeItem(self.drawingBox_spec)
                # Disconnect GrowBox listeners, leave the position listener
                self.p_spec.scene().sigMouseMoved.disconnect()
                if self.showPointerDetails.isChecked():
                    self.p_spec.scene().sigMouseMoved.connect(self.mouseMoved)
                # Reset the trackers
                self.started = not(self.started)
                self.startedInAmpl = False

                # Reset cursor to not drawing (or leave as drawing if LMB draws)
                if self.MouseDrawingButton==Qt.MouseButton.RightButton:
                    self.p_ampl.unsetCursor()
                    self.specPlot.unsetCursor()

                # Pass either default y coords or box limits:
                x1 = self.start_ampl_loc
                x2 = self.convertSpectoAmpl(max(mousePoint.x(), 0.0))
                # Could add this check if right edge seems dangerous:
                # endx = min(x2, np.shape(self.sg)[0]+1)
                if self.config['specMouseAction']>1:
                    y1 = self.start_spec_y
                    y2 = mousePoint.y()
                    miny = self.convertFreqtoY(self.sp.minFreqShow)
                    maxy = self.convertFreqtoY(self.sp.maxFreqShow)
                    y1 = min(max(miny, y1), maxy)
                    y2 = min(max(miny, y2), maxy)

                    # When dragging, can sometimes make boxes by mistake, which is annoying.
                    # To avoid, check that the box isn't too small
                    if np.abs((x2-x1)*(y2-y1)) < self.minboxsize:
                        print("Small box detected, ignoring")
                        return

                    y1 = self.convertYtoFreq(y1)
                    y2 = self.convertYtoFreq(y2)
                else:
                    y1 = 0
                    y2 = 0

                # If the user has pressed shift, copy the last species and don't use the context menu
                # If they pressed Control, add ? to the names
                # NOTE: Ctrl+Shift combo doesn't have a Qt modifier and is ignored.
                # TODO: Could make all these options
                modifiers = QApplication.keyboardModifiers()
                if modifiers == Qt.KeyboardModifier.ShiftModifier:
                    self.addSegment(x1, x2, y1, y2, species=self.lastSpecies)
                elif modifiers == Qt.KeyboardModifier.ControlModifier:
                    self.addSegment(x1, x2, y1, y2)
                    # Context menu
                    self.fillBirdList(unsure=True)
                    self.menuBirdList.popup(QPoint(int(evt.screenPos().x()), int(evt.screenPos().y())))
                elif modifiers == Qt.KeyboardModifier.MetaModifier:
                    # TODO: SRM: Check
                    # TODO: Check fillBirdList and toggleViewSp and whether they compete
                    self.addSegment(self.start_ampl_loc, max(mousePoint.x(),0.0),species=self.lastSpecies)
                    if self.viewCallType is False:
                        self.viewCallType = True
                        # Calltype context menu
                        self.fillBirdList()
                        self.menuBirdList.popup(QPoint(int(evt.screenPos().x()), int(evt.screenPos().y())))
                        self.viewCallType = False
                    else:
                        self.menuBirdList.popup(QPoint(int(evt.screenPos().x()), int(evt.screenPos().y())))
                else:
                    self.addSegment(x1, x2, y1, y2)
                    # Context menu
                    self.fillBirdList()
                    self.menuBirdList.popup(QPoint(int(evt.screenPos().x()), int(evt.screenPos().y())))
                self.p_spec.setFocus()

                # Select the new segment/box
                self.selectSegment(self.box1id)

            # If this is the first click:
            else:
                # If this is right click (drawing mode):
                if evt.button() == self.MouseDrawingButton:
                    if self.config['readOnly']:
                        return
                    nonebrush = self.ColourNone
                    self.start_ampl_loc = self.convertSpectoAmpl(mousePoint.x())
                    self.start_spec_y = mousePoint.y()

                    # Start a new box:
                    if self.config['specMouseAction']>1:
                        # spectrogram mouse follower box:
                        startpointS = QPointF(mousePoint.x(), mousePoint.y())
                        endpointS = QPointF(mousePoint.x(), mousePoint.y())

                        self.drawingBox_spec = SupportClasses_GUI.ShadedRectROI(startpointS, endpointS - startpointS, invertible=True)
                        self.drawingBox_spec.setBrush(nonebrush)
                        self.p_spec.addItem(self.drawingBox_spec, ignoreBounds=True)
                        self.p_spec.scene().sigMouseMoved.connect(self.GrowBox_spec)
                    # Start a new segment:
                    else:
                        # Spectrogram bar and mouse follower:
                        self.vLine_s = pg.InfiniteLine(angle=90, movable=False,pen={'color': 'r', 'width': 3})
                        self.p_spec.addItem(self.vLine_s, ignoreBounds=True)
                        self.vLine_s.setPos(mousePoint.x())

                        self.drawingBox_spec = pg.LinearRegionItem(brush=nonebrush)
                        self.p_spec.addItem(self.drawingBox_spec, ignoreBounds=True)
                        self.drawingBox_spec.setRegion([mousePoint.x(),mousePoint.x()])
                        self.p_spec.scene().sigMouseMoved.connect(self.GrowBox_spec)
                        # NOTE: only in segment mode react to movement over ampl plot:
                        self.p_ampl.scene().sigMouseMoved.connect(self.GrowBox_ampl)

                    # For box and segment - amplitude plot bar:
                    self.vLine_a = pg.InfiniteLine(angle=90, movable=False,pen={'color': 'r', 'width': 3})
                    self.p_ampl.addItem(self.vLine_a, ignoreBounds=True)
                    self.vLine_a.setPos(self.start_ampl_loc)

                    self.drawingBox_ampl = pg.LinearRegionItem(brush=nonebrush)
                    self.p_ampl.addItem(self.drawingBox_ampl, ignoreBounds=True)
                    self.drawingBox_ampl.setRegion([self.start_ampl_loc, self.start_ampl_loc])

                    self.started = not (self.started)
                    self.startedInAmpl = False

                    # Force cursor to drawing
                    self.p_ampl.setCursor(QtGui.QCursor(QPixmap('img/cursor.bmp'), 0, 0))
                    self.specPlot.setCursor(QtGui.QCursor(QPixmap('img/cursor.bmp'), 0, 0))
                # if this is left click (selection mode):
                else:
                    # Check if the user has clicked in a box
                    # NOTE: Returns the first one it finds, i.e. the newest
                    box1id = -1
                    for count in range(len(self.listRectanglesa2)):
                        if type(self.listRectanglesa2[count]) == self.ROItype and self.listRectanglesa2[count] is not None:
                            x1 = self.listRectanglesa2[count].pos().x()
                            y1 = self.listRectanglesa2[count].pos().y()
                            x2 = x1 + self.listRectanglesa2[count].size().x()
                            y2 = y1 + self.listRectanglesa2[count].size().y()
                            if x1 <= mousePoint.x() and x2 >= mousePoint.x() and y1 <= mousePoint.y() and y2 >= mousePoint.y():
                                box1id = count
                                break
                        elif self.listRectanglesa2[count] is not None:
                            x1, x2 = self.listRectanglesa2[count].getRegion()
                            if x1 <= mousePoint.x() and x2 >= mousePoint.x():
                                box1id = count
                                break

                    # User clicked in a segment:
                    if box1id > -1:
                        #print("segment selected")
                        # Select the segment:
                        self.selectSegment(box1id)
                        # If this segment is clicked again, pop up bird menu:
                        if wasSelected==box1id:
                            modifiers = QApplication.keyboardModifiers()
                            if modifiers == Qt.KeyboardModifier.ControlModifier:
                                self.fillBirdList(unsure=True)
                            elif modifiers == Qt.KeyboardModifier.MetaModifier:
                                # TODO: Check fillBirdList and toggleViewSp and whether they compete
                                self.addSegment(self.start_ampl_loc, max(mousePoint.x(),0.0),species=self.lastSpecies)
                                if self.viewCallType is False:
                                    self.viewCallType = True
                                    self.fillBirdList()
                                    self.menuBirdList.popup(QPoint(int(evt.screenPos().x()), int(evt.screenPos().y())))
                                    self.viewCallType = False
                                else:
                                    self.fillBirdList()
                                    self.menuBirdList.popup(QPoint(int(evt.screenPos().x()), int(evt.screenPos().y())))
                            else:
                                self.fillBirdList()
                                self.menuBirdList.popup(QPoint(int(evt.screenPos().x()), int(evt.screenPos().y())))

    def GrowBox_ampl(self,pos):
        """ Listener for when a segment is being made in the amplitude plot.
        Makes the blue box that follows the mouse change size. """
        if self.p_ampl.sceneBoundingRect().contains(pos):
            mousePoint = self.p_ampl.mapSceneToView(pos)
            self.drawingBox_ampl.setRegion([self.start_ampl_loc, mousePoint.x()])
            self.drawingBox_spec.setRegion([self.convertAmpltoSpec(self.start_ampl_loc), self.convertAmpltoSpec(mousePoint.x())])

    def GrowBox_spec(self, pos):
        """ Listener for when a segment is being made in the spectrogram plot.
        Makes the blue box that follows the mouse change size. """
        # When dragging spectrogram boxes near scene edges, we have special rules
        # to keep tracking the potential box
        if self.config['specMouseAction']==3:
            bounds = self.p_spec.sceneBoundingRect()
            if not bounds.contains(pos):
                newX = min(bounds.right(), max(bounds.left(), pos.x()))
                newY = min(bounds.bottom(), max(bounds.top(), pos.y()))
                pos.setX(newX)
                pos.setY(newY)

        if self.p_spec.sceneBoundingRect().contains(pos):
            mousePoint = self.p_spec.mapSceneToView(pos)
            self.drawingBox_ampl.setRegion([self.start_ampl_loc, self.convertSpectoAmpl(mousePoint.x())])
            if self.config['specMouseAction']>1 and not self.startedInAmpl:
                # Making a box
                posY = mousePoint.y() - self.start_spec_y
                self.drawingBox_spec.setSize([mousePoint.x()-self.convertAmpltoSpec(self.start_ampl_loc), posY])
            else:
                # Making a segment
                self.drawingBox_spec.setRegion([self.convertAmpltoSpec(self.start_ampl_loc), mousePoint.x()])

    def birdSelectedMenu(self,birditem):
        """ Collects the label for a bird from the context menu and processes it.
        Has to update the overview segments in case their colour should change.
        Copes with two level names (with a > in).
        Also handles getting the name through a message box if necessary.
        """
        if type(birditem) is not str:
            birdname = birditem.text()
        else:
            birdname = birditem
        if birdname is None or birdname=='':
            return
                
        # special dialog for manual name entry
        if birdname == 'Other':
            # Ask the user for the new name, and save it
            birdname, ok = QInputDialog.getText(self, 'Bird name', 'Enter the bird name as genus (species)')
            if not ok:
                return

            birdname = str(birdname).title()
            # splits "A (B)", with B optional, into groups A and B
            match = re.fullmatch(r'(.*?)(?: \((.*)\))?', birdname)
            if not match:
                print("ERROR: provided name %s does not match format requirements" % birdname)
                return

            #if birdname.lower()=="don't know" or birdname.lower()=="other":
            if birdname.lower()=="don't know" or birdname.lower()=="other" or birdname.lower()=="(other)":
                print("ERROR: provided name %s is reserved, cannot create" % birdname)
                return

            if "?" in birdname:
                print("ERROR: provided name %s contains reserved symbol '?'" % birdname)
                return

            if len(birdname)==0 or len(birdname)>150:
                print("ERROR: provided name appears to be too short or too long")
                return

            twolevelname = '>'.join(match.groups(default=''))
            if birdname in self.longBirdList or twolevelname in self.longBirdList:
                # bird is already listed
                print("Warning: not adding species %s as it is already present" % birdname)
                return

            # maybe the genus is already listed?
            index = self.model.findItems(match.group(1), Qt.MatchFlag.MatchFixedString)
            if len(index) == 0:
                # Genus isn't in list
                item = QStandardItem(match.group(1))
                item.setSelectable(True)
                self.model.appendRow(item)
                # store as typed
                nametostore = birdname
            else:
                # Get the species item
                item = index[0]
                if match.group(2) is None:
                    print("ERROR: genus %s already exists, please provide species as well" % match.group(1))
                    return
                # Store in two-level format
                nametostore = twolevelname
                subitem = QStandardItem(match.group(2))
                item.setSelectable(False)
                item.appendRow(subitem)
                subitem.setSelectable(True)

            # update the main list:
            self.longBirdList.append(nametostore)
            self.longBirdList.remove('Unidentifiable')
            self.longBirdList = sorted(self.longBirdList, key=str.lower)
            self.longBirdList.append('Unidentifiable')
            self.ConfigLoader.blwrite(self.longBirdList, self.config['BirdListLong'], self.configdir)

        # parse birdname to certainty
        if birdname=="Don't Know":
            species = birdname
            certainty = 0
            self.prevBoxCol = self.ColourNone
        elif birdname[-1] == '?':
            species = birdname[:-1]
            certainty = 50
            self.prevBoxCol = self.ColourPossible
        else:
            species = birdname
            certainty = 100
            self.prevBoxCol = self.ColourNamed

        workingSeg = self.segments[self.box1id]
        self.refreshOverviewWith(workingSeg, delete=True)

        # Toggle the actual label in the segment list
        if workingSeg.hasLabel(species, certainty):
            workingSeg.removeLabel(species, certainty)
        else:
            # in case the only label so far was Don't Know,
            # change it to the new bird (to not waste time unticking it)
            if workingSeg.keys == [("Don't Know", 0)]:
                workingSeg.addLabel(species, certainty, filter="M")
                workingSeg.removeLabel("Don't Know", 0)
                # also need to untick that context menu item manually
                for act in self.menuBirdList.actions() + self.menuBirdOther.actions():
                    if act.text()=="Don't Know":
                        act.setChecked(False)
            else:
                # in single-bird mode, just remove the current label:
                workingSeg.addLabel(species, certainty, filter="M")
                if not self.multipleBirds:
                    workingSeg.removeLabel(workingSeg[4][0]["species"], workingSeg[4][0]["certainty"])

        # Put the selected bird name at the top of the list
        if self.config['ReorderList']:
            if self.batmode:
                # Either move the label to the top of the list, or delete the last
                if species in self.batList:
                    self.batList.remove(species)
                else:
                    del self.batList[-1]
                self.batList.insert(0,species)
            else:
                # Either move the label to the top of the list, or delete the last
                if species in self.shortBirdList:
                    self.shortBirdList.remove(species)
                else:
                    del self.shortBirdList[-1]
                self.shortBirdList.insert(0,species)

        # Refresh overview boxes after all updates:
        self.refreshOverviewWith(workingSeg)

        # Store the species in case the user wants it for the next segment
        # TODO SRM: correct certainty and filter?
        self.lastSpecies = [{"species": species, "certainty": 100, "filter": "M"}]
        self.updateText()
        self.updateColour()
        self.segInfo.setText(workingSeg.infoString())
        self.segmentsToSave = True

        QApplication.processEvents()

    def callSelectedMenu(self, species, callname):
        """ Simplified version of the above for dealing with calltype selection
        from the popup context menu. """

        if callname is None or callname=="" or callname=="Any":
            return

        if callname == 'Other': 
            # Ask the user for the new name, and save it
            callname, ok = QInputDialog.getText(self, 'Call type', 'Enter a label for this call type ')
            if not ok:
                return

            callname = str(callname).title()
            # splits "A (B)", with B optional, into groups A and B
            match = re.fullmatch(r'(.*?)(?: \((.*)\))?', callname)
            if not match:
                print("ERROR: provided name %s does not match format requirements" % callname)
                return

            if callname.lower()=="don't know" or callname.lower()=="other" or callname.lower()=="(other)":
                print("ERROR: provided name %s is reserved, cannot create" % callname)
                return

            if "?" in callname:
                print("ERROR: provided name %s contains reserved symbol '?'" % callname)
                return

            if len(callname)==0 or len(callname)>150:
                print("ERROR: provided name appears to be too short or too long")
                return
            
            self.possibleCTs = set()
            self.filters = []
            for filt in self.FilterDicts.values():
                if filt["species"]==species:
                    self.possibleCTs.update([subf["calltype"] for subf in filt["Filters"]])
                    self.filters.append(filt)
            
            if callname in self.possibleCTs:
                # call is already listed
                print("Warning: not adding call type %s as it is already present" % callname)
                return

            # TODO: Needs a bit of thought, since need to find (or create) a filter. And there might be more than one.
            # SRM: I think this is OK-ish. Now for the DialogsTraining
            if len(self.filters) == 0:
                # There wasn't a filter. Make one. Ask for name, or just use default?
                speciesData = {"species": species, "method": None, "SampleRate": self.sp.audioFormat.sampleRate(), "Filters": []}
                #speciesData = {"species": spmenu, "method": None, "SampleRate": self.sp.sampleRate, "Filters": []}
                filename = os.path.join(self.filtersDir, species+'.txt')
                print("no filter",filename)
                newfilter = speciesData
            elif len(self.filters) == 1:
                # There is one filter, so add the new calltype there? Or ask?
                filename = os.path.join(self.filtersDir,self.filters[0]["species"]+'.txt')
                print("one filter: ",filename)
                newfilter = self.filters[0]
                #print(newfilter)
            else:
                # TODO !!
                # More than one, need to ask
                print("filters: ",self.filters[0]["species"])
                filename = os.path.join(self.filtersDir,self.filters[0]["species"]+'.txt')
                newfilter = self.filters[0]

            # If not, ask, then make it
            #for filt in self.FilterDicts.values():
            # Add the new subfilter, with just a call type name
            newSubfilt = {'calltype': callname}
            newfilter["Filters"].append(newSubfilt)

            print(filename)
            f = open(filename, 'w')
            f.write(json.dumps(newfilter))
            f.close()

            #self.speciesData["Filters"].append(newSubfilt)
            # Save it
            
        workingSeg = self.segments[self.box1id]

        # TODO: SRM: Might not be first label
        #for lab in workingSeg[4]:
            #if lab["species"] == species:
                #lab["calltype"] = callname
        #workingSeg.addLabel(species, 101, filter="M", calltype=callname)
        if 'calltype' not in workingSeg[4]:
            workingSeg.extendLabel(species,100,callname)
        else:
            for lab in workingSeg[4]:
                if lab["species"]==species:
                    workingSeg[4][lab].update({"filter": "M", "certainty": 100, "calltype": callname})

        # Store the species in case the user wants it for the next segment
        self.lastSpecies = [{"species": species, "certainty": 100, "filter": "M", "calltype": callname}]
        self.updateText()
        self.segInfo.setText(workingSeg.infoString())
        self.segmentsToSave = True
    
    def batSelected(self,species):
        self.birdSelectedMenu(species)
    
    def birdAndCallSelected(self,species,call):
        self.birdSelectedMenu(species)
        if call is None:
            raise ValueError("call has not been provided!")
        self.callSelectedMenu(species,call)

        print("self.multipleBirds",self.multipleBirds)
        if not self.multipleBirds:
            self.menuBirdList.hide()

    def updateText(self, segID=None):
        """ When the user sets or changes the name in a segment, update the text label.
            Only requires the segment ID, or defaults to the selected one, and
            will read the label from it."""
        if segID is None:
            segID = self.box1id
        seg = self.segments[segID]

        if not self.viewCallType:
            # produce text from list of dicts
            text = []
            for lab in seg[4]:
                if lab["certainty"] == 50:
                    text.append(lab["species"] + '?')
                else:
                    text.append(lab["species"])
            text = ','.join(text)
        else:
            text = []
            for lab in seg[4]:
                if "calltype" in lab:
                    text.append(lab["calltype"])
                else:
                    text.append("(Other)")
            text = ','.join(text)

        # update the label
        self.listLabels[segID].setText(text,'k')
        self.listLabels[segID].update()
        QApplication.processEvents()

    def updateColour(self, segID=None):
        """ Updates the colour of a segment (useful for reviewing segments, for example).
            Only requires the segment ID, or defaults to the selected one, and
            will determine the color from it.
        """
        if segID is None:
            segID = self.box1id
        cert = min([lab["certainty"] for lab in self.segments[segID][4]])

        if cert == 0:
            brush = self.ColourNone
        elif cert == 100:
            brush = self.ColourNamed
        else:
            brush = self.ColourPossible

        # If we're updating the currently selected segment,
        # we should just store the new color (it'll be used on deselecting)
        if self.box1id == segID:
            self.prevBoxCol = brush
            # Except in batmode where lines are still visible when selected:
            if self.batmode:
                self.listRectanglesa2[segID].setPen(brush, width=6)
        # Otherwise actually redraw the segment/box:
        else:
            if self.listRectanglesa2[segID] is None:
                return

            col = QtGui.QColor(brush)
            col.setAlpha(100)
            self.listRectanglesa1[segID].setBrush(col)
            if not self.batmode:
                self.listRectanglesa2[segID].setBrush(col)
            else:
                self.listRectanglesa2[segID].setBrush(None)
                self.listRectanglesa2[segID].setPen(brush, width=6)

            col.setAlpha(180)
            self.listRectanglesa1[segID].setHoverBrush(fn.mkBrush(col))
            self.listRectanglesa2[segID].setHoverBrush(fn.mkBrush(col))

            if type(self.listRectanglesa2[segID]) == self.ROItype:
                self.listRectanglesa2[segID].transparent = False
                self.listRectanglesa2[segID].setPen(None)
                if self.config['transparentBoxes']:
                    col.setAlpha(255)
                    self.listRectanglesa2[segID].transparent = True
                    self.listRectanglesa2[segID].setPen(col, width=1)
                    self.listRectanglesa2[segID].setBrush(None)
                    self.listRectanglesa2[segID].setHoverBrush(None)
                    col.setAlpha(100)
            self.listRectanglesa1[segID].update()
            self.listRectanglesa2[segID].update()
        QApplication.processEvents()

    def setColourMap(self,cmap):
        """ Listener for the menu item that chooses a colour map.
        Loads them from the file as appropriate and sets the lookup table.
        """
        if not self.CLI and not self.batmode:
            if hasattr(self, 'media_obj'):
                if self.media_obj.isPlayingorPaused():
                    self.stopPlayback()

        self.config['cmap'] = cmap
        lut = colourMaps.getLookupTable(self.config['cmap'])

        self.specPlot.setLookupTable(lut)
        self.overviewImage.setLookupTable(lut)

    def invertColourMap(self):
        """ Listener for the menu item that converts the colour map"""
        self.config['invertColourMap'] = self.invertcm.isChecked()
        self.setColourLevels()

    def setSpectrogram(self):
        """ Normalizes the raw spectrogram in self.sp (ndarray), puts it on self,
            and precalculates some cached properties from it.
            Does NOT update graphics - only internal objects.
        """
        # TODO: There are two things to think about here.
        # 1. The spectrogram, which is large, is basically stored twice
        # 2. The spectrogram you see isn't the one that is processed
        if self.batmode:
            # NOTE batmode kind of assumes spectrogram was already on 0-1 scale
            self.sg = self.sp.normalisedSpec("Batmode")
        else:
            self.sg = self.sp.normalisedSpec(self.config['sgNormMode'])

        self.sgMinimum = np.min(self.sg)
        self.sgMaximum = np.max(self.sg)
        noisefloor = self.noisefloor/100*(self.sgMaximum - self.sgMinimum)+ self.sgMinimum
        self.sg = np.where(self.sg<noisefloor,0,self.sg)
        #print("noisefloor: ",self.noisefloor,noisefloor,self.sgMinimum,self.sgMaximum,np.mean(self.sg))

    def setColourLevels(self, brightness=None, contrast=None):
        """ Listener for the brightness and contrast sliders being changed. Also called when spectrograms are loaded, etc.
        Translates the brightness and contrast values into appropriate image levels.
        """
        if not self.CLI and not self.batmode:
            if hasattr(self, 'media_obj'):
                if self.media_obj.isPlayingorPaused():
                    self.stopPlayback()

        if brightness is None:
            brightness = self.specControls.brightSlider.value()

        if contrast is None:
            contrast = self.specControls.contrSlider.value()

        if self.config['invertColourMap']:
            self.config['brightness'] = brightness
        else:
            self.config['brightness'] = 100-brightness
        self.config['contrast'] = contrast

        self.saveConfig = True

        colRange = colourMaps.getColourRange(self.sgMinimum, self.sgMaximum, self.config['brightness'], self.config['contrast'], self.config['invertColourMap'])

        self.overviewImage.setLevels(colRange)
        self.specPlot.setLevels(colRange)

    def moveLeft(self):
        """ When the left button is pressed (next to the overview plot), move everything along.
        Allows a 10% overlap """
        minX, maxX = self.overviewImageRegion.getRegion()
        newminX = max(0,minX-(maxX-minX)*0.9)
        self.overviewImageRegion.setRegion([newminX, newminX+maxX-minX])

    def moveRight(self):
        """ When the right button is pressed (next to the overview plot), move everything along.
        Allows a 10% overlap """
        minX, maxX = self.overviewImageRegion.getRegion()
        newminX = min(np.shape(self.sg)[0]-(maxX-minX),minX+(maxX-minX)*0.9)
        self.overviewImageRegion.setRegion([newminX, newminX+maxX-minX])

    def prepare5minMove(self):
        # Convenience for next two functions
        self.saveSegments()
        self.resetStorageArrays()
        self.loadFile()

    def movePrev5mins(self):
        """ When the button to move to the next 5 minutes is pressed, enable that.
        Have to check if the buttons should be disabled or not,
        save the segments and reset the arrays, then call loadFile.
        """
        self.currentFileSection -= 1
        self.next5mins.setEnabled(True)
        self.moveNext5minsKey.setEnabled(True)
        if self.currentFileSection <= 0:
            self.prev5mins.setEnabled(False)
            self.movePrev5minsKey.setEnabled(False)
        self.prepare5minMove()

    def moveNext5mins(self):
        """ When the button to move to the previous 5 minutes is pressed, enable that.
        Have to check if the buttons should be disabled or not,
        save the segments and reset the arrays, then call loadFile.
        """
        self.currentFileSection += 1
        self.prev5mins.setEnabled(True)
        self.movePrev5minsKey.setEnabled(True)
        if self.currentFileSection >= self.nFileSections-1:
            self.next5mins.setEnabled(False)
            self.moveNext5minsKey.setEnabled(False)
        self.prepare5minMove()

    def moveTo5mins(self, pagenum=None):
        """ Jumps to the requested 5 min page.
            pagenum can be specified if this is called manually
              Otherwise (None) it will be read from the page selector.
        """
        if pagenum is None:
            pagenum = self.placeInFileSelector.value()
        self.placeInFileSelector.findChild(QLineEdit).deselect()
        self.placeInFileSelector.clearFocus()
        if self.currentFileSection==pagenum-1:
            # no jump needed
            return
        self.currentFileSection = pagenum-1
        if self.currentFileSection >= self.nFileSections-1:
            self.next5mins.setEnabled(False)
            self.moveNext5minsKey.setEnabled(False)
        else:
            self.next5mins.setEnabled(True)
            self.moveNext5minsKey.setEnabled(True)

        if self.currentFileSection <= 0:
            self.prev5mins.setEnabled(False)
            self.movePrev5minsKey.setEnabled(False)
        else:
            self.prev5mins.setEnabled(True)
            self.movePrev5minsKey.setEnabled(True)
        self.prepare5minMove()

    def scroll(self):
        """ When the slider at the bottom of the screen is moved, move everything along. """
        newminX = self.scrollSlider.value()
        if not self.updateRequestedByOverview:
            minX, maxX = self.overviewImageRegion.getRegion()
            self.overviewImageRegion.setRegion([newminX, newminX+maxX-minX])

    def changeWidth(self, value):
        """ Listener for the spinbox that decides the width of the main window.
        It updates the top figure plots as the window width is changed.
        Slightly annoyingly, it gets called when the value gets reset, hence the first line. """
        if not hasattr(self,'overviewImageRegion'):
            return
        self.windowSize = value

        if not self.updateRequestedByOverview:
            # Redraw the highlight in the overview figure appropriately
            minX, maxX = self.overviewImageRegion.getRegion()
            newmaxX = self.convertAmpltoSpec(value)+minX
            self.overviewImageRegion.setRegion([minX, newmaxX])

        self.scrollSlider.setMaximum(int(np.shape(self.sg)[0]-self.convertAmpltoSpec(self.widthWindow.value())))

        # Decide whether or not to show milliseconds
        if value > 3:
            self.timeaxis.setShowMS(False)
        else:
            self.timeaxis.setShowMS(True)

    def annotJumper(self, maxcert):
        """ Scrolls to next annotation of no more than maxcert certainty. """
        # (This is just a manual pg.BusyCursor)
        QApplication.setOverrideCursor(QtGui.QCursor(Qt.CursorShape.WaitCursor))
        # Identify the "current" annotation: selected or whatever is on screen
        if self.box1id > -1:
            currx = self.segments[self.box1id][0]
            self.deselectSegment(self.box1id)
        else:
            minX, maxX = self.overviewImageRegion.getRegion()
            currx = self.convertSpectoAmpl(minX) + self.startRead

        # Find next annotation:
        targetix = None
        for segix in range(len(self.segments)):
            seg = self.segments[segix]
            if seg[0]<=currx:
                continue
            # Note that the segments are not sorted by time,
            # hence some extra mess to find the next one:
            if targetix is not None and seg[0]>=self.segments[targetix][0]:
                continue
            for lab in seg[4]:
                if lab["certainty"]<=maxcert:
                    targetix = segix
        if targetix is None:
            QApplication.restoreOverrideCursor()
            print("No further annotation to jump to found")
            msg = SupportClasses_GUI.MessagePopup("w", "No more annotations", "No further annotation to jump to found")
            msg.exec()
            return

        target = self.segments[targetix]

        if target[0]>self.startRead + self.datalengthSec:
            pagenum, relstart = divmod(target[0], self.config['maxFileShow'])
            pagenum = int(pagenum+1)
            if pagenum > self.nFileSections:
                print("Warning: annotation outside file bounds")
                QApplication.restoreOverrideCursor()
                msg = SupportClasses_GUI.MessagePopup("w", "No more annotations", "No further annotation to jump to found in this sound file")
                msg.exec()
                return
            self.moveTo5mins(pagenum)
        newminT = target[0] - self.startRead - self.windowSize / 2  # in s
        newminX = self.convertAmpltoSpec(newminT)  # in spec pixels
        newmaxX = self.convertAmpltoSpec(newminT + self.windowSize)
        # This will trigger update of the other views
        self.overviewImageRegion.setRegion([newminX, newmaxX])
        self.selectSegment(targetix)
        QApplication.restoreOverrideCursor()

# ===============
# Generate the various dialogs that match the menu items

    def showDiagnosticDialog(self):
        """ Create the dialog to set diagnostic plot parameters.  """
        if not hasattr(self, 'diagnosticDialog'):
            self.diagnosticDialog = Dialogs.Diagnostic(self.FilterDicts)
            self.diagnosticDialog.activate.clicked.connect(self.setDiagnostic)
            self.diagnosticDialog.clear.clicked.connect(self.clearDiagnostic)
        self.diagnosticDialog.show()
        self.diagnosticDialog.activateWindow()

    def showDiagnosticDialogCNN(self):
        """ Create the dialog to set diagnostic plot parameters.  """
        if not hasattr(self, 'diagnosticDialogCNN'):
            self.diagnosticDialogCNN = Dialogs.DiagnosticCNN(self.FilterDicts)
            self.diagnosticDialogCNN.filter.currentTextChanged.connect(self.setCTDiagnosticsCNN)
            self.diagnosticDialogCNN.activate.clicked.connect(self.setDiagnosticCNN)
            self.diagnosticDialogCNN.clear.clicked.connect(self.clearDiagnosticCNN)
        self.diagnosticDialogCNN.show()
        self.diagnosticDialogCNN.activateWindow()

    def clearDiagnostic(self):
        """ Cleans up diagnostic plot space. Should be called when loading new file/page, or from Diagnostic Dialog.  """
        try:
            self.p_plot.clear()
            if hasattr(self, "p_legend"):
                self.p_legend.scene().removeItem(self.p_legend)
            if hasattr(self, "diagnosticCalls"):
                for c in self.diagnosticCalls:
                    self.p_spec.removeItem(c)
            self.d_plot.hide()
            # This can be very slow with many items.
            # Uncomment and use the option below, if e.g. you are doing intense testing.
            # if len(self.diagnosticCalls)>0:
            #     p_spec_new = SupportClasses_GUI.DragViewBox(self, enableMouse=False,enableMenu=False,enableDrag=self.config['specMouseAction']==3, thisIsAmpl=False)
            #     p_spec_new.addItem(self.specPlot)
            #     self.w_spec.removeItem(self.p_spec)
            #     del self.p_spec
            #     self.p_spec = p_spec_new
            #     self.w_spec.addItem(self.p_spec,row=0,col=1)
        except Exception as e:
            print(e)
        self.diagnosticCalls = []

    def setDiagnostic(self):
        """ Takes parameters returned from DiagnosticDialog and draws the training diagnostic plots.  """
        with pg.BusyCursor():
            self.diagnosticDialog.activate.setEnabled(False)
            self.statusLeft.setText("Making diagnostic plots...")
            # Note: importing here so that main program could be run without ext/
            from ext import ce_denoise

            # take values: -2/-3/-4 for AA types, -2/-3 for En/Spec plot
            [filter, aaType, markSpec] = self.diagnosticDialog.getValues()
            spInfo = self.FilterDicts[filter]
            # For now, just using params from the first subfilter
            spSubf = spInfo["Filters"][0]
            print("Using subfilter", spSubf["calltype"])
            # spInfo = json.load(open(os.path.join(self.filtersDir, filter + '.txt')))
            WINSIZE = 0.5

            # clear plot box and add legend
            self.clearDiagnostic()
            self.p_legend = pg.LegendItem()
            self.p_legend.setParentItem(self.p_plot)
            # 1 sec in spectrogram units
            specs = self.convertAmpltoSpec(1)

            # plot things
            # 1. decompose
            # if needed, adjusting sampling rate to match filter
            if self.sp.audioFormat.sampleRate() != spInfo['SampleRate']:
            #if self.sp.sampleRate != spInfo['SampleRate']:
                datatoplot = self.sp.resample(spInfo['SampleRate'])
            else:
                datatoplot = self.sp.data

            WF = WaveletFunctions.WaveletFunctions(data=datatoplot, wavelet='dmey2', maxLevel=5, samplerate=spInfo['SampleRate'])
            WF.WaveletPacket(spSubf['WaveletParams']['nodes'], 'symmetric', aaType==-4, antialiasFilter=True)
            numNodes = len(spSubf['WaveletParams']['nodes'])
            xs = np.arange(0, self.datalengthSec, WINSIZE)
            Esep = np.zeros(( numNodes, len(xs) ))

            ### DENOISING reference: relative |amp| on rec signals from each WP node, when wind is present
            ### just handmade from some wind examples
            # noiseenv = np.array([0.54, 0.83, 0.84, 1, 0.32, 0.54, 0.70, 0.70,  0.16, 0.19, 0.24, 0.22, 0.28, 0.27, 0.25, 0.26,  0.01, 0.06, 0.14, 0.12, 0.15, 0.16, 0.14, 0.15,  0.16, 0.16, 0.15, 0.15, 0.14, 0.16, 0.15, 0.15])
            # # reconstruct wind signal
            # windC = WF.reconstructWP2(34, aaType != -2, False)
            # windC = np.abs(windC)
            # print("calculating wind strength")
            # windE = ce_denoise.EnergyCurve(windC, 4000)
            # # "noise strength index" - strength of wind at each second, on linear scale
            # windMaxE = np.zeros(int(self.datalengthSec))
            # for w in range(int(self.datalengthSec)):
            #     windMaxE[w] = np.max(windE[w*spInfo['SampleRate'] : (w+1)*spInfo['SampleRate']])
            # del windC
            # del windE

            # 2. reconstruct from bands
            r = 0
            # M = spSubf['WaveletParams']['M']
            for node in spSubf['WaveletParams']['nodes']:
                # reconstruction as in detectCalls:
                print("working on node", node)
                C = WF.reconstructWP2(node, aaType != -2, True)
                C = SignalProc.bandpassFilter(C, spInfo['SampleRate'], spSubf['FreqRange'][0], spSubf['FreqRange'][1])

                C = np.abs(C)
                #E = ce_denoise.EnergyCurve(C, int( M*spInfo['SampleRate']/2 ))
                E = C
                C = np.log(C)

                # some prep that doesn't need to be looped over t:
                meanC = np.mean(C)
                sdC = np.std(C)

                # For our new detection: print energy of the quietest second
                trimmedlength = math.floor(self.datalengthSec)*spInfo['SampleRate']
                persecE = np.reshape(E[0:trimmedlength], (math.floor(self.datalengthSec), spInfo['SampleRate'])).mean(axis=1)
                print("Node %i: mean %f, SD %f, range %f - %f" % (node, meanC, sdC, min(persecE), max(persecE)))

                # get true freqs of this band
                freqmin, freqmax = WaveletFunctions.getWCFreq(node, spInfo['SampleRate'])
                # convert freqs to spec Y units
                freqmin = self.convertFreqtoY(freqmin)
                freqmax = self.convertFreqtoY(freqmax)

                # basic divergent color palette
                plotcol = (255*r//numNodes, 127*(r % 2), 0)

                # get max (or mean) E for each second
                # and normalize, so that we don't need to hardcode thr
                for w in range(len(xs)):
                    start = int(w*WINSIZE*spInfo['SampleRate'])
                    end = int((w+1)*WINSIZE*spInfo['SampleRate'])
                    maxE = np.mean(E[start:end])
                    ### DENOISE:
                    # based on wind strength in this second, calculate estimated |wind| in this node
                    # and subtract from maxE
                    # maxE = max(meanC, maxE - windMaxE[w]*noiseenv[node-31]*1.1)
                    Esep[r,w] = (np.log(maxE) - meanC) / sdC

                    # mark detected calls on spectrogram
                    if markSpec and Esep[r,w] > spSubf['WaveletParams']['thr']:
                        diagCall = pg.ROI((specs*xs[w], freqmin),
                                          (specs*WINSIZE, freqmax-freqmin),
                                          pen=plotcol, movable=False)
                        self.diagnosticCalls.append(diagCall)
                        self.p_spec.addItem(diagCall)

                # plot
                self.plotDiag = pg.PlotDataItem(xs, Esep[r,:], pen=fn.mkPen(plotcol, width=2))
                self.p_plot.addItem(self.plotDiag)
                self.p_legend.addItem(self.plotDiag, str(node))
                r = r + 1

            ### DENOISE: add line of wind strength
            # self.p_plot.addItem(pg.PlotDataItem(np.arange(int(self.datalengthSec))+0.5, np.log(windMaxE),
            #             pen=fn.mkPen((0,130,0), width=2)))
            # add line corresponding to thr
            # self.p_plot.addItem(pg.InfiniteLine(-0.8, angle=0, pen=fn.mkPen(color=(40,40,40), width=1)))
            self.p_plot.addItem(pg.InfiniteLine(spSubf['WaveletParams']['thr'], angle=0, pen=fn.mkPen(color=(40,40,40), width=1)))
            minX, maxX = self.overviewImageRegion.getRegion()
            self.p_plot.setXRange(self.convertSpectoAmpl(minX), self.convertSpectoAmpl(maxX), update=True, padding=0)
            self.plotaxis.setLabel('Power Z-score')
            self.d_plot.show()
        self.diagnosticDialog.activate.setEnabled(True)
        self.statusLeft.setText("Ready")

    def setCTDiagnosticsCNN(self):
        from PyQt6.QtWidgets import QCheckBox
        filter = self.diagnosticDialogCNN.filter.currentText()
        speciesData = self.FilterDicts[filter]
        CTs = []
        for f in speciesData['Filters']:
            CTs.append(f['calltype'])
        CTs.append('Noise')
        for ch in self.diagnosticDialogCNN.chkboxes:
            ch.hide()
        self.diagnosticDialogCNN.chkboxes = []
        for ct in CTs:
            self.diagnosticDialogCNN.chkboxes.append(QCheckBox(ct))
        for cb in self.diagnosticDialogCNN.chkboxes:
            if cb.text() != 'Noise':
                cb.setChecked(True)
            self.diagnosticDialogCNN.ctbox.addWidget(cb)

    def clearDiagnosticCNN(self):
        """ Cleans up diagnostic plot space. Should be called
            when loading new file/page, or from Diagnostic Dialog.
        """
        try:
            self.p_plot.clear()
            if hasattr(self, "p_legend"):
                self.p_legend.scene().removeItem(self.p_legend)
            self.d_plot.hide()
        except Exception as e:
            print(e)

    def setDiagnosticCNN(self):
        """ Takes parameters returned from DiagnosticDialog
            and draws the training diagnostic plots.
        """
        from itertools import chain, repeat
        with pg.BusyCursor():
            self.diagnosticDialogCNN.activate.setEnabled(False)
            self.statusLeft.setText("Making CNN diagnostic plots...")

            # Skip Wavelet filter, and show the raw CNN probabilities for current page, block length depends on CNN input size
            # load target CNN model if exists
            [filtername, selectedCTs] = self.diagnosticDialogCNN.getValues()
            print(selectedCTs)
            speciesData = self.FilterDicts[filtername]
            CTs = []
            for f in speciesData['Filters']:
                CTs.append(f['calltype'])
            CTs.append('Noise')
            self.CNNDicts = self.ConfigLoader.CNNmodels(self.FilterDicts, self.filtersDir, [filtername])

            segment = [[self.startRead, self.startRead + self.datalengthSec]]
            CNNmodel = None
            probs = 0
            if filtername in self.CNNDicts.keys():
                CNNmodel = self.CNNDicts[filtername]
            post = Segment.PostProcess(configdir=self.configdir, audioData=self.sp.data,
                                       sampleRate=self.sp.audioFormat.sampleRate(),
                                       tgtsampleRate=speciesData["SampleRate"], segments=segment,
                                       subfilter=speciesData['Filters'][0], CNNmodel=CNNmodel, cert=50)
            if CNNmodel:
                CNNwindow, probs = post.CNNDiagnostic()
            if isinstance(probs, int):
                self.diagnosticDialogCNN.activate.setEnabled(True)
                return

            # clear plot box and add legend
            self.clearDiagnostic()
            self.p_legend = pg.LegendItem()
            self.p_legend.setParentItem(self.p_plot)

            Psep = np.zeros((len(CTs), len(probs[:, 0].tolist())))
            for i in range(len(CTs)):
                Psep[i, :] = probs[:, i].tolist()

            # plot
            for ct in range(len(CTs)):
                if not selectedCTs[ct]:
                    continue
                else:
                    # basic divergent color palette
                    plotcol = (255 * ct // len(CTs), 127 * (ct % 2), 0)
                    y = Psep[ct, :]
                    # x = np.linspace(0, CNNwindow*len(y), len(y))
                    x = np.linspace(CNNwindow/2, CNNwindow*len(y)-CNNwindow/2, len(y))
                    self.plotDiag = pg.PlotDataItem(x, y, pen=fn.mkPen(plotcol, width=2))
                    self.p_plot.addItem(self.plotDiag)
                    self.p_legend.addItem(self.plotDiag, CTs[ct])
            self.d_plot.show()
        self.diagnosticDialogCNN.activate.setEnabled(True)
        self.statusLeft.setText("Ready")

    def showSpectrogramDialog(self):
        """ Create spectrogram dialog when the button is pressed.
        """
        if not hasattr(self,'spectrogramDialog'):
            self.spectrogramDialog = Dialogs.SpectrogramDialog(self.config['window_width'],self.config['incr'],self.sp.minFreq,self.sp.maxFreq, self.sp.minFreqShow,self.sp.maxFreqShow, self.config['windowType'], self.config['sgType'], self.config['sgNormMode'], self.config['sgScale'], self.config['nfilters'],self.batmode)
            self.spectrogramDialog.activate.clicked.connect(self.spectrogram)
        # First save the annotations
        self.saveSegments()
        self.spectrogramDialog.show()
        self.spectrogramDialog.activateWindow()

    def spectrogram(self):
        """ Listener for the spectrogram dialog.
        Has to do quite a bit of work to make sure segments are in the correct place, etc."""
        [self.config['windowType'], self.config['sgType'], self.config['sgNormMode'], self.config['sgMeanNormalise'], self.config['sgEqualLoudness'], window_width, incr, self.config['minFreq'], self.config['maxFreq'],sgScale,self.config['nfilters']] = self.spectrogramDialog.getValues()
        if self.config['sgScale'] != sgScale:
            self.config['sgScale'] = sgScale
            changedY = True
        else:  
            changedY = False

        if (self.config['minFreq'] >= self.config['maxFreq']):
            msg = SupportClasses_GUI.MessagePopup("w", "Error", "Incorrect frequency range")
            msg.exec()
            return
        with pg.BusyCursor():
            self.statusLeft.setText("Updating the spectrogram...")
            if self.batmode:
                print("Warning: only spectrogram freq. range can be changed in BMP mode")
            else:
                self.sp.setWidth(window_width, incr)
                _ = self.sp.spectrogram(window_width=window_width, incr=incr,window=self.config['windowType'],sgType=self.config['sgType'],sgScale=self.config['sgScale'],nfilters=self.config['nfilters'],mean_normalise=self.config['sgMeanNormalise'],equal_loudness=self.config['sgEqualLoudness'],onesided=self.config['sgOneSided'])
                self.setSpectrogram()

                # If the size of the spectrogram has changed, need to update the positions of things
                if incr != self.config['incr'] or window_width != self.config['window_width']:
                    self.config['incr'] = incr
                    self.config['window_width'] = window_width
                    if hasattr(self, 'seg'):
                        self.seg.setNewData(self.sp)

                    self.loadFile(self.filename)

                    # These two are usually set by redoFreqAxis, but that is called only later in this case
                    self.spectrogramDialog.low.setValue(self.config['minFreq'])
                    self.spectrogramDialog.high.setValue(self.config['maxFreq'])

        self.redoFreqAxis(self.config['minFreq'],self.config['maxFreq'],changedY=changedY)

        self.statusLeft.setText("Ready")

    def calculateStats(self):
        """ Calculate and export summary statistics for the currently marked segments """

        import Features
        #print("segs", self.segments)

        cs = open(self.filename[:-4] + '_features.csv', "w")
        cs.write("Start Time (sec),End Time (sec),Avg Power,Delta Power,Energy,Agg Entropy,Avg Entropy,Max Power,Max Freq\n")

        for seg in self.segments:
            # Important because all manual mode functions should operate on the current page only:
            # skip segments that are not visible in this page
            if seg[1]<=self.startRead or seg[0]>=self.startRead + self.datalengthSec:
                continue

            # coordinates in seconds from current page start, bounded at page borders:
            starttime = max(0, seg[0]-self.startRead)
            endtime = min(seg[1]-self.startRead, self.datalengthSec)
            #print(starttime, endtime)

            # piece of audio/waveform corresponding to this segment
            # (note: coordinates in wav samples)
            data = self.sp.data[int(starttime*self.sp.audioFormat.sampleRate()):int(endtime*self.sp.audioFormat.sampleRate())]
            #data = self.sp.data[int(starttime*self.sp.sampleRate):int(endtime*self.sp.sampleRate)]

            # piece of spectrogram corresponding to this segment
            startInSpecPixels = self.convertAmpltoSpec(starttime)
            endInSpecPixels = self.convertAmpltoSpec(endtime)
            #print(startInSpecPixels, endInSpecPixels)
            # self.sg[startInSpecPixels:endInSpecPixels, ]

            # if needed, there's already a Spectrogram instance self.sp with the full data on it,
            # so can also do something like:
            # self.sp.calculateMagicStatistic(starttime, endtime)

            # do something with this segment now...
            print("Calculating statistics on this segment...")

            # TODO: Workout the units
            f = Features.Features(data=data, sampleRate=self.sp.audioFormat.sampleRate(), window_width=self.config['window_width'], incr=self.config['incr'])
            #f = Features.Features(data=data, sampleRate=self.sp.sampleRate, window_width=self.config['window_width'], incr=self.config['incr'])
            avgPower, deltaPower, energy, aggEntropy, avgEntropy, maxPower, maxFreq = f.get_Raven_spectrogram_measurements(f1=int(self.convertFreqtoY(500)), f2=int(self.convertFreqtoY(8000)))
            # quartile1, quartile2, quartile3, f5, f95, interquartileRange = f.get_Raven_robust_measurements(f1=int(self.convertFreqtoY(500)), f2=int(self.convertFreqtoY(8000)))
            print(avgPower, deltaPower, energy, aggEntropy, avgEntropy, maxPower, maxFreq)
            # print(quartile1, quartile2, quartile3, f5, f95, interquartileRange)
            # cs.write("%s\t%.4f\t%.4f\t%.2f\t%.2f\t%.2f\t%.2f\t%.2f\t%.2f\t%.2f\t%.2f\t%.2f\t%.2f\t%.2f\t%.2f\t%.2f\n" % (self.filename, starttime, endtime, avgPower, deltaPower, energy, aggEntropy, avgEntropy, maxPower, maxFreq, quartile1, quartile2, quartile3, f5, f95, interquartileRange))
            # cs.write("%s,%.4f,%.4f,%.2f,%.2f,%.2f,%.2f,%.2f,%.2f,%.2f\n" % (self.filename, starttime, endtime, avgPower, deltaPower, energy, aggEntropy, avgEntropy, maxPower, maxFreq))
            cs.write("%.4f,%.4f,%.2f,%.2f,%.2f,%.2f,%.2f,%.2f,%.2f\n" % (starttime, endtime, avgPower, deltaPower, energy, aggEntropy, avgEntropy, maxPower, maxFreq))

        cs.close()

    def detectShapes(self):
        method, IFsettings = self.shapesDialog.getValues()
        allshapes = []
        specxunit = self.convertSpectoAmpl(1)
        specyunit = self.sp.audioFormat.sampleRate()//2 / np.shape(self.sg)[1]
        #specyunit = self.sp.sampleRate//2 / np.shape(self.sg)[1]
        if self.batmode:
            incr = 512
        else:
            incr = self.config['incr']

        # TODO TODO not tested for bats at all,
        # no idea what here may need adapting
        # (specifically b/c they have a spec that starts at >0)
        # NOTE: resulting shape.tstart will be relative to the current page
        with pg.BusyCursor():
            for segm in self.segments:
                segshape = None
                # convert from absolute to relative-to-page times
                segRelativeStart = segm[0]-self.startRead
                segRelativeEnd = segm[1]-self.startRead

                # skip if they are not in this page:
                if segRelativeEnd<0 or segRelativeStart>self.datalengthSec:
                    print("skipping out of page segment", segm[0], "-", segm[1])
                    allshapes.append(segshape)
                    continue

                if method=="stupidShaper":
                    # placeholder method:
                    adjusted_segm = [segRelativeStart, segRelativeEnd, segm[2], segm[3], segm[4]]
                    segshape = Shapes.stupidShaper(adjusted_segm, specxunit, specyunit)
                elif method=="fundFreqShaper":
                    # Fundamental frequency:
                    data = self.sp.data[int(segRelativeStart*self.sp.audioFormat.sampleRate()):int(segRelativeEnd*self.sp.audioFormat.sampleRate())]
                    #data = self.sp.data[int(segRelativeStart*self.sp.sampleRate):int(segRelativeEnd*self.sp.sampleRate)]
                    W = 4*incr
                    segshape = Shapes.fundFreqShaper(data, W, thr=0.5, fs=self.sp.audioFormat.sampleRate())
                    #segshape = Shapes.fundFreqShaper(data, W, thr=0.5, fs=self.sp.sampleRate)
                    # shape.tstart is relative to segment start (0)
                    # so we also need to add the segment start
                    segshape.tstart += segRelativeStart
                elif method=="instantShaper1" or method=="instantShaper2":
                    # instantaneous frequency
                    IFmethod = int(method[-1])
                    spstart = math.floor(self.convertAmpltoSpec(segRelativeStart))
                    spend = math.ceil(self.convertAmpltoSpec(segRelativeEnd))
                    sg = np.copy(self.sp.sg[spstart:spend,:])
                    # mask freqs outside the currently marked segment
                    if segm[3]>0:
                        markedylow = math.floor(self.convertFreqtoY(segm[2]))
                        markedyupp = math.ceil(self.convertFreqtoY(segm[3]))
                        sg[:,:markedylow] = 0
                        sg[:,markedyupp:] = 0
                    segshape = Shapes.instantShaper(sg, self.sp.audioFormat.sampleRate(), incr, self.config['window_width'], self.config['windowType'], IFmethod, IFsettings)
                    #segshape = Shapes.instantShaper(sg, self.sp.sampleRate, incr, self.config['window_width'], self.config['windowType'], IFmethod, IFsettings)
                    # shape.tstart is relative to segment start (0)
                    # so we also need to add the segment start
                    segshape.tstart += segRelativeStart
                allshapes.append(segshape)

        if len(allshapes)!=len(self.segments):
            print("ERROR: something went wrong in shape analysis, produced %d shapes", len(allshapes))
            return

        # print, plot or export the results
        # clear any old plots:
        for sh in self.shapePlots:
            try:
                self.p_spec.removeItem(sh)
            except Exception:
                pass
        self.shapePlots = []

        # NOTE: this skips -1 and values below minFreqShow and connects
        # the reamining dots. Might not be what you want.
        # TODO not sure if this will work when spec sp.minFreqShow>0
        for shape in allshapes:
            if shape is None:
                continue
            # Convert coordinates to Hz/s and back to spec y/x, b/c
            # spacing used to calculate the shape may differ from current spec pixel size.
            numy = len(shape.y)
            seqx = [self.convertAmpltoSpec(x*shape.tunit + shape.tstart) for x in range(numy)]
            seqfreqs = shape.y*shape.yunit + shape.ystart  # convert to Hz
            # Hide any below minFreqShow
            visible = seqfreqs>=self.sp.minFreqShow
            seqfreqs = seqfreqs[visible]
            seqx = np.asarray(seqx)[visible]
            seqy = [self.convertFreqtoY(y) for y in seqfreqs]

            self.shapePlots.append(pg.PlotDataItem())
            self.shapePlots[-1].setData(seqx, seqy, pen=pg.mkPen('r', width=2))
            self.p_spec.addItem(self.shapePlots[-1])

    def showShapesDialog(self):
        """ Create the shape analysis dialog. """
        self.shapesDialog = Dialogs.Shapes()
        self.shapesDialog.show()
        self.shapesDialog.activateWindow()
        self.shapesDialog.activate.clicked.connect(self.detectShapes)

    def showDenoiseDialog(self):
        """ Create the denoising dialog when the relevant button is pressed.
        """
        self.denoiseDialog = Dialogs.Denoise(DOC=self.DOC,minFreq=self.sp.minFreq,maxFreq=self.sp.maxFreq)
        self.denoiseDialog.show()
        self.denoiseDialog.activateWindow()
        self.denoiseDialog.activate.clicked.connect(self.denoise)
        self.denoiseDialog.undo.clicked.connect(self.denoise_undo)
        self.denoiseDialog.save.clicked.connect(self.denoise_save)

    def backup(self):
        """ Enables denoising to be undone. """
        if hasattr(self, 'audiodata_backup'):
            if self.audiodata_backup is not None:
                audiodata_backup_new = np.empty(
                    (np.shape(self.audiodata_backup)[0], np.shape(self.audiodata_backup)[1] + 1))
                audiodata_backup_new[:, :-1] = np.copy(self.audiodata_backup)
                audiodata_backup_new[:, -1] = np.copy(self.sp.data)
                self.audiodata_backup = audiodata_backup_new
            else:
                self.audiodata_backup = np.empty((np.shape(self.sp.data)[0], 1))
                self.audiodata_backup[:, 0] = np.copy(self.sp.data)
        else:
            self.audiodata_backup = np.empty((np.shape(self.sp.data)[0], 1))
            self.audiodata_backup[:, 0] = np.copy(self.sp.data)
        self.showFreq_backup = [self.sp.minFreqShow, self.sp.maxFreqShow]

    def decomposeWP(self, x=None):
        """ Listener for quickWP control button.
            Takes DATA and produces a WP decomposition.
        """
        print("Decomposing to WP...")
        ot = time.time()
        self.WFinst = WaveletFunctions.WaveletFunctions(data=self.sp.data, wavelet="dmey2", maxLevel=self.config['maxSearchDepth'], samplerate=self.sp.audioFormat.sampleRate())
        #self.WFinst = WaveletFunctions.WaveletFunctions(data=self.sp.data, wavelet="dmey2", maxLevel=self.config['maxSearchDepth'], samplerate=self.sp.sampleRate)
        maxLevel = 5
        allnodes = range(2 ** (maxLevel + 1) - 1)
        self.WFinst.WaveletPacket(allnodes, mode='symmetric', antialias=False)
        print("Done")
        print(time.time() - ot)

    def denoiseSeg(self):
        """ Listener for quickDenoise control button.
            Extracts a segment from DATA between START and STOP (in ms),
            denoises that segment, concats with rest of original DATA,
            and updates the original DATA.
        """
        if self.box1id > -1:
            start, stop = self.listRectanglesa1[self.box1id].getRegion()
            #start = int(start*self.sp.sampleRate)
            start = int(start*self.sp.audioFormat.sampleRate())
            #stop = int(stop*self.sp.sampleRate)
            stop = int(stop*self.sp.audioFormat.sampleRate())
        else:
            print("Can't play, no segment selected")
            return

        if hasattr(self, 'media_obj'):
            if self.media_obj.isPlayingorPaused():
                self.stopPlayback()

        # Since there is no dialog menu, settings are preset constants here:
        noiseest = "ols" # or qr, or const
        thrType = "soft"
        depth = 6   # can also use 0 to autoset
        wavelet = "dmey2"
        aaRec = False  # True if nicer spectrogram is needed - but it's not very clean either way
        aaWP = False
        thr = 2.0  # this one is difficult to set universally...

        self.statusLeft.setText("Denoising...")
        with pg.BusyCursor():
            opstartingtime = time.time()
            print("Denoising requested at " + time.strftime('%H:%M:%S', time.gmtime(opstartingtime)))

            # extract the piece of audiodata under current segment
            denoised = self.sp.data[start : stop]

            WF = WaveletFunctions.WaveletFunctions(data=denoised, wavelet=wavelet, maxLevel=self.config['maxSearchDepth'], samplerate=self.sp.audioFormat.sampleRate())
            #WF = WaveletFunctions.WaveletFunctions(data=denoised, wavelet=wavelet, maxLevel=self.config['maxSearchDepth'], samplerate=self.sp.sampleRate)
            denoised = WF.waveletDenoise(thrType, thr, depth, aaRec=aaRec, aaWP=aaWP, noiseest=noiseest, costfn="fixed")

            # bandpass to selected zones, if it's a box
            # TODO this could be done faster: pass to waveletDenoise and
            # do not reconstruct from nodes outside the specified band
            if self.segments[self.box1id][3]>0:
                bottom = max(0.1, self.sp.minFreq, self.segments[self.box1id][2])
                top = min(self.segments[self.box1id][3], self.sp.maxFreq-0.1)
                print("Extracting samples between %d-%d Hz" % (bottom, top))
                denoised = SignalProc.bandpassFilter(denoised, sampleRate=self.sp.audioFormat.sampleRate(), start=bottom, end=top)
                #denoised = SignalProc.bandpassFilter(denoised, sampleRate=self.sp.sampleRate, start=bottom, end=top)

            print("Denoising calculations completed in %.4f seconds" % (time.time() - opstartingtime))

            # update full audiodata
            self.sp.data[start : stop] = denoised
            #self.audiodata[start : stop] = denoised

            # recalculate spectrogram
            #_ = self.sp.spectrogram(window=str(self.windowType),sgType=str(self.sgType),sgScale=str(self.sgScale),nfilters=int(str(self.nfilters)),mean_normalise=self.sgMeanNormalise,equal_loudness=self.sgEqualLoudness,onesided=self.config['sgOneSided'])
            _ = self.sp.spectrogram(window_width=self.config['window_width'], incr=self.config['incr'],window=self.config['windowType'],sgType=self.config['sgType'],sgScale=self.config['sgScale'],nfilters=self.config['nfilters'],mean_normalise=self.config['sgMeanNormalise'],equal_loudness=self.config['sgEqualLoudness'],onesided=self.config['sgOneSided'])
            self.setSpectrogram()

            # Update the ampl image
            self.amplPlot.setData(np.linspace(0.0,self.datalength/self.sp.audioFormat.sampleRate(),num=self.datalength,endpoint=True),self.sp.data)
            #self.amplPlot.setData(np.linspace(0.0,self.datalength/self.sp.sampleRate,num=self.datalength,endpoint=True),self.sp.data)

            # Update the spec & overview images.
            # Does not reset to start if the freqs aren't changed
            self.redoFreqAxis(self.sp.minFreqShow,self.sp.maxFreqShow, store=False)

            if hasattr(self,'spectrogramDialog'):
                self.spectrogramDialog.setValues(self.sp.minFreq,self.sp.maxFreq,self.sp.minFreqShow,self.sp.maxFreqShow)

            self.setColourLevels()

            print("Denoising completed in %s seconds" % round(time.time() - opstartingtime, 4))
        self.statusLeft.setText("Ready")

    def denoise(self):
        """ Listener for the denoising dialog.
        Calls the denoiser and then plots the updated data.
        """
        if self.CLI:
            # In CLI mode, default values will be retrieved from dialogs.
            self.denoiseDialog = Dialogs.Denoise(DOC=self.DOC,minFreq=self.sp.minFreq,maxFreq=self.sp.maxFreq)
            # values can be passed here explicitly, e.g.:
            # self.denoiseDialog.depth.setValue(10)
            # or could add an argument to pass custom defaults, e.g.:
            # self.denoiseDialog = Dialogs.Denoise(defaults=("wt", 1, 2, 'a')
        with pg.BusyCursor():
            opstartingtime = time.time()
            print("Denoising requested at " + time.strftime('%H:%M:%S', time.gmtime(opstartingtime)))
            self.statusLeft.setText("Denoising...")
            # Note: dialog returns all possible parameters
            if not self.DOC:
                [alg, depth, thrType, thr,wavelet,start,end,width,aaRec,aaWP,noiseest] = self.denoiseDialog.getValues()
            else:
                wavelet = "dmey2"
                [alg, start, end, width] = self.denoiseDialog.getValues()
            self.backup()

            if str(alg)=="Wavelets":
                # here we override default 0-Fs/2 returns
                start = self.sp.minFreqShow
                end = self.sp.maxFreqShow
                self.waveletDenoiser = WaveletFunctions.WaveletFunctions(data=self.sp.data, wavelet=wavelet, maxLevel=self.config['maxSearchDepth'], samplerate=self.sp.audioFormat.sampleRate())
                #self.waveletDenoiser = WaveletFunctions.WaveletFunctions(data=self.sp.data, wavelet=wavelet, maxLevel=self.config['maxSearchDepth'], samplerate=self.sp.sampleRate)
                if not self.DOC:
                    # pass dialog settings
                    # TODO set costfn determines which leaves will be used, by default 'threshold' (universal threshold).
                    # fixed = use all leaves up to selected level. 'Entropy' is also tested and possible
                    self.sp.data = self.waveletDenoiser.waveletDenoise(thrType,float(str(thr)), depth, aaRec=aaRec, aaWP=aaWP, noiseest=noiseest, costfn="fixed")
                else:
                    # go with defaults
                    self.sp.data = self.waveletDenoiser.waveletDenoise("soft", 3, aaRec=True, aaWP=False, costfn="fixed", noiseest="ols")

            else:
                # Spectrogram will deal with denoising
                self.sp.denoise(alg, start=start, end=end, width=width)
            #self.audiodata = self.sp.data

            print("Denoising calculations completed in %.4f seconds" % (time.time() - opstartingtime))

            #_ = self.sp.spectrogram(window=str(self.windowType),sgType=str(self.sgType),sgScale=str(self.sgScale),nfilters=int(str(self.nfilters)),mean_normalise=self.sgMeanNormalise,equal_loudness=self.sgEqualLoudness,onesided=self.config['sgOneSided'])
            _ = self.sp.spectrogram(window_width=self.config['window_width'], incr=self.config['incr'],window=self.config['windowType'],sgType=self.config['sgType'],sgScale=self.config['sgScale'],nfilters=self.config['nfilters'],mean_normalise=self.config['sgMeanNormalise'],equal_loudness=self.config['sgEqualLoudness'],onesided=self.config['sgOneSided'])
            self.setSpectrogram()

            self.amplPlot.setData(np.linspace(0.0,self.datalength/self.sp.audioFormat.sampleRate(),num=self.datalength,endpoint=True),self.sp.data)
            #self.amplPlot.setData(np.linspace(0.0,self.datalength/self.sp.sampleRate,num=self.datalength,endpoint=True),self.sp.data)

            # Update the frequency axis
            self.redoFreqAxis(start, end, store=False)

            if hasattr(self,'spectrogramDialog'):
                self.spectrogramDialog.setValues(self.sp.minFreq,self.sp.maxFreq,self.sp.minFreqShow,self.sp.maxFreqShow)

            self.setColourLevels()

            print("Denoising completed in %s seconds" % round(time.time() - opstartingtime, 4))
            self.statusLeft.setText("Ready")

    def denoise_undo(self):
        """ Listener for undo button in denoising dialog.
        """
        print("Undoing",np.shape(self.audiodata_backup))
        if hasattr(self,'audiodata_backup'):
            if self.audiodata_backup is not None:
                if np.shape(self.audiodata_backup)[1]>0:
                    self.sp.data = np.copy(self.audiodata_backup[:,-1])
                    self.audiodata_backup = self.audiodata_backup[:,:-1]
                    #self.sp.data = self.audiodata

                    #_ = self.sp.spectrogram(window=str(self.windowType),sgType=str(self.sgType),sgScale=str(self.sgScale),nfilters=int(str(self.nfilters)),mean_normalise=self.sgMeanNormalise,equal_loudness=self.sgEqualLoudness,onesided=self.config['sgOneSided'])
                    _ = self.sp.spectrogram(window_width=self.config['window_width'], incr=self.config['incr'],window=self.config['windowType'],sgType=self.config['sgType'],sgScale=self.config['sgScale'],nfilters=self.config['nfilters'],mean_normalise=self.config['sgMeanNormalise'],equal_loudness=self.config['sgEqualLoudness'],onesided=self.config['sgOneSided'])
                    self.setSpectrogram()

                    self.amplPlot.setData(
                        np.linspace(0.0, self.datalengthSec, num=self.datalength, endpoint=True),
                        self.sp.data)
                    if hasattr(self,'seg'):
                        self.seg.setNewData(self.sp)

                    if hasattr(self, 'showFreq_backup'):
                        self.redoFreqAxis(self.showFreq_backup[0], self.showFreq_backup[1])
                    else:
                        self.redoFreqAxis(self.sp.minFreq, self.sp.maxFreq)
                    self.setColourLevels()

    def denoise_save(self):
        """ Listener for save button in denoising dialog.
        Adds _d to the filename and saves it as a new sound file.
        """
        filename = self.filename[:-4] + '_d' + self.filename[-4:]
        wavio.write(filename,self.sp.data.astype('int16'),self.sp.audioFormat.sampleRate(),scale='dtype-limits', sampwidth=2)
        #wavio.write(filename,self.sp.data.astype('int16'),self.sp.sampleRate,scale='dtype-limits', sampwidth=2)
        self.statusLeft.setText("Saved")
        msg = SupportClasses_GUI.MessagePopup("d", "Saved", "Destination: " + '\n' + filename)
        msg.exec()
        return

    def saveSelectedSound(self,changespeed):
        """ Listener for 'Save selected (slow) sound' button.
            Chooses destination, file name, and exports.
            Speed comes from self.playSpeed
        """
        if self.box1id is None or self.box1id<0:
            print("No box selected")
            msg = SupportClasses_GUI.MessagePopup("w", "No segment", "No sound selected to save")
            msg.exec()
            return
        else:
            if type(self.listRectanglesa2[self.box1id]) == self.ROItype:
                x1 = self.listRectanglesa2[self.box1id].pos().x()
                x2 = x1 + self.listRectanglesa2[self.box1id].size().x()
                y1 = max(self.sp.minFreq, self.segments[self.box1id][2])
                y2 = min(self.segments[self.box1id][3], self.sp.maxFreq)
                y1 = self.sp.minFreq
                y2 = self.sp.maxFreq
            else:
                x1, x2 = self.listRectanglesa2[self.box1id].getRegion()
                y1 = self.sp.minFreq
                y2 = self.sp.maxFreq
            x1 = math.floor(x1 * self.config['incr'])
            x2 = math.floor(x2 * self.config['incr'])
            filename, drop = QFileDialog.getSaveFileName(self, 'Save File as', '', '*.wav')
            if filename:
                # filedialog doesn't attach extension
                filename = str(filename)
                if not filename.endswith('.wav'):
                    filename = filename + '.wav'
                #tosave = self.sp.data[int(x1):int(x2)]
                tosave = SignalProc.bandpassFilter(self.sp.data[int(x1):int(x2)], sampleRate=self.sp.audioFormat.sampleRate(),start=y1, end=y2)
                #tosave = SignalProc.bandpassFilter(self.sp.data[int(x1):int(x2)], sampleRate=self.sp.sampleRate,start=y1, end=y2)
                if changespeed:
                    tosave = SignalProc.wsola(tosave,self.playSpeed) 
                if self.sp.audioFormat.sampleFormat() == QAudioFormat.SampleFormat.Int16:
                    sampwidth = 2
                elif self.sp.audioFormat.sampleFormat() == QAudioFormat.SampleFormat.Int32:
                    sampwidth = 4
                elif self.sp.audioFormat.sampleFormat() == QAudioFormat.SampleFormat.UInt8:
                    sampwidth = 1
                else:
                    print("ERROR: sampleSize %d not supported" % self.audioFormat.sampleSize())
                wavio.write(filename, tosave, self.sp.audioFormat.sampleRate(), scale=None, sampwidth=sampwidth)
                #wavio.write(filename, tosave, self.sp.sampleRate, scale=None, sampwidth=sampwidth)
                #wavio.write(filename, tosave, self.sp.sampleRate, scale='dtype-limits', sampwidth=samplewidth)
            # update the file list box
            self.fillFileList(self.SoundFileDir, os.path.basename(self.filename))

    def redoFreqAxis(self,start,end, store=True,changedY=False):
        """ This is the listener for the menu option to make the frequency axis tight (after bandpass filtering or just spectrogram changes)
            On the same go updates spectrogram and overview plots.
                store: boolean, indicates whether changes should be stored in the config
        """
        changedY = (start!=self.sp.minFreqShow or end!=self.sp.maxFreqShow or changedY)
        # Lots of updating can be avoided if the Y freqs aren't changing:
        if changedY:
            self.sp.minFreqShow = max(start,self.sp.minFreq)
            self.sp.maxFreqShow = min(end,self.sp.maxFreq)

            if store:
                if self.batmode:
                    self.config['minFreqBats'] = start
                    self.config['maxFreqBats'] = end
                else:
                    self.config['minFreq'] = start
                    self.config['maxFreq'] = end

        # SRM: changed
        # draw a spectrogram of proper height:
        #height = self.sampleRate // 2 / np.shape(self.sg)[1]
        #pixelstart = int(self.sp.minFreqShow/height)
        #pixelend = int(self.sp.maxFreqShow/height)

        #self.overviewImage.setImage(self.sg[:,pixelstart:pixelend])
        #self.overviewImageRegion.setBounds([0, len(self.sg)])
        #self.specPlot.setImage(self.sg[:,pixelstart:pixelend])
        self.setfigs()

        # if Y freqs changed, some segments may appear/be dropped:
        if changedY:
            # Remove everything and redraw it
            self.removeSegments(delete=False)
            self.drawOverview()
            self.drawfigMain(remaking=True)

            try:
                for r in self.segmentPlots:
                    self.p_spec.removeItem(r)
                self.segmentPlots=[]
            except Exception:
                pass
            else:
                self.showFundamentalFreq()

            try:
                self.p_spec.removeItem(self.derivPlot)
            except Exception:
                pass
            else:
                self.showSpectralDeriv()

            if not self.DOC:
                try:
                    self.p_spec.removeItem(self.formantPlot)
                except Exception:
                    pass
                else:
                    self.showFormants()

            try:
                self.p_spec.removeItem(self.energyPlot)
            except Exception:
                pass
            else:
                self.showMaxEnergy()

        QApplication.processEvents()

    def buildChpRecogniser(self):
        """ Train a changepoint detector.
            Currently, takes nodes from a selected wavelet filter,
            and only trains alpha, length etc.
        """
        self.saveSegments()
        self.buildRecAdvWizard = DialogsTraining.BuildRecAdvWizard(self.filtersDir, self.config, method="chp")
        self.buildRecAdvWizard.button(QWizard.WizardButton.FinishButton).clicked.connect(lambda: self.saveRecogniser(test=False))
        self.buildRecAdvWizard.saveTestBtn.clicked.connect(lambda: self.saveRecogniser(test=True))
        self.buildRecAdvWizard.activateWindow()
        self.buildRecAdvWizard.exec()
        # reread filters list with the new one
        self.FilterDicts = self.ConfigLoader.filters(self.filtersDir)

    def buildRecogniser(self):
        """Listener for 'Build a recogniser'
           All training and file I/O are done in Dialogs.py currently.
        """
        self.saveSegments()
        self.buildRecAdvWizard = DialogsTraining.BuildRecAdvWizard(self.filtersDir, self.config, method="wv")
        self.buildRecAdvWizard.button(QWizard.WizardButton.FinishButton).clicked.connect(lambda: self.saveRecogniser(test=False))
        self.buildRecAdvWizard.saveTestBtn.clicked.connect(lambda: self.saveRecogniser(test=True))
        self.buildRecAdvWizard.activateWindow()
        self.buildRecAdvWizard.exec()
        # reread filters list with the new one
        self.FilterDicts = self.ConfigLoader.filters(self.filtersDir)

    def buildCNN(self):
        """Listener for 'Build a CNN'
        """
        self.saveSegments()
        self.buildCNNWizard = DialogsTraining.BuildCNNWizard(self.filtersDir, self.config, self.configdir)
        #self.buildCNNWizard.button(3).clicked.connect(lambda: self.RecogniserCNN(test=False))
        self.buildCNNWizard.saveTestBtn.clicked.connect(lambda: self.saveRecogniserCNN(test=True))
        self.buildCNNWizard.activateWindow()
        self.buildCNNWizard.exec()

    def testRecogniser(self, filter=None):
        """ Listener for the Test Recogniser action """
        self.testRecWizard = DialogsTraining.TestRecWizard(self.filtersDir, self.configdir, filter)
        self.testRecWizard.exec()

    def saveRecogniser(self, test=False):
        try:
            # actually write out the filter
            filename = os.path.join(self.filtersDir, self.buildRecAdvWizard.field("filtfile"))
            # also write ROC in to a file
            rocfilename = self.buildRecAdvWizard.speciesData["species"] + "_ROCWF" + time.strftime("_%H-%M-%S", time.gmtime())
            self.buildRecAdvWizard.speciesData["ROCWF"] = rocfilename
            rocfilename = os.path.join(self.filtersDir, rocfilename + '.json')
            print("Saving new recogniser to ", filename)
            with open(filename, 'w') as f:
                f.write(json.dumps(self.buildRecAdvWizard.speciesData, indent=4))
            with open(rocfilename, 'w') as f:
                f.write(json.dumps(self.buildRecAdvWizard.ROCData, indent=4))

            # prompt the user
            if test:
                msg = SupportClasses_GUI.MessagePopup("d", "Training completed!", "Training completed!\nProceeding to testing.")
            else:
                msg = SupportClasses_GUI.MessagePopup("d", "Training completed!", "Training completed!\nWe strongly recommend testing the recogniser on a separate dataset before actual use.")
            msg.exec()
            self.buildRecAdvWizard.done(1)
            if test:
                self.testRecogniser(filter=os.path.basename(filename))
        except Exception as e:
            print("ERROR: could not save recogniser because:", e)
            self.buildRecAdvWizard.done(0)

    def saveRecogniserCNN(self, test=False):
        # Actually write out the filter and CNN model
        modelsrc = os.path.join(self.buildCNNWizard.cnntrain.tmpdir2.name, 'model.json')
        CNN_name = self.buildCNNWizard.cnntrain.species + time.strftime("_%H-%M-%S", time.gmtime())
        self.buildCNNWizard.cnntrain.currfilt["CNN"]["CNN_name"] = CNN_name
        modelfile = os.path.join(self.filtersDir, CNN_name + '.json')
        weightsrc = self.buildCNNWizard.cnntrain.bestweight
        weightfile = os.path.join(self.filtersDir, CNN_name + '.h5')
        # Also write ROC in to a file
        rocfilename = self.buildCNNWizard.cnntrain.currfilt["species"] + "_ROCNN" + time.strftime("_%H-%M-%S", time.gmtime())
        self.buildCNNWizard.cnntrain.currfilt["ROCNN"] = rocfilename
        rocfilename = os.path.join(self.filtersDir, rocfilename + '.json')

        try:
            if self.buildCNNWizard.savePage.saveoption == 'New':
                filename = os.path.join(self.filtersDir, self.buildCNNWizard.savePage.enterFiltName.text())
                print("Saving a new recogniser", filename)
                # save ROC
                with open(rocfilename, 'w') as f:
                    f.write(json.dumps(self.buildCNNWizard.cnntrain.ROCdata, indent=4))
            else:
                filename = os.path.join(self.filtersDir, self.buildCNNWizard.cnntrain.filterName)
                print("Updating the existing recogniser ", filename)
                # save ROC
                # TODO this was disabled here for some reason,
                # not sure what is the intended behaviour
                # with open(rocfilename, 'w') as f:
                #     f.write(json.dumps(self.buildCNNWizard.cnntrain.ROCdata, indent=4))

            # Store the recognizer txt
            with open(filename, 'w') as f:
                f.write(json.dumps(self.buildCNNWizard.cnntrain.currfilt, indent=4))
            # Actually copy the model
            copyfile(modelsrc, modelfile)
            copyfile(weightsrc, weightfile)
            # And remove temp dirs
            self.buildCNNWizard.cnntrain.tmpdir1.cleanup()
            self.buildCNNWizard.cnntrain.tmpdir2.cleanup()
            # prompt the user
            if test:
                msg = SupportClasses_GUI.MessagePopup("d", "Training completed!", "Training completed!\nProceeding to testing.")
            else:
                msg = SupportClasses_GUI.MessagePopup("d", "Training completed!", "Training completed!\nWe strongly recommend testing the recogniser on a separate dataset before actual use.")
            msg.exec()
            self.buildCNNWizard.done(1)
            if test:
                self.testRecogniser(filter=os.path.basename(filename))
        except Exception as e:
            print("ERROR: could not save recogniser because:", e)

    def saveRecogniserROC(self):
        # nothing to worry about CNN files, they are untouched
        try:
            if self.filterManager.saveoption == 'New':
                filename = os.path.join(self.filtersDir, self.filterManager.enterFiltName.text())
                print("Saving a new recogniser", filename)
                msgtext = "Saved as a new recogniser: " + self.filterManager.enterFiltName.text() + "\n\nWe strongly recommend testing the recogniser on a test dataset before actual use."
            else:
                filename = os.path.join(self.filtersDir, self.filterManager.listFiles.currentItem().text() + '.txt')
                print("Updating the existing recogniser ", filename)
                msgtext = "Updated the recogniser: " + self.filterManager.listFiles.currentItem().text()+ "\n\nWe strongly recommend testing the recogniser on a test dataset before actual use."

            # store the changed recognizer txt
            with open(filename, 'w') as f:
                f.write(json.dumps(self.filterManager.newfilter, indent=4))
            # prompt the user
            msg = SupportClasses_GUI.MessagePopup("d", "Saved!", msgtext)
            msg.exec()
        except Exception as e:
            print("ERROR: could not save recogniser because:", e)
        self.filterManager.close()

    def excel2Annotation(self):
        """ Utility function dialog: Generate AviaNZ style annotations given the start-end of calls in excel format
        """
        self.excel2AnnotationDialog = Dialogs.Excel2Annotation()
        self.excel2AnnotationDialog.show()
        self.excel2AnnotationDialog.activateWindow()
        self.excel2AnnotationDialog.btnGenerateAnnot.clicked.connect(self.genExcel2Annot)

    def tag2Annotation(self):
        """ Utility function dialog: Generate AviaNZ style annotations given freebird style (XML) annotations
        """
        self.tag2AnnotationDialog = Dialogs.Tag2Annotation()
        self.tag2AnnotationDialog.show()
        self.tag2AnnotationDialog.activateWindow()
        self.tag2AnnotationDialog.btnGenerateAnnot.clicked.connect(self.genTag2Annot)

    def backupAnnotations(self):
        """ Utility function dialog: backup annotation files
        """
        self.backupAnnotationDialog = Dialogs.BackupAnnotation()
        self.backupAnnotationDialog.show()
        self.backupAnnotationDialog.activateWindow()
        self.backupAnnotationDialog.btnCopyAnnot.clicked.connect(self.backupAnnotation)

    def genExcel2Annot(self):
        """ Utility function: Generate AviaNZ style annotations given the start-end of calls in excel format"""

        values = self.excel2AnnotationDialog.getValues()
        if values:
            [excelfile, audiofile, species, colstart, colend, collow, colhigh] = values
        else:
            return

        try:
            # Read excel file
            book = openpyxl.load_workbook(excelfile)
            sheet = book.active
            starttime = sheet[colstart+'2': colstart + str(sheet.max_row)]
            endtime = sheet[colend+'2': colend + str(sheet.max_row)]
            flow = sheet[collow+'2': collow + str(sheet.max_row)]
            fhigh = sheet[colhigh+'2': colhigh + str(sheet.max_row)]

            #_, duration, _, _ = wavio.readFmt(audiofile)
            wavobj = wavio.read(filename, 0, 0)
            duration = wavobj.nseconds

            annotation = []
            for i in range(len(starttime)):
                annotation.append([float(starttime[i][0].value), float(endtime[i][0].value), float(flow[i][0].value),
                                   float(fhigh[i][0].value),
                                   [{"species": species, "certainty": 100.0, "filter": "M", "calltype": species}]])
            annotation.insert(0, {"Operator": "", "Reviewer": "", "Duration": duration})
            file = open(audiofile + '.data', 'w')
            json.dump(annotation, file)
            file.close()
            self.excel2AnnotationDialog.txtSpecies.setText('')
            self.excel2AnnotationDialog.txtAudio.setText('')
            self.excel2AnnotationDialog.txtExcel.setText('')
            msg = SupportClasses_GUI.MessagePopup("d", "Generated annotation",
                                              "Successfully saved the annotation file: " + '\n' + audiofile + '.data')
            msg.exec()
        except Exception as e:
            print("ERROR: Generating annotation failed with error:")
            print(e)
            return

    def genTag2Annot(self):
        """ Utility function: Generate AviaNZ style annotations given the freebird style annotations
        There are 3 parts to Freebird tags: x.tag, x.p, s.sample.
        x.p has time: StartTimeSecond and DurationSecond. What are they?
        x.setting has view info, which we ignore
        x.tag has species code, time, duration, freqlow and freqhigh
        There is also the species list. Which we need to store and copy into .avianz.
        """

        # TODO: sort out the frequencies
        # TODO: test that it saves to the right folder
        sessiondir = self.tag2AnnotationDialog.getValues()
        if sessiondir is None:
            return

        if sessiondir.endswith(".session"):
            sessiondir = sessiondir[:-8]
            #sessiondir = os.path.join(sessiondir,".session")
        print(sessiondir)

        spName = []
        spCode = []

        if not os.path.isabs(self.config['FreebirdList']):
            filename = os.path.join(self.configdir,self.config['FreebirdList'])
        else:
            filename = self.config['FreebirdList']

        if self.config['FreebirdList'][-4:] == '.csv':
            try:
                with open(filename, mode='r') as f:
                #with open(os.path.join(self.configdir,self.config['FreebirdList']), mode='r') as f:
                    cs = csv.DictReader(f)
                    for l in cs:
                        if l['FreebirdCode'] != '':
                            spName.append(l['SpeciesName'])
                            spCode.append(int(l['FreebirdCode']))

                f.close()
            except:
                print("Warning: Did not find Freebird species list")
        elif self.config['FreebirdList'][-5:] == '.xlsx':
            try:
                book = openpyxl.load_workbook(filename)
                sheet = book.active
            except:
                print("Warning: Did not find Freebird species list")

                name = sheet['A2': 'A' + str(sheet.max_row)]
                code = sheet['B2': 'B' + str(sheet.max_row)]
    
                for i in range(len(name)):
                    spName.append(str(name[i][0].value))
                for i in range(len(code)):
                    if code[i][0].value is not None:
                        spCode.append(int(code[i][0].value))
                    else:
                        spCode.append(-1)

        spDict = dict(zip(spCode, spName))

        # Go into each .session folder
        # Generate the .data files from .tag, read operator/reviewer from the corresponding .setting file
        for root, dirs, files in os.walk(sessiondir):
            for file in files:
                if file.endswith('.tag'):
                    tagFile = os.path.join(root, file)
                    #print(tagFile)
                    tagSegments = Segment.SegmentList()

                    # First get the metadata
                    operator = ""
                    reviewer = ""
                    duration = ""
                    try:
                        stree = ET.parse(tagFile[:-4] + '.setting')
                        stroot = stree.getroot()
                        for elem in stroot:
                            if elem.tag == 'Operator':
                                operator = elem.text
                            if elem.tag == 'Reviewer' and elem.text:
                                reviewer = elem.text
                    except:
                        print("Can't read %s.setting or missing data" %tagFile[:-4])
                    try:
                        # Read the duration from the sample if possible
                        ptree = ET.parse(tagFile[:-4] + '.p')
                        ptroot = ptree.getroot()
                        for elem in ptroot:
                            for elem2 in elem:
                                if elem2.tag == 'DurationSecond':
                                    duration = elem2.text
                    except:
                        print("Can't read %s.p or missing data" %tagFile[:-4])
                        # Otherwise, load the wav file
                        # TODO: Test
                        import Spectrogram 
                        sp = Spectrogram.Spectrogram(512,256, 0, 0)
                        sp.readWav(tagFile[:-4] + '.wav', 0, 0)
                        duration = sp.fileLength / sp.audioFormat.sampleRate()
                        #duration = sp.fileLength / sp.sampleRate
           
                    tagSegments.metadata = {"Operator": operator, "Reviewer": reviewer, "Duration": duration}
                                
                    try:
                        tree = ET.parse(tagFile)
                        troot = tree.getroot()
          
                        for elem in troot:
                            try:
                                species = [{"species": spDict[int(elem[0].text)], "certainty": 100, "filter": "M"}]
                                # TODO: Get the size right! Something weird about the freqs
                                newSegment = Segment.Segment([float(elem[1].text), float(elem[1].text) + float(elem[2].text), 0,0, species])
                                #newSegment = Segment.Segment([float(elem[1].text), float(elem[1].text) + float(elem[2].text), float(elem[3].text), float(elem[4].text), species])
                                tagSegments.append(newSegment)
                                #print(tagSegments)
                            except KeyError:
                                print("{0} not in bird list for file {1}".format(elem[0].text,tagFile))
                    except Exception as e:
                        print("Can't read %s or missing data" %tagFile)
                        print("Warning: Generating annotation from %s failed with error:" % (tagFile))
                        print(e)
                
                    # save .data, possible over-writing TODO
                    # Don't want in the .session folder
                    if root[-8:] == ".session":
                       di = root[:-8] 
                    else:
                        di = root
                    tagSegments.saveJSON(os.path.join(di,file[:-4] + '.wav.data'))
                    #print("saving to",os.path.join(di,file[:-4] + '.wav.data'))
         
        #self.tag2AnnotationDialog.txtDuration.setText('')
        self.tag2AnnotationDialog.txtSession.setText('')
        msg = SupportClasses_GUI.MessagePopup("d", "Generated annotation", "Successfully saved the annotations in: " + '\n' + sessiondir)
        msg.exec()
        
    def genTag2Annot_xlsx_TBD(self):
        """ Utility function: Generate AviaNZ style annotations given the freebird style annotations
        There are 3 parts to Freebird tags: x.tag, x.p, s.sample.
        x.p has time: StartTimeSecond and DurationSecond. What are they?
        x.setting has view info, which we ignore
        x.tag has species code, time, duration, freqlow and freqhigh
        There is also the species list. Which we need to store and copy into .avianz.
        """

        # TODO: Remove duration from dialog
        sessiondir = self.tag2AnnotationDialog.getValues()
        if sessiondir is None:
            return

        spName = []
        spCode = []

        if self.config['FreebirdList'][-4:] == '.csv':
            try:
                with open(self.config['FreebirdList'], mode='r') as f:
                    cs = csv.DictReader(f)
                    for l in cs:
                        if l['FreebirdCode'] != '':
                            spName.append(l['SpeciesName'])
                            spCode.append(int(l['FreebirdCode']))

                f.close()
            except:
                print("Warning: Did not find Freebird species list")
        elif self.config['FreebirdList'][-5:] == '.xlsx':
            try:
                book = openpyxl.load_workbook(os.path.join(self.config['FreebirdList']))
                sheet = book.active
            except:
                print("Warning: Did not find Freebird species list")

                name = sheet['A2': 'A' + str(sheet.max_row)]
                code = sheet['B2': 'B' + str(sheet.max_row)]
    
                for i in range(len(name)):
                    spName.append(str(name[i][0].value))
                for i in range(len(code)):
                    if code[i][0].value is not None:
                        spCode.append(int(code[i][0].value))
                    else:
                        spCode.append(-1)
                spDict = dict(zip(spCode, spName))

        # Generate the .data files from .tag, read operator/reviewer from the corresponding .setting file
        for root, dirs, files in os.walk('Freebird'):
            for file in files:
                if file.endswith('.tag'):
                    tagFile = os.path.join(root, file)
                    tagSegments = Segment.SegmentList()

                    # First get the metadata
                    operator = ""
                    reviewer = ""
                    duration = ""
                    try:
                        stree = ET.parse(tagFile[:-4] + '.setting')
                        stroot = stree.getroot()
                        for elem in stroot:
                            if elem.tag == 'Operator':
                                operator = elem.text
                            if elem.tag == 'Reviewer' and elem.text:
                                reviewer = elem.text
                    except:
                        print("Can't read %s.setting or missing data" %tagFile[:-4])
                    try:
                        # Read the duration from the sample if possible
                        ptree = ET.parse(tagFile[:-4] + '.p')
                        ptroot = ptree.getroot()
                        for elem in ptroot:
                            for elem2 in elem:
                                if elem2.tag == 'DurationSecond':
                                    duration = elem2.text
                    except:
                        print("Can't read %s.p or missing data" %tagFile[:-4])
                        # Otherwise, load the wav file
                        import Spectrogram 
                        sp = Spectrogram.Spectrogram(512,256, 0, 0)
                        sp.readWav(tagFile[:-4] + '.wav', 0, 0)
                        duration = sp.fileLength / sp.sampleRate
                        #duration = sp.fileLength / sp.sampleRate
        
                    tagSegments.metadata = {"Operator": operator, "Reviewer": reviewer, "Duration": duration}
                        
                    try:
                        tree = ET.parse(tagFile)
                        troot = tree.getroot()
        
                        for elem in troot:
                            try:
                                species = [{"species": spDict[int(elem[0].text)], "certainty": 100, "filter": "M"}]
                                # TODO: Get the size right! Something weird about the freqs
                                newSegment = Segment.Segment([float(elem[1].text), float(elem[1].text) + float(elem[2].text), 0,0, species])
                                #newSegment = Segment.Segment([float(elem[1].text), float(elem[1].text) + float(elem[2].text), float(elem[3].text), float(elem[4].text), species])
                                tagSegments.append(newSegment)
                                #print(tagSegments)
                            except KeyError:
                                print("{0} not in bird list for file %s" %elem[0],tagFile)
                    except Exception as e:
                        print("Can't read %s or missing data" %tagFile)
                        print("Warning: Generating annotation from %s failed with error:" % (tagFile))
                        print(e)
        
                    # save .data, possible over-writing *TODO
                    tagSegments.saveJSON('../' + tagFile[:-4] + '.wav.data')
        
            #self.tag2AnnotationDialog.txtDuration.setText('')
            self.tag2AnnotationDialog.txtSession.setText('')
            msg = SupportClasses_GUI.MessagePopup("d", "Generated annotation", "Successfully saved the annotations in: " + '\n' + sessiondir)
            msg.exec()
        


    def genTag2Annot_TBDeleted(self):
        """ Utility function: Generate AviaNZ style annotations given the freebird style annotations"""
        """ Holy shit this is an embarrassment, Nirosha."""
        """ There are 3 parts to Freebird tags: x.tag, x.p, s.sample.
        x.p has time: StartTimeSecond and DurationSecond. What are they?
        x.setting has view info, which we ignore
        x.tag has species code, time, duration, freqlow and freqhigh
        There is also the species list. Which we need to store and copy into .avianz.
        """

        # TODO: Remove duration from dialog
        values = self.tag2AnnotationDialog.getValues()
        if values:
            [sessiondir, duration] = values
        else:
            return

        # Read freebird bird list
        # TODO: make it part of the setup
        spName = []
        spCode = []
        try:
            book = openpyxl.load_workbook(os.path.join(self.configdir, "Freebird_species_list.xlsx"))
            sheet = book.active
        except:
            print("Warning: Did not find Freebird species list")
            return

        name = sheet['A2': 'A' + str(sheet.max_row)]
        code = sheet['B2': 'B' + str(sheet.max_row)]
        for i in range(len(name)):
            spName.append(str(name[i][0].value))
        for i in range(len(code)):
            if code[i][0].value is not None:
                spCode.append(int(code[i][0].value))
            else:
                spCode.append(-1)
        spDict = dict(zip(spCode, spName))

        # Generate the .data files from .tag, read operator/reviewer from the corresponding .setting file
        for root, dirs, files in os.walk(str(sessiondir)):
            for file in files:
                if file.endswith('.tag'):
                    tagFile = os.path.join(root, file)
                    tagSegments = Segment.SegmentList()
                    try:
                        # First get the metadata
                        operator = ""
                        reviewer = ""
                        stree = ET.parse(tagFile[:-4] + '.setting')
                        stroot = stree.getroot()
                        for elem in stroot:
                            if elem.tag == 'Operator':
                                operator = elem.text
                            if elem.tag == 'Reviewer' and elem.text:
                                reviewer = elem.text
                        annotation.insert(0, {"Operator": operator, "Reviewer": reviewer, "Duration": duration})
                        # Read the duration from the sample if possible
                        # TODO
                        # Otherwise, read file in
                        # TODO

                        tagSegments.metadata = {"Operator": operator, "Reviewer": reviewer, "Duration": self.datalength / self.sp.audioFormat.sampleRate()}
                        #tagSegments.metadata = {"Operator": operator, "Reviewer": reviewer, "Duration": self.sp.fileLength / self.sp.sampleRate}
                        
                        tree = ET.parse(tagFile)
                        troot = tree.getroot()

                        for elem in troot:
                            try:
                                species = spDict[int(elem[0].text)]
                                # TODO: Get the size right!
                                newSegment = Segment.Segment([float(elem[1].text), float(elem[1].text) + float(elem[2].text), elem[3].text, elem[4].text, species])
                                tagSegments.append(newSegment)
                            except KeyError:
                                print("{0} not in bird list for file %s" %elem[0],tagfile)

                        # save .data, possible over-writing
                        # TODO!!
                        self.segments.saveJSON(str(self.filename) + '.data')
                        #file = open(tagFile[:-4] + '.wav.data', 'w')
                        #json.dump(annotation, file)
                        #file.close()
                    except Exception as e:
                        print("Warning: Generating annotation from %s failed with error:" % (tagFile))
                        print(e)
                        return

            # What is this?
            self.tag2AnnotationDialog.txtDuration.setText('')
            self.tag2AnnotationDialog.txtSession.setText('')
            msg = SupportClasses_GUI.MessagePopup("d", "Generated annotation", "Successfully saved the annotations in: " + '\n' + sessiondir)
            msg.exec()

    def backupAnnotation(self):
        """ Utility function: Copy .data and corrections files while preserving directory hierarchy"""
        # TODO: Test!!!
        values = self.backupAnnotationDialog.getValues()
        if values:
            [src, dst] = values
            print(src,dst)
        else:
            return

        l = len(src)
        for root, dirs, files in os.walk(src):
            for d in dirs:
                #print(dst,root,dirs)
                os.mkdir(os.path.join(dst,root[l+1:],d))
            for f in files:
                if f[-5:].lower() == '.data' or 'corrections' in f or 'BatData' in f or 'BatPasses' in f:
                    shutil.copy2(os.path.join(root, f),os.path.join(dst,root[l+1:]))
        self.backupAnnotationDialog.close()
        
        #try:
            #if platform.system() == 'Windows':
                #subprocess.call(['xcopy', src+'\*.data', dst, '/s', '/e'])
            #elif platform.system() == 'Linux' or platform.system() == 'Darwin':     # TODO: zero testing!
        #except Exception as e:
            #print("Warning: Coping failed with error:")
            #print(e)
            #return

    def segmentationDialog(self):
        """ Create the segmentation dialog when the relevant button is pressed.
        """
        maxampl = 0.001
        if self.datalength>0:
            maxampl = np.max(self.sp.data)
        self.segmentDialog = Dialogs.Segmentation(maxampl,DOC=self.DOC, species=self.FilterDicts)
        self.segmentDialog.show()
        self.segmentDialog.activateWindow()
        self.segmentDialog.undo.clicked.connect(self.segment_undo)
        self.segmentDialog.activate.clicked.connect(self.segment)

    def segment(self):
        """ Listener for the segmentation dialog. Calls the relevant segmenter.
        """
        if self.CLI:
            maxampl = 0.001
            if self.datalength>0:
                maxampl = np.max(self.sp.data)
            self.segmentDialog = Dialogs.Segmentation(maxampl)

        opstartingtime = time.time()
        print('Segmenting requested at ' + time.strftime('%H:%M:%S', time.localtime()))
        # for undoing:
        self.prevSegments = copy.deepcopy(self.segments)

        self.segmentsToSave = True
        # settings is a dict with parameters for various possible methods
        alg, settings = self.segmentDialog.getValues()

        with pg.BusyCursor():
            filtname = str(settings["filtname"])
            self.statusLeft.setText('Segmenting...')
            # Delete old segments:
            # only this species, if using species-specific methods:
            if alg == 'Wavelet Filter' or alg == 'WV Changepoint':
                if filtname == 'Choose species...':
                    msg = SupportClasses_GUI.MessagePopup("w", "Species Error", 'Please select your species!')
                    msg.exec()
                    return

                filtspecies = self.FilterDicts[filtname]["species"]
                oldsegs = self.segments.getSpecies(filtspecies)
                # Only show segments which are at least partly visible in this page:
                for ix in reversed(oldsegs):
                    seg = self.segments[ix]
                    if seg[0] > self.startRead + self.datalengthSec or seg[1] < self.startRead:
                        oldsegs.remove(ix)

                todelete = []
                # deleting from the end, because deleteSegments shifts IDs:
                for si in reversed(oldsegs):
                    # DO NOT delete segments in other pages
                    if self.listRectanglesa1[si] is None:
                        continue
                    # clear these species from overview colors
                    self.refreshOverviewWith(self.segments[si], delete=True)
                    # remove all labels for the current species
                    wipedAll = self.segments[si].wipeSpecies(filtspecies)
                    self.refreshOverviewWith(self.segments[si])
                    # drop the segment if it's the only species, or just update the graphics
                    if wipedAll:
                        todelete.append(si)
                    else:
                        self.updateText(si)
                        self.updateColour(si)
                # reverse loop to allow deleting segments
                for dl in todelete:
                    self.deleteSegment(dl)
            else:
                self.removeSegments()

            # NON-SPECIFIC methods here (produce "Don't Know"):
            if alg == 'Default':
                newSegments = self.seg.bestSegments()
            elif alg == 'Median Clipping':
                newSegments = self.seg.medianClip(settings["medThr"], minSegment=self.config['minSegment'])
                newSegments = self.seg.checkSegmentOverlap(newSegments)
                # will also remove too short segments (medSize is set in ms because sliders limited to int)
                # print("before length", newSegments)
                # newSegments = self.seg.deleteShort(newSegments, minlength=medSize/1000)
            elif alg == 'Harma':
                newSegments = self.seg.Harma(float(str(settings["HarmaThr1"])),float(str(settings["HarmaThr2"])),minSegment=self.config['minSegment'])
                newSegments = self.seg.checkSegmentOverlap(newSegments)
            elif alg == 'Power':
                newSegments = self.seg.segmentByPower(float(str(settings["PowerThr"])))
                newSegments = self.seg.checkSegmentOverlap(newSegments)
            elif alg == 'Onsets':
                newSegments = self.seg.onsets()
                newSegments = self.seg.checkSegmentOverlap(newSegments)
            elif alg == 'Fundamental Frequency':
                newSegments = self.seg.yinSegs(int(str(settings["FFminfreq"])), int(str(settings["FFminperiods"])), float(str(settings["Yinthr"])),
                                                         int(str(settings["FFwindow"])))
                newSegments = self.seg.checkSegmentOverlap(newSegments)
            elif alg == 'FIR':
                newSegments = self.seg.segmentByFIR(float(str(settings["FIRThr1"])))
                newSegments = self.seg.checkSegmentOverlap(newSegments)
            # SPECIES-SPECIFIC methods from here:
            elif alg == 'Wavelet Filter':
                # Old WF filter, not compatible with wind removal:
                speciesData = self.FilterDicts[filtname]
                ws = WaveletSegment.WaveletSegment(speciesData)
                ws.readBatch(self.sp.data, self.sp.audioFormat.sampleRate(), d=False, spInfo=[speciesData], wpmode="new", wind=False)
                #ws.readBatch(self.sp.data, self.sp.sampleRate, d=False, spInfo=[speciesData], wpmode="new", wind=False)
                newSegments = ws.waveletSegment(0, wpmode="new")
                # this will produce a list of lists (over subfilters)
            elif alg == 'WV Changepoint':
                print("Changepoint detection requested")
                speciesData = self.FilterDicts[filtname]
                # this will produce a list of lists (over subfilters)
                ws = WaveletSegment.WaveletSegment(speciesData)
                ws.readBatch(self.sp.data, self.sp.audioFormat.sampleRate(), d=False, spInfo=[speciesData], wpmode="new", wind=settings["wind"]>0)
                #ws.readBatch(self.sp.data, self.sp.sampleRate, d=False, spInfo=[speciesData], wpmode="new", wind=settings["wind"]>0)
                # nuisance-signal changepoint detector (alg 2)
                # with all params passed:
                newSegments = ws.waveletSegmentChp(0, alpha=settings["chpalpha"], window=settings["chpwindow"], maxlen=settings["maxlen"], alg=2, silent=False, wind=settings["wind"])

            # TODO: make sure cross corr outputs lists of lists
            elif alg == 'Cross-Correlation':
                if filtname != 'Choose species...':
                    # need to load template/s
                    newSegments = self.findMatches(float(str(settings["CCThr1"])), filtname)
                else:
                    newSegments = self.findMatches(float(str(settings["CCThr1"])))
            else:
                print("ERROR: unrecognised algorithm", alg)
                return

            # Post-process
            # 1. Delete windy segments
            # 2. Delete rainy segments
            # 3. Check fundamental frq
            # 4. Merge neighbours
            # 5. Delete short segmentsost process to remove short segments, wind, rain, and use F0 check.
            if alg == 'Wavelet Filter' or alg == 'WV Changepoint':
                print('Segments detected: ', sum(isinstance(seg, list) for subf in newSegments for seg in subf))
                print(newSegments)
                print('Post-processing...')
                # load target CNN model if exists
                self.CNNDicts = self.ConfigLoader.CNNmodels(self.FilterDicts, self.filtersDir, [filtname])
                # postProcess currently operates on single-level list of segments,
                # so we run it over subfilters for wavelets:
                for filtix in range(len(speciesData['Filters'])):
                    subfilter = speciesData['Filters'][filtix]
                    CNNmodel = None
                    if 'CNN' in speciesData:
                        CNNmodel = self.CNNDicts.get(speciesData['CNN']['CNN_name'])

                    post = Segment.PostProcess(configdir=self.configdir, audioData=self.sp.data, sampleRate=self.sp.audioFormat.sampleRate(),
                    #post = Segment.PostProcess(configdir=self.configdir, audioData=self.sp.data, sampleRate=self.sp.sampleRate,
                                               tgtsampleRate=speciesData["SampleRate"], segments=newSegments[filtix],
                                               subfilter=subfilter, CNNmodel=CNNmodel, cert=50)
                    # Deprecated wind filter:
                    # if settings["windold"]:
                    #     post.wind()
                    #     print('After wind: segments: ', len(post.segments))
                    if CNNmodel:
                        print('Post-processing with CNN')
                        post.CNN()
                        print('After CNN: segments: ', len(post.segments))
                    if settings["rain"]:
                        post.rainClick()
                        print('After rain segments: ', len(post.segments))
                    if 'F0' in subfilter and 'F0Range' in subfilter:
                        if subfilter['F0']:
                            print("Checking for fundamental frequency...")
                            post.fundamentalFrq()
                            print("After FF segments:", len(post.segments))
                    if alg=='Wavelet Filter':
                        post.joinGaps(maxgap=subfilter['TimeRange'][3])
                    if subfilter['TimeRange'][0]>0:
                        post.deleteShort(minlength=subfilter['TimeRange'][0])

                    newSegments[filtix] = post.segments
            else:
                print('Segments detected: ', len(newSegments))
                print('Post-processing...')
                post = Segment.PostProcess(configdir=self.configdir, audioData=self.sp.data, sampleRate=self.sp.audioFormat.sampleRate(), segments=newSegments, subfilter={})
                #post = Segment.PostProcess(configdir=self.configdir, audioData=self.sp.data, sampleRate=self.sp.sampleRate, segments=newSegments, subfilter={})
                # if settings["windold"]:
                #     post.wind()
                #     print('After wind segments: ', len(post.segments))
                if settings["rain"]:
                    post.rainClick()
                    print('After rain segments: ', len(post.segments))
                post.joinGaps(maxgap=settings["maxgap"])
                post.deleteShort(minlength=settings["minlen"])
                newSegments = post.segments
            print("After post processing: ", newSegments)

            # Generate Segment-type output.
            if alg=='Wavelet Filter' or alg=='WV Changepoint':
                for filtix in range(len(speciesData['Filters'])):
                    speciesSubf = speciesData['Filters'][filtix]
                    y1 = speciesSubf['FreqRange'][0]
                    y2 = min(self.sp.audioFormat.sampleRate()//2, speciesSubf['FreqRange'][1])
                    #y2 = min(self.sp.sampleRate//2, speciesSubf['FreqRange'][1])
                    for seg in newSegments[filtix]:
                        self.addSegment(float(seg[0][0]), float(seg[0][1]), y1, y2,
                                [{"species": filtspecies, "certainty": seg[1], "filter": filtname, "calltype": speciesSubf["calltype"]}], index=-1)
                        self.segmentsToSave = True
            elif alg=='Cross-Correlation' and filtname != 'Choose species...':
                # TODO: this has not been updated for a while
                filtspecies = self.FilterDicts[filtname]["species"]
                for filtix in range(len(speciesData['Filters'])):
                    speciesSubf = speciesData['Filters'][filtix]
                    y1 = speciesSubf['FreqRange'][0]
                    y2 = min(self.sp.audioFormat.sampleRate()//2, speciesSubf['FreqRange'][1])
                    #y2 = min(self.sp.sampleRate//2, speciesSubf['FreqRange'][1])
                    for seg in newSegments[filtix]:
                        self.addSegment(float(seg[0]), float(seg[1]), y1, y2,
                                [{"species": filtspecies, "certainty": seg[1]}], index=-1)
                        self.segmentsToSave = True
            else:
                for seg in newSegments:
                    self.addSegment(seg[0][0],seg[0][1])
                    self.segmentsToSave = True

            self.segmentDialog.undo.setEnabled(True)
            self.statusLeft.setText('Ready')
        print('Segmentation finished at %s' % (time.time() - opstartingtime))

    def segment_undo(self):
        """ Listener for undo button in segmentation dialog.
            Deletes everything, and re-adds segments from a backup.
        """
        # just in case:
        self.segmentDialog.undo.setEnabled(False)
        if not hasattr(self, 'prevSegments'):
            print("Nothing to undo!")
            return

        self.removeSegments()
        # This recreates the previous segments (from all pages):
        self.segments = copy.deepcopy(self.prevSegments)
        # So here we only need to show them:
        for seg in self.prevSegments:
            self.addSegment(seg[0], seg[1], seg[2], seg[3], seg[4], saveSeg=False, coordsAbsolute=True)
        self.segmentsToSave = True

    def exportSeg(self):
        # TODO: Used?
        # First, deal with older xls if present:
        foundxls = []
        for f in os.listdir(self.SoundFileDir):
            if f.startswith("DetectionSummary_") and f.endswith(".xlsx"):
                foundxls.append(f)

        if len(foundxls)>0:
            # check with user
            msg = SupportClasses_GUI.MessagePopup("w", "Excel file exists", "Detection summaries already present in " + self.SoundFileDir + ". Overwrite them, append to them, or cancel the operation?")
            msg.setStandardButtons(QMessageBox.StandardButton.Cancel)
            msg.addButton("Overwrite", QMessageBox.ButtonRole.YesRole)
            msg.addButton("Append", QMessageBox.ButtonRole.YesRole)
            # cancelBtn = msg.addButton(QMessageBox.Cancel)
            reply = msg.exec()
            # print(reply)
            if reply == 4194304:  # weird const for Cancel
                return
            elif reply == 1:
                action = "append"
            elif reply == 0:
                action = "overwrite"
            else:
                print("ERROR: Unrecognised reply", reply)
                return

            # remove all the old excels:
            if action == "overwrite":
                for f in foundxls:
                    try:
                        os.remove(os.path.join(self.SoundFileDir, f))
                    except Exception as e:
                        print("Could not remove file", os.path.join(self.SoundFileDir, f))
                        print(e)
        else:
            # create new workbook, in effect
            action = "overwrite"

        # sort the segments into increasing time order (to make neater output)
        sortOrder = self.segments.orderTime()
        self.listRectanglesa1 = [self.listRectanglesa1[i] for i in sortOrder]
        self.listRectanglesa2 = [self.listRectanglesa2[i] for i in sortOrder]
        self.listLabels = [self.listLabels[i] for i in sortOrder]

        # excel should be split by page size, but for short files just give the file size
        datalen = self.config['maxFileShow'] if self.nFileSections>1 else self.datalengthSec
        excel = SupportClasses.ExcelIO()
        self.segments.filename = self.filename
        success = excel.export([self.segments], self.SoundFileDir, action=action, pagelenarg=datalen, numpages=self.nFileSections, startTime=self.startTime, precisionMS=self.batmode)
        # add user notification
        if success==0:
            print("Warning: Excel output was not saved")
            return
        else:
            msg = SupportClasses_GUI.MessagePopup("d", "Segments Exported", "Check this directory for the Excel output: " + '\n' + self.SoundFileDir)
            msg.exec()
            return

    def findMatches(self,thr=0.4, species='Choose species...'):
        """ Calls the cross-correlation function to find matches like the currently highlighted box.
        It also check if you have selected a species, then allow to read those templates and match.
        """
        # TODO: Remove?
        # print ("inside find Matches: ", species)
        segments = []
        if species != 'Choose species...' and os.path.exists('Sound Files/' + species):
            self.statusLeft.setText("Finding matches...")
            print("Reading template/s")
            # Todo: do more than one template and merge result?
            sp_temp = Spectrogram.Spectrogram(self.config['window_width'], self.config['incr'])
            sp_temp.readWav('Sound Files/'+species+'/train1_1.wav')

            # Parse wav format details based on file header:
            sampleRate = sp_temp.audioFormat.sampleRate()
            #sampleRate = sp_temp.sampleRate
            audiodata = sp_temp.data

            # downsample
            print("fs: ", sampleRate, self.sppInfo[str(species)][4])
            #print("fs: ", sampleRate, self.sppInfo[str(species)][4])
            if sampleRate != self.sppInfo[str(species)][4]:
                sp_temp.resample(self.sppInfo[str(species)][4])
            datalength = np.shape(audiodata)[0]
            len_seg = datalength / sampleRate

            sgRaw_temp = sp_temp.spectrogram(mean_normalise=self.sgMeanNormalise,
                                        equal_loudness=self.sgEqualLoudness, onesided=self.config['sgOneSided'])

            # Get the data for the spectrogram
            if self.sp.audioFormat.sampleRate() != self.sppInfo[str(species)][4]:
            #if self.sp.sampleRate != self.sppInfo[str(species)][4]:
                data1 = self.sp.resample(self.sppInfo[str(species)][4])
                sampleRate1 = self.sppInfo[str(species)][4]
            else:
                data1 = self.sp.data
                sampleRate1 = self.sp.audioFormat.sampleRate()
                #sampleRate1 = self.sp.sampleRate
            # TODO utilize self.sp / Spectrogram more here
            sp_temp.data = data1
            sp_temp.audioFormat.setSampleRate(sampleRate1)
            #sp_temp.sampleRate = sampleRate1
            #sgRaw = self.sp.spectrogram(window=str(self.windowType),sgType=str(self.sgType),sgScale=str(self.sgScale),nfilters=int(str(self.nfilters)),mean_normalise=self.sgMeanNormalise,equal_loudness=self.sgEqualLoudness,onesided=self.config['sgOneSided'])
            sgRaw = self.sp.spectrogram(window_width=self.config['window_width'], incr=self.config['incr'],window=self.config['windowType'],sgType=self.config['sgType'],sgScale=self.config['sgScale'],nfilters=self.config['nfilters'],mean_normalise=self.config['sgMeanNormalise'],equal_loudness=self.config['sgEqualLoudness'],onesided=self.config['sgOneSided'])
            indices = self.seg.findCCMatches(sgRaw_temp,sgRaw,thr)
            # scale indices to match with self.samplerate
            indices = [i*self.sp.audioFormat.sampleRate()/sampleRate1 for i in indices]
            #indices = [i*self.sp.sampleRate/sampleRate1 for i in indices]
            # print('indices:', indices)
            # identifySegments(seg=indices, minlength=10)
            # indices are in spectrogram pixels, need to turn into times
            y1 = self.convertFreqtoY(self.sppInfo[str(species)][2]/2)
            if self.sppInfo[str(species)][4]/2 > self.sp.audioFormat.sampleRate():
            #if self.sppInfo[str(species)][4]/2 > self.sp.sampleRate:
                y2 = self.convertFreqtoY(self.sp.audioFormat.sampleRate() / 2 - self.sp.audioFormat.sampleRate() * 0.01)
                #y2 = self.convertFreqtoY(self.sp.sampleRate / 2 - self.sp.sampleRate * 0.01)
            else:
                y2 = self.convertFreqtoY(self.sppInfo[str(species)][4] / 2)
            for i in indices:
                if np.abs(i) > self.config['overlap_allowed']:
                    time = i*self.config['incr'] / self.sp.audioFormat.sampleRate()
                    #time = i*self.config['incr'] / self.sp.sampleRate
                    # print(time, time + len_seg,self.segments)
                    # self.addSegment(time, time+len_seg,y1,y2,[species+'?'])
                    segments.append([time, time+len_seg])
        elif self.box1id is None or self.box1id<0:
            print("No box selected")
            msg = SupportClasses_GUI.MessagePopup("w", "No segment", "No segment selected to match")
            msg.exec()
            return []
        else:
            self.statusLeft.setText("Finding matches...")
            # Only want to draw new segments, so find out how many there are now
            seglen = len(self.segments)
            # Get the segment -- note that takes the full y range
            if type(self.listRectanglesa2[self.box1id]) == self.ROItype:
                x1 = self.listRectanglesa2[self.box1id].pos().x()
                x2 = x1 + self.listRectanglesa2[self.box1id].size().x()
            else:
                x1, x2 = self.listRectanglesa2[self.box1id].getRegion()
            # Get the data for the spectrogram
            # TODO utilize self.sp / Spectrogram more here
            #sgRaw = self.sp.spectrogram(window=str(self.windowType),sgType=str(self.sgType),sgScale=str(self.sgScale),nfilters=int(str(self.nfilters)),mean_normalise=self.sgMeanNormalise,equal_loudness=self.sgEqualLoudness,onesided=self.config['sgOneSided'])
            sgRaw = self.sp.spectrogram(window_width=self.config['window_width'], incr=self.config['incr'],window=self.config['windowType'],sgType=self.config['sgType'],sgScale=self.config['sgScale'],nfilters=self.config['nfilters'],mean_normalise=self.config['sgMeanNormalise'],equal_loudness=self.config['sgEqualLoudness'],onesided=self.config['sgOneSided'])
            segment = sgRaw[int(x1):int(x2),:]
            len_seg = (x2-x1) * self.config['incr'] / self.sp.audioFormat.sampleRate()
            #len_seg = (x2-x1) * self.config['incr'] / self.sp.sampleRate
            indices = self.seg.findCCMatches(segment,sgRaw,thr)
            # indices are in spectrogram pixels, need to turn into times
            for i in indices:
                # Miss out the one selected: note the hack parameter
                if np.abs(i-x1) > self.config['overlap_allowed']:
                    time = i*self.config['incr'] / self.sp.audioFormat.sampleRate()
                    #time = i*self.config['incr'] / self.sp.sampleRate
                    segments.append([time, time+len_seg])
                    # self.addSegment(time, time+len_seg,0,0,self.segments[self.box1id][4])
            self.statusLeft.setText("Ready")
        return segments

    def classifySegments(self):
        """Listener for Action->Cluster segments menu item, cluster segments marked in the current file. Only to display
            the auto generated clusters
        """
        # TODO: Probably broken!
        if len(self.segments) > 1:
            cl = Clustering.Clustering([], [], 5)
            # TODO: This is the signature
            #def cluster(self, dataset, fs, species, feature='we', n_mels=24, minlen=0.2, denoise=False, alg='agglomerative'):
            segments, nclasses, duration = cl.cluster(dataset,self.sp.audioFormat.sampleRate(), None, feature='we')
            #segments, nclasses, duration = cl.cluster(dataset,self.sp.sampleRate, None, feature='we')
            self.clusterD = Dialogs.Cluster(segments, self.sp.audioFormat.sampleRate(), nclasses, self.config)
            #self.clusterD = Dialogs.Cluster(segments, self.sp.sampleRate, nclasses, self.config)
            self.clusterD.show()
        else:
            print('need segments to cluster!')
            return

# ===============
# Code for playing sounds
    def playVisible(self):
        """ Listener for button to play the visible area.
        On PLAY, turns to PAUSE and two other buttons turn to STOPs.
        If PAUSED, we just want to unpause.
        """
        if self.batmode:
            # Currently playback disabled in this mode - also takes care of spacebar signal
            return

        if self.media_obj.isPlaying():
            self.pausePlayback()
        else:
            self.bar.setMovable(False)
            self.swapPlayButtonState(True)

            if self.media_obj.isPlayingorPaused():
                self.media_obj.pressedPlay()
            else:
                self.segmentStart = self.p_ampl.viewRange()[0][0]*1000
                self.segmentStop = self.p_ampl.viewRange()[0][1]*1000

                # if bar was moved under pause, update the playback start position based on the bar:
                if self.bar.value()>0:
                    start = self.convertSpectoAmpl(self.bar.value())*1000  
                    print("found bar at %d ms" % start)
                else:
                    start = self.segmentStart

                self.media_obj.pressedPlay(start=start, stop=self.segmentStop)
            
            self.speedButton.setEnabled(False)
            #self.NotifyTimer.start(30)

    def playSelectedSegment(self,low=None,high=None):
        """ Listener for PlaySegment button (also called by listener for PlayBandlimitedSegment).
        Get selected segment start and end (or return if no segment selected).
        On PLAY, all three buttons turn to STOPs.
        """
        # Something is passed in from the event, trap it
        if not low:
            low = None

        if self.media_obj.isPlayingorPaused():
            self.stopPlayback()
        elif self.box1id > -1:
            self.segmentStart = self.listRectanglesa1[self.box1id].getRegion()[0] * 1000
            self.segmentStop = self.listRectanglesa1[self.box1id].getRegion()[1] * 1000

            self.bar.setMovable(False)
            self.swapPlayButtonState(True)

            self.media_obj.playSeg(self.segmentStart, self.segmentStop, speed=self.playSpeed, low=low, high=high)
            self.speedButton.setEnabled(False)
            #self.NotifyTimer.start(30)
        #else:
            #print("Can't play, no segment selected")

    def playBandLimitedSegment(self):
        """ Listener for PlayBandlimitedSegment button.
        Gets the band limits of the segment and passes them on.
        """
        if self.media_obj.isPlayingorPaused():
            self.stopPlayback()
        elif self.box1id > -1:
            low = max(0.1, self.sp.minFreq, self.segments[self.box1id][2])
            high = min(self.segments[self.box1id][3], self.sp.maxFreq-0.1)
            self.playSelectedSegment(low,high)
            self.speedButton.setEnabled(False)

    def pausePlayback(self):
        """ Restores the PLAY buttons, calls media_obj to pause playing."""
        self.media_obj.pressedPause()
        #self.NotifyTimer.stop()
        self.bar.setMovable(True)
        #self.swapPlayButtonState(False)
        self.playButton.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_MediaPlay))

    def stopPlayback(self):
        """ Restores the PLAY buttons, slider, text, calls media_obj to stop playing."""
        self.media_obj.pressedStop()
        #self.NotifyTimer.stop()
        self.bar.setMovable(True)
        if not hasattr(self, 'segmentStart') or self.segmentStart is None:
            self.segmentStart = 0
        self.bar.setValue(-1000)
        self.swapPlayButtonState(False)
        self.speedButton.setEnabled(True)

    def movePlaySlider(self):
        """ Listener called on sound notify (every 30 ms).
        Controls the slider, text timer, and listens for playback finish.
        """
        eltime = self.media_obj.processedUSecs() // 1000 // self.playSpeed + self.media_obj.timeoffset
            
        # listener for playback finish. Note small buffer for catching up
        if eltime > (self.segmentStop-10):
            print("Stopped at %d ms" % eltime)
            self.stopPlayback()
        else:
            # Note small buffer 
            self.bar.setValue(int(self.convertAmpltoSpec(eltime / 1000.0 - 0.02)))

    def volSliderMoved(self, value):
        self.media_obj.applyVolSlider(value)

    def barMoved(self, evt):
        """ Listener for when the bar showing playback position moves.
            Resets player so that it won't try to resume
        """
        print("Resetting playback")
        self.media_obj.pressedStop()

    def swapPlayButtonState(self,newStateisPlay):
        # Swap all button icons between play and pause/stop
        if newStateisPlay:
            self.playButton.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_MediaPause))
            self.playSegButton.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_MediaStop))
            self.playBandLimitedSegButton.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_MediaStop))
        else:
            self.playButton.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_MediaPlay))
            self.playSegButton.setIcon(QIcon('img/playsegment.png'))
            self.playBandLimitedSegButton.setIcon(QIcon('img/playBandLimited.png'))

        # OS X doesn't repaint them by default smh
        self.playButton.repaint()
        self.playSegButton.repaint()
        self.playBandLimitedSegButton.repaint()
        QApplication.processEvents()

    def floorSliderMoved(self,value):
        self.noisefloor = value
        self.setSpectrogram()
        self.setfigs()

    def setOperatorReviewerDialog(self):
        """ Listener for Set Operator/Reviewer menu item.
        """
        if hasattr(self, 'operator') and hasattr(self, 'reviewer') :
            self.operatorReviewerDialog = Dialogs.OperatorReviewer(operator=self.operator,reviewer=self.reviewer)
        else:
            self.operatorReviewerDialog = Dialogs.OperatorReviewer(operator='', reviewer='')
        self.operatorReviewerDialog.activate.clicked.connect(self.changeOperator)
        self.operatorReviewerDialog.exec()

    def changeOperator(self):
        """ Listener for the operator/reviewer dialog.
        """
        name1, name2 = self.operatorReviewerDialog.getValues()
        self.operator = str(name1)
        self.reviewer = str(name2)
        self.statusRight.setText("Operator: " + self.operator + ", Reviewer: "+self.reviewer)
        self.operatorReviewerDialog.close()
        self.segmentsToSave = True

    def manageFilters(self):
        filterManagerSimple = Dialogs.FilterManager(self.filtersDir)
        filterManagerSimple.exec()

    def customiseFiltersROC(self):
        self.filterManager = DialogsTraining.FilterCustomiseROC(self.filtersDir)
        self.filterManager.btnSave.clicked.connect(self.saveRecogniserROC)
        self.filterManager.exec()

    def addNoiseData(self):
        """ Listener for the adding metadata about noise action """
        # this field isn't required and may not be present at all
        if "noiseLevel" not in self.segments.metadata:
            self.segments.metadata["noiseLevel"] = None
        if "noiseTypes" not in self.segments.metadata:
            self.segments.metadata["noiseTypes"] = []

        self.getNoiseDataDialog = Dialogs.addNoiseData(self.segments.metadata["noiseLevel"], self.segments.metadata["noiseTypes"])
        self.getNoiseDataDialog.activate.clicked.connect(self.getNoiseData)
        self.getNoiseDataDialog.exec()

    def getNoiseData(self):
        """ Collect data about the noise from the dialog """
        self.segments.metadata["noiseLevel"], self.segments.metadata["noiseTypes"] = self.getNoiseDataDialog.getNoiseData()
        self.getNoiseDataDialog.close()
        self.segmentsToSave = True

    def saveImage(self, imageFile=''):
        if self.cheatsheet:
            self.showMaximized() # for nice spec images

        exporter = pge.ImageExporter(self.w_spec.scene())

        if imageFile=='':
            imageFile, drop = QFileDialog.getSaveFileName(self, "Save Image", "", "Images (*.png *.xpm *.jpg)")
            if not (imageFile.endswith('.png') or imageFile.endswith('.xpm') or imageFile.endswith('.jpg')):
                # Exporter won't be able to deduce file type and will quit silently
                imageFile = imageFile + '.png'
        try:
            exporter.export(imageFile)
            print("Exporting spectrogram to file %s" % imageFile)
        except Exception as e:
            print("Warning: failed to save image")
            print(e)

    def saveImageRaw(self):
        imageFile = self.filename[:-4] + '.png'
        print("Exporting raw spectrogram to file %s" % imageFile)
        self.specPlot.save(imageFile)

    def changeSettings(self):
        """ Create the parameter tree when the Interface settings menu is pressed.
        """
        self.saveSegments()
        fn1 = self.config['BirdListShort']
        if '/' in fn1:
            fn1 = os.path.basename(fn1)
        fn2 = self.config['BirdListLong']
        if fn2 is not None and '/' in fn2:
            fn2 = os.path.basename(fn2)
        fn3 = self.config['BatList']
        if fn3 is not None and '/' in fn3:
            fn3 = os.path.basename(fn3)
        fn4 = self.config['FreebirdList']
        if fn4 is not None and '/' in fn4:
            fn4 = os.path.basename(fn4)
        hasMultipleSegments = False
        for s in self.segments:
            if len(s[4])>1:
                hasMultipleSegments=True

        params = [
            {'name': 'Mouse settings', 'type' : 'group', 'children': [
                {'name': 'Use right button to make segments', 'type': 'bool', 'tip': 'If true, segments are drawn with right clicking.',
                 'value': self.config['drawingRightBtn']},
                {'name': 'Spectrogram mouse action', 'type': 'list', 'values':
                    {'Mark segments by clicking' : 1, 'Mark boxes by clicking' : 2, 'Mark boxes by dragging' : 3},
                 'value': self.config['specMouseAction']}
            ]},

            {'name': 'Paging', 'type': 'group', 'children': [
                {'name': 'Page size', 'type': 'float', 'value': self.config['maxFileShow'], 'limits': (5, 3600),
                 'step': 5,
                 'suffix': ' sec'},
                {'name': 'Page overlap', 'type': 'float', 'value': self.config['fileOverlap'], 'limits': (0, 20),
                 'step': 2,
                 'suffix': ' sec'},
            ]},

            {'name': 'Annotation', 'type': 'group', 'children': [
                {'name': 'Annotation overview cell length', 'type': 'float',
                 'value': self.config['widthOverviewSegment'],
                 'limits': (5, 300), 'step': 5,
                 'suffix': ' sec'},
                {'name': 'Make boxes transparent', 'type': 'bool',
                 'value': self.config['transparentBoxes']},
                {'name': 'Auto save segments every', 'type': 'float', 'value': self.config['secsSave'],
                 'step': 5,
                 'limits': (5, 900),
                 'suffix': ' sec'},
                {'name': 'Segment colours', 'type': 'group', 'children': [
                    {'name': 'Confirmed segments', 'type': 'color', 'value': self.config['ColourNamed'],
                     'tip': "Correctly labeled segments"},
                    {'name': 'Possible', 'type': 'color', 'value': self.config['ColourPossible'],
                     'tip': "Segments that need further approval"},
                    {'name': "Don't know", 'type': 'color', 'value': self.config['ColourNone'],
                     'tip': "Segments that are not labelled"},
                    {'name': 'Currently selected', 'type': 'color', 'value': self.config['ColourSelected'],
                     'tip': "Currently selected segment"},
                ]},
                {'name': 'Guidelines', 'type': 'group', 'children': [
                    {'name': 'Show frequency guides', 'type': 'list', 'values':
                        {'Always': 'always', 'For bats only': 'bat', 'Never': 'never'},
                        'value': self.config['guidelinesOn']},
                    {'name': 'Guideline 1 frequency', 'type': 'float', 'value': self.config['guidepos'][0]/1000, 'limits': (0, 1000), 'suffix': ' kHz'},
                    {'name': 'Guideline 1 colour', 'type': 'color', 'value': self.config['guidecol'][0]},
                    {'name': 'Guideline 2 frequency', 'type': 'float', 'value': self.config['guidepos'][1]/1000, 'limits': (0, 1000), 'suffix': ' kHz'},
                    {'name': 'Guideline 2 colour', 'type': 'color', 'value': self.config['guidecol'][1]},
                    {'name': 'Guideline 3 frequency', 'type': 'float', 'value': self.config['guidepos'][2]/1000, 'limits': (0, 1000), 'suffix': ' kHz'},
                    {'name': 'Guideline 3 colour', 'type': 'color', 'value': self.config['guidecol'][2]},
                    {'name': 'Guideline 4 frequency', 'type': 'float', 'value': self.config['guidepos'][3]/1000, 'limits': (0, 1000), 'suffix': ' kHz'},
                    {'name': 'Guideline 4 colour', 'type': 'color', 'value': self.config['guidecol'][3]},
                ]},
                {'name': 'Check-ignore protocol', 'type': 'group', 'children': [
                    {'name': 'Show check-ignore marks', 'type': 'bool', 'value': self.config['protocolOn']},
                    {'name': 'Length of checking zone', 'type': 'float', 'value': self.config['protocolSize'],
                     'limits': (1, 300), 'step': 1, 'suffix': ' sec'},
                    {'name': 'Repeat zones every', 'type': 'float', 'value': self.config['protocolInterval'],
                     'limits': (1, 300), 'step': 1, 'suffix': ' sec'},
                    {'name': 'Line colour', 'type': 'color', 'value': self.config['protocolLineCol']},
                    {'name': 'Line width', 'type': 'int', 'value': self.config['protocolLineWidth'],
                     'limits': (1, 10), 'step': 1},
                ]}
            ]},

            {'name': 'Bird List', 'type': 'group', 'children': [
                {'name': 'Common Bird List', 'type': 'group', 'children': [
                    # {'name': 'Filename', 'type': 'text', 'value': self.config['BirdListShort']},
                    {'name': 'Filename', 'type': 'str', 'value': fn1, 'readonly': True},
                    {'name': 'Choose File', 'type': 'action'},
                ]},
                {'name': 'Full Bird List', 'type': 'group', 'children': [
                    # {'name': 'Filename', 'type': 'str', 'value': fn2,'readonly':True, 'tip': "Can be None"},
                    {'name': 'Filename', 'type': 'str', 'value': fn2, 'readonly': True},
                    #{'name': 'No long list', 'type': 'bool',
                     #'value': self.config['BirdListLong'] is None or self.config['BirdListLong'] == 'None',
                     #'tip': "If you don't have a long list of birds"},
                    {'name': 'Choose File', 'type': 'action'}
                ]},
                {'name': 'Bat List', 'type': 'group', 'children': [
                    {'name': 'Filename', 'type': 'str', 'value': fn3, 'readonly': True},
                    {'name': 'Choose File', 'type': 'action'}
                ]},
                {'name': 'Freebird List', 'type': 'group', 'children': [
                    {'name': 'Filename', 'type': 'str', 'value': fn4, 'readonly': True},
                    {'name': 'Choose File', 'type': 'action'}
                ]},
                {'name': 'Dynamically reorder bird list', 'type': 'bool', 'value': self.config['ReorderList']},
                {'name': 'Default to multiple species', 'type': 'bool', 'value': self.config['MultipleSpecies'],
                 'readonly': hasMultipleSegments},
            ]},
            {'name': 'User', 'type': 'group', 'children': [
                {'name': 'Operator', 'type': 'str', 'value': self.config['operator'],
                 'tip': "Person name"},

                {'name': 'Reviewer', 'type': 'str', 'value': self.config['reviewer'],
                 'tip': "Person name"},
            ]},
            {'name': 'Maximise window on startup', 'type': 'bool', 'value': self.config['StartMaximized']},
            {'name': 'Require noise data', 'type': 'bool', 'value': self.config['RequireNoiseData']},
        ]

        ## Create tree of Parameter objects
        self.p = Parameter.create(name='params', type='group', children=params)
        self.p.sigTreeStateChanged.connect(self.changeParams)
        ## Create ParameterTree widget
        self.t = ParameterTree()
        self.t.setParameters(self.p, showTop=False)
        self.t.show()
        self.t.setWindowTitle('AviaNZ - Interface Settings')
        self.t.setWindowIcon(QIcon('img/Avianz.ico'))
        self.t.setFixedHeight(900)
        self.t.setMinimumWidth(520)

    def changeParams(self,param, changes):
        """ Update the config and the interface if anything changes in the tree
        """
        # first save the annotations
        self.saveSegments()

        # some regexes to parse guideline settings
        rgx_guide_pos = re.compile(r"Annotation.Guidelines.Guideline ([0-9]) frequency")
        rgx_guide_col = re.compile(r"Annotation.Guidelines.Guideline ([0-9]) colour")

        for param, change, data in changes:
            path = self.p.childPath(param)
            if path is not None:
                childName = '.'.join(path)
            else:
                childName = param.name()

            if childName=='Output parameters.Auto save segments every':
                self.config['secsSave']=data
            elif childName=='Annotation.Annotation overview cell length':
                self.config['widthOverviewSegment']=data

            elif childName=='Annotation.Make boxes transparent':
                self.config['transparentBoxes']=data
                self.dragRectsTransparent()
            elif childName == 'Mouse settings.Use right button to make segments':
                self.config['drawingRightBtn'] = data
                if self.config['drawingRightBtn']:
                    self.MouseDrawingButton = Qt.MouseButton.RightButton
                    self.specPlot.unsetCursor()
                    self.p_ampl.unsetCursor()
                    self.bar.setCursor(Qt.CursorShape.OpenHandCursor)
                else:
                    self.MouseDrawingButton = Qt.MouseButton.LeftButton
                    self.bar.unsetCursor()
                    self.specPlot.setCursor(QtGui.QCursor(QPixmap('img/cursor.bmp'), 0, 0))
                    self.p_ampl.setCursor(QtGui.QCursor(QPixmap('img/cursor.bmp'), 0, 0))
                self.bar.btn = self.MouseDrawingButton
            elif childName == 'Mouse settings.Spectrogram mouse action':
                self.config['specMouseAction'] = data
                self.p_spec.enableDrag = data==3 and not self.readonly.isChecked()
            elif childName == 'Paging.Page size':
                self.config['maxFileShow'] = data
            elif childName=='Paging.Page overlap':
                self.config['fileOverlap'] = data
            elif childName == 'Maximise window on startup':
                self.config['StartMaximized'] = data
                if data:
                    self.showMaximized()
            elif childName == 'Bird List.Dynamically reorder bird list':
                self.config['ReorderList'] = data
            elif childName == 'Bird List.Default to multiple species':
                self.config['MultipleSpecies'] = data
            elif childName == 'Require noise data':
                self.config['RequireNoiseData'] = data
            elif childName=='Bird List.Common Bird List.Filename':
                self.config['BirdListShort'] = data
            elif childName=='Bird List.Full Bird List.Filename':
                self.config['BirdListLong'] = data
            elif childName=='Bird List.Bat List.Filename':
                self.config['BatList'] = data
            elif childName=='Bird List.Freebird List.Filename':
                self.config['FreebirdList'] = data
            elif childName=='Annotation.Segment colours.Confirmed segments':
                rgbaNamed = list(data.getRgb())
                if rgbaNamed[3] > 100:
                    rgbaNamed[3] = 100
                self.config['ColourNamed'] = rgbaNamed
                self.ColourNamed = QtGui.QColor(self.config['ColourNamed'][0], self.config['ColourNamed'][1],
                                                self.config['ColourNamed'][2], self.config['ColourNamed'][3])
                self.ColourNamedDark = QtGui.QColor(self.config['ColourNamed'][0], self.config['ColourNamed'][1],
                                                    self.config['ColourNamed'][2], 255)
                self.listFiles.ColourNamed = self.ColourNamed
            elif childName=='Annotation.Segment colours.Possible':
                rgbaVal = list(data.getRgb())
                if rgbaVal[3] > 100:
                    rgbaVal[3] = 100
                self.config['ColourPossible'] = rgbaVal
                self.ColourPossible = QtGui.QColor(self.config['ColourPossible'][0], self.config['ColourPossible'][1],
                                                   self.config['ColourPossible'][2], self.config['ColourPossible'][3])
                self.ColourPossibleDark = QtGui.QColor(self.config['ColourPossible'][0],
                                                       self.config['ColourPossible'][1],
                                                       self.config['ColourPossible'][2], 255)
                self.listFiles.ColourPossibleDark = self.ColourPossibleDark
            elif childName=="Annotation.Segment colours.Don't know":
                rgbaVal = list(data.getRgb())
                if rgbaVal[3] > 100:
                    rgbaVal[3] = 100
                self.config['ColourNone'] = rgbaVal
                self.ColourNone = QtGui.QColor(self.config['ColourNone'][0], self.config['ColourNone'][1],
                                               self.config['ColourNone'][2], self.config['ColourNone'][3])
                self.ColourNoneDark = QtGui.QColor(self.config['ColourNone'][0], self.config['ColourNone'][1],
                                                   self.config['ColourNone'][2], 255)
                self.listFiles.ColourNone = self.ColourNone
            elif childName=='Annotation.Segment colours.Currently selected':
                rgbaVal = list(data.getRgb())
                if rgbaVal[3] > 100:
                    rgbaVal[3] = 100
                self.config['ColourSelected'] = rgbaVal
                # update the interface
                self.ColourSelected = QtGui.QColor(self.config['ColourSelected'][0], self.config['ColourSelected'][1],
                                                   self.config['ColourSelected'][2], self.config['ColourSelected'][3])
                self.ColourSelectedDark = QtGui.QColor(self.config['ColourSelected'][0], self.config['ColourSelected'][1],
                                                   self.config['ColourSelected'][2], 255)
            elif childName=='Annotation.Guidelines.Show frequency guides':
                self.config['guidelinesOn'] = data
                self.drawGuidelines()
            elif rgx_guide_pos.match(childName): # childName=='Annotation.Guidelines.Guideline 1 frequency':
                guideid = int(rgx_guide_pos.search(childName).group(1))-1
                self.config['guidepos'][guideid] = float(data)*1000
                self.drawGuidelines()
            elif rgx_guide_col.match(childName): # childName=='Annotation.Guidelines.Guideline 1 colour':
                guideid = int(rgx_guide_col.search(childName).group(1))-1
                self.config['guidecol'][guideid] = data
                self.drawGuidelines()
            elif childName=='Annotation.Check-ignore protocol.Show check-ignore marks':
                self.config['protocolOn'] = data
                self.drawProtocolMarks()
            elif childName=='Annotation.Check-ignore protocol.Length of checking zone':
                self.config['protocolSize'] = data
                self.drawProtocolMarks()
            elif childName=='Annotation.Check-ignore protocol.Repeat zones every':
                self.config['protocolInterval'] = data
                self.drawProtocolMarks()
            elif childName=='Annotation.Check-ignore protocol.Line colour':
                rgbaVal = list(data.getRgb())
                self.config['protocolLineCol'] = rgbaVal
                self.drawProtocolMarks()
            elif childName=='Annotation.Check-ignore protocol.Line width':
                self.config['protocolLineWidth'] = data
                self.drawProtocolMarks()
            elif childName=='User.Operator':
                self.config['operator'] = data
                self.operator = data
                self.statusRight.setText("Operator: " + str(self.operator) + ", Reviewer: " + str(self.reviewer))
            elif childName=='User.Reviewer':
                self.config['reviewer'] = data
                self.reviewer = data
                self.statusRight.setText("Operator: " + str(self.operator) + ", Reviewer: " + str(self.reviewer))
            elif childName=='Bird List.Common Bird List.Choose File':
                filename, drop = QFileDialog.getOpenFileName(self, 'Choose Common Bird List', self.SoundFileDir, "Text files (*.txt)")
                if filename == '':
                    print("no list file selected")
                    return
                else:
                    self.shortBirdList = self.ConfigLoader.shortbl(filename, self.configdir)
                    if self.shortBirdList is not None:
                        self.config['BirdListShort'] = filename
                        self.p['Bird List','Common Bird List', 'Filename'] = filename
                    else:
                        self.shortBirdList = self.ConfigLoader.shortbl(self.config['BirdListShort'], self.configdir)
            elif childName=='Bird List.Full Bird List.Choose File':
                filename, drop = QFileDialog.getOpenFileName(self, 'Choose Full Bird List', self.SoundFileDir, "Text files (*.txt)")
                if filename == '':
                    print("no list file selected")
                    return
                else:
                    self.longBirdList = self.ConfigLoader.longbl(filename, self.configdir)
                    if self.longBirdList is not None:
                        self.config['BirdListLong'] = filename
                        self.p['Bird List','Full Bird List','Filename'] = filename
                    else:
                        self.longBirdList = self.ConfigLoader.longbl(self.config['BirdListLong'], self.configdir)
            elif childName=='Bird List.Bat List.Choose File':
                filename, drop = QFileDialog.getOpenFileName(self, 'Choose Bat List', self.SoundFileDir, "Text files (*.txt)")
                if filename == '':
                    print("no list file selected")
                    return
                else:
                    self.batList = self.ConfigLoader.batl(filename, self.configdir)
                    if self.batList is not None:
                        self.config['BatList'] = filename
                        self.p['Bird List','Bat List','Filename'] = filename
                    else:
                        self.batList = self.ConfigLoader.batl(self.config['BatList'], self.configdir)
            elif childName=='Bird List.Freebird List.Choose File':
                filename, drop = QFileDialog.getOpenFileName(self, 'Choose Freebird List', self.configdir, "*.csv *.xlsx")
                if filename == '':
                    print("no list file selected")
                    return
                else:
                    self.config['FreebirdList'] = filename
                    #self.p['Bird List','Full Bird List','No long list'] = False
            #elif childName=='Bird List.Full Bird List.No long list':
                #if param.value():
                    #self.config['BirdListLong'] = 'None'
                    #self.p['Bird List','Full Bird List','Filename'] = 'None'
                    #self.longBirdList = None
                #else:
                    #if self.p['Bird List','Full Bird List','Filename'] is None or self.p['Bird List','Full Bird List','Filename'] == '' or self.p['Bird List','Full Bird List','Filename'] == 'None':
                        #filename, drop = QFileDialog.getOpenFileName(self, 'Choose File', self.SoundFileDir, "Text files (*.txt)")
                        #if filename == '':
                            #print("no list file selected")
                            #return
                        #else:
                            #self.p['Bird List','Full Bird List','Filename'] = filename
                            #self.config['BirdListLong'] = filename
                            #self.longBirdList = self.ConfigLoader.longbl(self.config['BirdListLong'], self.configdir)

        self.saveConfig = True

        self.resetStorageArrays()
        # pass the file name to reset interface properly
        self.loadFile(self.filename)

# ============
# Various actions: deleting segments, saving, quitting
    def confirmSegment(self):
        """ Listener for the Confirm segment button.
            Ups the certainty to 100 on the current segment.
            DO NOT use for All Sp Review, as that one may also change species and
            needs to call refreshOverview with old species.
        """
        id = self.box1id
        #print("confirming id:", id)

        if id>-1:
            # force wipe old overview to empty
            self.refreshOverviewWith(self.segments[id], delete=True)

            # raise certainty to 100 on all labels in this seg
            self.segments[id].confirmLabels()

            self.refreshOverviewWith(self.segments[id])
            self.updateText(id)
            self.updateColour(id)
            self.segInfo.setText(self.segments[id].infoString())
            self.segmentsToSave = True

    def deleteSegment(self,id=-1,hr=False):
        """ Listener for delete segment button, or backspace key. Also called when segments are deleted by the
        human classify dialogs.
        Stops playback immediately in all cases.
        Deletes the segment that is selected, otherwise does nothing.
        Updates the overview segments as well.
        """
        #print("deleting id:", id)
        if hasattr(self, 'media_obj'):
            if self.media_obj.isPlayingorPaused():
                self.stopPlayback()

        if not hr and id<0:
            id = self.box1id

        if id>-1:
            self.refreshOverviewWith(self.segments[id], delete=True)

            if self.listRectanglesa1[id] is not None:
                try:
                    self.listRectanglesa1[id].sigRegionChangeFinished.disconnect()
                    self.listRectanglesa2[id].sigRegionChangeFinished.disconnect()
                except:
                    pass
                self.p_ampl.removeItem(self.listRectanglesa1[id])
                self.p_spec.removeItem(self.listRectanglesa2[id])
                self.p_spec.removeItem(self.listLabels[id])
            del self.listLabels[id]
            del self.segments[id]
            del self.listRectanglesa1[id]
            del self.listRectanglesa2[id]
            self.segmentsToSave = True
            self.refreshFileColor()

            self.box1id = -1
            self.segInfo.setText("")
            # reset segment playback buttons
            self.refreshSegmentControls()

    def deleteAll(self):
        """ Listener for delete all button.
        Checks if the user meant to do it, then calls removeSegments()
        """
        if len(self.segments) == 0:
            msg = SupportClasses_GUI.MessagePopup("w", "No segments", "No segments to delete")
            msg.exec()
            return
        else:
            msg = SupportClasses_GUI.MessagePopup("t", "Delete All Segments?", "Are you sure you want to delete all segments?")
            msg.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            reply = msg.exec()
            if reply == QMessageBox.StandardButton.Yes:
                self.removeSegments()
                self.segmentsToSave = True

            # reset segment playback buttons
            self.refreshSegmentControls()

    def removeSegments(self,delete=True):
        """ Remove all the segments in response to the menu selection, or when a new file is loaded. """
        for r in self.listLabels:
            if r is not None:
                self.p_spec.removeItem(r)
        for r in self.listRectanglesa1:
            if r is not None:
                try:
                    r.sigRegionChangeFinished.disconnect()
                    self.p_ampl.removeItem(r)
                except:
                    pass
        for r in self.listRectanglesa2:
            if r is not None:
                try:
                    r.sigRegionChangeFinished.disconnect()
                    self.p_spec.removeItem(r)
                except:
                    pass

        # clear overview boxes and their count trackers
        for ovid in range(len(self.SegmentRects)):
            self.overviewSegments[ovid, :] = 0
            self.SegmentRects[ovid].setBrush(pg.mkBrush('w'))
            self.SegmentRects[ovid].update()

        self.segInfo.setText("")
        if delete:
            if hasattr(self, "segments"):
                self.segments.clear()
            self.listRectanglesa1 = []
            self.listRectanglesa2 = []
            self.listLabels = []
            self.box1id = -1

    def refreshFileColor(self):
        """ Extracts the minimum certainty and updates the color
            of this file in the file list. """
        if len(self.segments)==0:
            mincert = -1
        else:
            mincert = min([lab["certainty"] for seg in self.segments for lab in seg[4]])
        self.listFiles.refreshFile(os.path.basename(self.filename), mincert)

    def saveSegments(self):
        """ Save the segmentation data as a json file.
        Name of the file is the name of the wave file + .data"""

        # def checkSave():
        #     msg = QMessageBox()
        #     msg.setIcon(QMessageBox.Information)
        #     msg.setText("Do you want to save?")
        #     msg.setInformativeText("You didn't identify any segments, are you sure you want to save this annotation?")
        #     msg.setWindowTitle("No segments")
        #     msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        #     msg.buttonClicked.connect(msgbtn)
        #     retval = msg.exec_()
        #     print "value of pressed message box button:", retval
        #     return retval

        if self.segmentsToSave:
            self.segments.metadata["Operator"] = self.operator
            self.segments.metadata["Reviewer"] = self.reviewer

            print("SAVING JSON")
            print(self.segments.metadata)

            self.segments.saveJSON(str(self.filename) + '.data')

            # Refresh this file's icon in file list dock
            self.refreshFileColor()
            self.segmentsToSave = False
            self.statusLeft.setText("Segments saved at " + time.strftime("%X", time.localtime()))

    def closeFile(self):
        """ Calls the appropriate functions when a file is gently closed (on quit or change of file). """
        # stop playing the file
        if hasattr(self, 'media_obj'): # check in case we are looking at a bat
            self.media_obj.pressedStop()

        # Save noise data if the user requires it
        if self.config['RequireNoiseData']:
            if "noiseLevel" not in self.segments.metadata or self.segments.metadata["noiseLevel"] is None:
                self.addNoiseData()

        self.saveSegments()
        #print("Closing", self.filename)

        # Update recent files list
        if self.filename is not None and self.filename not in self.config['RecentFiles']:
            self.config['RecentFiles'].append(self.filename)
            if len(self.config['RecentFiles'])>4:
                del self.config['RecentFiles'][0]
            # NOTE: we're making this flag useless as every new file open will update the config
            self.saveConfig = True

        # Add in the operator and reviewer at the top, and then save the segments and the config file.
        if self.saveConfig:
            self.ConfigLoader.configwrite(self.config, self.configfile)

        # Save the shortBirdList
        self.ConfigLoader.blwrite(self.shortBirdList, self.config['BirdListShort'], self.configdir)

    def restart(self):
        """ Listener for the restart option, which uses exit(1) to restart the program at the splash screen """
        print("Restarting")
        if hasattr(self, 'media_obj'):
            if self.media_obj.isPlayingorPaused():
                self.stopPlayback()
                #self.media_thread.quit()
                #self.media_thread.wait()
        QApplication.exit(1)

    def closeEvent(self, event=None):
        """ Catch the user closing the window by clicking the Close button or otherwise."""
        print("Quitting")
        self.closeFile()
        QApplication.exit(0)

    def backupDatafiles(self):
        # TODO: Can probably be removed
        print("Backing up files in ", self.SoundFileDir)
        listOfDataFiles = QDir(self.SoundFileDir).entryList(['*.data'])
        for file in listOfDataFiles:
            source = self.SoundFileDir + '/' + file
            destination = source[:-5]+".backup"
            if os.path.isfile(destination):
                pass
                #print(destination," exists, not backing up")
            else:
                #print(source)
                #print(destination," doesn't exist")
                copyfile(source, destination)

    def eventFilter(self, obj, event):
        """ Handles two types of events:
            1) Clicks for the context menu. It allows the user to select
            multiple birds by stopping the menu being closed on first click.
            2) Keyboard presses for spec/ampl plots:
              backspace to delete a segment
              escape to pause playback
              ctrl on Mac to detect right clicks
        """
        if isinstance(obj, QMenu) and event.type() in [QtCore.QEvent.Type.MouseButtonRelease]:
            if hasattr(self, 'multipleBirds') and self.multipleBirds:
                if obj.activeAction():
                    if not obj.activeAction().menu():
                        #if the selected action does not have a submenu
                        #eat the event, but trigger the function
                        obj.activeAction().trigger()
                        return True
            return QMenu.eventFilter(self,obj, event)
        if isinstance(obj, pg.GraphicsLayoutWidget):
            if event.type()==QtCore.QEvent.Type.KeyPress:
                key = event.key()
                if key == Qt.Key.Key_Backspace or key == Qt.Key.Key_Delete:
                    self.deleteSegment()
                    return True
                elif key == Qt.Key.Key_Escape:
                    if hasattr(self, 'media_obj'):
                        if self.media_obj.isPlaying():
                            self.stopPlayback()
                    return True
                elif key == Qt.Key.Key_Meta and platform.system() == 'Darwin':
                    # flip to rightMB cursors
                    if self.MouseDrawingButton==Qt.MouseButton.RightButton:
                        self.p_ampl.setCursor(QtGui.QCursor(QPixmap('img/cursor.bmp'), 0, 0))
                        self.specPlot.setCursor(QtGui.QCursor(QPixmap('img/cursor.bmp'), 0, 0))
                        self.bar.unsetCursor()
                    else:
                        self.p_ampl.unsetCursor()
                        self.specPlot.unsetCursor()
                        self.bar.setCursor(Qt.CursorShape.OpenHandCursor)
                    return True
            elif event.type()==QtCore.QEvent.Type.KeyRelease:
                if event.key() == Qt.Key.Key_Meta and platform.system() == 'Darwin':
                    # revert to standard cursors (for leftMB)
                    if self.MouseDrawingButton==Qt.MouseButton.RightButton:
                        self.p_ampl.unsetCursor()
                        self.specPlot.unsetCursor()
                        self.bar.setCursor(Qt.CursorShape.OpenHandCursor)
                    else:
                        self.p_ampl.setCursor(QtGui.QCursor(QPixmap('img/cursor.bmp'), 0, 0))
                        self.specPlot.setCursor(QtGui.QCursor(QPixmap('img/cursor.bmp'), 0, 0))
                        self.bar.unsetCursor()
                    return True
        return False
