
# Segment.py
#
# A variety of segmentation algorithms for AviaNZ

# Version 2.0 18/11/19
# Authors: Stephen Marsland, Nirosha Priyadarshani, Julius Juodakis

#    AviaNZ bioacoustic analysis program
#    Copyright (C) 2017--2019

#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.

#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.

#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.
import SignalProc
import WaveletFunctions

import numpy as np
import scipy.ndimage as spi
from scipy import signal
import librosa
import time
from ext import ce_denoise as ce
import json
import os
import re
import math
import copy
import wavio
from itertools import chain, repeat
from scipy.interpolate import interp1d
from scipy.signal import medfilt
import skimage.measure as skm

from PyQt5.QtCore import QTime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font


class Segment(list):
    """ A single AviaNZ annotation ("segment" or "box" type).
        Deals with identifying the right Label from this list.

        Labels should be added either when initiating Segment,
        or through Segment.addLabel.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if len(self) != 5:
            print("ERROR: incorrect number of args provided to Segment (need 5, not %d)" % len(self))
            return
        if self[0]<0 or self[1]<0:
            print("ERROR: Segment times must be positive or 0")
            return
        if self[2]<0 or self[3]<0:
            print("ERROR: Segment frequencies must be positive or 0")
            return
        if not isinstance(self[4], list):
            print("ERROR: Segment labels must be a list")
            return

        # check if labels have the right structure
        for lab in self[4]:
            if not isinstance(lab, dict):
                print("ERROR: Segment label must be a dict")
                return
            if "species" not in lab or not isinstance(lab["species"], str):
                print("ERROR: species bad or missing from label")
                return
            if "certainty" not in lab or not isinstance(lab["certainty"], (int, float)):
                print("ERROR: certainty bad or missing from label")
                return
            if "filter" in lab and lab["filter"]!="M" and "calltype" not in lab:
                print("ERROR: calltype required when automated filter provided in label")
                return

        # fix types to avoid numpy types etc
        self[0] = float(self[0])
        self[1] = float(self[1])
        self[2] = int(self[2])
        self[3] = int(self[3])

        self.keys = [(lab['species'], lab['certainty']) for lab in self[4]]
        if len(self.keys)>len(set(self.keys)):
            print("ERROR: non-unique species/certainty combo detected")
            return

    def hasLabel(self, species, certainty):
        """ Check if label identified by species-cert combo is present in this segment. """
        return (species, certainty) in self.keys

    def addLabel(self, species, certainty, **label):
        """ Adds a label to this segment.
            Species and certainty are required and passed positionally.
            Any further label properties (filter, calltype...) must be passed as keyword args:
              addLabel("LSK", 100, filter="M"...)
        """
        if not isinstance(species, str):
            print("ERROR: bad species provided")
            return
        if not isinstance(certainty, (int, float)):
            print("ERROR: bad certainty provided")
            return
        if "filter" in label and label["filter"]!="M" and "calltype" not in label:
            print("ERROR: calltype required when automated filter provided in label")
            return
        if self.hasLabel(species, certainty):
            print("ERROR: this species-certainty label already present")
            return
        label["species"] = species
        label["certainty"] = certainty

        self[4].append(label)
        self.keys.append((species, certainty))

    ### --- couple functions to process all labels for a given species ---

    def wipeSpecies(self, species):
        """ Remove all labels for species, return True if all labels were wiped
            (and the interface should delete the segment).
        """
        deletedAll = list(set([lab["species"] for lab in self[4]])) == [species]
        # note that removeLabel will re-add a Don't Know in the end, so can't just check the final label.
        for lab in self[4]:
            self.removeLabel(lab["species"], lab["certainty"])
        return deletedAll

    def confirmLabels(self, species=None):
        """ Raise the certainty of this segment's uncertain labels to 100.
            Affects all species (if None) or indicated species.
        """
        for labix in range(len(self[4])):
            lab = self[4][labix]
            if (species is None or lab["species"]==species) and lab["certainty"] < 100:
                lab["certainty"] = 100
                self.keys[labix] = (lab["species"], lab["certainty"])

    def removeLabel(self, species, certainty):
        """ Removes label from this segment.
            Does not delete the actual segment - that's left for the interface to take care of.
        """
        deleted = False
        for lab in self[4]:
            if lab["species"]==species and lab["certainty"]==certainty:
                self[4].remove(lab)
                try:
                    self.keys.remove((species, certainty))
                except Exception as e:
                    text = "************ WARNING ************\n"
                    text += str(e)
                    text += "\nWhile trying to remove key"+str(species)+"-"+str(certainty) + "from"+ str(self[4])
                    text += "\nWhich had keys" + str(self.keys)
                    import SupportClasses
                    msg = SupportClasses.MessagePopup("w", "ERROR - please report", text)
                    msg.exec_()
                # if that was the last label, flip to Don't Know
                if len(self[4])==0:
                    self.addLabel("Don't Know", 0)
                deleted = True
                break

        if not deleted:
            print("ERROR: could not find species-certainty combo to remove:", species, certainty)
            return

    def infoString(self):
        """ Returns a nicely-formatted string of this segment's info."""
        s = []
        for lab in self[4]:
            labs = "sp.: {}, cert.: {}%".format(lab["species"], lab["certainty"])
            if "filter" in lab and lab["filter"]!="M":
                labs += ", filter: " + lab["filter"]
            if "calltype" in lab:
                labs += ", call: " + lab["calltype"]
            s.append(labs)
        return "; ".join(s)


class SegmentList(list):
    """ List of Segments. Deals with I/O - parsing JSON,
        and retrieving the right Segment from this list.
    """

    def parseJSON(self, file, duration=0):
        """ Takes in a filename and reads metadata to self.metadata,
            and other segments to just the main body of self.
            If wav file is loaded, pass the true duration in s to check
            (it will override any duration read from the JSON).
        """
        try:
            f = open(file, 'r')
            annots = json.load(f)
            f.close()
        except Exception as e:
            print("ERROR: file %s failed to load with error:" % file)
            print(e)
            return

        # first segment stores metadata
        self.metadata = dict()
        if isinstance(annots[0], list) and annots[0][0] == -1:
            print("old format metadata detected")
            self.metadata = {"Operator": annots[0][2], "Reviewer": annots[0][3]}
            # when file is loaded, true duration can be passed. Otherwise,
            # some old files have duration in samples, so need a rough check
            if duration>0:
                self.metadata["Duration"] = duration
            elif isinstance(annots[0][1], (int, float)) and annots[0][1]>0 and annots[0][1]<100000:
                self.metadata["Duration"] = annots[0][1]
            else:
                # fallback to reading the wav:
                try:
                    self.metadata["Duration"] = wavio.readFmt(file[:-5])[1]
                except Exception as e:
                    print("ERROR: duration not found in metadata, arguments, or read from wav")
                    print(file)
                    print(e)
                return
            # noise metadata
            if len(annots[0])<5 or not isinstance(annots[0][4], list):
                self.metadata["noiseLevel"] = None
                self.metadata["noiseTypes"] = []
            else:
                self.metadata["noiseLevel"] = annots[0][4][0]
                self.metadata["noiseTypes"] = annots[0][4][1]
            del annots[0]

        elif isinstance(annots[0], dict):
            # New format has 3 required fields:
            self.metadata = annots[0]
            if duration>0:
                self.metadata["Duration"] = duration
            if "Operator" not in self.metadata or "Reviewer" not in self.metadata or "Duration" not in self.metadata:
                print("ERROR: required metadata fields not found")
                return
            del annots[0]

        # read the segments
        self.clear()
        for annot in annots:
            if not isinstance(annot, list) or len(annot)!=5:
                print("ERROR: annotation in wrong format:", annot)
                return

            # This could be turned on to skip segments outside Duration bounds, 
            # but may result in deleting actually useful annotations if Duration was wrong
            # if annot[0] > self.metadata["Duration"] and annot[1] > self.metadata["Duration"]:
            #     print("Warning: ignoring segment outside set duration", annot)
            #     continue

            # deal with old formats here, so that the Segment class
            # could require (and validate) clean input

            # Early version of AviaNZ stored freqs as values between 0 and 1.
            # The .1 is to take care of rounding errors
            if 0 < annot[2] < 1.1 and 0 < annot[3] < 1.1:
                print("Warning: ignoring old-format frequency marks")
                annot[2] = 0
                annot[3] = 0

            # single string-type species labels
            if isinstance(annot[4], str):
                annot[4] = [annot[4]]
            # for list-type labels, parse each into certainty and species
            if isinstance(annot[4], list):
                listofdicts = []
                for lab in annot[4]:
                    # new format:
                    if isinstance(lab, dict):
                        labdict = lab
                    # old format parsing:
                    elif lab == "Don't Know":
                        labdict = {"species": "Don't Know", "certainty": 0}
                    elif lab.endswith('?'):
                        labdict = {"species": lab[:-1], "certainty": 50}
                    else:
                        labdict = {"species": lab, "certainty": 100}
                    listofdicts.append(labdict)
                # if no labels were present, i.e. "[]", addSegment will create a Don't Know
                annot[4] = listofdicts

            self.addSegment(annot)
        print("%d segments read" % len(self))

    def addSegment(self, segment):
        """ Just a cleaner wrapper to allow adding segments quicker.
            Passes a list "segment" to the Segment class.
        """
        # allows passing empty label list - creates "Don't Know" then
        if len(segment[4]) == 0:
            segment[4] = [{"species": "Don't Know", "certainty": 0}]
        self.append(Segment(segment))

    def addBasicSegments(self, seglist, freq=[0,0], **kwd):
        """ Allows to add bunch of basic segments from segmentation
            with identical species/certainty/freq values.
            seglist - list of 2-col segments [t1, t2]
            label is built from kwd.
            These will be converted to [t1, t2, freq[0], freq[1], label]
            and stored.
        """
        if not isinstance(freq, list) or freq[0]<0 or freq[1]<0:
            print("ERROR: cannot use frequencies", freq)
            return

        for seg in seglist:
            newseg = [seg[0], seg[1], freq[0], freq[1], [kwd]]
            self.addSegment(newseg)

    def getSpecies(self, species):
        """ Returns indices of all segments that have the indicated species in label. """
        out = []
        for segi in range(len(self)):
            # check each label in this segment:
            labs = self[segi][4]
            for lab in labs:
                if lab["species"] == species:
                    out.append(segi)
                    # go to next seg
                    break
        return(out)

    def saveJSON(self, file, reviewer=""):
        """ Returns 1 on succesful save."""
        if reviewer != "":
            self.metadata["Reviewer"] = reviewer
        annots = [self.metadata]
        for seg in self:
            annots.append(seg)

        file = open(file, 'w')
        json.dump(annots, file)
        file.write("\n")
        file.close()
        return 1

    def orderTime(self):
        """ Returns the order of segments in this list sorted by start time.
            Sorts itself using the order. Can then be used to sort any additional lists
            in matching order (graphics etc). """
        sttimes = [s[0] for s in self]
        sttimes = np.argsort(sttimes)
        self.sort(key=lambda s: s[0])

        return(sttimes)

    def getSummaries(self):
        """ Calculates some summary parameters relevant for populating training dialogs.
            and returns other parameters for populating the training dialogs.
        """
        if len(self)==0:
            print("ERROR: no annotations for this calltype found")
            return

        # get parameter limits for populating training dialogs:
        # FreqRange, in Hz
        fLow = np.min([seg[2] for seg in self])
        fHigh = np.max([seg[3] for seg in self])
        # TimeRange, in s
        lenMin = np.min([seg[1] - seg[0] for seg in self])
        lenMax = np.max([seg[1] - seg[0] for seg in self])

        return(lenMin, lenMax, fLow, fHigh)

    def exportGT(self, filename, species, window=1, inc=None):
        """ Given the AviaNZ annotations, exports a 0/1 ground truth as a txt file,
            and returns other parameters for populating the training dialogs.
        filename - current wav file name.
        species - string, will export the annotations for it.
        Window and inc defined as in waveletSegment.
        """

        if inc is None:
            inc = window
        resolution = math.gcd(int(100*window), int(100*inc)) / 100

        # number of segments of width window at inc overlap
        duration = int(np.ceil(self.metadata["Duration"] / resolution))
        eFile = filename[:-4] + '-res' + str(float(resolution)) + 'sec.txt'

        # deal with empty files
        thisSpSegs = self.getSpecies(species)
        # if len(thisSpSegs)==0:
        #     print("Warning: no annotations for this species found in file", filename)
        #     # delete the file to avoid problems with old GT files
        #     try:
        #         os.remove(eFile)
        #     except Exception:
        #         pass
        #     return

        GT = np.tile([0, 0, None], (duration,1))
        # fill first column with "time"
        GT[:,0] = range(1, duration+1)
        GT[:,0] = GT[:,0] * resolution

        for segix in thisSpSegs:
            seg = self[segix]
            # start and end in resolution base
            s = int(max(0, math.floor(seg[0] / resolution)))
            e = int(min(duration, math.ceil(seg[1] / resolution)))
            for i in range(s, e):
                GT[i,1] = 1
                GT[i,2] = species
        GT = GT.tolist()

        # now save the resulting txt:
        with open(eFile, "w") as f:
            for l, el in enumerate(GT):
                string = '\t'.join(map(str,el))
                for item in string:
                    f.write(item)
                f.write('\n')
            f.write('\n')
            print("output successfully saved to file", eFile)

    def exportExcel(self, dirName, filename, action, pagelen, numpages=1, speciesList=[], startTime=0, resolution=1):
        """ Exports the annotations to xlsx, with three sheets:
        time stamps, presence/absence, and per second presence/absence.
        Saves each species into a separate workbook,
        + an extra workbook for all species (to function as a readable segment printout).
        It makes the workbook if necessary.

        Inputs
            dirName:    xlsx will be stored here
            filename:   name of the wav file, to be recorded inside the xlsx
            action:     "append" or "overwrite" any found Excels
            pagelen:    page length, seconds (for filling out absence)
            numpages:   number of pages in this file (of size pagelen)
            speciesList:    list of species that are currently processed -- will force an xlsx output even if none were detected
            startTime:  timestamp for cell names
            resolution: output resolution on excel (sheet 3) in seconds. Default is 1
        """

        pagelen = math.ceil(pagelen)

        # will export species present in self, + passed as arg, + "all species" excel
        speciesList = set(speciesList)
        for seg in self:
            speciesList.update([lab["species"] for lab in seg[4]])
        speciesList.add("Any sound")
        print("The following species were detected for export:", speciesList)

        # ideally, we store relative paths, but that's not possible across drives:
        try:
            relfname = str(os.path.relpath(filename, dirName))
        except Exception as e:
            print("Falling back to absolute paths. Encountered exception:")
            print(e)
            relfname = str(os.path.abspath(filename))

        # functions for filling out the excel sheets:
        def writeToExcelp1(wb, segs, currsp):
            ws = wb['Time Stamps']
            r = ws.max_row + 1
            # Print the filename
            ws.cell(row=r, column=1, value=relfname)
            # Loop over the segments
            for seg in segs:
                ws.cell(row=r, column=2, value=str(QTime(0,0,0).addSecs(seg[0]+startTime).toString('hh:mm:ss')))
                ws.cell(row=r, column=3, value=str(QTime(0,0,0).addSecs(seg[1]+startTime).toString('hh:mm:ss')))
                if seg[3]!=0:
                    ws.cell(row=r, column=4, value=int(seg[2]))
                    ws.cell(row=r, column=5, value=int(seg[3]))
                if currsp=="Any sound":
                    # print species and certainty
                    text = [lab["species"] for lab in seg[4]]
                    ws.cell(row=r, column=6, value=", ".join(text))
                    text = [str(lab["certainty"]) for lab in seg[4]]
                    ws.cell(row=r, column=7, value=", ".join(text))
                else:
                    # only print certainty
                    text = []
                    for lab in seg[4]:
                        if lab["species"]==currsp:
                            text.append(str(lab["certainty"]))
                    ws.cell(row=r, column=6, value=", ".join(text))
                r += 1

        def writeToExcelp2(wb, segs):
            # segs: a 2D list of [start, end, certainty] for each seg
            ws = wb['Presence Absence']
            r = ws.max_row + 1
            ws.cell(row=r, column=1, value=relfname)
            ws.cell(row=r, column=2, value='_')
            if len(segs)>0:
                pres = "Yes"
                certainty = [lab[2] for lab in segs]
                certainty = max(certainty)
            else:
                pres = "No"
                certainty = 0
            ws.cell(row=r, column=2, value=pres)
            ws.cell(row=r, column=3, value=certainty)

        def writeToExcelp3(wb, detected, pagenum):
            # writes binary output DETECTED (per s) from page PAGENUM of length PAGELEN
            starttime = pagenum * pagelen
            ws = wb['Per Time Period']
            # print resolution "header"
            r = ws.max_row + 1
            ws.cell(row=r, column=1, value=str(resolution) + ' secs resolution')
            ft = Font(color=colors.DARKYELLOW)
            ws.cell(row=r, column=1).font=ft
            # print file name and page number
            ws.cell(row=r+1, column=1, value=relfname)
            ws.cell(row=r+1, column=2, value=str(pagenum+1))
            # fill the header and detection columns
            c = 3
            for t in range(0, len(detected), resolution):
                # absolute (within-file) times:
                win_start = starttime + t
                win_end = min(win_start+resolution, int(pagelen * numpages))
                ws.cell(row=r, column=c, value=str(win_start) + '-' + str(win_end))
                ws.cell(row=r, column=c).font = ft
                # within-page times:
                det = detected[t:win_end - starttime]
                # max certainty on 0-100 scale:
                det = max(det)
                ws.cell(row=r+1, column=c, value=det)
                c += 1

        # now, generate the actual files, SEPARATELY FOR EACH SPECIES:
        for species in speciesList:
            print("Exporting species %s" % species)
            # clean version for filename
            speciesClean = re.sub(r'\W', "_", species)

            # setup output files:
            # if an Excel exists, append (so multiple files go into one worksheet)
            # if not, create new
            eFile = dirName + '/DetectionSummary_' + speciesClean + '.xlsx'

            if action == "overwrite" or not os.path.isfile(eFile):
                # make a new workbook:
                wb = Workbook()

                # First sheet
                wb.create_sheet(title='Time Stamps', index=1)
                ws = wb['Time Stamps']
                ws.cell(row=1, column=1, value="File Name")
                ws.cell(row=1, column=2, value="start (hh:mm:ss)")
                ws.cell(row=1, column=3, value="end (hh:mm:ss)")
                ws.cell(row=1, column=4, value="min freq. (Hz)")
                ws.cell(row=1, column=5, value="max freq. (Hz)")
                if species=="Any sound":
                    ws.cell(row=1, column=6, value="species")
                    ws.cell(row=1, column=7, value="certainty")
                else:
                    ws.cell(row=1, column=6, value="certainty")

                    # Second sheet
                    wb.create_sheet(title='Presence Absence', index=2)
                    ws = wb['Presence Absence']
                    ws.cell(row=1, column=1, value="File Name")
                    ws.cell(row=1, column=2, value="Present?")
                    ws.cell(row=1, column=3, value="Certainty, %")

                    # Third sheet
                    wb.create_sheet(title='Per Time Period', index=3)
                    ws = wb['Per Time Period']
                    ws.cell(row=1, column=1, value="File Name")
                    ws.cell(row=1, column=2, value="Page")
                    ws.cell(row=1, column=3, value="Maximum certainty of species presence (0 = absent)")

                # Hack to delete original sheet
                del wb['Sheet']
            elif action == "append":
                try:
                    wb = load_workbook(eFile)
                except Exception as e:
                    print("ERROR: cannot open file %s to append" % eFile)  # no read permissions or smth
                    print(e)
                    return 0
            else:
                print("ERROR: unrecognised action", action)
                return 0

            # extract segments for the current species
            # if species=="All", take ALL segments.
            if species=="Any sound":
                speciesSegs = self
            else:
                speciesSegs = [self[ix] for ix in self.getSpecies(species)]

            # export segments
            writeToExcelp1(wb, speciesSegs, species)

            if species!="Any sound":
                # extract the certainty from each label for current species
                # to a 2D list of segs x [start, end, certainty]
                speciesCerts = []
                for seg in speciesSegs:
                    for lab in seg[4]:
                        if lab["species"]==species:
                            speciesCerts.append([seg[0], seg[1], lab["certainty"]])

                # export presence/absence and max certainty
                writeToExcelp2(wb, speciesCerts)

                # Generate per second binary output
                # (assuming all pages are of same length as current data)
                for p in range(0, numpages):
                    detected = np.zeros(pagelen)
                    # convert segs to max certainty at each second
                    for seg in speciesCerts:
                        for t in range(pagelen):
                            # convert within-page time to segment (within-file) time
                            truet = t + p*pagelen
                            if math.floor(seg[0]) <= truet and truet < math.ceil(seg[1]):
                                # store certainty if it's larger
                                detected[t] = max(detected[t], seg[2])
                    # write page p to xlsx
                    writeToExcelp3(wb, detected, p)

            # Save the file
            try:
                wb.save(eFile)
            except Exception as e:
                print("ERROR: could not create new file %s" % eFile)  # no read permissions or smth
                print(e)
                return 0
        return 1


class Segmenter:
    """ This class implements six forms of segmentation for the AviaNZ interface:
    Amplitude threshold (rubbish)
    Energy threshold
    Harma
    Median clipping of spectrogram
    Fundamental frequency using yin
    FIR

    It also computes ways to merge them

    Important parameters:
        mingap: the smallest space between two segments (otherwise merge them)
        minlength: the smallest size of a segment (otherwise delete it)
        ignoreInsideEnvelope: whether you keep the superset of a set of segments or the individuals when merging
        maxlength: the largest size of a segment (currently unused)
        threshold: generally this is of the form mean + threshold * std dev and provides a way to filter

    And two forms of recognition:
    Cross-correlation
    DTW

    Each returns start and stop times for each segment (in seconds) as a Python list of pairs.
    It is up to the caller to convert these to a true SegmentList.
    See also the species-specific segmentation in WaveletSegment
    """

    def __init__(self, sp=None, fs=0, mingap=0.3, minlength=0.2):
        # This is the reference to SignalProc
        self.sp = sp
        self.fs = fs
        # Spectrogram
        if hasattr(sp, 'sg'):
            self.sg = sp.sg
        else:
            self.sg = None
        # These are the spectrogram params. Needed to compute times.
        if sp:
            self.data = sp.data
            self.fs = sp.sampleRate
            self.window_width = sp.window_width
            self.incr = sp.incr
        self.mingap = mingap
        self.minlength = minlength

    def setNewData(self, sp):
        # To be called when a new sound file is loaded
        self.sp = sp
        self.data = sp.data
        self.fs = sp.sampleRate
        self.sg = sp.sg
        self.window_width = sp.window_width
        self.incr = sp.incr

    def bestSegments(self,FIRthr=0.7, medianClipthr=3.0, yinthr=0.9):
        """ A reasonably good segmentaion - a merged version of FIR, median clipping, and fundamental frequency using yin
        """
        segs1 = self.segmentByFIR(FIRthr)
        segs2 = self.medianClip(medianClipthr)
        segs3, p, t = self.yin(100, thr=yinthr, returnSegs=True)
        segs1 = self.mergeSegments(segs1, segs2)
        segs = self.mergeSegments(segs1, segs3)
        # mergeSegments also sorts, so not needed:
        # segs = sorted(segs, key=lambda x: x[0])
        return segs

    def mergeSegments(self, segs1, segs2=None):
        """ Given two segmentations of the same file, return the merged set of them:
        [[1,3] [2,4] [5,7] [7,8]] -> [[1,4] [5,7] [7,8]]
        Can take in one or two lists. """
        if segs1 == [] and segs2 == []:
            return []
        elif segs1 == []:
            return segs2
        elif segs2 == []:
            return segs1

        if segs2 is not None:
            segs1.extend(segs2)
        segs1.sort(key=lambda seg: seg[0])
        out = [segs1[0]]
        for curr in segs1:
            lastout = out[-1]
            if curr[0] < lastout[1]:
                lastout[1] = max(lastout[1], curr[1])
            else:
                # note that segments that only touch ([1,2][2,3])
                # will not be merged
                out.append(curr)
        return out

    def convert01(self, presabs, window=1):
        """ Turns a list of presence/absence [0 1 1 1]
            into a list of start-end segments [[1,4]].
            Can use non-1 s units of pres/abs.
        """
        presabs = np.squeeze(presabs)
        out = []
        t = 0
        while t < len(presabs):
            if presabs[t]==1:
                start = t
                while t<len(presabs) and presabs[t]!=0:
                    t += 1
                out.append([start*window, t*window])
            t += 1
        return out

    def deleteShort(self, segs, minlength=0.25):
        """ Checks whether segments are long enough.
            Operates on start-end list [[1,3], [4,5]] -> [[1,3]].
        """
        out = []
        if minlength == 0:
            minlength = self.minlength
        for seg in segs:
            if seg[1]-seg[0] >= minlength:
                out.append(seg)
        return out

    def joinGaps(self, segs, maxgap=3):
        """ Merges segments within maxgap units.
            Operates on start-end list [[1,2], [3,4]] -> [[1,4]].
        """
        out = []
        i = 0
        while i < len(segs):
            start = segs[i][0]
            while i+1 < len(segs) and segs[i+1][0]-segs[i][1] <= maxgap:
                i += 1
            out.append([start, segs[i][1]])
            i += 1
        return out

    def splitLong(self, segs, maxlen=10):
        """
        Splits long segments (> maxlen) evenly
        Operates on start-end list [[1,2], [3,4]] -> [[1,4]].
        """
        out = []
        for seg in segs:
            l = seg[1]-seg[0]
            if l > maxlen:
                n = int(np.ceil(l/maxlen))
                d = l/n
                for i in range(n):
                    out.append([seg[0] + d*i, seg[0] + d * (i+1)])
            else:
                out.append(seg)
        return out

    def checkSegmentOverlap(self, segs):
        # Needs to be python array, not np array
        # Sort by increasing start times
        if isinstance(segs, np.ndarray):
            segs = segs.tolist()
        segs = sorted(segs)
        segs = np.array(segs)

        newsegs = []
        # Loop over segs until the start value of 1 is not inside the end value of the previous
        s=0
        while s<len(segs):
            i = s
            end = segs[i,1]
            while i < len(segs)-1 and segs[i+1,0] < end:
                i += 1
                end = max(end, segs[i,1])
            newsegs.append([segs[s,0],end])
            s = i+1

        return newsegs

    def segmentByFIR(self, threshold):
        """ Segmentation using FIR envelope.
        """
        nsecs = len(self.data) / float(self.fs)
        fftrate = int(np.shape(self.sg)[0]) / nsecs
        upperlimit = 100
        FIR = [0.078573000000000004, 0.053921000000000004, 0.041607999999999999, 0.036006000000000003, 0.031521,
               0.029435000000000003, 0.028122000000000001, 0.027286999999999999, 0.026241000000000004,
               0.025225999999999998, 0.024076, 0.022926999999999999, 0.021703999999999998, 0.020487000000000002,
               0.019721000000000002, 0.019015000000000001, 0.018563999999999997, 0.017953, 0.01753,
               0.017077000000000002, 0.016544, 0.015762000000000002, 0.015056, 0.014456999999999999, 0.013913,
               0.013299, 0.012879, 0.012568000000000001, 0.012454999999999999, 0.012056000000000001, 0.011634,
               0.011077, 0.010707, 0.010217, 0.0098840000000000004, 0.0095959999999999986, 0.0093607000000000013,
               0.0090197999999999997, 0.0086908999999999997, 0.0083841000000000002, 0.0081481999999999995,
               0.0079185000000000002, 0.0076363000000000004, 0.0073406000000000009, 0.0070686999999999998,
               0.0068438999999999991, 0.0065873000000000008, 0.0063688999999999994, 0.0061700000000000001,
               0.0059743000000000001, 0.0057561999999999995, 0.0055351000000000003, 0.0053633999999999991,
               0.0051801, 0.0049743000000000001, 0.0047431000000000001, 0.0045648999999999993,
               0.0043972000000000004, 0.0042459999999999998, 0.0041016000000000004, 0.0039503000000000003,
               0.0038013000000000005, 0.0036351, 0.0034856000000000002, 0.0033270999999999999,
               0.0032066999999999998, 0.0030569999999999998, 0.0029206999999999996, 0.0027760000000000003,
               0.0026561999999999996, 0.0025301999999999998, 0.0024185000000000001, 0.0022967,
               0.0021860999999999998, 0.0020696999999999998, 0.0019551999999999998, 0.0018563,
               0.0017562000000000001, 0.0016605000000000001, 0.0015522000000000001, 0.0014482999999999998,
               0.0013492000000000001, 0.0012600000000000001, 0.0011788, 0.0010909000000000001, 0.0010049,
               0.00091527999999999998, 0.00082061999999999999, 0.00074465000000000002, 0.00067159000000000001,
               0.00060258999999999996, 0.00053370999999999996, 0.00046135000000000002, 0.00039071,
               0.00032736000000000001, 0.00026183000000000001, 0.00018987999999999999, 0.00011976000000000001,
               6.0781000000000006e-05, 0.0]
        f = interp1d(np.arange(0, len(FIR)), np.squeeze(FIR))
        samples = f(np.arange(1, upperlimit, float(upperlimit) / int(fftrate / 10.)))
        padded = np.concatenate((np.zeros(int(fftrate / 10.)), np.mean(self.sg, axis=1), np.zeros(int(fftrate / 10.))))
        envelope = spi.filters.convolve(padded, samples, mode='constant')[:-int(fftrate / 10.)]
        ind = envelope > np.median(envelope) + threshold * np.std(envelope)
        segs = self.convert01(ind, self.incr / self.fs)
        return segs

    def segmentByAmplitude(self, threshold, usePercent=True):
        """ Bog standard amplitude segmentation.
        A straw man, do not use.
        """
        if usePercent:
            threshold = threshold*np.max(self.data)
        seg = np.abs(self.data)>threshold
        seg = self.convert01(seg, self.fs)
        return seg

    def segmentByEnergy(self, thr, width, min_width=450):
        """ Based on description in Jinnai et al. 2012 paper in Acoustics
        Computes the 'energy curve' as windowed sum of absolute values of amplitude
        I median filter it, 'cos it's very noisy
        And then threshold it (no info on their threshold) and find max in each bit above threshold
        I also check for width of those (they don't say anything)
        They then return the max-width:max+width segments for each max
        """
        data = np.abs(self.data)
        E = np.zeros(len(data))
        E[width] = np.sum(data[:2*width+1])
        for i in range(width+1,len(data)-width):
            E[i] = E[i-1] - data[i-width-1] + data[i+width]
        E = E/(2*width)

        # TODO: Automatic energy gain (normalisation method)

        # This thing is noisy, so I'm going to median filter it. SoundID doesn't seem to?
        Em = np.zeros(len(data))
        for i in range(width,len(data)-width):
            Em[i] = np.median(E[i-width:i+width])
        for i in range(width):
            Em[i] = np.median(E[0:2*i])
            Em[-i] = np.median(E[-2 * i:])

        # TODO: Better way to do this?
        threshold = np.mean(Em) + thr*np.std(Em)

        # Pick out the regions above threshold and the argmax of each, assuming they are wide enough
        starts = []
        ends = []
        insegment = False
        for i in range(len(data)-1):
            if not insegment:
                if Em[i]<threshold and Em[i+1]>threshold:
                    starts.append(i)
                    insegment = True
            if insegment:
                if Em[i]>threshold and Em[i+1]<threshold:
                    ends.append(i)
                    insegment = False
        if insegment:
            ends.append(len(data))
        maxpoints = []
        Emm = np.zeros(len(data))
        for i in range(len(starts)):
            if ends[i] - starts[i] > min_width:
                maxpoints.append(np.argmax(Em[starts[i]:ends[i]]))
                Emm[starts[i]:ends[i]] = Em[starts[i]:ends[i]]

        # TODO: SoundID appears to now compute the 44 LPC coeffs for each [midpoint-width:midpoint+width]
        # TODO: And then compute the geometric distance to templates

        segs = []
        for i in range(len(starts)):
            segs.append([float(starts[i])/self.fs,float(ends[i])/self.fs])
        return segs

    def Harma(self, thr=10., stop_thr=0.8, minSegment=50):
        """ Harma's method, but with a different stopping criterion
        # Assumes that spectrogram is not normalised
        maxFreqs = 10. * np.log10(np.max(self.sg, axis = 1))
        """
        maxFreqs = 10. * np.log10(np.max(self.sg, axis=1))
        maxFreqs = medfilt(maxFreqs,21)
        biggest = np.max(maxFreqs)
        segs = []

        while np.max(maxFreqs)>stop_thr*biggest:
            t0 = np.argmax(maxFreqs)
            a_n = maxFreqs[t0]

            # Go backwards looking for where the syllable stops
            t = t0
            while maxFreqs[t] > a_n - thr and t>0:
                t -= 1
            t_start = t

            # And forwards
            t = t0
            while maxFreqs[t] > a_n - thr and t<len(maxFreqs)-1:
                t += 1
            t_end = t

            # Set the syllable just found to 0
            maxFreqs[t_start:t_end] = 0
            if float(t_end - t_start)*self.incr/self.fs*1000.0 > minSegment:
                segs.append([float(t_start)* self.incr / self.fs,float(t_end)* self.incr / self.fs])

        return segs

    def segmentByPower(self, thr=1.):
        """ Segmentation simply on the power
        """
        maxFreqs = 10. * np.log10(np.max(self.sg, axis=1))
        maxFreqs = medfilt(maxFreqs, 21)
        ind = maxFreqs > (np.mean(maxFreqs)+thr*np.std(maxFreqs))
        segs = self.convert01(ind, self.incr / self.fs)
        return segs

    def medianClip(self, thr=3.0, medfiltersize=5, minaxislength=5, minSegment=70):
        """ Median clipping for segmentation
        Based on Lasseck's method
        minaxislength - min "length of the minor axis of the ellipse that has the same normalized second central moments as the region", based on skm.
        minSegment - min number of pixels exceeding thr to declare an area as segment.
        This version only clips in time, ignoring frequency
        And it opens up the segments to be maximal (so assumes no overlap).
        The multitaper spectrogram helps a lot

        """
        tt = time.time()
        sg = self.sg/np.max(self.sg)

        # This next line gives an exact match to Lasseck, but screws up bitterns!
        #sg = sg[4:232, :]

        rowmedians = np.median(sg, axis=1)
        colmedians = np.median(sg, axis=0)

        clipped = np.zeros(np.shape(sg),dtype=int)
        for i in range(np.shape(sg)[0]):
            for j in range(np.shape(sg)[1]):
                if (sg[i, j] > thr * rowmedians[i]) and (sg[i, j] > thr * colmedians[j]):
                    clipped[i, j] = 1

        # This is the stencil for the closing and dilation. It's a 5x5 diamond. Can also use a 3x3 diamond
        diamond = np.zeros((5,5),dtype=int)
        diamond[2,:] = 1
        diamond[:,2] = 1
        diamond[1,1] = diamond[1,3] = diamond[3,1] = diamond[3,3] = 1
        #diamond[2, 1:4] = 1
        #diamond[1:4, 2] = 1

        clipped = spi.binary_closing(clipped,structure=diamond).astype(int)
        clipped = spi.binary_dilation(clipped,structure=diamond).astype(int)
        clipped = spi.median_filter(clipped,size=medfiltersize)
        clipped = spi.binary_fill_holes(clipped)

        blobs = skm.regionprops(skm.label(clipped.astype(int)))

        # Delete blobs that are too small
        keep = []
        for i in range(len(blobs)):
            if blobs[i].filled_area > minSegment and blobs[i].minor_axis_length > minaxislength:
                keep.append(i)

        list = []
        blobs = [blobs[i] for i in keep]

        # convert bounding box pixels to milliseconds:
        for l in blobs:
            list.append([float(l.bbox[0] * self.incr / self.fs), float(l.bbox[2] * self.incr / self.fs)])
        return list

    def checkSegmentOverlapCentroids(self, blobs, minSegment=50):
        # Delete overlapping boxes by computing the centroids and picking out overlaps
        # Could also look at width and so just merge boxes that are about the same size
        # Note: no mingap parameter is used right now
        centroids = []
        for i in blobs:
            centroids.append((i[1] - i[0])/2)
        centroids = np.array(centroids)
        ind = np.argsort(centroids)
        centroids = centroids[ind]
        blobs = np.array(blobs)[ind]

        current = 0
        centroid = centroids[0]
        count = 0
        list = []
        list.append(blobs[0])
        for i in centroids:
            # TODO: replace this with simple overlap?
            if (i - centroid)*1000 < minSegment / 2. * 10:
                if blobs[ind[count]][0] < list[current][0]:
                    list[current][0] = blobs[ind[count]][0]
                if blobs[ind[count]][1] > list[current][1]:
                    list[current][1] = blobs[ind[count]][1]
            else:
                current += 1
                centroid = centroids[count]
                list.append([blobs[ind[count]][0], blobs[ind[count]][1]])
            count += 1

        segments = []
        for i in list:
            if float(i[1] - i[0])*1000 > minSegment:
                segments.append([i[0], i[1]])
        return segments

    def onsets(self, thr=3.0):
        """ Segmentation using the onset times from librosa.
        There are no offset times -- compute an energy drop?
        A straw man really.
        """
        o_env = librosa.onset.onset_strength(self.data, sr=self.fs, aggregate=np.median)
        cutoff = np.mean(o_env) + thr * np.std(o_env)
        o_env = np.where(o_env > cutoff, o_env, 0)
        onsets = librosa.onset.onset_detect(onset_envelope=o_env, sr=self.fs)
        times = librosa.frames_to_time(np.arange(len(o_env)), sr=self.fs)

        segments = []
        for i in range(len(onsets)):
            segments.append([times[onsets[i]],times[onsets[i]]+0.2])
        return segments

    def yin(self, minfreq=100, minperiods=3, thr=0.5, W=1000, returnSegs=False):
        """ Segmentation by computing the fundamental frequency.
        Uses the Yin algorithm of de Cheveigne and Kawahara (2002)
        """
        if self.data.dtype == 'int16':
            data = self.data.astype(float)/32768.0
        else:
            data = self.data

        # The threshold is necessarily higher than the 0.1 in the paper

        # Window width W should be at least 3*period.
        # A sample rate of 16000 and a min fundamental frequency of 100Hz would then therefore suggest reasonably short windows
        minwin = float(self.fs) / minfreq * minperiods
        if minwin > W:
            print("Extending window width to ", minwin)
            W = minwin

        # Make life easier, and make W be a function of the spectrogram window width
        W = int(round(W/self.window_width)*self.window_width)
        # work on a copy of the main data, just to be safer
        data2 = np.zeros(len(data))
        data2[:] = data[:]

        # Now, compute squared diffs between signal and shifted signal
        # (Yin fund freq estimator)
        # over all tau<W, for each start position.
        # (starts are shifted by half a window size), i.e.:
        starts = range(0, len(data) - 2 * W, W // 2)

        # Will return pitch as self.fs/besttau for each window
        if len(data2) <= 2*W:
            print("ERROR: data too short for F0: must be %d, is %d" % (2*W, len(data2)))
            if returnSegs:
                return np.array([]), np.array([]), np.array([])
            else:
                return np.array([]), np.array([]), minfreq, W
        pitch = ce.FundFreqYin(data2, W, thr, self.fs)

        if returnSegs:
            ind = pitch > minfreq
            # ffreq is calculated over windows of size W
            # 1/len(pitch) maps is to 0-1
            # * np.shape(sg) maps it to spec windows
            # * self.incr/self.fs maps it to actual seconds
            units = self.incr/self.fs * np.shape(self.sg)[0] / len(pitch)
            segs = self.convert01(ind, units)
            return segs, pitch, np.array(starts)
        else:
            return pitch, np.array(starts), minfreq, W

    def findCCMatches(self, seg, sg, thr):
        """ Cross-correlation. Takes a segment and looks for others that match it to within thr.
        match_template computes fast normalised cross-correlation
        """
        from skimage.feature import match_template

        # seg and sg have the same $y$ size, so the result of match_template is 1D
        #m = match_template(sg,seg)
        matches = np.squeeze(match_template(sg, seg))

        import peakutils
        md = np.shape(seg)[0]/2
        threshold = thr*np.max(matches)
        indices = peakutils.indexes(matches, thres=threshold, min_dist=md)
        return indices

    def findDTWMatches(self, seg, data):
        # TODO: This is slow and crap. Note all the same length, for a start, and the fact that it takes forever!
        # Use MFCC first?
        d = np.zeros(len(data))
        for i in range(len(data)):
            d[i] = self.dtw(seg, data[i:i+len(seg)])
        return d

    def dtw(self, x, y, wantDistMatrix=False):
        # Compute the dynamic time warp between two 1D arrays
        dist = np.zeros((len(x)+1,len(y)+1))
        dist[1:, :] = np.inf
        dist[:, 1:] = np.inf
        for i in range(len(x)):
            for j in range(len(y)):
                dist[i+1, j+1] = np.abs(x[i]-y[j]) + min(dist[i, j+1], dist[i+1, j], dist[i, j])
        if wantDistMatrix:
            return dist
        else:
            return dist[-1, -1]

    def dtw_path(self, d):
        # Shortest path through DTW matrix
        i = np.shape(d)[0]-2
        j = np.shape(d)[1]-2
        xpath = [i]
        ypath = [j]
        while i>0 or j>0:
                next = np.argmin((d[i,j],d[i+1,j],d[i,j+1]))
                if next == 0:
                    i -= 1
                    j -= 1
                elif next == 1:
                    j -= 1
                else:
                    i -= 1
                xpath.insert(0,i)
                ypath.insert(0,j)
        return xpath, ypath

    # def testDTW(self):
    #     x = [0, 0, 1, 1, 2, 4, 2, 1, 2, 0]
    #     y = [1, 1, 1, 2, 2, 2, 2, 3, 2, 0]
    #
    #     d = self.dtw(x,y,wantDistMatrix=True)
    #     print self.dtw_path(d)

    def impMask(self, engp=90, fp=0.75):
        """
        Impulse mask
        :param engp: energy percentile (for rows of the spectrogram)
        :param fp: frequency proportion to consider it as an impulse (cols of the spectrogram)
        :return: audiodata
        """
        print('Impulse masking...')
        postp = PostProcess(audioData=self.data, sampleRate=self.fs, segments=[], subfilter={})
        imps = postp.impulse_cal(fs=self.fs, engp=engp, fp=fp)    # 0 - presence of impulse noise
        print('Samples to mask: ', len(self.data) - np.sum(imps))
        # Mask only the affected samples
        return np.multiply(self.data, imps)

class PostProcess:
    """ This class implements few post processing methods basically to avoid false positives.
    Operates on detections from a single subfilter.

    segments:   detected segments in form of [[s1,e1], [s2,e2],...
    subfilter:  AviaNZ format subfilter
    """

    def __init__(self, audioData=None, sampleRate=0, tgtsampleRate=0, segments=[], subfilter={}, CNNmodel=None):
        self.audioData = audioData
        self.sampleRate = sampleRate
        self.segments = segments
        self.subfilter = subfilter

        if CNNmodel:
            self.CNNmodel = CNNmodel[0]    # CNNmodel is a list [model, win, inputdim, outputdict]
            self.CNNwindow = CNNmodel[1]
            self.CNNinputdim = CNNmodel[2]
            self.CNNoutputs = CNNmodel[3]
            self.tgtsampleRate = tgtsampleRate
        else:
            self.CNNmodel = None

        if subfilter != {}:
            self.minLen = subfilter['TimeRange'][0]
            if 'F0Range' in subfilter:
                self.F0 = subfilter['F0Range']
            self.fLow = subfilter['FreqRange'][0]
            self.fHigh = subfilter['FreqRange'][1]
            self.minLen = subfilter['TimeRange'][0]
        else:
            self.minLen = 0.25
            self.fLow = 0
            self.fHigh = 0

    def generateFeaturesCNN(self, seg, data, fs):
        '''
        Prepare a syllable to input to the CNN model
        Returns the features (currently the spectrogram)
        '''
        featuress = []
        n = (seg[1] - seg[0]) // self.CNNwindow
        for i in range(int(n)):
            audiodata = data[int(self.CNNwindow * i * fs):int(self.CNNwindow * (i + 1) * fs)]
            sp = SignalProc.SignalProc(256, 128)
            sp.data = audiodata
            sp.sampleRate = fs
            sgRaw = sp.spectrogram(256, 128)
            maxg = np.max(sgRaw)
            featuress.append([np.rot90(sgRaw / maxg).tolist()])
        return featuress

    def CNN(self):
        """
        Post-proc with CNN model, self.segments get updated
        """
        if self.CNNmodel:
            if len(self.segments) == 0 or len(self.segments) == 1 and self.segments[0][0] == -1:
                pass
            else:
                newSegments = copy.deepcopy(self.segments)
                for seg in self.segments:
                    if seg[0] == -1:
                        continue
                    else:
                        print('\n--- Segment', seg)
                        data = self.audioData[int(seg[0]*self.sampleRate):int(seg[1]*self.sampleRate)]
                        # find the syllables from the seg and generate features for CNN
                        sp = SignalProc.SignalProc(256, 128)
                        sp.data = data
                        sp.sampleRate = self.sampleRate
                        if self.sampleRate != self.tgtsampleRate:
                            sp.resample(self.tgtsampleRate)
                        #     data = sp.data
                        # _ = sp.spectrogram(256, 128)
                        # segment = Segmenter(sp, self.sampleRate)
                        # syls = segment.medianClip(thr=3, medfiltersize=5, minaxislength=9, minSegment=50)
                        # if len(syls) == 0:
                        #     syls = segment.medianClip(thr=1.5, medfiltersize=5, minaxislength=9, minSegment=50)
                        # syls = segment.checkSegmentOverlap(syls)
                        # probs = 0
                        # i = 0
                        # for syl in syls:
                        #     if syl[1] - syl[0] < self.CNNwindow:
                        #         if (syl[0] - (self.CNNwindow - syl[1] + syl[0]) / 2) >= 0 and \
                        #                 (syl[1] + (self.CNNwindow - syl[1] + syl[0]) / 2) <= seg[1]-seg[0]:
                        #             syl = [syl[0] - (self.CNNwindow - syl[1] + syl[0]) / 2,
                        #                    syl[0] - (self.CNNwindow - syl[1] + syl[0]) / 2 + self.CNNwindow + 0.05]
                        #         else:
                        #             continue
                        #     featuress = self.generateFeaturesCNN(seg=[syl[0], syl[1]],
                        #                                          data=data[int(syl[0]*sp.sampleRate):int(syl[1]*sp.sampleRate)], fs=sp.sampleRate)
                        #     featuress = np.array(featuress)
                        #     # print(np.shape(featuress))
                        #     featuress = featuress.reshape(featuress.shape[0], self.CNNinputdim[0], self.CNNinputdim[1], 1)
                        #     print('\t# images in syllable ', i+1, ': ', np.shape(featuress)[0])
                        #     featuress = featuress.astype('float32')
                        #     # predict with CNN
                        #     p = self.CNNmodel.predict_proba(featuress)
                        #     # print('\t', p)
                        #     if isinstance(probs, int):
                        #         probs = p
                        #     else:
                        #         probs = np.vstack((probs, p))
                        #     i += 1

                        featuress = self.generateFeaturesCNN(seg=seg,
                                                             data=sp.data, fs=sp.sampleRate)
                        featuress = np.array(featuress)
                        # print(np.shape(featuress))
                        featuress = featuress.reshape(featuress.shape[0], self.CNNinputdim[0], self.CNNinputdim[1], 1)
                        featuress = featuress.astype('float32')
                        # predict with CNN
                        probs = 0
                        if np.shape(featuress)[0] > 0:
                            probs = self.CNNmodel.predict_proba(featuress)
                        else:
                            probs = 0
                        if isinstance(probs, int):
                            prediction = len(self.CNNoutputs) - 1     # Remember that the noise class is always the last, male-0, femle-1, noise-2
                        else:
                            # # Method I - mix of max and mean
                            # p = np.max(probs, axis=0).tolist()[:-1]
                            # p_n = np.mean(probs, axis=0)[-1].tolist()
                            # p.insert(len(p), p_n)
                            # print('\tSegment ', seg, '->', np.shape(probs)[0], ' images ->', p)
                            # print(probs)
                            # prediction = np.argmax(p)
                            # if prediction == len(self.CNNoutputs) - 1 and any(x > 0.6 for x in p[:-1]):
                            #     prediction = np.argmax(p[:-1])
                            #     # prediction = self.CNNoutputs[str(prediction)]  # actual call type  # TODO

                            # Method II - mean of best n
                            ind0 = np.argsort(probs[:, 0])
                            ind1 = np.argsort(probs[:, 1])
                            ind2 = np.argsort(probs[:, 2])
                            meanprob = [np.mean(probs[ind0[-5:], 0]), np.mean(probs[ind1[-5:], 1]), np.mean(probs[ind2[-5:], 2])]
                            prediction = np.argmax(meanprob)
                            if prediction == len(self.CNNoutputs) - 1 and any(x > 0.6 for x in meanprob[:-1]):
                                prediction = np.argmax(meanprob[:-1])
                            print(probs)
                            print(np.shape(probs)[0], ' total images -> mean prob of best n (=<5)', meanprob)
                        if prediction == len(self.CNNoutputs) - 1:
                            print('Deleted by CNN')
                            newSegments.remove(seg)
                        else:
                            print('Not deleted by CNN')
                self.segments = newSegments

    def wind_cal(self, data, sampleRate, fn_peak=0.35):
        """ Calculate wind
        Adopted from Automatic Identification of Rainfall in Acoustic Recordings by Carol Bedoya et al.
        :param data: audio data
        :param sampleRate: sample rate
        :param fn_peak: min height of a peak to be considered it as a significant peak
        :return: mean and std of wind, a binary indicator of false negative
        """
        wind_lower = 2.0 * 50 / sampleRate
        wind_upper = 2.0 * 500 / sampleRate
        f, p = signal.welch(data, fs=sampleRate, window='hamming', nperseg=512, detrend=False)
        p = np.log10(p)

        limite_inf = int(round(p.__len__() * wind_lower))
        limite_sup = int(round(
            p.__len__() * wind_upper))
        a_wind = p[limite_inf:limite_sup]  # section of interest of the power spectral density

        fn = False  # is this a false negative?
        if self.fLow > 500 or self.fLow == 0:   # fLow==0 when non-species-specific
            # Check the presence/absence of the target species/call type > 500 Hz.
            ind = np.abs(f - 500).argmin()
            if self.fLow == 0:
                ind_fLow = ind
            else:
                ind_fLow = np.abs(f - self.fLow).argmin() - ind
            if self.fHigh == 0:
                ind_fHigh = len(f) - 1
            else:
                ind_fHigh = np.abs(f - self.fHigh).argmin() - ind
            p = p[ind:]

            peaks, _ = signal.find_peaks(p)
            peaks = [i for i in peaks if (ind_fLow <= i <= ind_fHigh)]
            prominences = signal.peak_prominences(p, peaks)[0]
            # If there is at least one significant prominence in the target frequency band, then it could be a FN
            if len(prominences) > 0 and np.max(prominences) > fn_peak:
                fn = True

        return np.mean(a_wind), np.std(a_wind), fn    # mean of the PSD in the frequency band of interest.Upper part of
                                                      # the step 3 in Algorithm 2.1

    def wind(self, windT=2.5, fn_peak=0.35):
        """
        Delete wind corrupted segments, mainly wind gust
        :param windT: wind threshold
        :param fn_peak: min height of a peak to be considered it as a significant peak
        :return: self.segments get updated
        """
        if len(self.segments) == 0 or len(self.segments) == 1 and self.segments[0][0] == -1:
            pass
        else:
            newSegments = copy.deepcopy(self.segments)
            for seg in self.segments:
                if seg[0] == -1:
                    continue
                else:
                    data = self.audioData[int(seg[0]*self.sampleRate):int(seg[1]*self.sampleRate)]
                    ind = np.flatnonzero(data).tolist()     # eliminate impulse masked sections
                    data = np.asarray(data)[ind].tolist()
                    if len(data) == 0:
                        continue
                    m, _, fn = self.wind_cal(data=data, sampleRate=self.sampleRate, fn_peak=fn_peak)
                    if m > windT and not fn:
                        print(seg, m, 'windy, deleted')
                        newSegments.remove(seg)
                    elif m > windT and fn:
                        print(seg, m, 'windy, but possible bird call')
                    else:
                        print(seg, m, 'not windy/possible bird call')
            self.segments = newSegments

    def impulse_cal(self, fs, engp=90, fp=0.75, blocksize=10):
        """
        Find sections where impulse sounds occur e.g. clicks
        window  -   window length (no overlap)
        engp    -   energy percentile (thr), the percentile of energy to inform that a section got high energy across
                    frequency bands
        fp      -   frequency percentage (thr), the percentage of frequency bands to have high energy to mark a section
                    as having impulse noise
        blocksize - max number of consecutive blocks, 10 consecutive blocks (~1/25 sec) is a good value, to not to mask
                    very close-range calls
        :return: a binary list of length len(audioData) indicating presence of impulsive noise (0) otherwise (1)
        """

        # Calculate window length
        w1 = np.floor(fs/250)      # Window length of 1/250 sec selected experimentally
        arr = [2 ** i for i in range(5, 11)]
        pos = np.abs(arr - w1).argmin()
        window = arr[pos]

        sp = SignalProc.SignalProc(window, window)     # No overlap
        sp.data = self.audioData
        sp.sampleRate = self.sampleRate
        sg = sp.spectrogram(multitaper=False)

        # For each frq band get sections where energy exceeds some (90%) percentile, engp
        # and generate a binary spectrogram
        sgb = np.zeros((np.shape(sg)))
        ep = np.percentile(sg, engp, axis=0)    # note thr - 90% for energy percentile
        for y in range(np.shape(sg)[1]):
            ey = sg[:, y]
            sgb[np.where(ey > ep[y]), y] = 1

        # If lots of frq bands got 1 then predict a click
        # 1 - presence of impulse noise, 0 - otherwise here
        impulse = np.where(np.count_nonzero(sgb, axis=1) > np.shape(sgb)[1] * fp, 1, 0)     # Note thr fp

        # When an impulsive noise detected, it's better to check neighbours to make sure its not a bird call
        # very close to the microphone.
        imp_inds = np.where(impulse > 0)[0].tolist()
        imp = self.countConsecutive(imp_inds, len(impulse))

        impulse = []
        for item in imp:
            if item > blocksize or item == 0:        # Note threshold - blocksize, 10 consecutive blocks ~1/25 sec
                impulse.append(1)
            else:
                impulse.append(0)

        impulse = list(chain.from_iterable(repeat(e, window) for e in impulse))  # Make it same length as self.audioData

        if len(impulse) > len(self.audioData):      # Sanity check
            impulse = impulse[0:len(self.audioData)]
        elif len(impulse) < len(self.audioData):
            gap = len(self.audioData) - len(impulse)
            impulse = np.pad(impulse, (0, gap), 'constant')

        return impulse

    def countConsecutive(self, nums, length):
        gaps = [[s, e] for s, e in zip(nums, nums[1:]) if s + 1 < e]
        edges = iter(nums[:1] + sum(gaps, []) + nums[-1:])
        edges = list(zip(edges, edges))
        edges_reps = [item[1] - item[0] + 1 for item in edges]
        res = np.zeros((length)).tolist()
        t = 0
        for item in edges:
            for i in range(item[0], item[1]+1):
                res[i] = edges_reps[t]
            t += 1
        return res

    def rainClick(self):
        """
        delete random clicks e.g. rain.
        """
        # TODO
        return
        newSegments = copy.deepcopy(self.segments)
        if newSegments.__len__() > 1:
            # Get avg energy
            sp = SignalProc.SignalProc()
            sp.data = self.audioData
            sp.sampleRate = self.sampleRate
            rawsg = sp.spectrogram()
            # Normalise
            rawsg -= np.mean(rawsg, axis=0)
            rawsg /= np.max(np.abs(rawsg), axis=0)
            mean = np.mean(rawsg)
            std = np.std(rawsg)
            thr = mean - 2 * std  # thr for the recording

            for seg in self.segments:
                if seg[0] == -1:
                    continue
                else:
                    data = self.audioData[int(seg[0]*self.sampleRate):int(seg[1]*self.sampleRate)]
                    rawsg = sp.spectrogram(data)
                    # Normalise
                    rawsg -= np.mean(rawsg, axis=0)
                    rawsg /= np.max(np.abs(rawsg), axis=0)
                    if np.min(rawsg) < thr:
                        newSegments.remove(seg)
        self.segments = newSegments

    def fundamentalFrq(self, fileName=None):
        '''
        Check for fundamental frequency of the segments, discard the segments that do not indicate the species.
        '''
        for segix in reversed(range(len(self.segments))):
            seg = self.segments[segix]

            # read the sound segment and check fundamental frq.
            secs = int(seg[1] - seg[0])
            # Got to read from the source instead of using self.audioData - ff is wrong if you use self.audioData somehow
            sp = SignalProc.SignalProc(256, 128)
            if fileName:
                sp.readWav(fileName, secs, seg[0])
                self.sampleRate = sp.sampleRate
                self.audioData = sp.data
            else:
                sp.data = self.audioData
                sp.sampleRate = self.sampleRate

            # denoise before fundamental frq. extraction
            # data = self.denoise_filter(level=8, d=True, f=False, f1=self.fLow, f2=self.fHigh)
            # sp.data = data
            # sp.sampleRate = self.sampleRate
            _ = sp.spectrogram(mean_normalise=True, onesided=True, multitaper=False)

            segment = Segmenter(sp, fs=sp.sampleRate)
            pitch, y, minfreq, W = segment.yin(minfreq=100)
            ind = np.squeeze(np.where(pitch > minfreq))
            pitch = pitch[ind]

            if pitch.size == 0:
                print('Segment ', seg, ' *++ no fundamental freq detected, could be faded call or noise')
                del self.segments[segix]
            elif pitch.size == 1:
                if (pitch < self.F0[0]) or (pitch > self.F0[1]):
                    print('segment ', seg, round(pitch), ' *-- fundamental freq is out of range, could be noise')
                    del self.segments[segix]
            elif (np.mean(pitch) < self.F0[0]) or (np.mean(pitch) > self.F0[1]):
                print('segment* ', seg, round(np.mean(pitch)), pitch, np.median(pitch), ' *-- fundamental freq is out of range, could be noise')
                del self.segments[segix]

    def denoise_filter(self, level=5, d=False, f=False, f1=0, f2=0):
        """ Wavelet denoise and band-pass filter for fundamental frquency
        """
        # TODO: only used in fundamental frq calculation. Remove when sorting out FF.
        if d:
            wf = WaveletFunctions.WaveletFunctions(data=self.audioData, wavelet='dmey2', maxLevel=level, samplerate=self.sampleRate)
            denoisedData = wf.waveletDenoise(thresholdType="soft", maxLevel=level)
        else:
            denoisedData=self.audioData

        if f:
            filteredDenoisedData = self.sp.ButterworthBandpass(denoisedData, self.sampleRate, low=f1, high=f2)
        else:
            filteredDenoisedData = denoisedData

        return filteredDenoisedData
