# AviaNZ.py
#
# This is the main class for the AviaNZ interface
# Version 1.3 23/10/18
# Authors: Stephen Marsland, Nirosha Priyadarshani, Julius Juodakis

#    AviaNZ birdsong analysis program
#    Copyright (C) 2017--2018

#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.

#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.

#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.

# TODO: Automate some of the training options, also the extra filters
# TODO: And the false positive graph
# TODO: Think about the filter dictionary a bit more for option checking, and adding new options
# TODO: Sort out the segmentation code, particularly wrt merging sets of segments
# TODO: Full list of next steps

import sys, os, json, platform, re, shutil
from shutil import copyfile
from openpyxl import load_workbook, Workbook

from PyQt5.QtGui import QIcon, QStandardItemModel, QStandardItem, QKeySequence
from PyQt5.QtWidgets import QApplication, QInputDialog, QFileDialog, QMainWindow, QActionGroup, QToolButton, QLabel, QSlider, QScrollBar, QDoubleSpinBox, QPushButton, QListWidget, QListWidgetItem, QMenu, QFrame, QMessageBox, QWidgetAction, QComboBox, QTreeView, QShortcut
from PyQt5.QtCore import Qt, QDir, QTimer, QPoint, QPointF, QLocale, QModelIndex
from PyQt5.QtMultimedia import QAudio, QAudioFormat

import wavio
import numpy as np
from scipy import signal
from scipy.signal import medfilt
from scipy.ndimage.filters import median_filter

import pyqtgraph as pg
pg.setConfigOption('background','w')
pg.setConfigOption('foreground','k')
pg.setConfigOption('antialias',True)
from pyqtgraph.Qt import QtCore, QtGui
from pyqtgraph.dockarea import DockArea, Dock
import pyqtgraph.functions as fn
import pyqtgraph.exporters as pge

import SupportClasses as SupportClasses
import Dialogs as Dialogs
import SignalProc
import Segment
import WaveletSegment
import WaveletFunctions
#import Features
#import Learning
import AviaNZ_batch
import fnmatch
import librosa

from openpyxl import load_workbook, Workbook
import matplotlib.markers as mks
import matplotlib.pyplot as plt
import matplotlib.ticker as mtick

from pyqtgraph.parametertree import Parameter, ParameterTree 

import locale, time

import click, webbrowser, colourMaps, copy, math

print("Package import complete.")

# import pdb
# from PyQt5.QtCore import pyqtRemoveInputHook
# from pdb import set_trace
# 
# def debug_trace():
#     pyqtRemoveInputHook()
#     set_trace()
class AviaNZ(QMainWindow):
    """Main class for the user interface.
    Contains most of the user interface and plotting code"""

    def __init__(self,root=None,configdir=None,CLI=False,cheatsheet=False,zooniverse=False,firstFile='', imageFile='', command=''):
        """Initialisation of the class. Load main config and bird lists from configdir.
        Also initialises the data structures and loads an initial file (specified explicitly)
        and sets up the window.
        One interesting configuration point is the DOC setting, which hides the more 'research' functions."""
        print("Starting AviaNZ...")

        super(AviaNZ, self).__init__()
        self.root = root
        self.CLI = CLI
        self.cheatsheet = cheatsheet
        self.zooniverse = zooniverse
        self.trainPerFile = True

        # configdir passes the standard user app dir based on OS.
        # At this point, the main config file should already be ensured to exist.
        self.configdir = configdir
        self.configfile = os.path.join(configdir, "AviaNZconfig.txt")
        self.ConfigLoader = SupportClasses.ConfigLoader()
        self.config = self.ConfigLoader.config(self.configfile)
        self.saveConfig = True

        # Load filters
        self.filtersDir = os.path.join(configdir, self.config['FiltersDir'])
        self.FilterFiles = self.ConfigLoader.filters(self.filtersDir)

        # Load the birdlists:
        # short list is necessary, long list can be None
        self.shortBirdList = self.ConfigLoader.shortbl(self.config['BirdListShort'], configdir)
        if self.shortBirdList is None:
            sys.exit()
        # Will be None if fails to load or filename was "None"
        self.longBirdList = self.ConfigLoader.longbl(self.config['BirdListLong'], configdir)

        if self.longBirdList is None:
            # If don't have a long bird list,
            # check that the length of the short bird list is OK, and otherwise split it
            # 40 is a bit random, but 20 in a list is long enough!
            if len(self.shortBirdList) > 40:
                self.longBirdList = self.shortBirdList.copy()
                self.shortBirdList = self.shortBirdList[:40]
            else:
                self.longBirdList = None

        # Noise level details
        self.noiseLevel = None
        self.noiseTypes = []

        # avoid comma/point problem in number parsing
        QLocale.setDefault(QLocale(QLocale.English, QLocale.NewZealand))
        print('Locale is set to ' + QLocale().name())

        # The data structures for the segments
        self.listLabels = []
        self.listRectanglesa1 = []
        self.listRectanglesa2 = []
        self.SegmentRects = []
        self.segmentPlots=[]
        self.box1id = -1

        self.started = False
        self.startedInAmpl = False
        self.startTime = 0
        self.segmentsToSave = False

        self.lastSpecies = ["Don't Know"]
        self.DOC = self.config['DOC']
        self.Hartley = self.config['Hartley']
        self.extra = "none"

        # Whether or not the context menu allows multiple birds.
        self.multipleBirds = self.config['MultipleSpecies']

        self.SoundFileDir = self.config['SoundFileDir']
        self.previousFile = None
        self.focusRegion = None
        self.operator = self.config['operator']
        self.reviewer = self.config['reviewer']

        # audio things
        self.audioFormat = QAudioFormat()
        self.audioFormat.setCodec("audio/pcm")
        self.audioFormat.setByteOrder(QAudioFormat.LittleEndian)
        self.audioFormat.setSampleType(QAudioFormat.SignedInt)

        # Spectrogram
        self.sgOneSided = True
        self.sgMeanNormalise = True
        self.sgMultitaper = False
        self.sgEqualLoudness = False

        # working directory
        if not os.path.isdir(self.SoundFileDir):
            print("Directory doesn't exist: making it")
            os.makedirs(self.SoundFileDir)

        self.backupDatafiles()

        # INPUT FILE LOADING
        # search order: infile -> firstFile -> dialog
        # Make life easier for now: preload a birdsong
        print("Loaded")
        if not os.path.isfile(firstFile) and not cheatsheet and not zooniverse:
            firstFile = self.SoundFileDir + '/' + 'tril1.wav' #'male1.wav' # 'kiwi.wav'
            #firstFile = "/home/julius/Documents/kiwis/rec/birds1.wav"

        if not os.path.isfile(firstFile) and not cheatsheet and not zooniverse:
            if self.CLI:
                print("file %s not found, exiting" % firstFile)
                sys.exit()
            else:
                # pop up a dialog to select file
                firstFile, drop = QtGui.QFileDialog.getOpenFileName(self, 'Choose File', self.SoundFileDir, "Wav files (*.wav)")
                while firstFile == '':
                    msg = SupportClasses.MessagePopup("w", "Select Sound File", "Choose a sound file to proceed.\nDo you want to continue?")
                    msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
                    reply = msg.exec_()
                    if reply == QMessageBox.Yes:
                        firstFile, drop = QFileDialog.getOpenFileName(self, 'Choose File', self.SoundFileDir, "Wav files (*.wav)")
                    else:
                        sys.exit()

        # parse firstFile to dir and file parts
        if not cheatsheet and not zooniverse:
            self.SoundFileDir = os.path.dirname(firstFile)
            firstFile = os.path.basename(firstFile)
            print("Working dir set to %s" % self.SoundFileDir)
            print("Opening file %s" % firstFile)

        # to keep code simpler, graphic options are created even in CLI mode
        # they're just not shown because QMainWindow.__init__ is skipped
        if not self.CLI:
            QMainWindow.__init__(self, root)

        # parse mouse settings
        if self.config['drawingRightBtn']:
            self.MouseDrawingButton = QtCore.Qt.RightButton
        else:
            self.MouseDrawingButton = QtCore.Qt.LeftButton
        # SRM: min size of box to add
        # TODO: Currently only hard-coded here!
        self.minboxsize = 0.1
        self.createMenu()
        self.createFrame()

        self.resetStorageArrays()
        if self.CLI:
            if cheatsheet or zooniverse:
                # use infile and imagefile as directories 
                print(firstFile)
                self.SoundFileDir = firstFile
                # Read folders and sub-folders
                for root, dirs, files in os.walk(firstFile):
                    for f in files:
                        if f[-4:] == '.wav':
                            print(os.path.join(root, f))
                            self.loadFile(os.path.join(root, f))
                            self.widthWindow.setValue(60)  # self.datalengthSec)
                            self.saveImage(os.path.join(root, f[:-4]))
            else:
                self.loadFile(firstFile)
                while command!=():
                    c = command[0]
                    command = command[1:]
                    print("Next command to execute is %s" % c)
                    if c=="denoise":
                        self.denoise()
                    elif c=="segment":
                        self.segment()
                    else:
                        print("ERROR: %s is not a valid command" % c)
                        sys.exit()
                if imageFile!='':
                    # reset images to show full width if in CLI:
                    self.widthWindow.setValue(self.datalengthSec)
                    # looks unnecessary:
                    # self.p_spec.setXRange(0, self.convertAmpltoSpec(self.datalengthSec), update=True, padding=0)
                    self.saveImage(imageFile)
        else:
            # Make the window and associated widgets
            self.setWindowTitle('AviaNZ')
            # Make the window full screen
            if self.config['StartMaximized']:
                self.showMaximized()
            keyPressed = QtCore.Signal(int)


            # Save the segments every minute
            self.timer = QTimer()
            #QObject.connect(self.timer, SIGNAL("timeout()"), self.saveSegments)
            self.timer.timeout.connect(self.saveSegments)
            self.timer.start(self.config['secsSave']*1000)
            
            self.fillFileList(firstFile)
            self.listLoadFile(firstFile)
            #self.previousFile = firstFile

        if self.DOC and not cheatsheet and not zooniverse:
            self.setOperatorReviewerDialog()


    def createMenu(self):
        """ Create the menu entries at the top of the screen and link them as appropriate.
        Some of them are initialised according to the data in the configuration file."""

        fileMenu = self.menuBar().addMenu("&File")
        fileMenu.addAction("&Open sound file", self.openFile, "Ctrl+O")
        # fileMenu.addAction("&Change Directory", self.chDir)
        fileMenu.addAction("&Set Operator/Reviewer (Current File)", self.setOperatorReviewerDialog)
        fileMenu.addSeparator()
        fileMenu.addAction("Restart Program",self.restart,"Ctrl+R")
        fileMenu.addAction("Quit",self.quit,"Ctrl+Q")

        # This is a very bad way to do this, but I haven't worked anything else out (setMenuRole() didn't work)
        # Add it a second time, then it appears!
        if platform.system() == 'Darwin':
            fileMenu.addAction("&Quit",self.quit,"Ctrl+Q")

        specMenu = self.menuBar().addMenu("&Appearance")

        self.useAmplitudeTick = specMenu.addAction("Show amplitude plot", self.useAmplitudeCheck)
        self.useAmplitudeTick.setCheckable(True)
        self.useAmplitudeTick.setChecked(self.config['showAmplitudePlot'])
        self.useAmplitude = True

        self.useFilesTick = specMenu.addAction("Show file list", self.useFilesCheck)
        self.useFilesTick.setCheckable(True)
        self.useFilesTick.setChecked(self.config['showListofFiles'])

        # this can go under "Change interface settings"
        self.showOverviewSegsTick = specMenu.addAction("Show annotation overview", self.showOverviewSegsCheck)
        self.showOverviewSegsTick.setCheckable(True)
        self.showOverviewSegsTick.setChecked(self.config['showAnnotationOverview'])

        self.showPointerDetails = specMenu.addAction("Show pointer details in spectrogram", self.showPointerDetailsCheck)
        self.showPointerDetails.setCheckable(True)
        self.showPointerDetails.setChecked(self.config['showPointerDetails'])

        specMenu.addSeparator()

        colMenu = specMenu.addMenu("&Choose colour map")
        colGroup = QActionGroup(self)
        for colour in self.config['ColourList']:
            cm = colMenu.addAction(colour)
            cm.setCheckable(True)
            if colour==self.config['cmap']:
                cm.setChecked(True)
            receiver = lambda checked, cmap=colour: self.setColourMap(cmap)
            cm.triggered.connect(receiver)
            colGroup.addAction(cm)
        self.invertcm = specMenu.addAction("Invert colour map",self.invertColourMap)
        self.invertcm.setCheckable(True)
        self.invertcm.setChecked(self.config['invertColourMap'])

        # specMenu.addSeparator()
        specMenu.addAction("Change spectrogram parameters",self.showSpectrogramDialog)

        if not self.DOC and not self.Hartley:
            self.showInvSpec = specMenu.addAction("Show inverted spectrogram", self.showInvertedSpectrogram)
            self.showInvSpec.setCheckable(True)
            self.showInvSpec.setChecked(False)

            specMenu.addSeparator()
            specMenu.addAction("Show training diagnostics",self.showDiagnosticDialog)
            extraMenu = specMenu.addMenu("Diagnostic plots")
            extraGroup = QActionGroup(self)
            for ename in ["none", "Wavelet scalogram", "Wavelet correlations", "Wind energy", "Rain", "Filtered spectrogram, new + AA", "Filtered spectrogram, new", "Filtered spectrogram, old"]:
                em = extraMenu.addAction(ename)
                em.setCheckable(True)
                if ename == self.extra:
                    em.setChecked(True)
                receiver = lambda checked, ename=ename: self.setExtraPlot(ename)
                em.triggered.connect(receiver)
                extraGroup.addAction(em)

        specMenu.addSeparator()

        self.readonly = specMenu.addAction("Make read only",self.makeReadOnly,"Ctrl+R")
        self.readonly.setCheckable(True)
        self.readonly.setChecked(self.config['readOnly'])
        
        specMenu.addSeparator()
        specMenu.addAction("Interface settings", self.changeSettings)

        actionMenu = self.menuBar().addMenu("&Actions")
        actionMenu.addAction("&Delete all segments", self.deleteAll, "Ctrl+D")
        if not self.Hartley:
            actionMenu.addSeparator()
            actionMenu.addAction("Denoise",self.showDenoiseDialog)
            actionMenu.addAction("Add metadata about noise", self.addNoiseData, "Ctrl+N")
            #actionMenu.addAction("Find matches",self.findMatches)
            actionMenu.addSeparator()
            self.showFundamental = actionMenu.addAction("Show fundamental frequency", self.showFundamentalFreq,"Ctrl+F")
            self.showFundamental.setCheckable(True)
            self.showFundamental.setChecked(False)
            self.showSpectral = actionMenu.addAction("Show spectral derivative", self.showSpectralDeriv)
            self.showSpectral.setCheckable(True)
            self.showSpectral.setChecked(False)

        if not self.DOC and not self.Hartley:
            actionMenu.addAction("Filter spectrogram",self.medianFilterSpec)
            actionMenu.addAction("Denoise spectrogram",self.denoiseImage)

        if not self.Hartley:
            actionMenu.addSeparator()
            actionMenu.addAction("Segment",self.segmentationDialog,"Ctrl+S")

        if not self.DOC and not self.Hartley:
            actionMenu.addAction("Classify segments",self.classifySegments,"Ctrl+C")
            actionMenu.addSeparator()
        #self.showAllTick = actionMenu.addAction("Show all pages", self.showAllCheck)
        #self.showAllTick.setCheckable(True)
        #self.showAllTick.setChecked(self.config['showAllPages'])

        if not self.Hartley:
            actionMenu.addAction("Human Review [All segments]",self.humanClassifyDialog1,"Ctrl+1")
            actionMenu.addAction("Human Review [Choose species]",self.humanRevDialog2,"Ctrl+2")
            actionMenu.addSeparator()
            actionMenu.addAction("Export segments to Excel",self.exportSeg)
            actionMenu.addSeparator()

        actionMenu.addAction("Train a species detector", self.trainWaveletDialog)
        actionMenu.addSeparator()
        actionMenu.addAction("Save as image",self.saveImage,"Ctrl+I")
        actionMenu.addAction("Save selected sound", self.save_selected_sound)
        actionMenu.addSeparator()
        actionMenu.addAction("Put docks back",self.dockReplace)

        helpMenu = self.menuBar().addMenu("&Help")
        #aboutAction = QAction("About")
        helpMenu.addAction("Help",self.showHelp,"Ctrl+H")
        helpMenu.addAction("Cheat Sheet", self.showCheatSheet)
        helpMenu.addSeparator()
        helpMenu.addAction("About",self.showAbout,"Ctrl+A")
        if platform.system() == 'Darwin':
            helpMenu.addAction("About",self.showAbout,"Ctrl+A")

    def showAbout(self):
        """ Create the About Message Box"""
        msg = SupportClasses.MessagePopup("a", "About", ".")
        msg.exec_()
        return

    def showHelp(self):
        """ Show the user manual (a pdf file)"""
        # webbrowser.open_new(r'file://' + os.path.realpath('./Docs/AviaNZManual.pdf'))
        webbrowser.open_new(r'http://avianz.net/docs/AviaNZManual_v1.4.pdf')

    def showCheatSheet(self):
        """ Show the cheatsheet of sample spectrograms (a pdf file)"""
        # webbrowser.open_new(r'file://' + os.path.realpath('./Docs/CheatSheet.pdf'))
        webbrowser.open_new(r'http://avianz.net/index.php/cheat-sheet')

    def createFrame(self):
        """ Creates the main window.
        This consists of a set of pyqtgraph docks with widgets in.
         d_ for docks, w_ for widgets, p_ for plots"""

        # Make the window and set its size
        self.area = DockArea()
        self.setCentralWidget(self.area)
        self.resize(1240,600)
        self.move(100,50)

        # Make the docks and lay them out
        self.d_overview = Dock("Overview",size=(1200,150))
        self.d_ampl = Dock("Amplitude",size=(1200,150))
        self.d_spec = Dock("Spectrogram",size=(1200,300))
        self.d_controls = Dock("Controls",size=(40,100))
        self.d_files = Dock("Files",size=(40,200))
        self.d_plot = Dock("Plots",size=(1200,150))
        self.d_controls.setSizePolicy(1,1)

        self.area.addDock(self.d_files,'left')
        self.area.addDock(self.d_overview,'right',self.d_files)
        self.area.addDock(self.d_ampl,'bottom',self.d_overview)
        self.area.addDock(self.d_spec,'bottom',self.d_ampl)
        self.area.addDock(self.d_controls,'bottom',self.d_files)
        self.area.addDock(self.d_plot,'bottom',self.d_spec)

        # Store the state of the docks in case the user wants to reset it
        self.state = self.area.saveState()
        containers, docks = self.area.findAll()
        self.state_cont = [cont.sizes() for cont in containers]

        # Put content widgets in the docks
        self.w_overview = pg.LayoutWidget()
        self.d_overview.addWidget(self.w_overview)
        self.w_overview1 = pg.GraphicsLayoutWidget()
        self.w_overview1.ci.layout.setContentsMargins(0.5, 1, 0.5, 1)
        self.w_overview.addWidget(self.w_overview1,row=0, col=2,rowspan=3)

        self.p_overview = self.w_overview1.addViewBox(enableMouse=False,enableMenu=False,row=0,col=0)
        self.p_overview2 = SupportClasses.ChildInfoViewBox(enableMouse=False, enableMenu=False)
        self.w_overview1.addItem(self.p_overview2,row=1,col=0)
        self.p_overview2.setXLink(self.p_overview)

        self.w_ampl = pg.GraphicsLayoutWidget()
        self.p_ampl = SupportClasses.DragViewBox(self, enableMouse=False,enableMenu=False,enableDrag=False, thisIsAmpl=True)
        self.w_ampl.addItem(self.p_ampl,row=0,col=1)
        self.d_ampl.addWidget(self.w_ampl)

        self.w_spec = pg.GraphicsLayoutWidget()
        self.p_spec = SupportClasses.DragViewBox(self, enableMouse=False,enableMenu=False,enableDrag=self.config['specMouseAction']==3, thisIsAmpl=False)
        self.w_spec.addItem(self.p_spec,row=0,col=1)
        self.d_spec.addWidget(self.w_spec)

        self.w_plot = pg.GraphicsLayoutWidget()
        self.p_plot = self.w_plot.addViewBox(enableMouse=False,enableMenu=False)
        self.w_plot.addItem(self.p_plot,row=0,col=1)
        self.d_plot.addWidget(self.w_plot)

        # The axes
        # Time axis has to go separately in loadFile
        self.ampaxis = pg.AxisItem(orientation='left')
        self.w_ampl.addItem(self.ampaxis,row=0,col=0)
        self.ampaxis.linkToView(self.p_ampl)
        self.ampaxis.setWidth(w=65)
        self.ampaxis.setLabel('')

        self.specaxis = pg.AxisItem(orientation='left')
        if not self.zooniverse:
            self.w_spec.addItem(self.specaxis,row=0,col=0)
        self.specaxis.linkToView(self.p_spec)
        self.specaxis.setWidth(w=65)

        # Plot window also needs an axis to make them line up
        self.plotaxis = pg.AxisItem(orientation='left')
        self.w_plot.addItem(self.plotaxis,row=0,col=0)
        self.plotaxis.linkToView(self.p_plot)
        self.plotaxis.setWidth(w=65)
        self.plotaxis.setLabel('')

        # Hide diagnostic plot window until requested
        self.d_plot.hide()

        # The print out at the bottom of the spectrogram with data in
        self.pointData = pg.TextItem(color=(255,0,0),anchor=(0,0))
        #self.p_spec.addItem(self.pointData)

        # The various plots
        self.overviewImage = pg.ImageItem(enableMouse=False)
        self.p_overview.addItem(self.overviewImage)
        self.overviewImageRegion = pg.LinearRegionItem()
        # this is needed for compatibility with other shaded rectangles:
        self.overviewImageRegion.lines[0].btn = QtCore.Qt.RightButton
        self.overviewImageRegion.lines[1].btn = QtCore.Qt.RightButton
        self.p_overview.addItem(self.overviewImageRegion, ignoreBounds=True)
        self.amplPlot = pg.PlotDataItem()
        self.p_ampl.addItem(self.amplPlot)
        self.specPlot = pg.ImageItem()
        self.p_spec.addItem(self.specPlot)

        # Connect up the listeners
        self.p_ampl.scene().sigMouseClicked.connect(self.mouseClicked_ampl)
        self.p_spec.scene().sigMouseClicked.connect(self.mouseClicked_spec)

        # Connect up so can disconnect if not selected...
        self.p_spec.scene().sigMouseMoved.connect(self.mouseMoved)
        self.p_spec.addItem(self.pointData)

        # The content of the other two docks
        self.w_controls = pg.LayoutWidget()
        self.d_controls.addWidget(self.w_controls)
        self.w_files = pg.LayoutWidget()
        self.d_files.addWidget(self.w_files)

        # The buttons to move through the overview
        self.leftBtn = QToolButton()
        self.leftBtn.setArrowType(Qt.LeftArrow)
        #self.connect(self.leftBtn, SIGNAL('clicked()'), self.moveLeft)
        self.leftBtn.clicked.connect(self.moveLeft)
        self.w_overview.addWidget(self.leftBtn,row=0,col=0)
        self.rightBtn = QToolButton()
        self.rightBtn.setArrowType(Qt.RightArrow)
        #self.connect(self.rightBtn, SIGNAL('clicked()'), self.moveRight)
        self.rightBtn.clicked.connect(self.moveRight)
        self.w_overview.addWidget(self.rightBtn,row=0,col=1)

        # Buttons to move to next/previous five minutes
        self.prev5mins=QToolButton()
        self.prev5mins.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaSeekBackward))
        self.prev5mins.setToolTip("Previous page")
        #self.connect(self.prev5mins, SIGNAL('clicked()'), self.movePrev5mins)
        self.prev5mins.clicked.connect(self.movePrev5mins)
        self.w_overview.addWidget(self.prev5mins,row=2,col=0)
        self.next5mins=QToolButton()
        self.next5mins.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaSeekForward))
        self.next5mins.setToolTip("Next page")
        #self.connect(self.next5mins, SIGNAL('clicked()'), self.moveNext5mins)
        self.next5mins.clicked.connect(self.moveNext5mins)
        self.w_overview.addWidget(self.next5mins,row=2,col=1)
        self.placeInFileLabel = QLabel('')
        self.w_overview.addWidget(self.placeInFileLabel,row=1,colspan=2)

        # Corresponding keyboard shortcuts:
        self.moveLeftKey = QShortcut(QKeySequence(Qt.Key_Left), self)
        self.moveLeftKey.activated.connect(self.moveLeft)
        self.moveRightKey = QShortcut(QKeySequence(Qt.Key_Right), self)
        self.moveRightKey.activated.connect(self.moveRight)
        self.movePrev5minsKey = QShortcut(QKeySequence("Shift+Left"), self)
        self.movePrev5minsKey.activated.connect(self.movePrev5mins)
        self.moveNext5minsKey = QShortcut(QKeySequence("Shift+Right"), self)
        self.moveNext5minsKey.activated.connect(self.moveNext5mins)

        # Button to move to the next file in the list
        self.nextFileBtn=QToolButton()
        self.nextFileBtn.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaSkipForward))
        self.nextFileBtn.clicked.connect(self.openNextFile)
        self.nextFileBtn.setToolTip("Open next file")
        self.w_files.addWidget(self.nextFileBtn,row=0,col=1)

        # The buttons inside the controls dock
        self.playButton = QtGui.QToolButton()
        self.playButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaPlay))
        self.playButton.setIconSize(QtCore.QSize(20, 20))
        self.playButton.setToolTip("Play visible")
        self.playButton.clicked.connect(self.playVisible)
        self.playKey = QShortcut(QKeySequence("Space"), self)
        self.playKey.activated.connect(self.playVisible)

        self.stopButton = QtGui.QToolButton()
        self.stopButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaStop))
        self.stopButton.setIconSize(QtCore.QSize(20, 20))
        self.stopButton.setToolTip("Stop playback")
        self.stopButton.clicked.connect(self.stopPlayback)

        self.playSegButton = QtGui.QToolButton()
        self.playSegButton.setIcon(QtGui.QIcon('img/playsegment.png'))
        self.playSegButton.setIconSize(QtCore.QSize(20, 20))
        self.playSegButton.setToolTip("Play selected")
        self.playSegButton.clicked.connect(self.playSelectedSegment)
        self.playSegButton.setEnabled(False)

        self.quickDenButton = QtGui.QToolButton()
        self.quickDenButton.setIcon(QtGui.QIcon('img/denoisesegment.png'))
        self.quickDenButton.setIconSize(QtCore.QSize(20, 20))
        self.quickDenButton.setToolTip("Denoise segment")
        self.quickDenButton.clicked.connect(self.denoiseSeg)
        self.quickDenButton.setEnabled(False)

        # self.quickDenNButton = QtGui.QToolButton()
        # self.quickDenNButton.setIcon(QtGui.QIcon('img/denoisesegment.png'))
        # self.quickDenNButton.setIconSize(QtCore.QSize(20, 20))
        # self.quickDenNButton.setToolTip("Denoise segment, node-specific")
        # self.quickDenNButton.clicked.connect(self.denoiseSegN)
        # self.quickDenNButton.setEnabled(False)

        # self.quickWPButton = QtGui.QToolButton()
        # self.quickWPButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_ArrowDown))
        # self.quickWPButton.setIconSize(QtCore.QSize(20, 20))
        # self.quickWPButton.setToolTip("Decompose to WP")
        # self.quickWPButton.clicked.connect(self.decomposeWP)

        self.playBandLimitedSegButton = QtGui.QToolButton()
        self.playBandLimitedSegButton.setIcon(QtGui.QIcon('img/playBandLimited.png'))
        self.playBandLimitedSegButton.setIconSize(QtCore.QSize(20, 20))
        self.playBandLimitedSegButton.setToolTip("Play selected-band limited")
        self.playBandLimitedSegButton.clicked.connect(self.playBandLimitedSegment)
        self.playBandLimitedSegButton.setEnabled(False)

        self.timePlayed = QLabel()

        # Volume control
        self.volSlider = QSlider(Qt.Horizontal)
        self.volSlider.sliderMoved.connect(self.volSliderMoved)
        self.volSlider.setRange(0,100)
        self.volSlider.setValue(50)
        self.volIcon = QLabel()
        self.volIcon.setPixmap(self.style().standardIcon(QtGui.QStyle.SP_MediaVolume).pixmap(32))

        # Brightness, and contrast sliders
        self.brightnessSlider = QSlider(Qt.Horizontal)
        self.brightnessSlider.setMinimum(0)
        self.brightnessSlider.setMaximum(100)
        self.brightnessSlider.setValue(self.config['brightness'])
        self.brightnessSlider.setTickInterval(1)
        self.brightnessSlider.valueChanged.connect(self.setColourLevels)

        self.contrastSlider = QSlider(Qt.Horizontal)
        self.contrastSlider.setMinimum(0)
        self.contrastSlider.setMaximum(100)
        self.contrastSlider.setValue(self.config['contrast'])
        self.contrastSlider.setTickInterval(1)
        self.contrastSlider.valueChanged.connect(self.setColourLevels)

        # Delete segment button. We have to get rid of the extra event args
        deleteButton = QPushButton("&Delete Current Segment")
        deleteButton.clicked.connect(lambda _ : self.deleteSegment())

        # The spinbox for changing the width shown in the controls dock
        self.widthWindow = QDoubleSpinBox()
        self.widthWindow.setSingleStep(1.0)
        self.widthWindow.setDecimals(2)
        self.widthWindow.setValue(self.config['windowWidth'])
        self.widthWindow.valueChanged[float].connect(self.changeWidth)

        # Place all these widgets in the Controls dock
        self.w_controls.addWidget(self.playButton,row=0,col=0)
        self.w_controls.addWidget(self.stopButton,row=0,col=1)
        self.w_controls.addWidget(self.playSegButton,row=0,col=2)
        self.w_controls.addWidget(self.playBandLimitedSegButton,row=0,col=3)
        self.w_controls.addWidget(self.quickDenButton,row=1,col=0)
        # self.w_controls.addWidget(self.quickDenNButton,row=1,col=1)
        # self.w_controls.addWidget(self.quickWPButton,row=1,col=0)
        self.w_controls.addWidget(self.timePlayed,row=2,col=0, colspan=4)
        self.w_controls.addWidget(self.volIcon, row=3, col=0)
        self.w_controls.addWidget(self.volSlider, row=3, col=1, colspan=3)
        self.w_controls.addWidget(QLabel("Brightness"),row=4,col=0,colspan=4)
        self.w_controls.addWidget(self.brightnessSlider,row=5,col=0,colspan=4)
        self.w_controls.addWidget(QLabel("Contrast"),row=6,col=0,colspan=4)
        self.w_controls.addWidget(self.contrastSlider,row=7,col=0,colspan=4)
        self.w_controls.addWidget(deleteButton,row=8,col=0,colspan=4)
        self.w_controls.addWidget(QLabel('Visible window (seconds)'),row=9,col=0,colspan=4)
        self.w_controls.addWidget(self.widthWindow,row=10,col=0,colspan=4)

        # The slider to show playback position
        # This is hidden, but controls the moving bar
        self.playSlider = QSlider(Qt.Horizontal)
        # self.playSlider.sliderReleased.connect(self.playSliderMoved)
        self.playSlider.setVisible(False)
        self.d_spec.addWidget(self.playSlider)
        self.bar = pg.InfiniteLine(angle=90, movable=True, pen={'color': 'c', 'width': 3})
        self.bar.btn = self.MouseDrawingButton

        # A slider to move through the file easily
        self.scrollSlider = QScrollBar(Qt.Horizontal)
        self.scrollSlider.valueChanged.connect(self.scroll)
        self.d_spec.addWidget(self.scrollSlider)

        # List to hold the list of files
        self.listFiles = QListWidget()
        self.listFiles.setMinimumWidth(150)
        self.listFiles.itemDoubleClicked.connect(self.listLoadFile)

        self.w_files.addWidget(QLabel('Double click to open'),row=0,col=0)
        self.w_files.addWidget(QLabel('Red names have been viewed'),row=1,col=0)
        self.w_files.addWidget(self.listFiles,row=2,colspan=2)

        # The context menu (drops down on mouse click) to select birds
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.menuBirdList = QMenu()
        self.menuBird2 = QMenu('Other')
        #self.menuBird2 = self.menuBirdList.addMenu('Other')
        # New line to allow multiple selections
        self.menuBirdList.installEventFilter(self)
        self.menuBird2.installEventFilter(self)
        self.fillBirdList()
        self.menuBirdList.triggered.connect(self.birdSelectedMenu)
        self.menuBird2.triggered.connect(self.birdSelectedMenu)
        #self.menuBirdList.aboutToHide.connect(self.processMultipleBirdSelections)

        # Make the colours that are used in the interface
        # The dark ones are to draw lines instead of boxes
        self.ColourSelected = QtGui.QColor(self.config['ColourSelected'][0], self.config['ColourSelected'][1], self.config['ColourSelected'][2], self.config['ColourSelected'][3])
        self.ColourNamed = QtGui.QColor(self.config['ColourNamed'][0], self.config['ColourNamed'][1], self.config['ColourNamed'][2], self.config['ColourNamed'][3])
        self.ColourNone = QtGui.QColor(self.config['ColourNone'][0], self.config['ColourNone'][1], self.config['ColourNone'][2], self.config['ColourNone'][3])
        self.ColourPossible = QtGui.QColor(self.config['ColourPossible'][0], self.config['ColourPossible'][1], self.config['ColourPossible'][2], self.config['ColourPossible'][3])

        self.ColourSelectedDark = QtGui.QColor(self.config['ColourSelected'][0], self.config['ColourSelected'][1], self.config['ColourSelected'][2], 255)
        self.ColourNamedDark = QtGui.QColor(self.config['ColourNamed'][0], self.config['ColourNamed'][1], self.config['ColourNamed'][2], 255)
        self.ColourNoneDark = QtGui.QColor(self.config['ColourNone'][0], self.config['ColourNone'][1], self.config['ColourNone'][2], 255)
        self.ColourPossibleDark = QtGui.QColor(self.config['ColourPossible'][0], self.config['ColourPossible'][1], self.config['ColourPossible'][2], 255)

        # Hack to get the type of an ROI
        p_spec_r = SupportClasses.ShadedRectROI(0, 0)
        self.ROItype = type(p_spec_r)

        # Listener for key presses
        self.p_ampl.keyPressed.connect(self.handleKey)
        self.p_spec.keyPressed.connect(self.handleKey)

        # Function calls to check if should show various parts of the interface, whether dragging boxes or not
        self.makeReadOnly()
        self.useAmplitudeCheck()
        self.useFilesCheck()
        self.showOverviewSegsCheck()
        self.dragRectsTransparent()
        self.showPointerDetailsCheck()

        # add statusbar
        self.statusLeft = QLabel("Left")
        self.statusLeft.setFrameStyle(QFrame.Panel) #,QFrame.Sunken)
        self.statusMid = QLabel("????")
        self.statusMid.setFrameStyle(QFrame.Panel) #,QFrame.Sunken)
        self.statusRight = QLabel("")
        self.statusRight.setAlignment(Qt.AlignRight)
        self.statusRight.setFrameStyle(QFrame.Panel) #,QFrame.Sunken)
        # Style
        statusStyle='QLabel {border:transparent}'
        self.statusLeft.setStyleSheet(statusStyle)
        # self.statusMid.setStyleSheet(statusStyle)
        self.statusRight.setStyleSheet(statusStyle)
        self.statusBar().addPermanentWidget(self.statusLeft,1)
        # self.statusBar().addPermanentWidget(self.statusMid,1)
        self.statusBar().addPermanentWidget(self.statusRight,1)

        # Set the message in the status bar
        self.statusLeft.setText("Ready")

        # Plot everything
        if not self.CLI:
            self.show()

    def handleKey(self,ev):
        """ Handle keys pressed during program use.
        These are:
            backspace to delete a segment
            escape to pause playback """
        if ev == Qt.Key_Backspace or ev == Qt.Key_Delete:
            self.deleteSegment()
        elif ev == Qt.Key_Escape and self.media_obj.isPlaying():
            self.stopPlayback()

    def makeFullBirdList(self):
        """ Makes a combo box holding the complete list of birds.
        Some work is needed to keep track of the indices since it's a two column
        list: species and subspecies in most cases.
        Also parses the DOC files, which use > to mark the subspecies. """
        fullbirdlist = QComboBox()
        fullbirdlist.setView(QTreeView())
        fullbirdlist.setRootModelIndex(QModelIndex())

        fullbirdlist.view().setHeaderHidden(True)
        fullbirdlist.view().setItemsExpandable(True)
        fullbirdlist.setMouseTracking(True)

        self.model = QStandardItemModel()
        headlist = []
        if self.longBirdList is not None:
            for bird in self.longBirdList:
                ind = bird.find('>')
                if ind == -1:
                    ind = len(bird)
                if bird[:ind] not in headlist:
                    headlist.append(bird[:ind])
                    item = QStandardItem(bird[:ind])
                    item.setSelectable(True)
                    self.model.appendRow(item)
                if ind < len(bird):
                    subitem = QStandardItem(bird[ind+1:])
                    item.setSelectable(False)
                    item.appendRow(subitem)
                    subitem.setSelectable(True)
        item = QStandardItem("Other")
        item.setSelectable(True)
        self.model.appendRow(item)

        fullbirdlist.setModel(self.model)
        return fullbirdlist

    def fillBirdList(self,unsure=False):
        """ Sets the contents of the context menu.
        The first 20 items are in the first menu, the next in a second menu.
        Any extras go into the combobox at the end of the second list.
        This is called a lot because the order of birds in the list changes since the last choice
        is moved to the top of the list. """
        self.menuBirdList.clear()
        self.menuBird2.clear()

        for item in self.shortBirdList[:20]:
            # Add ? marks if Ctrl menu is called
            if unsure and item != "Don't Know":
                item = item+'?'

            # Transform > marks
            pos = item.find('>')
            if pos > -1:
                item = item[:pos] + ' (' + item[pos+1:] + ')'

            bird = self.menuBirdList.addAction(item)
            bird.setCheckable(True)
            if hasattr(self,'segments') and item in self.segments[self.box1id][4]:
                bird.setChecked(True)
            self.menuBirdList.addAction(bird)
        self.menuBirdList.addMenu(self.menuBird2)
        if self.Hartley:
            for item in self.shortBirdList[20:]:
                # Add ? marks if Ctrl menu is called
                if unsure and item != "Don't Know" and item != "Other":
                    item = item+'?'
                # Transform > marks
                pos = item.find('>')
                if pos > -1:
                    item = item[:pos] + ' (' + item[pos+1:] + ')'

                bird = self.menuBird2.addAction(item)
                bird.setCheckable(True)
                if hasattr(self,'segments') and item in self.segments[self.box1id][4]:
                    bird.setChecked(True)
                self.menuBird2.addAction(bird)
        else:
            for item in self.shortBirdList[20:40]:
                # Add ? marks if Ctrl menu is called
                if unsure and item != "Don't Know" and item != "Other":
                    item = item+'?'
                # Transform > marks
                pos = item.find('>')
                if pos > -1:
                    item = item[:pos] + ' (' + item[pos+1:] + ')'

                bird = self.menuBird2.addAction(item)
                bird.setCheckable(True)
                if hasattr(self,'segments') and item in self.segments[self.box1id][4]:
                    bird.setChecked(True)
                self.menuBird2.addAction(bird)

            self.fullbirdlist = self.makeFullBirdList()  # a QComboBox
            self.showFullbirdlist = QWidgetAction(self.menuBirdList)
            self.showFullbirdlist.setDefaultWidget(self.fullbirdlist)
            self.menuBird2.addAction(self.showFullbirdlist)
            self.fullbirdlist.activated.connect(self.birdSelectedList)

    def fillFileList(self,fileName):
        """ Generates the list of files for the file listbox.
        fileName - currently opened file (marks it in the list).
        Most of the work is to deal with directories in that list.
        It only sees *.wav files. Picks up *.data to make the filenames
        red in the list."""

        if not os.path.isdir(self.SoundFileDir):
            print("ERROR: directory %s doesn't exist" % self.soundFileDir)
            return

        # clear file listbox
        self.listFiles.clearSelection()
        self.listFiles.clearFocus()
        self.listFiles.clear()

        self.listOfFiles = QDir(self.SoundFileDir).entryInfoList(['..','*.wav'],filters=QDir.AllDirs|QDir.NoDot|QDir.Files,sort=QDir.DirsFirst)
        listOfDataFiles = QDir(self.SoundFileDir).entryList(['*.data'])
        for file in self.listOfFiles:
            # If there is a .data version, colour the name red to show it has been labelled
            item = QListWidgetItem(self.listFiles)
            self.listitemtype = type(item)
            if file.isDir():
                item.setText(file.fileName() + "/")
            else:
                item.setText(file.fileName())
            if file.fileName()+'.data' in listOfDataFiles:
                item.setForeground(Qt.red)
        # mark the current file
        if fileName:
            index = self.listFiles.findItems(fileName+"\/?",Qt.MatchRegExp)
            if len(index)>0:
                self.listFiles.setCurrentItem(index[0])
            else:
                self.listFiles.setCurrentRow(0)

    def resetStorageArrays(self):
        """ Called when new files are loaded.
        Resets the variables that hold the data to be saved and/or plotted. 
        """

        # Remove the segments
        self.removeSegments()

        # Check if media is playing and stop it if so
        if hasattr(self,'media_obj'):
            if self.media_obj.isPlaying():
                self.stopPlayback()

        # This is a flag to say if the next thing that the user clicks on should be a start or a stop for segmentation
        if self.started:
            # This is the second click, so should pay attention and close the segment
            # Stop the mouse motion connection, remove the drawing boxes
            if self.started_window=='a':
                try:
                    self.p_ampl.scene().sigMouseMoved.disconnect()
                except Exception:
                    pass
                self.p_ampl.removeItem(self.vLine_a)
            else:
                try:
                    self.p_spec.scene().sigMouseMoved.disconnect()
                except Exception:
                    pass
                # Add the other mouse move listener back
                if self.showPointerDetails.isChecked():
                    self.p_spec.scene().sigMouseMoved.connect(self.mouseMoved)
                self.p_spec.removeItem(self.vLine_s)
            self.p_ampl.removeItem(self.drawingBox_ampl)
            self.p_spec.removeItem(self.drawingBox_spec)
        self.started = False
        self.startedInAmpl = False
        self.segmentsToSave = False

        # Keep track of start points and selected buttons
        self.windowStart = 0
        self.playPosition = self.windowStart
        self.prevBoxCol = self.config['ColourNone']
        self.bar.setValue(0)

        # Reset the MultipleSpecies option
        self.multipleBirds = self.config['MultipleSpecies']

        # reset playback buttons
        self.playSegButton.setEnabled(False)
        self.playBandLimitedSegButton.setEnabled(False)
        self.quickDenButton.setEnabled(False)

        # Remove any fundamental frequencies drawn
        for r in self.segmentPlots:
            self.p_spec.removeItem(r)
        self.segmentPlots=[]

        # Remove spectral derivatives
        try:
            self.p_spec.removeItem(self.derivPlot)
        except Exception:
            pass
        # Cheatsheet: remove the freq labels
        if self.zooniverse and hasattr(self,'label1'):
            self.p_spec.removeItem(self.label1)
            self.p_spec.removeItem(self.label2)
            self.p_spec.removeItem(self.label3)
            self.p_spec.removeItem(self.label4)
            self.p_spec.removeItem(self.label5)

    def openFile(self):
        """ This handles the menu item for opening a file.
        Splits the directory name and filename out, and then passes the filename to the loader."""
        fileName, drop = QtGui.QFileDialog.getOpenFileName(self, 'Choose File', self.SoundFileDir,"Wav files (*.wav)")
        success = 1
        SoundFileDirOld = self.SoundFileDir
        fileNameOld = os.path.basename(self.filename)
        if fileName != '':
            print("Opening file %s" % fileName)
            self.SoundFileDir = os.path.dirname(fileName)
            success = self.listLoadFile(os.path.basename(fileName))
        if success==1:
            print("Error loading file, reloading current file")
            self.SoundFileDir = SoundFileDirOld
            self.filename = fileNameOld
            self.listLoadFile(fileNameOld)


    def listLoadFile(self,current):
        """ Listener for when the user clicks on a filename (also called by openFile() )
        Prepares the program for a new file.
        Saves the segments of the current file, resets flags and calls loadFile() """

        # Need name of file
        if type(current) is self.listitemtype:
            current = current.text()
            current = re.sub('\/.*', '', current)

        fullcurrent = os.path.join(self.SoundFileDir, current)
        if not os.path.isdir(fullcurrent):
            if not os.path.isfile(fullcurrent):
                print("File %s does not exist!" % fullcurrent)
                return(1)
            # avoid files with no data (Tier 1 has 0Kb .wavs
            if os.stat(fullcurrent).st_size == 0:
                print("Cannot open file %s of size 0!" % fullcurrent)
                return(1)
            elif os.stat(fullcurrent).st_size < 100:
                print("File %s appears to have only header" % fullcurrent)
                return(1)

        # If there was a previous file, make sure the type of its name is OK. This is because you can get these
        # names from the file listwidget, or from the openFile dialog.
        # - is this needed at all??
        if self.previousFile is not None:
            if type(self.previousFile) is not self.listitemtype:
                self.previousFile = self.listFiles.findItems(os.path.basename(str(self.previousFile)), Qt.MatchExactly)
                if len(self.previousFile)>0:
                    self.previousFile = self.previousFile[0]

            # Check if user requires noise data
            if self.config['RequireNoiseData'] and self.noiseLevel is None:
                self.addNoiseData()

            # setting this to True forces initial save
            # self.segmentsToSave = True
            self.saveSegments()

        self.previousFile = current
        #if type(current) is self.listitemtype:
        #    current = current.text()

        # Update the file list to show the right one
        i=0
        while i<len(self.listOfFiles)-1 and self.listOfFiles[i].fileName() != current:
            i+=1
        if self.listOfFiles[i].isDir() or (i == len(self.listOfFiles)-1 and self.listOfFiles[i].fileName() != current):
            dir = QDir(self.SoundFileDir)
            dir.cd(self.listOfFiles[i].fileName())
            # Now repopulate the listbox
            self.SoundFileDir=str(dir.absolutePath())
            self.previousFile = None
            if (i == len(self.listOfFiles)-1) and (self.listOfFiles[i].fileName() != current):
                self.loadFile(current)
            self.fillFileList(current)
        else:
            self.loadFile(current)
        return(0)

    def loadFile(self,name=None):
        """ This does the work of loading a file.
        We are using wavio to do the reading. We turn the data into a float, but do not normalise it (/2^(15)).
        For 2 channels, just take the first one.
        Normalisation can cause problems for some segmentations, e.g. Harma.

        If no name is specified, loads the next section of the current file

        This method also gets the spectrogram to plot it, loads the segments from a *.data file, and
        passes the new data to any of the other classes that need it.
        Then it sets up the audio player and fills in the appropriate time data in the window, and makes
        the scroll bar and overview the appropriate lengths.
        """
        self.resetStorageArrays()

        with pg.ProgressDialog("Loading..", 0, 7) as dlg:
            dlg.setCancelButton(None)
            dlg.setWindowIcon(QIcon('img/Avianz.ico'))
            dlg.setWindowTitle('AviaNZ')
            if name is not None:
                if not self.cheatsheet:
                    self.filename = self.SoundFileDir+'/'+name
                else:
                    self.filename = name
                dlg += 1

                # Create an instance of the Signal Processing class
                if not hasattr(self, 'sp'):
                    if self.cheatsheet:
                        self.sp = SignalProc.SignalProc([],0,512,256)
                    else:
                        self.sp = SignalProc.SignalProc([], 0, self.config['window_width'], self.config['incr'])

                self.currentFileSection = 0

                if hasattr(self, 'timeaxis') and not self.zooniverse:
                    self.w_spec.removeItem(self.timeaxis)

                # Check if the filename is in standard DOC format
                # Which is xxxxxx_xxxxxx.wav or ccxx_cccc_xxxxxx_xxxxxx.wav (c=char, x=0-9), could have _ afterward
                # So this checks for the 6 ints _ 6 ints part anywhere in string
                DOCRecording = re.search('(\d{6})_(\d{6})',name[-17:-4])

                if DOCRecording:
                    self.startTime = DOCRecording.group(2)

                    #if int(self.startTime[:2]) > 8 or int(self.startTime[:2]) < 8:
                    if int(self.startTime[:2]) > 17 or int(self.startTime[:2]) < 7: # 6pm to 6am
                        print("Night time DOC recording")
                    else:
                        print("Day time DOC recording")
                        # TODO: And modify the order of the bird list
                    self.startTime = int(self.startTime[:2]) * 3600 + int(self.startTime[2:4]) * 60 + int(self.startTime[4:6])
                    self.timeaxis = SupportClasses.TimeAxisHour(orientation='bottom',linkView=self.p_ampl)
                else:
                    self.startTime = 0
                    if self.cheatsheet:
                        self.timeaxis = SupportClasses.TimeAxisSec(orientation='bottom',linkView=self.p_ampl)
                    else:
                        self.timeaxis = SupportClasses.TimeAxisMin(orientation='bottom',linkView=self.p_ampl)

                if not self.zooniverse:
                    self.w_spec.addItem(self.timeaxis, row=1, col=1)

                # This next line is a hack to make the axis update
                #self.changeWidth(self.widthWindow.value())

                dlg += 1
            else:
                dlg += 2

            # Read in the file and make the spectrogram
            self.startRead = max(0,self.currentFileSection*self.config['maxFileShow']-self.config['fileOverlap'])
            if self.startRead == 0:
                self.lenRead = self.config['maxFileShow']+self.config['fileOverlap']
            else:
                self.lenRead = self.config['maxFileShow'] + 2*self.config['fileOverlap']

            if os.stat(self.filename).st_size != 0: # avoid files with no data (Tier 1 has 0Kb .wavs)
                wavobj = wavio.read(self.filename,self.lenRead,self.startRead)

                # Parse wav format details based on file header:
                self.sampleRate = wavobj.rate
                self.audiodata = wavobj.data
                self.minFreq = 0
                self.maxFreq = self.sampleRate / 2.
                self.fileLength = wavobj.nframes
                self.audioFormat.setChannelCount(np.shape(self.audiodata)[1])
                self.audioFormat.setSampleRate(self.sampleRate)
                self.audioFormat.setSampleSize(wavobj.sampwidth*8)
                print("Detected format: %d channels, %d Hz, %d bit samples" % (self.audioFormat.channelCount(), self.audioFormat.sampleRate(), self.audioFormat.sampleSize()))

                self.minFreqShow = max(self.minFreq, self.config['minFreq'])
                self.maxFreqShow = min(self.maxFreq, self.config['maxFreq'])

                dlg += 1

                if self.audiodata.dtype is not 'float':
                    self.audiodata = self.audiodata.astype('float')  # / 32768.0

                if np.shape(np.shape(self.audiodata))[0] > 1:
                    self.audiodata = self.audiodata[:, 0]
                self.datalength = np.shape(self.audiodata)[0]
                self.datalengthSec = self.datalength / self.sampleRate
                print("Length of file is ", self.datalengthSec, " seconds (", self.datalength, "samples) loaded from ", self.fileLength / self.sampleRate, "seconds (", self.fileLength, " samples) with sample rate ",self.sampleRate, " Hz.")

                if name is not None: # i.e. starting a new file, not next section
                    if self.datalength != self.fileLength:
                        #print("not all of file loaded")
                        self.nFileSections = int(np.ceil(self.fileLength/self.datalength))
                        self.prev5mins.setEnabled(False)
                        self.next5mins.setEnabled(True)
                        self.movePrev5minsKey.setEnabled(False)
                        self.moveNext5minsKey.setEnabled(True)
                    else:
                        self.nFileSections = 1
                        self.prev5mins.setEnabled(False)
                        self.next5mins.setEnabled(False)
                        self.movePrev5minsKey.setEnabled(False)
                        self.moveNext5minsKey.setEnabled(False)

                if self.nFileSections == 1:
                    self.placeInFileLabel.setText('')
                else:
                    self.placeInFileLabel.setText("Page "+ str(self.currentFileSection+1) + " of " + str(self.nFileSections))

                # Get the data for the main spectrogram
                sgRaw = self.sp.spectrogram(self.audiodata, self.config['window_width'],
                                            self.config['incr'], mean_normalise=self.sgMeanNormalise, equal_loudness=self.sgEqualLoudness, onesided=self.sgOneSided, multitaper=self.sgMultitaper)
                maxsg = np.min(sgRaw)
                self.sg = np.abs(np.where(sgRaw == 0, 0.0, 10.0 * np.log10(sgRaw / maxsg)))

                # Load any previous segments stored
                if os.path.isfile(self.filename + '.data'):
                    file = open(self.filename + '.data', 'r')
                    self.segments = json.load(file)
                    file.close()
                    if len(self.segments) > 0:
                        if self.segments[0][0] == -1:
                            self.operator = self.segments[0][2]
                            self.reviewer = self.segments[0][3]
                            # This could probably do with more error checking
                            if type(self.segments[0][4]) is int:
                                self.noiseLevel = None
                                self.noiseTypes = []
                            else:
                                self.noiseLevel = self.segments[0][4][0]
                                self.noiseTypes = self.segments[0][4][1]
                            del self.segments[0]
                    if len(self.segments) > 0:
                        for s in self.segments:
                            if 0 < s[2] < 1.1 and 0 < s[3] < 1.1:
                                # *** Potential for major cockups here. First version didn't normalise the segment data for dragged boxes.
                                # The second version did, storing them as values between 0 and 1. It modified the original versions by assuming that the spectrogram was 128 pixels high (256 width window).
                                # This version does what it should have done in the first place, which is to record actual frequencies
                                # The .1 is to take care of rounding errors
                                # TODO: Because of this change (23/8/18) I run a backup on the datafiles in the init
                                s[2] = self.convertYtoFreq(s[2])
                                s[3] = self.convertYtoFreq(s[3])
                                self.segmentsToSave = True
                            if type(s[4]) is not list:
                                s[4] = [s[4]]
                            # Check if there are any multiple birds in the list, and switch the option on regardless of user preference
                            if len(s[4])>1:
                                self.multipleBirds = True

                if self.Hartley and not os.path.isfile(self.filename + '.data'):
                    self.addRegularSegments()

                self.statusRight.setText("Operator: " + str(self.operator) + ", Reviewer: " + str(self.reviewer))

                # Update the data that is seen by the other classes

                self.sp.setNewData(self.audiodata,self.sampleRate)

                if hasattr(self,'seg'):
                    self.seg.setNewData(self.audiodata,sgRaw,self.sampleRate,self.config['window_width'],self.config['incr'])
                else:
                    self.seg = Segment.Segment(self.audiodata, sgRaw, self.sp, self.sampleRate,
                                               self.config['window_width'], self.config['incr'])
                self.sp.setNewData(self.audiodata,self.sampleRate)

                # Update the Dialogs
                if hasattr(self,'spectrogramDialog'):
                    self.spectrogramDialog.setValues(self.minFreq,self.maxFreq,self.minFreqShow,self.maxFreqShow)
                if hasattr(self,'denoiseDialog'):
                    #print(self.denoiseDialog)
                    self.denoiseDialog.setValues(self.minFreq,self.maxFreq)

                # Delete any denoising backups from the previous file
                if hasattr(self,'audiodata_backup'):
                    self.audiodata_backup = None
                if not self.Hartley:
                    self.showFundamental.setChecked(False)
                    self.showSpectral.setChecked(False)
                if not self.DOC and not self.Hartley:
                    self.showInvSpec.setChecked(False)

                self.timeaxis.setOffset(self.startRead+self.startTime)

                # Set the window size
                self.windowSize = self.config['windowWidth']
                self.timeaxis.setRange(0, self.windowSize)
                self.widthWindow.setRange(0.5, self.datalengthSec)
    
                # Reset it if the file is shorter than the window
                if self.datalengthSec < self.windowSize:
                    self.windowSize = self.datalengthSec
                self.widthWindow.setValue(self.windowSize)
    
                self.totalTime = self.convertMillisecs(1000*self.datalengthSec)
                self.timePlayed.setText(self.convertMillisecs(0) + "/" + self.totalTime)

                # Load the file for playback
                self.media_obj = SupportClasses.ControllableAudio(self.audioFormat)
                # this responds to audio output timer
                self.media_obj.notify.connect(self.movePlaySlider)

                # Reset the media player
                self.stopPlayback()
                self.volSliderMoved(0)
                self.segmentStop = 50
                self.media_obj.filterSeg(0, 50, self.audiodata)
                self.volSliderMoved(self.volSlider.value())

                # Set the length of the scrollbar.
                self.scrollSlider.setRange(0,np.shape(self.sg)[0] - self.convertAmpltoSpec(self.widthWindow.value()))
                self.scrollSlider.setValue(0)
    
                # Get the height of the amplitude for plotting the box
                self.minampl = np.min(self.audiodata)+0.1*(np.max(self.audiodata)+np.abs(np.min(self.audiodata)))
                self.drawOverview()
                dlg += 1
                self.drawfigMain()
                self.setWindowTitle('AviaNZ - ' + self.filename)
                dlg += 1
                self.statusLeft.setText("Ready")

    def openNextFile(self):
        """ Listener for next file >> button.
        Get the next file in the list and call the loader. """
        i=self.listFiles.currentRow()
        if i+1<len(self.listFiles):
            self.listFiles.setCurrentRow(i+1)
            self.listLoadFile(self.listFiles.currentItem())
        else:
            # Tell the user they've finished
            msg = SupportClasses.MessagePopup("d", "Last file", "You've finished processing the folder")
            msg.exec_()

    def showPointerDetailsCheck(self):
        """ Listener for the menuitem that sets if detailed info should be shown when hovering over spectrogram.
        Turning this off saves lots of CPU performance."""
        self.config['showPointerDetails'] = self.showPointerDetails.isChecked()
        if self.showPointerDetails.isChecked():
            self.p_spec.scene().sigMouseMoved.connect(self.mouseMoved)
            self.p_spec.addItem(self.pointData)
        else:
            self.p_spec.scene().sigMouseMoved.disconnect()
            self.p_spec.removeItem(self.pointData)
            #self.pointData.setText("")

    def dragRectsTransparent(self):
        """ Listener for the check menu item that decides if the user wants the dragged rectangles to have colour or not.
        It's a switch from Brush to Pen or vice versa.
        """
        if self.config['transparentBoxes']:
            for box in self.listRectanglesa2:
                if type(box) == self.ROItype:
                    col = box.brush.color()
                    col.setAlpha(255)
                    box.transparent = True
                    box.setBrush(pg.mkBrush(None))
                    box.setPen(pg.mkPen(col,width=1))
                    box.update()
                    col.setAlpha(100)
        else:
            for box in self.listRectanglesa2:
                if type(box) == self.ROItype:
                    col = box.pen.color()
                    col.setAlpha(self.ColourNamed.alpha())
                    box.transparent = False
                    box.setBrush(pg.mkBrush(col))
                    box.setPen(pg.mkPen(None))
                    box.update()
                    col.setAlpha(100)

    def useAmplitudeCheck(self):
        """ Listener for the check menu item saying if the user wants to see the waveform.
        Does not remove the dock, just hides it. It's therefore easy to replace, but could have some performance overhead.
        """
        if self.useAmplitudeTick.isChecked():
            self.useAmplitude = True
            self.d_ampl.show()
        else:
            self.useAmplitude = False
            self.d_ampl.hide()
        self.config['showAmplitudePlot'] = self.useAmplitudeTick.isChecked()

    def useFilesCheck(self):
        """ Listener to process if the user swaps the check menu item to see the file list. """
        if self.useFilesTick.isChecked():
            self.d_files.show()
        else:
            self.d_files.hide()
        self.config['showListofFiles'] = self.useFilesTick.isChecked()

    def showOverviewSegsCheck(self):
        """ Listener to process if the user swaps the check menu item to see the overview segment boxes. """
        if self.showOverviewSegsTick.isChecked():
            self.p_overview2.show()
        else:
            self.p_overview2.hide()
        self.config['showAnnotationOverview'] = self.showOverviewSegsTick.isChecked()

    def makeReadOnly(self):
        """ Listener to process the check menu item to make the plots read only.
        Turns off the listeners for the amplitude and spectrogram plots.
        Also has to go through all of the segments, turn off the listeners, and make them unmovable.
        """
        self.config['readOnly'] = self.readonly.isChecked()
        if self.readonly.isChecked():
            # this is for accepting drag boxes or not
            self.p_spec.enableDrag = False

            # when clicking is used to draw segments/boxes,
            # read-only changes are implemented in the button signals.
            # Because connecting-disconnecting slots is very dirty.

            # this will re-make segment boxes with correct moving abilities:
            if hasattr(self, 'audiodata'):
                self.removeSegments(delete=False)
                self.drawfigMain(remaking=True)
                self.statusLeft.setText("Ready. Read-only mode!")
        else:
            self.p_spec.enableDrag = self.config['specMouseAction']==3
            if hasattr(self, 'audiodata'):
                self.removeSegments(delete=False)
                self.drawfigMain(remaking=True)
                self.statusLeft.setText("Ready")

    def dockReplace(self):
        """ Listener for if the docks should be replaced menu item.
            A rewrite of pyqtgraph.dockarea.restoreState.
            """
        containers, docks = self.area.findAll()
        # main recursion of restoreState:
        self.area.buildFromState(self.state['main'], docks, self.area, missing='error')
        # restoreState doesn't restore non-floating window sizes smh
        self.d_plot.hide()
        containers, docks = self.area.findAll()
        # basically say that left panel and controls should be as small as possible:
        self.d_controls.setSizePolicy(1,1)
        containers[1].setSizePolicy(1,1)
        # for cont in range(len(containers)):
        #     containers[cont].setSizes(self.state_cont[cont])


    def showFundamentalFreq(self):
        """ Calls the SignalProc class to compute, and then draws the fundamental frequency.
        Uses the yin algorithm. """
        with pg.BusyCursor():
            if self.showFundamental.isChecked():
                self.statusLeft.setText("Drawing fundamental frequency...")
                pitch, y, minfreq, W = self.seg.yin()
                #print("pitch: ", np.min(pitch), np.max(pitch))
                ind = np.squeeze(np.where(pitch>minfreq))
                pitch = pitch[ind]
                ind = ind*W/(self.config['window_width'])
                x = (pitch*2/self.sampleRate*np.shape(self.sg)[1]).astype('int')

                x = medfilt(x, 15)

                # Get the individual pieces
                segs = self.seg.identifySegments(ind, maxgap=10, minlength=5)
                count = 0
                self.segmentPlots = []
                for s in segs:
                    count += 1
                    s[0] = s[0] * self.sampleRate / self.config['incr']
                    s[1] = s[1] * self.sampleRate / self.config['incr']
                    i = np.where((ind>s[0]) & (ind<s[1]))
                    self.segmentPlots.append(pg.PlotDataItem())
                    self.segmentPlots[-1].setData(ind[i], x[i], pen=pg.mkPen('r', width=3))
                    self.p_spec.addItem(self.segmentPlots[-1])
            else:
                self.statusLeft.setText("Removing fundamental frequency...")
                for r in self.segmentPlots:
                    self.p_spec.removeItem(r)
            self.statusLeft.setText("Ready")

    def showSpectralDeriv(self):
        with pg.BusyCursor():
            if self.showSpectral.isChecked():
                self.statusLeft.setText("Drawing spectral derivative...")
                sd = self.sp.spectral_derivative(self.audiodata, self.sampleRate, self.config['window_width'], self.config['incr'], 2, 5.0)

                self.derivPlot = pg.ScatterPlotItem() 
                x, y = np.where(sd > 0)
                self.derivPlot.setData(x, y, pen=pg.mkPen('b', width=5))

                self.p_spec.addItem(self.derivPlot)
            else:
                self.statusLeft.setText("Removing spectral derivative...")
                self.p_spec.removeItem(self.derivPlot)
            self.statusLeft.setText("Ready")

    def showInvertedSpectrogram(self):
        """ Listener for the menu item that draws the spectrogram of the waveform of the inverted spectrogram."""
        with pg.BusyCursor():
            if self.showInvSpec.isChecked():
                sgRaw = self.sp.show_invS()
            else:
                sgRaw = self.sp.spectrogram(self.audiodata, mean_normalise=self.sgMeanNormalise, equal_loudness=self.sgEqualLoudness, onesided=self.sgOneSided, multitaper=self.sgMultitaper)
            maxsg = np.min(sgRaw)
            self.sg = np.abs(np.where(sgRaw == 0, 0.0, 10.0 * np.log10(sgRaw / maxsg)))
            self.overviewImage.setImage(self.sg)
            self.specPlot.setImage(self.sg)

    def medianFilterSpec(self):
        """ Median filter the spectrogram. To be used in conjunction with spectrogram inversion. """
        # TODO: Play with this
        with pg.BusyCursor():
            self.statusLeft.setText("Filtering...")
            median_filter(self.sg,size=(100,20))
            self.specPlot.setImage(self.sg)
            self.statusLeft.setText("Ready")

    def denoiseImage(self):
        # TODO
        """ Denoise the spectrogram. To be used in conjunction with spectrogram inversion. """
        #from cv2 import fastNlMeansDenoising
        #sg = np.array(self.sg/np.max(self.sg)*255,dtype = np.uint8)
        #sg = fastNlMeansDenoising(sg,10,7,21)
        #self.specPlot.setImage(sg)
# ==============
# Code for drawing and using the main figure

    def convertAmpltoSpec(self,x):
        """ Unit conversion """
        return x*self.sampleRate/self.config['incr']

    def convertSpectoAmpl(self,x):
        """ Unit conversion """
        return x*self.config['incr']/self.sampleRate

    def convertMillisecs(self,millisecs):
        """ Unit conversion """
        seconds = (millisecs / 1000) % 60
        minutes = (millisecs / (1000 * 60)) % 60
        return "%02d" % minutes+":"+"%02d" % seconds

    def convertYtoFreq(self,y,sgy=None):
        """ Unit conversion """
        if sgy is None:
            sgy = np.shape(self.sg)[1]
        return y * self.sampleRate//2 / sgy + self.minFreqShow

    def convertFreqtoY(self,f,sgy=None):
        """ Unit conversion """
        if sgy is None:
            sgy = np.shape(self.sg)[1]
        return (f-self.minFreqShow) * sgy / (self.sampleRate//2)

    def drawOverview(self):
        """ On loading a new file, update the overview figure to show where you are up to in the file.
        Also, compute the new segments for the overview, and make sure that the listeners are connected
        for clicks on them. """
        self.overviewImage.setImage(self.sg)
        #self.overviewImageRegion = pg.LinearRegionItem()
        # this is needed for compatibility with other shaded rectangles:
        #self.overviewImageRegion.lines[0].btn = QtCore.Qt.RightButton
        #self.overviewImageRegion.lines[1].btn = QtCore.Qt.RightButton
        #self.p_overview.addItem(self.overviewImageRegion, ignoreBounds=True)
        self.overviewImageRegion.setRegion([0, self.convertAmpltoSpec(self.widthWindow.value())])
        self.overviewImageRegion.sigRegionChangeFinished.connect(self.updateOverview)

        # Three y values are No. not known, No. known, No. possible
        # widthOverviewSegment is in seconds
        numSegments = int(np.ceil(np.shape(self.sg)[0]/self.convertAmpltoSpec(self.config['widthOverviewSegment'])))
        self.widthOverviewSegment = np.shape(self.sg)[0]//numSegments

        self.overviewSegments = np.zeros((numSegments,3))

        # Delete the overview segments
        for r in self.SegmentRects:
            self.p_overview2.removeItem(r)
        self.SegmentRects = []

        # Add new overview segments
        for i in range(numSegments):
            r = SupportClasses.ClickableRectItem(i*self.widthOverviewSegment, 0, self.widthOverviewSegment, 0.5)
            r.setPen(pg.mkPen('k'))
            r.setBrush(pg.mkBrush('w'))
            self.SegmentRects.append(r)
            self.p_overview2.addItem(r)
        self.p_overview2.sigChildMessage.connect(self.overviewSegmentClicked)

    def overviewSegmentClicked(self,x):
        """ Listener for an overview segment being clicked on.
        Work out which one, and move the region appropriately. Calls updateOverview to do the work. """
        minX, maxX = self.overviewImageRegion.getRegion()
        self.overviewImageRegion.setRegion([x, x+maxX-minX])
        self.updateOverview()
        self.playPosition = int(self.convertSpectoAmpl(x)*1000.0)

    def updateOverview(self, preserveLength=True):
        """ Listener for when the overview box is changed. Also called by overviewSegmentClicked().
        Does the work of keeping all the plots in the right place as the overview moves.
        It sometimes updates a bit slowly. """
        if hasattr(self, 'media_obj'):
            if self.media_obj.state() == QAudio.ActiveState or self.media_obj.state() == QAudio.SuspendedState:
                self.stopPlayback()
        #3/4/18: Want to stop it moving past either end
        # Need to disconnect the listener and reconnect it to avoid a recursive call
        minX, maxX = self.overviewImageRegion.getRegion()
        #print minX, maxX
        if minX<0:
            l = maxX-minX
            minX=0.0
            maxX=minX+l
            try:
                self.overviewImageRegion.sigRegionChangeFinished.disconnect()
            except:
                pass
            self.overviewImageRegion.setRegion([minX,maxX])
            self.overviewImageRegion.sigRegionChangeFinished.connect(self.updateOverview)
        if maxX>len(self.sg):
            l = maxX-minX
            maxX=float(len(self.sg))
            minX=max(0, maxX-l)
            try:
                self.overviewImageRegion.sigRegionChangeFinished.disconnect()
            except:
                pass
            self.overviewImageRegion.setRegion([minX,maxX])
            self.overviewImageRegion.sigRegionChangeFinished.connect(self.updateOverview)

        self.widthWindow.setValue(self.convertSpectoAmpl(maxX-minX))
        self.p_ampl.setXRange(self.convertSpectoAmpl(minX), self.convertSpectoAmpl(maxX), update=True, padding=0)
        self.p_spec.setXRange(minX, maxX, update=True, padding=0)
        self.p_plot.setXRange(self.convertSpectoAmpl(minX), self.convertSpectoAmpl(maxX), update=True, padding=0)

        # I know the next two lines SHOULD be unnecessary. But they aren't!
        self.p_ampl.setXRange(self.convertSpectoAmpl(minX), self.convertSpectoAmpl(maxX), padding=0)
        self.p_spec.setXRange(minX, maxX, padding=0)

        if self.extra != "none" and "Filtered spectrogram" not in self.extra:
            self.p_plot.setXRange(self.convertSpectoAmpl(minX), self.convertSpectoAmpl(maxX), padding=0)
        if "Filtered spectrogram" in self.extra:
            self.p_plot.setXRange(minX, maxX, padding=0)
        # self.setPlaySliderLimits(1000.0*self.convertSpectoAmpl(minX),1000.0*self.convertSpectoAmpl(maxX))
        self.scrollSlider.setValue(minX)
        self.pointData.setPos(minX,0)
        self.config['windowWidth'] = self.convertSpectoAmpl(maxX-minX)
        # self.saveConfig = True
        self.timeaxis.update()
        pg.QtGui.QApplication.processEvents()

    def drawfigMain(self,remaking=False):
        """ Draws the main amplitude and spectrogram plots and any segments on them.
        Has to do some work to get the axis labels correct.
        """
        self.amplPlot.setData(np.linspace(0.0,self.datalengthSec,num=self.datalength,endpoint=True),self.audiodata)
        self.timeaxis.setLabel('')

        height = self.sampleRate // 2 / np.shape(self.sg)[1]
        pixelstart = int(self.minFreqShow/height)
        pixelend = int(self.maxFreqShow/height)

        self.overviewImage.setImage(self.sg[:,pixelstart:pixelend])
        self.specPlot.setImage(self.sg[:,pixelstart:pixelend])
        #self.specPlot.setImage(self.sg)
        self.setExtraPlot(self.extra)

        self.setColourMap(self.config['cmap'])
        self.setColourLevels()

        # Sort out the spectrogram frequency axis
        # The constants here are divided by 1000 to get kHz, and then remember the top is sampleRate/2
        FreqRange = self.maxFreqShow-self.minFreqShow
        height = self.sampleRate // 2 / np.shape(self.sg)[1]
        SpecRange = FreqRange/height
        
        if self.zooniverse:
            offset=6
            txt='<span style="color: #0F0; font-size:20pt">%s</div>'%str(0)
            self.label1 = pg.TextItem(html=txt, color='g', anchor=(0,0))
            self.p_spec.addItem(self.label1)
            self.label1.setPos(0,0+offset)

            txt='<span style="color: #0F0; font-size:20pt">%s</div>'%str(int(FreqRange//4000))
            self.label2 = pg.TextItem(html=txt, color='g', anchor=(0,0))
            self.p_spec.addItem(self.label2)
            self.label2.setPos(0,SpecRange/4+offset)

            txt='<span style="color: #0F0; font-size:20pt">%s</div>'%str(int(FreqRange//2000))
            self.label3 = pg.TextItem(html=txt, color='g', anchor=(0,0))
            self.p_spec.addItem(self.label3)
            self.label3.setPos(0,SpecRange/2+offset)

            txt='<span style="color: #0F0; font-size:20pt">%s</div>'%str(int(3*FreqRange//4000))
            self.label4 = pg.TextItem(html=txt, color='g', anchor=(0,0))
            self.p_spec.addItem(self.label4)
            self.label4.setPos(0,3*SpecRange/4+offset)

            txt='<span style="color: #0F0; font-size:20pt">%s</div>'%str(int(FreqRange//1000))
            self.label5 = pg.TextItem(html=txt, color='g', anchor=(0,0))
            self.p_spec.addItem(self.label5)
            self.label5.setPos(0,SpecRange+offset)
        else:
            self.specaxis.setTicks([[(0,round(self.minFreqShow/1000, 2)),
                                 (SpecRange/4,round(self.minFreqShow/1000+FreqRange/4000, 2)),
                                 (SpecRange/2,round(self.minFreqShow/1000+FreqRange/2000, 2)),
                                 (3*SpecRange/4,round(self.minFreqShow/1000+3*FreqRange/4000, 2)),
                                 (SpecRange,round(self.minFreqShow/1000+FreqRange/1000, 2))]])
            self.specaxis.setLabel('kHz')

        self.updateOverview()
        self.textpos = int((self.maxFreqShow-self.minFreqShow)/height) #+ self.config['textoffset']

        # If there are segments, show them
        if not self.cheatsheet and not self.zooniverse:
            for count in range(len(self.segments)):
                if self.segments[count][2] == 0 and self.segments[count][3] == 0:
                    self.addSegment(self.segments[count][0], self.segments[count][1],0,0,self.segments[count][4],False,count,remaking)
                else:
                    self.addSegment(self.segments[count][0], self.segments[count][1],self.convertFreqtoY(self.segments[count][2]),self.convertFreqtoY(self.segments[count][3]),self.segments[count][4],False,count,remaking)

            #self.drawProtocolMarks()

            # This is the moving bar for the playback
            if not hasattr(self,'bar'):
                self.bar = pg.InfiniteLine(angle=90, movable=True, pen={'color': 'c', 'width': 3})
            self.p_spec.addItem(self.bar, ignoreBounds=True)
            self.bar.sigPositionChangeFinished.connect(self.barMoved)

        QApplication.processEvents()

    def setExtraPlot(self, plotname):
        """ Reacts to menu clicks and updates or hides diagnostic plot window."""
        self.extra = plotname

        # clear plot before updating/proceeding
        self.clearDiagnostic()

        if self.extra != "none":
            self.d_plot.show()
        else:
            self.d_plot.hide()

        # plot wavelet scalogram
        if self.extra == "Wavelet scalogram":
            self.plotExtra = pg.ImageItem()
            self.p_plot.addItem(self.plotExtra)

            # passing dummy spInfo because we only use this for a function
            ws = WaveletSegment.WaveletSegment(spInfo=[], wavelet='dmey2')
            e = ws.computeWaveletEnergy(self.audiodata, self.sampleRate)
            # e is 2^nlevels x nseconds

            pos, colour, mode = colourMaps.colourMaps("Inferno")
            cmap = pg.ColorMap(pos, colour, mode)
            lut = cmap.getLookupTable(0.0, 1.0, 256)
            self.plotExtra.setLookupTable(lut)
            self.plotExtra.setImage(e.T)
            self.plotaxis.setLabel('Wavelet node')

        # plot wc correlations
        if self.extra == "Wavelet correlations":
            self.plotExtra = pg.ImageItem()
            self.p_plot.addItem(self.plotExtra)

            # preprocess
            data = librosa.core.audio.resample(self.audiodata, self.sampleRate, 16000)
            data = self.sp.ButterworthBandpass(data, self.sampleRate, 100, 16000)

            # passing dummy spInfo because we only use this for a function
            ws = WaveletSegment.WaveletSegment(spInfo=[], wavelet='dmey2')
            e = ws.computeWaveletEnergy(self.audiodata, self.sampleRate)
            annotation = np.zeros(np.shape(e)[1])
            for s in self.segments:
                annotation[math.floor(s[0]):math.ceil(s[1])] = 1
            w0 = np.where(annotation == 0)[0]
            w1 = np.where(annotation == 1)[0]

            r = np.zeros((64, np.shape(e)[1]))
            # dummy parameters b/c we're only using this for WF.graycode
            WF = WaveletFunctions.WaveletFunctions(data=[0], wavelet='dmey2', maxLevel=5, samplerate=1)
            for count in range(62):
                # just compute_r from WaveletSegment
                corr = (np.mean(e[count, w1]) - np.mean(e[count,w0]))/np.std(e[count, :]) * np.sqrt(len(w0)*len(w1))/len(annotation)
                # map a long vector of rs to different image areas
                level = int(math.log(count+2, 2))
                node = count+2 - 2**level
                node = WF.graycode(node)
                r[node * 2**(6-level) : (node+1) * 2**(6-level), level] = corr
            r[:, 0] = np.linspace(np.min(r), np.max(r), num=64)
            # propagate along x
            for tmult in range(10, len(annotation)):
                r[:, tmult] = r[:, tmult-10]

            pos, colour, mode = colourMaps.colourMaps("Viridis")
            cmap = pg.ColorMap(pos, colour,mode)
            lut = cmap.getLookupTable(0.0, 1.0, 256)
            self.plotExtra.setLookupTable(lut)
            self.plotExtra.setImage(r.T)
            self.plotaxis.setLabel('Frequency bin')

        # plot energy in "wind" band
        if self.extra == "Wind energy":
            # compute following PostProcess.wind()
            we_mean = np.zeros(int(self.datalengthSec))
            we_std = np.zeros(int(self.datalengthSec))
            for w in range(int(self.datalengthSec)):
                data = self.audiodata[int(w*self.sampleRate):int((w+1)*self.sampleRate)]
                f, p = signal.welch(data, fs=self.sampleRate, window='hamming', nperseg=512, detrend=False)
                p = np.log10(p)
                limsup = int(p.__len__() * 2 * 500 / self.sampleRate)
                liminf = int(p.__len__() * 2 * 50 / self.sampleRate)
                a_wind = p[liminf:limsup]
                we_mean[w] = np.mean(a_wind)
                we_std[w] = np.std(a_wind)

            self.plotExtra = pg.PlotDataItem(np.arange(self.datalengthSec), we_mean)
            self.plotExtra.setPen(fn.mkPen(color='k', width=2))
            self.plotExtra2 = pg.PlotDataItem(np.arange(self.datalengthSec), we_mean+we_std)
            self.plotExtra2.setPen(fn.mkPen('c'))
            self.plotExtra3 = pg.PlotDataItem(np.arange(self.datalengthSec), we_mean-we_std)
            self.plotExtra3.setPen(fn.mkPen('c'))
            self.plotExtra4 = pg.PlotDataItem(np.arange(self.datalengthSec), np.ones((int(self.datalengthSec)))*2.5)
            self.plotExtra4.setPen(fn.mkPen('r'))
            self.plotaxis.setLabel('Mean (SD) power, V^2/Hz')
            self.p_plot.addItem(self.plotExtra)
            self.p_plot.addItem(self.plotExtra2)
            self.p_plot.addItem(self.plotExtra3)
            self.p_plot.addItem(self.plotExtra4)

        # plot energy in "rain"
        if self.extra == "Rain":
            # compute following PostProcess.rain()
            we_mean = np.zeros(int(self.datalengthSec))
            we_std = np.zeros(int(self.datalengthSec))
            for w in range(int(self.datalength/self.sampleRate)):
                data = self.audiodata[int(w*self.sampleRate):int((w+1)*self.sampleRate)]
                mfcc = librosa.feature.mfcc(data, self.sampleRate)
                # mfcc_delta = librosa.feature.delta(mfcc, width=3)
                # Normalise
                mfcc -= np.mean(mfcc, axis=0)
                mfcc /= np.max(np.abs(mfcc), axis=0)
                we_mean[w] = np.mean(mfcc[1, :])
                we_std[w] = np.std(mfcc[1, :])
            self.plotExtra = pg.PlotDataItem(np.arange(self.datalengthSec), we_mean)
            self.plotExtra.setPen(fn.mkPen(color='k', width=2))
            self.plotExtra2 = pg.PlotDataItem(np.arange(self.datalengthSec), we_mean + we_std)
            self.plotExtra2.setPen(fn.mkPen('r'))
            self.plotExtra3 = pg.PlotDataItem(np.arange(self.datalengthSec), we_mean - we_std)
            self.plotExtra3.setPen(fn.mkPen('r'))
            self.plotaxis.setLabel('Mean (MFCC)')
            self.p_plot.addItem(self.plotExtra)
            self.p_plot.addItem(self.plotExtra2)
            self.p_plot.addItem(self.plotExtra3)

        # plot spectrogram of only the filtered band:
        if self.extra == "Filtered spectrogram, new + AA" or self.extra == "Filtered spectrogram, new" or self.extra == "Filtered spectrogram, old":
            self.plotExtra = pg.ImageItem()
            self.p_plot.addItem(self.plotExtra)

            try:
                plotNodes = json.load(open(os.path.join(self.filtersDir, 'plotnodes.txt')))
            except:
                print("Couldn't load file, using default list")
                plotNodes = [35, 36, 8, 41, 43, 45]

            # resample
            if self.sampleRate != 16000:
                audiodata = librosa.core.audio.resample(self.audiodata, self.sampleRate, 16000)
            else:
                audiodata = self.audiodata

            WF = WaveletFunctions.WaveletFunctions(data=audiodata, wavelet='dmey2', maxLevel=5, samplerate=16000)

            # For now, not using antialiasFilter in the reconstructions as it's quick anyway
            if self.extra == "Filtered spectrogram, new + AA":
                WF.WaveletPacket(5, 'symmetric', True, True)
                C = WF.reconstructWP2(plotNodes[0], True, False)[:len(self.audiodata)]
                for node in plotNodes[1:]:
                    C = C + WF.reconstructWP2(node, True, False)[:len(C)]
            if self.extra == "Filtered spectrogram, new":
                WF.WaveletPacket(5, 'symmetric', False)
                C = WF.reconstructWP2(plotNodes[0], True, False)[:len(self.audiodata)]
                for node in plotNodes[1:]:
                    C = C + WF.reconstructWP2(node, True, False)[:len(C)]
            if self.extra == "Filtered spectrogram, old":
                WF.WaveletPacket(5, 'symmetric', False)
                C = WF.reconstructWP2(plotNodes[0], False)[:len(self.audiodata)]
                for node in plotNodes[1:]:
                    C = C + WF.reconstructWP2(node, False)[:len(C)]

            # reconstructed signal was @ 16 kHz,
            # so we upsample to get equal sized spectrograms
            if self.sampleRate != 16000:
                C = librosa.core.audio.resample(C, 16000, self.sampleRate)
            sgRaw = self.sp.spectrogram(C)
            sgHeightReduction = np.shape(sgRaw)[1]*16000//self.sampleRate
            sgRaw = sgRaw[:, :sgHeightReduction]
            maxsg = np.min(sgRaw)
            tempsp = np.abs(np.where(sgRaw == 0, 0.0, 10.0 * np.log10(sgRaw / maxsg)))

            pos, colour, mode = colourMaps.colourMaps("Inferno")
            cmap = pg.ColorMap(pos, colour,mode)
            lut = cmap.getLookupTable(0.0, 1.0, 256)
            self.plotExtra.setLookupTable(lut)
            self.plotExtra.setImage(tempsp)

            # set axis. Always at 0:sampleRate//2
            #minX, maxX = self.overviewImageRegion.getRegion()
            #self.p_plot.setXRange(minX, maxX, padding=0)
            MaxFreq = 8000
            height = 16000 // 2 / np.shape(tempsp)[1]
            SpecRange = MaxFreq/height
            self.plotaxis.setTicks([[(0, 0.0),
                                 (SpecRange/4,round(MaxFreq/4000, 2)),
                                 (SpecRange/2,round(MaxFreq/2000, 2)),
                                 (3*SpecRange/4,round(3*MaxFreq/4000, 2)),
                                 (SpecRange,round(MaxFreq/1000, 2))]])
            self.plotaxis.setLabel('kHz')


            # pproc = SupportClasses.postProcess(self.audiodata,self.sampleRate)
            #energy, e = pproc.detectClicks()
            #energy, e = pproc.eRatioConfd()
            #if len(clicks)>0:
            #self.plotPlot.setData(np.linspace(0.0,float(self.datalength)/self.sampleRate,num=np.shape(self.sg)[0],endpoint=True),energy)
            #self.plotPlot2.setData(np.linspace(0.0,float(self.datalength)/self.sampleRate,num=np.shape(self.sg)[0],endpoint=True),e*np.ones(np.shape(self.sg)[0]))
            #self.plotPlot2.setData(np.linspace(0.0,float(self.datalength)/self.sampleRate,num=np.shape(self.sg)[0],endpoint=True),e2)

            #ws = WaveletSegment.WaveletSegment(species='kiwi')
            #e = ws.computeWaveletEnergy(self.audiodata,self.sampleRate)

            # # Call MFCC in Features and plot some of them :)
            # ff = Features.Features(self.audiodata,self.sampleRate)
            # e = ff.get_mfcc()
            #
            # # self.plotPlot.setData(np.linspace(0.0,float(self.datalength)/self.sampleRate,num=np.shape(e)[1],endpoint=True),np.sum(e,axis=0))
            # # self.plotPlot.setPen(fn.mkPen('k'))
            # # self.plotPlot2.setData(np.linspace(0.0, float(self.datalength) / self.sampleRate, num=np.shape(e)[1], endpoint=True), e[0,:])
            # # self.plotPlot2.setPen(fn.mkPen('c'))
            # e1 = e[1,:]
            # e1 = (e[1,:]- np.mean(e[1,:]))/np.std(e[1,:])
            # self.plotPlot2.setData(np.linspace(0.0,float(self.datalength)/self.sampleRate,num=np.shape(e)[1],endpoint=True),e1)
            # self.plotPlot2.setPen(fn.mkPen('r'))
            # mean = np.mean(e1)
            # std = np.std(e1)
            # thr = mean - 2 * std
            # thr = np.ones((1, 100)) * thr
            # self.plotPlot7.setData(np.linspace(0.0,float(self.datalength)/self.sampleRate,num=100, endpoint=True), thr[0,:])
            # self.plotPlot7.setPen(fn.mkPen('c'))
            # # self.plotPlot3.setData(np.linspace(0.0,float(self.datalength)/self.sampleRate,num=np.shape(e)[1],endpoint=True),e[2,:])
            # # self.plotPlot3.setPen(fn.mkPen('c'))
            # # self.plotPlot4.setData(np.linspace(0.0,float(self.datalength)/self.sampleRate,num=np.shape(e)[1],endpoint=True),e[3,:])
            # # self.plotPlot4.setPen(fn.mkPen('r'))
            # # self.plotPlot5.setData(np.linspace(0.0,float(self.datalength)/self.sampleRate,num=np.shape(e)[1],endpoint=True),e[4,:])
            # # self.plotPlot5.setPen(fn.mkPen('g'))
            # # self.plotPlot6.setData(np.linspace(0.0,float(self.datalength)/self.sampleRate,num=np.shape(e)[1],endpoint=True),e[5,:])
            # # self.plotPlot6.setPen(fn.mkPen('g'))
            # # self.plotPlot7.setData(np.linspace(0.0,float(self.datalength)/self.sampleRate,num=np.shape(e)[1],endpoint=True),e[6,:])
            # # self.plotPlot7.setPen(fn.mkPen('g'))
            # # self.plotPlot8.setData(np.linspace(0.0,float(self.datalength)/self.sampleRate,num=np.shape(e)[1],endpoint=True),e[7,:])
            # # self.plotPlot8.setPen(fn.mkPen('g'))
            # # self.plotPlot9.setData(np.linspace(0.0,float(self.datalength)/self.sampleRate,num=np.shape(e)[1],endpoint=True),e[8,:])
            # # self.plotPlot9.setPen(fn.mkPen('g'))
            # # self.plotPlot10.setData(np.linspace(0.0,float(self.datalength)/self.sampleRate,num=np.shape(e)[1],endpoint=True),e[9,:])
            # # self.plotPlot10.setPen(fn.mkPen('g'))
            # # self.plotPlot11.setData(np.linspace(0.0,float(self.datalength)/self.sampleRate,num=np.shape(e)[1],endpoint=True),e[10,:])
            # # self.plotPlot11.setPen(fn.mkPen('c'))

            # # plot eRatio
            # post = SupportClasses.postProcess(self.audiodata, self.sampleRate, [])
            # e = post.eRatioConfd([], AviaNZ_extra=True)
            # # print np.shape(e)
            # # print e[0]
            # # print np.shape(e[0])[0]
            # self.plotPlot.setData(np.linspace(0.0,float(self.datalength)/self.sampleRate,num=np.shape(e[0])[0],endpoint=True),e[0])
            # self.plotPlot.setPen(fn.mkPen('b'))



    def updateRegion_spec(self):
        """ This is the listener for when a segment box is changed in the spectrogram.
        It updates the position of the matching box, and also the text within it.
        """
        sender = self.sender()
        i = 0
        while self.listRectanglesa2[i] != sender and i<len(self.listRectanglesa2):
            i = i+1
        if i==len(self.listRectanglesa2):
            print("Segment not found!")
        else:
            if type(sender) == self.ROItype:
                # update the box visual
                x1 = self.convertSpectoAmpl(sender.pos()[0])
                x2 = self.convertSpectoAmpl(sender.pos()[0]+sender.size()[0])
                self.segments[i][2] = self.convertYtoFreq(sender.pos()[1])#/np.shape(self.sg)[1]
                self.segments[i][3] = self.convertYtoFreq(sender.pos()[1]+sender.size()[1])#/np.shape(self.sg)[1]
                self.listLabels[i].setPos(sender.pos()[0], self.textpos)
            else:
                # update the segment visual
                x1 = self.convertSpectoAmpl(sender.getRegion()[0])
                x2 = self.convertSpectoAmpl(sender.getRegion()[1])
                self.listLabels[i].setPos(sender.getRegion()[0], self.textpos)
            # update the amplitude visual
            self.listRectanglesa1[i].setRegion([x1,x2])
            self.segmentsToSave = True

            # update the actual segment list and overview boxes
            startold = self.segments[i][0]
            endold = self.segments[i][1]
            species = self.segments[i][4]
            self.refreshOverviewWith(startold, endold, species, delete=True)

            self.segments[i][0] = x1 + self.startRead
            self.segments[i][1] = x2 + self.startRead
            self.refreshOverviewWith(self.segments[i][0], self.segments[i][1], species)

    def updateRegion_ampl(self):
        """ This is the listener for when a segment box is changed in the waveform plot.
        It updates the position of the matching box, and also the text within it.
        """
        sender = self.sender()
        i = 0
        while self.listRectanglesa1[i] != sender and i<len(self.listRectanglesa1):
            i = i+1
        if i==len(self.listRectanglesa1):
            print("Segment not found!")
        else:
            x1 = self.convertAmpltoSpec(sender.getRegion()[0])
            x2 = self.convertAmpltoSpec(sender.getRegion()[1])

            # if self.listRectanglesa2[i] is not None: - this shouldn't happen anyway
            if type(self.listRectanglesa2[i]) == self.ROItype:
                # update the box visual
                y1 = self.listRectanglesa2[i].pos().y()
                y2 = self.listRectanglesa2[i].size().y()
                self.listRectanglesa2[i].setPos(pg.Point(x1,y1))
                self.listRectanglesa2[i].setSize(pg.Point(x2-x1,y2))
            else:
                # update the segment visual
                self.listRectanglesa2[i].setRegion([x1,x2])
            self.listLabels[i].setPos(x1,self.textpos)
            self.segmentsToSave = True
            # how does this update amplitude visual??

            # update the actual segment list and overview boxes
            startold = self.segments[i][0]
            endold = self.segments[i][1]
            species = self.segments[i][4]
            self.refreshOverviewWith(startold, endold, species, delete=True)

            self.segments[i][0] = sender.getRegion()[0] + self.startRead
            self.segments[i][1] = sender.getRegion()[1] + self.startRead
            self.refreshOverviewWith(self.segments[i][0], self.segments[i][1], species)

    def addRegularSegments(self):
        """ Perform the Hartley bodge: make a file with 10s segments every minute """
        if len(self.segments) > 1:
            if self.segments[0][0] == 0 and self.segments[0][1] == 10:
                print("Not adding segments")
            else:
                i = 0
                while i < self.fileLength / self.sampleRate:
                    self.segments.append([i, i + self.config['protocolSize'], 0, 0, []])
                    i += self.config['protocolInterval']
                self.segmentsToSave = True
        else:
            i = 0
            while i < self.fileLength / self.sampleRate:
                self.segments.append([i, i + self.config['protocolSize'], 0, 0, []])
                i += self.config['protocolInterval']
            self.segmentsToSave = True

    """def drawProtocolMarks(self):
        # if check-ignore protocol is used, mark check-ignore limits.
        # Also called when the relevant parameters are changed in interface settings.

        # Clean old marks, if any
        if hasattr(self, 'protocolMarks'):
            for m in self.protocolMarks:
                self.p_spec.removeItem(m)
        self.protocolMarks = []

        if self.config['protocolOn']:
            linePen = pg.mkPen((148, 0, 211), width=5)
            lnum = 0
            linestart = 0
            # pages >1 start with an overlap zone, so need to offset marks:
            if self.currentFileSection > 0:
                linestart += self.config['fileOverlap']
            while linestart < self.datalength/self.sampleRate:
                lineend = min(self.datalength/self.sampleRate, linestart + self.config['protocolSize'])
                line = SupportClasses.FixedLineROI(((self.convertAmpltoSpec(linestart),0),
                                      (self.convertAmpltoSpec(lineend),0)), movable=False, pen=linePen)
                self.protocolMarks.append(line)
                self.p_spec.addItem(line)
                line.clearHandles()
                linestart += self.config['protocolInterval']
    """

    def refreshOverviewWith(self, startpoint, endpoint, species, delete=False):
        """Recalculates the overview box colours and refreshes their display.
        To be used when segments are added, deleted or moved."""
        # Work out which overview segment this segment is in (could be more than one)
        # min is to remove possible rounding error
        inds = int(self.convertAmpltoSpec(startpoint) / self.widthOverviewSegment)
        inde = min(int(self.convertAmpltoSpec(endpoint) / self.widthOverviewSegment),len(self.overviewSegments)-1)

        if species is None or "Don't Know" in species or type(species) is int or len(species)==0:
            brush = self.ColourNone
            if delete:
                self.overviewSegments[inds:inde+1,0] -= 1
            else:
                self.overviewSegments[inds:inde+1,0] += 1
        elif '?' in ''.join(species):
            brush = self.ColourPossible
            if delete:
                self.overviewSegments[inds:inde + 1, 2] -= 1
            else:
                self.overviewSegments[inds:inde + 1, 2] += 1
        else:
            brush = self.ColourNamed
            if delete:
                self.overviewSegments[inds:inde + 1, 1] -= 1
            else:
                self.overviewSegments[inds:inde + 1, 1] += 1

        # set the colour of these boxes in the overview
        for box in range(inds, inde + 1):
            if self.overviewSegments[box,0] > 0:
                self.SegmentRects[box].setBrush(self.ColourNone)
            elif self.overviewSegments[box,2] > 0:
                self.SegmentRects[box].setBrush(self.ColourPossible)
            elif self.overviewSegments[box,1] > 0:
                self.SegmentRects[box].setBrush(self.ColourNamed)
            else:
                # boxes w/o segments
                self.SegmentRects[box].setBrush(pg.mkBrush('w'))
            self.SegmentRects[box].update()

    def addSegment(self,startpoint,endpoint,y1=0,y2=0,species=None,saveSeg=True,index=-1,remaking=False):
        """ When a new segment is created, does the work of creating it and connecting its
        listeners. Also updates the relevant overview segment.
        startpoint, endpoint are in amplitude coordinates, while y1, y2 should be standard y coordinates (between 0 and 1)
        saveSeg means that we are drawing the saved ones. Need to check that those ones fit into
        the current window, can assume the other do, but have to save their times correctly.
        If a segment is too long for the current section, truncates it.
        """
        print("Segment added at %d-%d, %d-%d" % (startpoint, endpoint, self.convertYtoFreq(y1), self.convertYtoFreq(y2)))
        miny = self.convertFreqtoY(self.minFreqShow)
        maxy = self.convertFreqtoY(self.maxFreqShow)
        if not saveSeg:
            timeRangeStart = self.startRead
            timeRangeEnd = min(self.startRead + self.lenRead, self.fileLength / self.sampleRate)

            if (startpoint < timeRangeStart and endpoint < timeRangeStart) or (startpoint > timeRangeEnd and endpoint > timeRangeEnd):
                print("Warning: a segment was not shown")
                show = False
            else:
                startpoint = startpoint - timeRangeStart
                endpoint = endpoint - timeRangeStart
                show = True
        else:
            self.segmentsToSave = True
            show = True

        if show and ((y1 <= maxy and y2 >= miny) or (y1==0 and y2==0)):
            # This is one we want to show

            # Get the name and colour sorted
            if species is None or species==["Don't Know"] or len(species) == 0:
                if self.Hartley:
                    species = []
                else:
                    species = ["Don't Know"]
                brush = self.ColourNone
            elif "Don't Know" in species:
                brush = self.ColourNone
            elif '?' in ''.join(species):
                brush = self.ColourPossible
            else:
                brush = self.ColourNamed

            self.refreshOverviewWith(startpoint, endpoint, species)
            self.prevBoxCol = brush

            segsMovable = not (self.Hartley or self.config['readOnly'])

            # Make sure startpoint and endpoint are in the right order
            if startpoint > endpoint:
                temp = startpoint
                startpoint = endpoint
                endpoint = temp

            # Add the segment in both plots and connect up the listeners
            p_ampl_r = SupportClasses.LinearRegionItem2(self, brush=brush, movable=segsMovable)
            self.p_ampl.addItem(p_ampl_r, ignoreBounds=True)
            p_ampl_r.setRegion([startpoint, endpoint])
            p_ampl_r.sigRegionChangeFinished.connect(self.updateRegion_ampl)

            # full-height segments:
            if y1==0 and y2==0:
                p_spec_r = SupportClasses.LinearRegionItem2(self, brush=brush, movable=segsMovable)
                p_spec_r.setRegion([self.convertAmpltoSpec(startpoint), self.convertAmpltoSpec(endpoint)])
            # rectangle boxes:
            else:
                if y1 > y2:
                    temp = y1
                    y1 = y2
                    y2 = temp
                startpointS = QPointF(self.convertAmpltoSpec(startpoint),max(y1,miny))
                endpointS = QPointF(self.convertAmpltoSpec(endpoint),min(y2,maxy))
                p_spec_r = SupportClasses.ShadedRectROI(startpointS, endpointS - startpointS, movable=segsMovable, parent=self)
                if self.config['transparentBoxes']:
                    col = self.prevBoxCol.rgb()
                    col = QtGui.QColor(col)
                    col.setAlpha(255)
                    p_spec_r.transparent = True
                    p_spec_r.setBrush(None)
                    p_spec_r.setHoverBrush(None)
                    p_spec_r.setPen(pg.mkPen(col,width=1))
                    col.setAlpha(100)
                else:
                    p_spec_r.setBrush(pg.mkBrush(self.prevBoxCol))
                    col = self.prevBoxCol
                    col.setAlpha(min(col.alpha()*2, 255))
                    p_spec_r.transparent = False
                    p_spec_r.setHoverBrush(pg.mkBrush(col))
                    p_spec_r.setPen(pg.mkPen(None))
                    col.setAlpha(100)
            self.p_spec.addItem(p_spec_r, ignoreBounds=True)
            p_spec_r.sigRegionChangeFinished.connect(self.updateRegion_spec)

            # Put the text into the box
            label = pg.TextItem(text=','.join(species), color='k', anchor=(0,1))
            # label = pg.TextItem(text=species, color='k')
            self.p_spec.addItem(label)
            label.setPos(self.convertAmpltoSpec(startpoint), self.textpos)

            # Add the segments to the relevent lists
            if remaking:
                self.listRectanglesa1[index] = p_ampl_r
                self.listRectanglesa2[index] = p_spec_r
                self.listLabels[index] = label
            else:
                self.listRectanglesa1.append(p_ampl_r)
                self.listRectanglesa2.append(p_spec_r)
                self.listLabels.append(label)

            if saveSeg:
                # Add the segment to the data
                # Increment the time to be correct for the current section of the file
                if y1==0 and y2==0:
                    self.segments.append([startpoint+self.startRead, endpoint+self.startRead, 0, 0, species])
                else:
                    self.segments.append([startpoint+self.startRead, endpoint+self.startRead, self.convertYtoFreq(y1), self.convertYtoFreq(y2), species])

            # mark this as the current segment
            if index>-1:
                self.box1id = index
            else:
                self.box1id = len(self.segments) - 1
        else:
            # Add a None element into the array so that the correct boxids work
            if remaking:
                self.listRectanglesa1[index] = None
                self.listRectanglesa2[index] = None
                self.listLabels[index] = None
            else:
                self.listRectanglesa1.append(None)
                self.listRectanglesa2.append(None)
                self.listLabels.append(None)

    def selectSegment(self, boxid):
        """ Changes the segment colors and enables playback buttons."""
        #print("selected %d" % boxid)
        self.playSegButton.setEnabled(True)
        self.quickDenButton.setEnabled(True)
        # self.quickDenNButton.setEnabled(True)
        self.box1id = boxid

        brush = fn.mkBrush(self.ColourSelected)
        if self.listRectanglesa1[boxid] is not None and self.listRectanglesa2[boxid] is not None:
            self.listRectanglesa1[boxid].setBrush(brush)
            self.listRectanglesa2[boxid].setBrush(brush)
            self.listRectanglesa1[boxid].setHoverBrush(brush)
            self.listRectanglesa2[boxid].setHoverBrush(brush)

            self.listRectanglesa1[boxid].update()
            self.listRectanglesa2[boxid].update()
        
        # self.listRectanglesa2[boxid].setPen(fn.mkPen(self.ColourSelectedDark,width=1))
        # if it's a rectangle:
        if type(self.listRectanglesa2[boxid]) == self.ROItype:
            self.playBandLimitedSegButton.setEnabled(True)

    def deselectSegment(self, boxid):
        """ Restores the segment colors and disables playback buttons."""
        #print("deselected %d" % boxid)
        self.playSegButton.setEnabled(False)
        self.quickDenButton.setEnabled(False)
        self.playBandLimitedSegButton.setEnabled(False)
        self.box1id = -1

        col = self.prevBoxCol
        col.setAlpha(100)
        self.listRectanglesa1[boxid].setBrush(fn.mkBrush(col))
        self.listRectanglesa2[boxid].setBrush(fn.mkBrush(col))
        col.setAlpha(200)
        self.listRectanglesa1[boxid].setHoverBrush(fn.mkBrush(col))
        self.listRectanglesa2[boxid].setHoverBrush(fn.mkBrush(col))
        col.setAlpha(100)
        if self.config['transparentBoxes'] and type(self.listRectanglesa2[boxid]) == self.ROItype:
            col = self.prevBoxCol.rgb()
            col = QtGui.QColor(col)
            col.setAlpha(255)
            self.listRectanglesa2[boxid].setBrush(pg.mkBrush(None))
            self.listRectanglesa2[boxid].setHoverBrush(pg.mkBrush(None))
            self.listRectanglesa2[boxid].setPen(col,width=1)
            col.setAlpha(100)

        self.listRectanglesa1[boxid].update()
        self.listRectanglesa2[boxid].update()

### mouse management

    def mouseMoved(self,evt):
        """ Listener for mouse moves.
        If the user moves the mouse in the spectrogram, print the time, frequency, power for the mouse location. """
        if not self.showPointerDetails.isChecked():
            return
        elif self.p_spec.sceneBoundingRect().contains(evt):
            mousePoint = self.p_spec.mapSceneToView(evt)
            indexx = int(mousePoint.x())
            indexy = int(mousePoint.y())
            if indexx > 0 and indexx < np.shape(self.sg)[0] and indexy > 0 and indexy < np.shape(self.sg)[1]:
                time = self.convertSpectoAmpl(mousePoint.x()) + self.currentFileSection * self.config['maxFileShow'] - (self.currentFileSection>0)*self.config['fileOverlap'] + self.startTime
                seconds = time % 60
                minutes = (time//60) % 60
                hours = (time//3600) % 24
                if hours>0:
                    self.pointData.setText('time=%.2d:%.2d:%05.2f (hh:mm:ss.ms), freq=%0.1f (Hz),power=%0.1f (dB)' % (hours,minutes,seconds, mousePoint.y() * self.sampleRate//2 / np.shape(self.sg)[1] + self.minFreqShow, self.sg[indexx, indexy]))
                else:
                    self.pointData.setText('time=%.2d:%05.2f (mm:ss.ms), freq=%0.1f (Hz),power=%0.1f (dB)' % (minutes,seconds, mousePoint.y() * self.sampleRate//2 / np.shape(self.sg)[1] + self.minFreqShow, self.sg[indexx, indexy]))

    def mouseClicked_ampl(self,evt):
        """ Listener for if the user clicks on the amplitude plot.
        If there is a box selected, get its colour.
        If the user has clicked inside the scene, they could be
        (1) clicking in an already existing box -> select it
        (2) clicking anywhere else -> start a box
        (3) clicking a second time to finish a box -> create the segment
        """
        pos = evt.scenePos()

        # if any box is selected, deselect (wherever clicked)
        if self.box1id>-1:
            self.deselectSegment(self.box1id)

        # if clicked inside scene:
        if self.p_ampl.sceneBoundingRect().contains(pos):
            mousePoint = self.p_ampl.mapSceneToView(pos)

            # if this is the second click and not a box, close the segment
            if self.started:
                # can't finish boxes in ampl plot
                if self.config['specMouseAction']>1:
                    if self.startedInAmpl:
                        # started in ampl and finish in ampl,
                        # so continue as usual to draw a segment
                        pass
                    else:
                        # started in spec so ignore this bullshit
                        return

                # remove the drawing box:
                self.p_spec.removeItem(self.vLine_s)
                self.p_ampl.removeItem(self.vLine_a)
                self.p_ampl.removeItem(self.drawingBox_ampl)
                self.p_spec.removeItem(self.drawingBox_spec)
                # disconnect GrowBox listeners, leave the position listener
                self.p_ampl.scene().sigMouseMoved.disconnect()
                self.p_spec.scene().sigMouseMoved.disconnect()
                if self.showPointerDetails.isChecked():
                    self.p_spec.scene().sigMouseMoved.connect(self.mouseMoved)

                # If the user has pressed shift, copy the last species and don't use the context menu
                # If they pressed Control, add ? to the names
                modifiers = QtGui.QApplication.keyboardModifiers()
                if modifiers == QtCore.Qt.ShiftModifier:
                    self.addSegment(self.start_ampl_loc, max(mousePoint.x(),0.0),species=self.lastSpecies)
                elif modifiers == QtCore.Qt.ControlModifier:
                    self.addSegment(self.start_ampl_loc,max(mousePoint.x(),0.0))
                    # Context menu
                    self.fillBirdList(unsure=True)
                    self.menuBirdList.popup(QPoint(evt.screenPos().x(), evt.screenPos().y()))
                else:
                    self.addSegment(self.start_ampl_loc,max(mousePoint.x(),0.0))
                    # Context menu
                    self.fillBirdList()
                    self.menuBirdList.popup(QPoint(evt.screenPos().x(), evt.screenPos().y()))

                # the new segment is now selected and can be played
                self.selectSegment(self.box1id)
                self.started = not(self.started)
                self.startedInAmpl = False

            # if this is the first click:
            else:
                # if this is right click (drawing mode):
                # (or whatever you want)
                if evt.button() == self.MouseDrawingButton:
                    if self.Hartley or self.config['readOnly']:
                        return
                    # this would prevent starting boxes in ampl plot
                    # if self.config['specMouseAction']>1:
                    #    return

                    nonebrush = self.ColourNone
                    self.start_ampl_loc = mousePoint.x()

                    # spectrogram plot bar and mouse followers:
                    self.vLine_s = pg.InfiniteLine(angle=90, movable=False,pen={'color': 'r', 'width': 3})
                    self.p_spec.addItem(self.vLine_s, ignoreBounds=True)
                    self.vLine_s.setPos(self.convertAmpltoSpec(self.start_ampl_loc))

                    self.drawingBox_spec = pg.LinearRegionItem(brush=nonebrush)
                    self.p_spec.addItem(self.drawingBox_spec, ignoreBounds=True)
                    self.drawingBox_spec.setRegion([self.convertAmpltoSpec(self.start_ampl_loc), self.convertAmpltoSpec(self.start_ampl_loc)])
                    self.p_spec.scene().sigMouseMoved.connect(self.GrowBox_spec)

                    # amplitude plot bar and mouse followers:
                    self.vLine_a = pg.InfiniteLine(angle=90, movable=False,pen={'color': 'r', 'width': 3})
                    self.p_ampl.addItem(self.vLine_a, ignoreBounds=True)
                    self.vLine_a.setPos(self.start_ampl_loc)

                    self.drawingBox_ampl = pg.LinearRegionItem(brush=nonebrush)
                    self.p_ampl.addItem(self.drawingBox_ampl, ignoreBounds=True)
                    self.drawingBox_ampl.setRegion([self.start_ampl_loc, self.start_ampl_loc])
                    self.p_ampl.scene().sigMouseMoved.connect(self.GrowBox_ampl)

                    self.started = not (self.started)
                    self.startedInAmpl = True

                # if this is left click (selection mode):
                else:
                    # Check if the user has clicked in a box
                    # Note: Returns the first one it finds, i.e. the newest
                    box1id = -1
                    for count in range(len(self.listRectanglesa1)):
                        if self.listRectanglesa1[count] is not None:
                            x1, x2 = self.listRectanglesa1[count].getRegion()
                            if x1 <= mousePoint.x() and x2 >= mousePoint.x():
                                box1id = count
                                break

                    # User clicked in a segment:
                    if box1id > -1:
                        # select the segment:
                        self.prevBoxCol = self.listRectanglesa1[box1id].brush.color()
                        self.selectSegment(box1id)

                        # popup dialog
                        modifiers = QtGui.QApplication.keyboardModifiers()
                        if modifiers == QtCore.Qt.ControlModifier:
                            self.fillBirdList(unsure=True)
                        else:
                            self.fillBirdList()
                        self.menuBirdList.popup(QPoint(evt.screenPos().x(), evt.screenPos().y()))
                    else:
                        # TODO: pan the view
                        pass

    def mouseClicked_spec(self,evt):
        """ Listener for if the user clicks on the spectrogram plot.
        See the amplitude version (mouseClicked_ampl()) for details. Although much of the code is a repeat,
        it is separated for clarity.
        """
        pos = evt.scenePos()

        # if any box is selected, deselect (wherever clicked)
        if self.box1id>-1:
            self.deselectSegment(self.box1id)

        # if clicked inside scene:
        if self.p_spec.sceneBoundingRect().contains(pos):
            mousePoint = self.p_spec.mapSceneToView(pos)

            # if this is the second click, close the segment/box
            # note: can finish segment with either left or right click
            if self.started:
                if self.config['specMouseAction']>1 and self.startedInAmpl:
                    # started in ampl, and spec is used for boxes, so can't continue here
                    return

                # remove the drawing box:
                if not self.config['specMouseAction']>1:
                    self.p_spec.removeItem(self.vLine_s)
                    self.p_ampl.scene().sigMouseMoved.disconnect()
                self.p_ampl.removeItem(self.vLine_a)
                self.p_ampl.removeItem(self.drawingBox_ampl)
                self.p_spec.removeItem(self.drawingBox_spec)
                # disconnect GrowBox listeners, leave the position listener
                self.p_spec.scene().sigMouseMoved.disconnect()
                if self.showPointerDetails.isChecked():
                    self.p_spec.scene().sigMouseMoved.connect(self.mouseMoved)

                # Pass either default y coords or box limits:
                x1 = self.start_ampl_loc
                x2 = self.convertSpectoAmpl(max(mousePoint.x(), 0.0))
                # Could add this check if right edge seems dangerous:
                # endx = min(x2, np.shape(self.sg)[0]+1)
                if self.config['specMouseAction']>1:
                    y1 = self.start_spec_y
                    y2 = mousePoint.y()
                    miny = self.convertFreqtoY(self.minFreqShow)
                    maxy = self.convertFreqtoY(self.maxFreqShow)
                    y1 = min(max(miny, y1), maxy)
                    y2 = min(max(miny, y2), maxy)
                else:
                    y1 = 0
                    y2 = 0

                # SRM: When dragging, can sometimes make boxes by mistake, which is annoying.
                # To avoid, check that the box isn't too small
                #print(x1,x2,y1,y2,(x2-x1)*(y2-y1))
                if (np.abs((x2-x1)*(y2-y1)) > self.minboxsize):
                    # If the user has pressed shift, copy the last species and don't use the context menu
                    # If they pressed Control, add ? to the names
                    # note: Ctrl+Shift combo doesn't have a Qt modifier and is ignored.
                    modifiers = QtGui.QApplication.keyboardModifiers()
                    if modifiers == QtCore.Qt.ShiftModifier:
                        self.addSegment(x1, x2, y1, y2, species=self.lastSpecies)
                    elif modifiers == QtCore.Qt.ControlModifier:
                        self.addSegment(x1, x2, y1, y2)
                        # Context menu
                        self.fillBirdList(unsure=True)
                        self.menuBirdList.popup(QPoint(evt.screenPos().x(), evt.screenPos().y()))
                    else:
                        self.addSegment(x1, x2, y1, y2)
                        # Context menu
                        self.fillBirdList()
                        self.menuBirdList.popup(QPoint(evt.screenPos().x(), evt.screenPos().y()))

                    # select the new segment/box
                    self.selectSegment(self.box1id)
                self.started = not(self.started)
                self.startedInAmpl = False

            # if this is the first click:
            else:
                # if this is right click (drawing mode):
                if evt.button() == self.MouseDrawingButton:
                    if self.Hartley or self.config['readOnly']:
                        return
                    nonebrush = self.ColourNone
                    self.start_ampl_loc = self.convertSpectoAmpl(mousePoint.x())
                    self.start_spec_y = mousePoint.y()

                    # start a new box:
                    if self.config['specMouseAction']>1:
                        # spectrogram mouse follower box:
                        startpointS = QPointF(mousePoint.x(), mousePoint.y())
                        endpointS = QPointF(mousePoint.x(), mousePoint.y())

                        self.drawingBox_spec = SupportClasses.ShadedRectROI(startpointS, endpointS - startpointS, invertible=True)
                        self.drawingBox_spec.setBrush(nonebrush)
                        self.p_spec.addItem(self.drawingBox_spec, ignoreBounds=True)
                        self.p_spec.scene().sigMouseMoved.connect(self.GrowBox_spec)
                    # start a new segment:
                    else:
                        # spectrogram bar and mouse follower:
                        self.vLine_s = pg.InfiniteLine(angle=90, movable=False,pen={'color': 'r', 'width': 3})
                        self.p_spec.addItem(self.vLine_s, ignoreBounds=True)
                        self.vLine_s.setPos(mousePoint.x())

                        self.drawingBox_spec = pg.LinearRegionItem(brush=nonebrush)
                        self.p_spec.addItem(self.drawingBox_spec, ignoreBounds=True)
                        self.drawingBox_spec.setRegion([mousePoint.x(),mousePoint.x()])
                        self.p_spec.scene().sigMouseMoved.connect(self.GrowBox_spec)
                        # note - only in segment mode react to movement over ampl plot:
                        self.p_ampl.scene().sigMouseMoved.connect(self.GrowBox_ampl)

                    # for box and segment - amplitude plot bar:
                    self.vLine_a = pg.InfiniteLine(angle=90, movable=False,pen={'color': 'r', 'width': 3})
                    self.p_ampl.addItem(self.vLine_a, ignoreBounds=True)
                    self.vLine_a.setPos(self.start_ampl_loc)

                    self.drawingBox_ampl = pg.LinearRegionItem(brush=nonebrush)
                    self.p_ampl.addItem(self.drawingBox_ampl, ignoreBounds=True)
                    self.drawingBox_ampl.setRegion([self.start_ampl_loc, self.start_ampl_loc])

                    self.started = not (self.started)
                    self.startedInAmpl = False

                # if this is left click (selection mode):
                else:
                    # Check if the user has clicked in a box
                    # Note: Returns the first one it finds, i.e. the newest
                    box1id = -1
                    for count in range(len(self.listRectanglesa2)):
                        if type(self.listRectanglesa2[count]) == self.ROItype and self.listRectanglesa2[count] is not None:
                            x1 = self.listRectanglesa2[count].pos().x()
                            y1 = self.listRectanglesa2[count].pos().y()
                            x2 = x1 + self.listRectanglesa2[count].size().x()
                            y2 = y1 + self.listRectanglesa2[count].size().y()
                            if x1 <= mousePoint.x() and x2 >= mousePoint.x() and y1 <= mousePoint.y() and y2 >= mousePoint.y():
                                box1id = count
                                break
                        elif self.listRectanglesa2[count] is not None:
                            x1, x2 = self.listRectanglesa2[count].getRegion()
                            if x1 <= mousePoint.x() and x2 >= mousePoint.x():
                                box1id = count
                                break

                    # User clicked in a segment:
                    if box1id > -1:
                        # select the segment:
                        self.prevBoxCol = self.listRectanglesa1[box1id].brush.color()
                        self.selectSegment(box1id)

                        modifiers = QtGui.QApplication.keyboardModifiers()
                        if modifiers == QtCore.Qt.ControlModifier:
                            self.fillBirdList(unsure=True)
                        else:
                            self.fillBirdList()
                        self.menuBirdList.popup(QPoint(evt.screenPos().x(), evt.screenPos().y()))
                    else:
                        # TODO: pan the view
                        pass

    def GrowBox_ampl(self,pos):
        """ Listener for when a segment is being made in the amplitude plot.
        Makes the blue box that follows the mouse change size. """
        if self.p_ampl.sceneBoundingRect().contains(pos):
            mousePoint = self.p_ampl.mapSceneToView(pos)
            self.drawingBox_ampl.setRegion([self.start_ampl_loc, mousePoint.x()])
            self.drawingBox_spec.setRegion([self.convertAmpltoSpec(self.start_ampl_loc), self.convertAmpltoSpec(mousePoint.x())])

    def GrowBox_spec(self, pos):
        """ Listener for when a segment is being made in the spectrogram plot.
        Makes the blue box that follows the mouse change size. """
        if self.p_spec.sceneBoundingRect().contains(pos):
            mousePoint = self.p_spec.mapSceneToView(pos)
            self.drawingBox_ampl.setRegion([self.start_ampl_loc, self.convertSpectoAmpl(mousePoint.x())])
            if self.config['specMouseAction']>1 and not self.startedInAmpl:
                # making a box
                posY = mousePoint.y() - self.start_spec_y
                self.drawingBox_spec.setSize([mousePoint.x()-self.convertAmpltoSpec(self.start_ampl_loc), posY])
            else:
                # making a segment
                self.drawingBox_spec.setRegion([self.convertAmpltoSpec(self.start_ampl_loc), mousePoint.x()])

    def birdSelectedList(self,index):
        """ If the user clicks in the full bird list, update the text, and copy the species into the short list """
        birdname = self.fullbirdlist.view().currentIndex().parent().data(0)
        if birdname is None:
            birdname = self.fullbirdlist.currentText()
        else:
            birdname = birdname + ' (' + self.fullbirdlist.currentText() + ')'
        self.birdSelectedMenu(birdname)
        if not self.multipleBirds:
            self.menuBirdList.hide()

    def birdSelectedMenu(self,birditem):
        """ Collects the label for a bird from the context menu and processes it.
        Has to update the overview segments in case their colour should change.
        Also handles getting the name through a message box if necessary.
        """
        if type(birditem) is not str:
            birdname = birditem.text()
        else:
            birdname = birditem
        if birdname is None or birdname=='':
            return

        startpoint = self.segments[self.box1id][0]-self.startRead
        endpoint = self.segments[self.box1id][1]-self.startRead
        oldname = self.segments[self.box1id][4]

        # if it was checked, uncheck
        # basically re-create all names here:
        self.refreshOverviewWith(startpoint, endpoint, oldname, delete=True)
        self.segments[self.box1id][4] = []
        for t in oldname:
            # don't add the current species and Don't Know
            if t != birdname and t != "Don't Know":
                self.updateText(t)

        # if species wasn't in the list before, means it is now ticked, so add it:
        if birdname not in oldname and birdname != "Other":
            self.updateText(birdname)

        # patch for unticking initial "Don't Know"
        if "Don't Know" not in self.segments[self.box1id][4]:
            for act in self.menuBirdList.actions() + self.menuBird2.actions():
                if act.text()=="Don't Know":
                    act.setChecked(False)

        # patch for updating if all names were deleted:
        if self.segments[self.box1id][4] == []:
            if self.Hartley:
                self.updateText(text=[])
            else:
                self.updateText(text="Don't Know")

        # Now update the text
        if birdname == 'Other':
            # This allows textual name entry
            # Ask the user for the new name, and save it
            text, ok = QInputDialog.getText(self, 'Bird name', 'Enter the bird name as genus (species)')
            if ok:
                text = str(text).title()
                # splits "A (B)", with B optional, into groups A and B
                match = re.fullmatch(r'(.*?)(?: \((.*)\))?', text)
                if match:
                    twolevelname = '>'.join(match.groups(default=''))
                    if text in self.longBirdList or twolevelname in self.longBirdList:
                        # bird is already listed
                        print("Warning: not adding species %s as it is already present" % text)
                        return

                    # maybe the genus is already listed?
                    index = self.model.findItems(match.group(1), QtCore.Qt.MatchFixedString)
                    if len(index) == 0:
                        # Genus isn't in list
                        item = QStandardItem(match.group(1))
                        item.setSelectable(True)
                        self.model.appendRow(item)
                        # store as typed
                        nametostore = text
                    else:
                        # Get the species item
                        item = index[0]
                        if match.group(2) is None:
                            print("ERROR: genus %s already exists, please provide species as well" % match.group(1))
                            return
                        # store in two-level format
                        nametostore = twolevelname
                        subitem = QStandardItem(match.group(2))
                        item.setSelectable(False)
                        item.appendRow(subitem)
                        subitem.setSelectable(True)

                    # either way, update the lists:
                    if self.config['ReorderList']:
                        self.shortBirdList.insert(0,text)
                        del self.shortBirdList[-1]
                    self.longBirdList.append(nametostore)
                    self.longBirdList.remove('Unidentifiable')
                    self.longBirdList = sorted(self.longBirdList, key=str.lower)
                    self.longBirdList.append('Unidentifiable')
                    self.ConfigLoader.blwrite(self.longBirdList, self.config['BirdListLong'], self.configdir)
                else:
                    print("ERROR: provided name %s does not match format requirements" % text)
                    return

                self.updateText(text)

        # Put the selected bird name at the top of the list
        if len(birdname) > 0 and birdname[-1] == '?':
            birdname = birdname[:-1]
        if self.config['ReorderList']:
            # Either move the label to the top of the list, or delete the last
            if birdname in self.shortBirdList:
                self.shortBirdList.remove(birdname)
            else:
                del self.shortBirdList[-1]
            self.shortBirdList.insert(0,birdname)

        # refresh overview boxes after all updates:
        self.refreshOverviewWith(startpoint, endpoint, self.segments[self.box1id][4])

        if not self.multipleBirds:
            # select the bird and close
            self.menuBirdList.hide()

    def updateText(self,text,segID=None):
        """ When the user sets or changes the name in a segment, update the text and the colour. """
        if segID is None:
            segID = self.box1id

        # produce list from text
        if self.multipleBirds:
            if type(text) is list:
                self.segments[segID][4].extend(text)
                self.lastSpecies = text
            else:
                self.segments[segID][4].append(text)
                self.lastSpecies = [text]
            # get the unique elements:
            self.segments[segID][4] = list(set(self.segments[segID][4]))
        else:
            if type(text) is list:
                self.segments[segID][4] = text
            else:
                self.segments[segID][4] = [text]

        # Store the species in case the user wants it for the next segment
        self.segmentsToSave = True

        # produce text from list, to update the label
        text = ','.join(self.segments[segID][4])
        self.listLabels[segID].setText(text,'k')

        # Update the colour
        if "Don't Know" not in self.segments[segID][4] and len(self.segments[segID][4]) > 0:
            if '?' in text:
                self.prevBoxCol = self.ColourPossible
            else:
                self.prevBoxCol = self.ColourNamed
        else:
            self.prevBoxCol = self.ColourNone

    def setColourMap(self,cmap):
        """ Listener for the menu item that chooses a colour map.
        Loads them from the file as appropriate and sets the lookup table.
        """
        if self.media_obj.isPlaying():
            self.stopPlayback()
        self.config['cmap'] = cmap

        pos, colour, mode = colourMaps.colourMaps(cmap)

        cmap = pg.ColorMap(pos, colour,mode)
        self.lut = cmap.getLookupTable(0.0, 1.0, 256)

        self.specPlot.setLookupTable(self.lut)
        self.overviewImage.setLookupTable(self.lut)

    def invertColourMap(self):
        """ Listener for the menu item that converts the colour map"""
        # self.config['invertColourMap'] = not self.config['invertColourMap']
        self.config['invertColourMap'] = self.invertcm.isChecked()
        self.setColourLevels()

    def setColourLevels(self):
        """ Listener for the brightness and contrast sliders being changed. Also called when spectrograms are loaded, etc.
        Translates the brightness and contrast values into appropriate image levels.
        Calculation is simple.
        """
        if self.media_obj.isPlaying():
            self.stopPlayback()
        minsg = np.min(self.sg)
        maxsg = np.max(self.sg)

        if self.config['invertColourMap']:
            self.config['brightness'] = self.brightnessSlider.value()
        else:
            self.config['brightness'] = 100-self.brightnessSlider.value()
        self.config['contrast'] = self.contrastSlider.value()

        self.colourStart = (self.config['brightness'] / 100.0 * self.config['contrast'] / 100.0) * (maxsg - minsg) + minsg
        self.colourEnd = (maxsg - minsg) * (1.0 - self.config['contrast'] / 100.0) + self.colourStart

        if self.config['invertColourMap']:
            self.overviewImage.setLevels([self.colourEnd, self.colourStart])
            self.specPlot.setLevels([self.colourEnd, self.colourStart])
        else:
            self.overviewImage.setLevels([self.colourStart, self.colourEnd])
            self.specPlot.setLevels([self.colourStart, self.colourEnd])

    def moveLeft(self):
        """ When the left button is pressed (next to the overview plot), move everything along
        Allows a 10% overlap """
        minX, maxX = self.overviewImageRegion.getRegion()
        newminX = max(0,minX-(maxX-minX)*0.9)
        self.overviewImageRegion.setRegion([newminX, newminX+maxX-minX])
        self.updateOverview()
        self.playPosition = int(self.convertSpectoAmpl(newminX)*1000.0)

    def moveRight(self):
        """ When the right button is pressed (next to the overview plot), move everything along
        Allows a 10% overlap """
        minX, maxX = self.overviewImageRegion.getRegion()
        newminX = min(np.shape(self.sg)[0]-(maxX-minX),minX+(maxX-minX)*0.9)
        self.overviewImageRegion.setRegion([newminX, newminX+maxX-minX])
        self.updateOverview()
        self.playPosition = int(self.convertSpectoAmpl(newminX)*1000.0)

    def prepare5minMove(self):
        self.saveSegments()
        self.resetStorageArrays()
        self.loadFile()

    def movePrev5mins(self):
        """ When the button to move to the next 5 minutes is pressed, enable that.
        Have to check if the buttons should be disabled or not,
        save the segments and reset the arrays, then call loadFile.
        """
        self.currentFileSection -= 1
        self.next5mins.setEnabled(True)
        self.moveNext5minsKey.setEnabled(True)
        if self.currentFileSection <= 0:
            self.prev5mins.setEnabled(False)
            self.movePrev5minsKey.setEnabled(False)
        self.prepare5minMove()

    def moveNext5mins(self):
        """ When the button to move to the previous 5 minutes is pressed, enable that.
        Have to check if the buttons should be disabled or not,
        save the segments and reset the arrays, then call loadFile.
        """
        self.currentFileSection += 1
        self.prev5mins.setEnabled(True)
        self.movePrev5minsKey.setEnabled(True)
        if self.currentFileSection >= self.nFileSections-1:
            self.next5mins.setEnabled(False)
            self.moveNext5minsKey.setEnabled(False)
        self.prepare5minMove()

    def scroll(self):
        """ When the slider at the bottom of the screen is moved, move everything along. """
        newminX = self.scrollSlider.value()
        minX, maxX = self.overviewImageRegion.getRegion()
        self.overviewImageRegion.setRegion([newminX, newminX+maxX-minX])
        self.updateOverview()
        self.playPosition = int(self.convertSpectoAmpl(newminX)*1000.0)

    def changeWidth(self, value):
        """ Listener for the spinbox that decides the width of the main window.
        It updates the top figure plots as the window width is changed.
        Slightly annoyingly, it gets called when the value gets reset, hence the first line. """
        if not hasattr(self,'overviewImageRegion'):
            return
        self.windowSize = value

        # Redraw the highlight in the overview figure appropriately
        minX, maxX = self.overviewImageRegion.getRegion()
        newmaxX = self.convertAmpltoSpec(value)+minX
        self.overviewImageRegion.setRegion([minX, newmaxX])
        self.scrollSlider.setMaximum(np.shape(self.sg)[0]-self.convertAmpltoSpec(self.widthWindow.value()))
        # self.updateOverview()

# ===============
# Generate the various dialogs that match the menu items

    def loadSegment(self, hr2=False):
        """ Loads a segment for the HumanClassify dialogs """
        if hr2:
            wavobj = wavio.read(self.filename)
        else:
            wavobj = wavio.read(self.filename, self.config['maxFileShow'], self.startRead)
        self.audiodata = wavobj.data

        if self.audiodata.dtype is not 'float':
            self.audiodata = self.audiodata.astype('float')  # / 32768.0

        if np.shape(np.shape(self.audiodata))[0] > 1:
            self.audiodata = self.audiodata[:, 0]

        # Get the data for the spectrogram
        sgRaw = self.sp.spectrogram(self.audiodata, self.config['window_width'],
                                    self.config['incr'], mean_normalise=self.sgMeanNormalise, equal_loudness=self.sgEqualLoudness, onesided=self.sgOneSided, multitaper=self.sgMultitaper)
        maxsg = np.min(sgRaw)
        self.sg = np.abs(np.where(sgRaw == 0, 0.0, 10.0 * np.log10(sgRaw / maxsg)))

    def showFirstPage(self):
        """ After the HumanClassify dialogs have closed, need to show the correct data on the screen
         Returns to the page user started with """
        if self.config['maxFileShow']<self.datalengthSec:
            self.currentFileSection = self.currentPage
            self.prepare5minMove()
            self.next5mins.setEnabled(True)
            self.prev5mins.setEnabled(False)
            self.moveNext5minsKey.setEnabled(True)
            self.movePrev5minsKey.setEnabled(False)

    def humanClassifyDialog1(self):
        """ Create the dialog that shows calls to the user for verification.
        There are two versions in here depending on whether you wish to show them from all the pages (5 min sections), the default, or not.
        The only difference is that that version requires sorting the segments and loading of new file sections and making the spectrogram intermittently.
        """

        # Store the current page to return to
        self.currentPage = self.currentFileSection
        self.segmentsDone = 0

        # Check there are segments to show on this page
        if not self.config['showAllPages']:
            if len(self.segments)>0:
                self.box1id = 0
                while self.box1id<len(self.segments) and self.listRectanglesa2[self.box1id] is None:
                    self.box1id += 1
        else:
            self.box1id = 0

        if (self.config['showAllPages'] and len(self.segments)==0) or (not self.config['showAllPages'] and (self.box1id == len(self.segments) or len(self.listRectanglesa2)==0)):
            msg = SupportClasses.MessagePopup("w", "No segments", "No segments to check")
            msg.exec_()
            return
        else:
            if self.config['showAllPages']:
                # Showing them on all pages is a bit more of a pain
                # Sort the segments into increasing time order, apply same order to listRects and labels
                sortOrder = sorted(range(len(self.segments)), key=self.segments.__getitem__)
                self.segments = [self.segments[i] for i in sortOrder]
                self.listRectanglesa1 = [self.listRectanglesa1[i] for i in sortOrder]
                self.listRectanglesa2 = [self.listRectanglesa2[i] for i in sortOrder]
                self.listLabels = [self.listLabels[i] for i in sortOrder]

                # Check which page is first to have segments on
                self.currentFileSection = -1

            self.humanClassifyDialog1 = Dialogs.HumanClassify1(self.lut,self.colourStart,self.colourEnd,self.config['invertColourMap'], self.brightnessSlider.value(), self.contrastSlider.value(), self.shortBirdList, self.longBirdList, self.multipleBirds, self)
            # load the first image:
            self.box1id = -1
            self.humanClassifyDialog1.setSegNumbers(0, len(self.segments))
            if hasattr(self, 'humanClassifyDialogSize'):
                self.humanClassifyDialog1.resize(self.humanClassifyDialogSize)

            self.humanClassifyNextImage1()
            self.humanClassifyDialog1.show()
            self.humanClassifyDialog1.activateWindow()
            #self.humanClassifyDialog1.close.clicked.connect(self.humanClassifyClose1)
            self.humanClassifyDialog1.buttonPrev.clicked.connect(self.humanClassifyPrevImage)
            self.humanClassifyDialog1.buttonNext.clicked.connect(self.humanClassifyNextImage1)
            self.humanClassifyDialog1.correct.clicked.connect(self.humanClassifyCorrect1)
            self.humanClassifyDialog1.delete.clicked.connect(self.humanClassifyDelete1)
            # self.statusLeft.setText("Ready")

    def humanClassifyClose1(self):
        """ Listener for the human verification dialog. """
        self.humanClassifyDialogSize = self.humanClassifyDialog1.size()
        self.humanClassifyDialog1.done(1)
        self.box1id = -1
        # Want to show a page at the end, so make it the first one
        if self.config['showAllPages']:
            self.showFirstPage()

    def humanClassifyNextImage1(self):
        """ Get the next image """
        self.humanClassifyDialogSize = self.humanClassifyDialog1.size()
        if self.box1id < len(self.segments)-1:
            self.box1id += 1
            species = list(self.segments[self.box1id][4])

            # Update the colour in case of sudden dialog closing
            text = ','.join(species)
            if "Don't Know" not in text and len(species) > 0:
                if '?' in text:
                    self.prevBoxCol = self.ColourPossible
                else:
                    self.prevBoxCol = self.ColourNamed
            else:
                self.prevBoxCol = self.ColourNone

            # update "done/to go" numbers:
            #print("Next image",self.segments[self.box1id])
            self.humanClassifyDialog1.setSegNumbers(self.box1id, len(self.segments))
            if not self.config['showAllPages']:
                # Different calls for the two types of region
                if self.listRectanglesa2[self.box1id] is not None:
                    if type(self.listRectanglesa2[self.box1id]) == self.ROItype:
                        x1nob = self.listRectanglesa2[self.box1id].pos()[0]
                        x2nob = x1nob + self.listRectanglesa2[self.box1id].size()[0]
                    else:
                        x1nob, x2nob = self.listRectanglesa2[self.box1id].getRegion()
                    x1 = int(x1nob - self.config['reviewSpecBuffer'])
                    x1 = max(x1, 0)
                    x2 = int(x2nob + self.config['reviewSpecBuffer'])
                    x2 = min(x2, len(self.sg))
                    x3 = int((self.listRectanglesa1[self.box1id].getRegion()[0] - self.config['reviewSpecBuffer']) * self.sampleRate)
                    x3 = max(x3, 0)
                    x4 = int((self.listRectanglesa1[self.box1id].getRegion()[1] + self.config['reviewSpecBuffer']) * self.sampleRate)
                    x4 = min(x4, len(self.audiodata))
                    self.humanClassifyDialog1.setImage(self.sg[x1:x2, :], self.audiodata[x3:x4], self.sampleRate, self.config['incr'],
                                                       species, self.convertAmpltoSpec(x1nob)-x1, self.convertAmpltoSpec(x2nob)-x1, 
                                                       self.segments[self.box1id][0], self.segments[self.box1id][1], self.minFreq, self.maxFreq)
            else:
                # Check if have moved to next segment, and if so load it
                # If there was a section without segments this would be a bit inefficient, actually no, it was wrong!
                if self.segments[self.box1id][0] > (self.currentFileSection+1)*self.config['maxFileShow']:
                    while self.segments[self.box1id][0] > (self.currentFileSection+1)*self.config['maxFileShow']:
                        self.currentFileSection += 1
                    self.startRead = self.currentFileSection * self.config['maxFileShow']
                    with pg.BusyCursor():
                        print("Loading next page", self.currentFileSection)
                        self.loadSegment()
                self.humanClassifyDialog1.setWindowTitle('Check Classifications: page ' + str(self.currentFileSection+1))

                # Show the next segment
                if self.segments[self.box1id] is not None:
                    x1nob = self.segments[self.box1id][0] - self.startRead
                    x2nob = self.segments[self.box1id][1] - self.startRead
                    x1 = int(self.convertAmpltoSpec(x1nob - self.config['reviewSpecBuffer']))
                    x1 = max(x1, 0)
                    x2 = int(self.convertAmpltoSpec(x2nob + self.config['reviewSpecBuffer']))
                    x2 = min(x2, len(self.sg))
                    x3 = int((x1nob - self.config['reviewSpecBuffer']) * self.sampleRate)
                    x3 = max(x3, 0)
                    x4 = int((x2nob + self.config['reviewSpecBuffer']) * self.sampleRate)
                    x4 = min(x4, len(self.audiodata))
                    self.humanClassifyDialog1.setImage(self.sg[x1:x2, :], self.audiodata[x3:x4], self.sampleRate, self.config['incr'],
                                                   species, self.convertAmpltoSpec(x1nob)-x1, self.convertAmpltoSpec(x2nob)-x1,
                                                   self.segments[self.box1id][0], self.segments[self.box1id][1], self.minFreq, self.maxFreq)
                else:
                    print("Segment %s missing for some reason" % self.box1id)

        else:
            msg = SupportClasses.MessagePopup("d", "Finished", "All segmentations checked")
            msg.exec_()
            self.humanClassifyClose1()

    def humanClassifyPrevImage(self):
        """ Go back one image by changing boxid and calling NextImage.
        Note: won't undo deleted segments."""
        if self.box1id>0:
            self.box1id -= 2
            self.humanClassifyNextImage1()

    def updateLabel(self,label):
        """ Update the label on a segment that is currently shown in the display. """
        # birdSelectedMenu flips the state of each label
        # so need to pass all labels for deletion, or clean before updating

        # Need to keep track of self.multipleBirds
        multipleTemp = self.multipleBirds
        self.multipleBirds = True
        self.segments[self.box1id][4] = []
        for l in label:
            self.birdSelectedMenu(l)
        if label==[]:
            self.birdSelectedMenu("Don't Know")
        self.multipleBirds = multipleTemp

        if self.listRectanglesa2[self.box1id] is not None:
            col = self.prevBoxCol.rgb()
            col = QtGui.QColor(col)
            col.setAlpha(100)
            self.listRectanglesa1[self.box1id].setBrush(self.prevBoxCol)
            self.listRectanglesa2[self.box1id].setBrush(self.prevBoxCol)
            col.setAlpha(200)
            self.listRectanglesa1[self.box1id].setHoverBrush(fn.mkBrush(col))
            self.listRectanglesa2[self.box1id].setHoverBrush(fn.mkBrush(col))
            if self.config['transparentBoxes'] and type(self.listRectanglesa2[self.box1id]) == self.ROItype:
                col.setAlpha(255)
                self.listRectanglesa2[self.box1id].setPen(col, width=1)
                self.listRectanglesa2[self.box1id].setBrush(pg.mkBrush(None))
                self.listRectanglesa2[self.box1id].setHoverBrush(pg.mkBrush(None))
                col.setAlpha(100)
            self.listRectanglesa1[self.box1id].update()
            self.listRectanglesa2[self.box1id].update()
            self.segmentsToSave = True

    def humanClassifyCorrect1(self):
        """ Correct segment labels, save the old ones if necessary """
        startpoint = self.segments[self.box1id][0]-self.startRead
        endpoint = self.segments[self.box1id][1]-self.startRead
        oldname = list(self.segments[self.box1id][4])

        self.humanClassifyDialog1.stopPlayback()
        self.segmentsDone += 1
        label, self.saveConfig, checkText = self.humanClassifyDialog1.getValues()

        if len(checkText) > 0:
            if checkText in self.longBirdList:
                pass
            else:
                self.longBirdList.append(checkText)
                self.longBirdList = sorted(self.longBirdList, key=str.lower)
                self.longBirdList.remove('Unidentifiable')
                self.longBirdList.append('Unidentifiable')
                self.ConfigLoader.blwrite(self.longBirdList, self.config['BirdListLong'], self.configdir)

        # Todo: boxid[4] has been updated so this if doesn't effect? added update label to else but not the ideal sol
        if label != self.segments[self.box1id][4]:
            if self.config['saveCorrections']:
                # Save the correction
                outputError = [self.segments[self.box1id], label]
                file = open(self.filename + '.corrections', 'a')
                json.dump(outputError, file,indent=1)
                file.close()

            # force wipe old overview to empty,
            # because it's difficult to maintain old species properly through dialogs
            self.refreshOverviewWith(startpoint, endpoint, oldname, delete=True)
            self.refreshOverviewWith(startpoint, endpoint, "Don't Know")
            self.updateLabel(label)

            if self.saveConfig:
                self.longBirdList.append(checkText)
                self.longBirdList = sorted(self.longBirdList, key=str.lower)
                self.longBirdList.remove('Unidentifiable')
                self.longBirdList.append('Unidentifiable')
                self.ConfigLoader.blwrite(self.longBirdList, self.config['BirdListLong'], self.configdir)
        elif '?' in ''.join(label):
            # Remove the question mark, since the user has agreed
            for i in range(len(self.segments[self.box1id][4])):
                if self.segments[self.box1id][4][i][-1] == '?':
                    self.segments[self.box1id][4][i] = self.segments[self.box1id][4][i][:-1]

            # force wipe old overview to empty
            self.refreshOverviewWith(startpoint, endpoint, oldname, delete=True)
            self.refreshOverviewWith(startpoint, endpoint, "Don't Know")
            self.updateLabel(self.segments[self.box1id][4])
        else:
            # segment info matches, so don't do anything
            pass

        self.humanClassifyDialog1.tbox.setText('')
        self.humanClassifyDialog1.tbox.setEnabled(False)
        self.humanClassifyNextImage1()

    def humanClassifyDelete1(self):
        """ If the user has deleted a segment in the review, delete it from the main display """
        id = self.box1id
        self.humanClassifyDialog1.stopPlayback()
        self.deleteSegment(self.box1id, hr=True)

        self.box1id = id-1
        self.segmentsToSave = True
        self.humanClassifyNextImage1()
        self.segmentsDone += 1

    def humanRevDialog2(self):
        """ Create the dialog that shows sets of calls to the user for verification.
        """
        if len(self.segments)==0 or self.box1id == len(self.segments) or len(self.listRectanglesa2)==0:
            msg = SupportClasses.MessagePopup("w", "No segments", "No segments to check")
            msg.exec_()
            return
        self.statusLeft.setText("Checking...")

        # Get all labels into a single list
        names = [sp for seg in self.segments for sp in seg[4]]

        # TODO: at the moment, we're showing all "yellow" and "green" segments together.
        names = [re.sub('\?','',item) for item in names]
        # Make them unique
        names = list(set(names))
        try:
            # can't use single-species review on "Don't Know" segments
            names.remove("Don't Know")
        except Exception:
            pass
        if len(names) == 0:
            msg = SupportClasses.MessagePopup("w", "No segments", "No segments to check")
            msg.exec_()
            return

        # mini dialog to select species:
        self.humanClassifyDialog2a = Dialogs.HumanClassify2a(names)

        if self.humanClassifyDialog2a.exec_() == 1:
            self.revLabel = self.humanClassifyDialog2a.getValues()

            # main dialog:
            self.humanClassifyDialog2 = Dialogs.HumanClassify2(self.sg, self.audiodata, self.segments,
                                                               self.revLabel, self.sampleRate, self.audioFormat,
                                                               self.config['incr'], self.lut, self.colourStart,
                                                               self.colourEnd, self.config['invertColourMap'],
                                                               self.brightnessSlider.value(), self.contrastSlider.value())
            self.humanClassifyDialog2.finish.clicked.connect(self.humanClassifyClose2)
            self.humanClassifyDialog2.exec_()

    def humanClassifyClose2(self):
        self.segmentsToSave = True
        todelete = []
        # initialize correction file. All "downgraded" segments will be stored
        outputErrors = []

        for btn in self.humanClassifyDialog2.buttons:
            btn.stopPlayback()
            self.box1id = btn.index
            currSeg = self.segments[btn.index]
            # clear these species from overview colors
            self.refreshOverviewWith(currSeg[0], currSeg[1], currSeg[4], delete=True)
            # btn.index carries the index of segment shown on btn
            if btn.mark=="red":
                outputErrors.append(currSeg)
                if len(currSeg[4])==1:
                    # delete if this was the only species label:
                    todelete.append(btn.index)
                else:
                    # otherwise just delete this species from the label
                    if self.revLabel in currSeg[4]:
                        currSeg[4].remove(self.revLabel)
                    if self.revLabel+'?' in currSeg[4]:
                        currSeg[4].remove(self.revLabel+'?')
                # this will also update the overview colors
                self.updateLabel(currSeg[4])
            # fix name or name+? of the analyzed species
            elif btn.mark=="yellow":
                for lbindex in range(len(currSeg[4])):
                    label = currSeg[4][lbindex]
                    # find "greens", swap to "yellows"
                    if label==self.revLabel:
                        outputErrors.append(currSeg)
                        currSeg[4][lbindex] = self.revLabel+'?'
                # this will also update the overview colors
                self.updateLabel(currSeg[4])
            elif btn.mark=="green":
                for lbindex in range(len(currSeg[4])):
                    label = currSeg[4][lbindex]
                    # find "yellows", swap to "greens"
                    if label==self.revLabel+'?':
                        currSeg[4][lbindex] = self.revLabel
                # this will also update the overview colors
                self.updateLabel(currSeg[4])

        self.humanClassifyDialog2.done(1)

        # Save the errors in a file
        if self.config['saveCorrections'] and len(outputErrors)>0:
            speciesClean = re.sub(r'\W', "_", self.revLabel)
            file = open(self.filename + '.corrections_' + speciesClean, 'a')
            json.dump(outputErrors, file,indent=1)
            file.close()

        # reverse loop to allow deleting segments
        for dl in reversed(todelete):
            self.deleteSegment(dl)
        self.saveSegments()
        self.statusLeft.setText("Ready")
        return

    def showDiagnosticDialog(self):
        """ Create the dialog to set diagnostic plot parameters.
        """
        if not hasattr(self, 'diagnosticDialog'):
            self.diagnosticDialog = Dialogs.Diagnostic(self.FilterFiles)
            self.diagnosticDialog.activate.clicked.connect(self.setDiagnostic)
            self.diagnosticDialog.clear.clicked.connect(self.clearDiagnostic)
        self.diagnosticDialog.show()
        self.diagnosticDialog.activateWindow()

    def clearDiagnostic(self):
        """ Cleans up diagnostic plot space. Should be called
            when loading new file/page, or from Diagnostic Dialog.
        """
        try:
            self.p_plot.clear()
            if hasattr(self, "p_legend"):
                self.p_legend.scene().removeItem(self.p_legend)
            if hasattr(self, "diagnosticCalls"):
                for c in self.diagnosticCalls:
                    self.p_spec.removeItem(c)
            self.d_plot.hide()
        except Exception as e:
            print(e)
        self.diagnosticCalls = []


    def setDiagnostic(self):
        """ Takes parameters returned from DiagnosticDialog
            and draws the training diagnostic plots.
        """
        with pg.BusyCursor():
            self.diagnosticDialog.activate.setEnabled(False)
            self.statusLeft.setText("Making diagnostic plots...")
            # Note: importing here so that main program could be run without ext/
            from ext import ce_denoise

            # take values: -2/-3/-4 for AA types, -2/-3 for En/Spec plot
            [filter, aaType, markSpec] = self.diagnosticDialog.getValues()
            spInfo = json.load(open(os.path.join(self.filtersDir, filter + '.txt')))

            # clear plot box and add legend
            self.clearDiagnostic()
            self.p_legend = pg.LegendItem()
            self.p_legend.setParentItem(self.p_plot)
            # 1 sec in spectrogram units
            specs = self.convertAmpltoSpec(1)

            # plot things
            # 1. decompose
            # if needed, adjusting sampling rate to match filter
            if self.sampleRate != spInfo['SampleRate']:
                datatoplot = librosa.core.audio.resample(self.audiodata, self.sampleRate, spInfo['SampleRate'])
            else:
                datatoplot = self.audiodata

            WF = WaveletFunctions.WaveletFunctions(data=datatoplot, wavelet='dmey2', maxLevel=5, samplerate=spInfo['SampleRate'])
            WF.WaveletPacket(5, 'symmetric', aaType==-4, antialiasFilter=True)
            numNodes = len(spInfo['WaveletParams'][2])
            Esep = np.zeros(( numNodes, int(self.datalengthSec) ))

            ### DENOISING reference: relative |amp| on rec signals from each WP node, when wind is present
            ### just handmade from some wind examples
            # noiseenv = np.array([0.54, 0.83, 0.84, 1, 0.32, 0.54, 0.70, 0.70,  0.16, 0.19, 0.24, 0.22, 0.28, 0.27, 0.25, 0.26,  0.01, 0.06, 0.14, 0.12, 0.15, 0.16, 0.14, 0.15,  0.16, 0.16, 0.15, 0.15, 0.14, 0.16, 0.15, 0.15])
            # # reconstruct wind signal
            # windC = WF.reconstructWP2(34, aaType != -2, False)
            # windC = np.abs(windC)
            # print("calculating wind strength")
            # windE = ce_denoise.EnergyCurve(windC, 4000)
            # # "noise strength index" - strength of wind at each second, on linear scale
            # windMaxE = np.zeros(int(self.datalengthSec))
            # for w in range(int(self.datalengthSec)):
            #     windMaxE[w] = np.max(windE[w*spInfo['SampleRate'] : (w+1)*spInfo['SampleRate']])
            # del windC
            # del windE

            # 2. reconstruct from bands
            r = 0
            M = spInfo['WaveletParams'][1]
            for node in spInfo['WaveletParams'][2]:
                # reconstruction as in detectCalls:
                print("working on node", node)
                C = WF.reconstructWP2(node, aaType != -2, True)
                C = self.sp.ButterworthBandpass(C, spInfo['SampleRate'],
                        low=spInfo['FreqRange'][0], high=spInfo['FreqRange'][1])

                C = np.abs(C)
                E = ce_denoise.EnergyCurve(C, int( M*spInfo['SampleRate']/2 ))
                C = np.log(C)

                # some prep that doesn't need to be looped over t:
                meanC = np.mean(C)
                sdC = np.std(C)

                # get true freqs of this band
                freqmin, freqmax = WF.getWCFreq(node, spInfo['SampleRate'])
                # convert freqs to spec Y units
                freqmin = self.convertFreqtoY(freqmin)
                freqmax = self.convertFreqtoY(freqmax)

                # get max E for each second
                # and normalize, so that we don't need to hardcode thr 
                for w in range(int(self.datalengthSec)):
                    maxE = np.max(E[w*spInfo['SampleRate'] : (w+1)*spInfo['SampleRate']])
                    ### DENOISE:
                    # based on wind strength in this second, calculate estimated |wind| in this node
                    # and subtract from maxE
                    # maxE = max(meanC, maxE - windMaxE[w]*noiseenv[node-31]*1.1)
                    Esep[r,w] = (np.log(maxE) - meanC) / sdC

                    # mark detected calls on spectrogram
                    if markSpec and Esep[r,w] > spInfo['WaveletParams'][0]:
                        diagCall = pg.ROI((specs*w, (freqmin+freqmax)/2),
                                (specs, freqmax-freqmin),
                                pen=(255*r//numNodes,0,0), movable=False)
                        self.diagnosticCalls.append(diagCall)
                        self.p_spec.addItem(diagCall)

                # plot
                self.plotDiag = pg.PlotDataItem(np.arange(int(self.datalengthSec))+0.5, Esep[r,:],
                        pen=fn.mkPen((255*r//numNodes,0,0), width=2))
                self.p_plot.addItem(self.plotDiag)
                self.p_legend.addItem(self.plotDiag, str(node))
                r = r + 1 

            ### DENOISE: add line of wind strength
            # self.p_plot.addItem(pg.PlotDataItem(np.arange(int(self.datalengthSec))+0.5, np.log(windMaxE),
            #             pen=fn.mkPen((0,130,0), width=2)))
            # add line corresponding to thr
            # self.p_plot.addItem(pg.InfiniteLine(-0.8, angle=0, pen=fn.mkPen(color=(40,40,40), width=1)))
            self.p_plot.addItem(pg.InfiniteLine(spInfo['WaveletParams'][0], angle=0, pen=fn.mkPen(color=(40,40,40), width=1)))
            minX, maxX = self.overviewImageRegion.getRegion()
            self.p_plot.setXRange(self.convertSpectoAmpl(minX), self.convertSpectoAmpl(maxX), update=True, padding=0)
            self.plotaxis.setLabel('Power Z-score')
            self.d_plot.show()
        self.diagnosticDialog.activate.setEnabled(True)
        self.statusLeft.setText("Ready")

    def showSpectrogramDialog(self):
        """ Create spectrogram dialog when the button is pressed.
        """
        if not hasattr(self,'spectrogramDialog'):
            self.spectrogramDialog = Dialogs.Spectrogram(self.config['window_width'],self.config['incr'],self.minFreq,self.maxFreq, self.minFreqShow,self.maxFreqShow)
        self.spectrogramDialog.show()
        self.spectrogramDialog.activateWindow()
        self.spectrogramDialog.activate.clicked.connect(self.spectrogram)

    def spectrogram(self):
        """ Listener for the spectrogram dialog.
        Has to do quite a bit of work to make sure segments are in the correct place, etc."""
        [windowType, self.sgMeanNormalise, self.sgEqualLoudness, self.sgMultitaper, window_width, incr, minFreq, maxFreq] = self.spectrogramDialog.getValues()
        if (minFreq >= maxFreq):
            msg = SupportClasses.MessagePopup("w", "Error", "Incorrect frequency range")
            msg.exec_()
        with pg.BusyCursor():
            self.statusLeft.setText("Updating the spectrogram...")
            self.sp.setWidth(int(str(window_width)), int(str(incr)))
            oldSpecy = np.shape(self.sg)[1]
            sgRaw = self.sp.spectrogram(self.audiodata,window=str(windowType),mean_normalise=self.sgMeanNormalise,equal_loudness=self.sgEqualLoudness,onesided=self.sgOneSided,multitaper=self.sgMultitaper)
            maxsg = np.min(sgRaw)
            self.sg = np.abs(np.where(sgRaw==0,0.0,10.0 * np.log10(sgRaw/maxsg)))

            # If the size of the spectrogram has changed, need to update the positions of things
            if int(str(incr)) != self.config['incr'] or int(str(window_width)) != self.config['window_width']:
                self.config['incr'] = int(str(incr))
                self.config['window_width'] = int(str(window_width))
                if hasattr(self, 'seg'):
                    self.seg.setNewData(self.audiodata, sgRaw, self.sampleRate, self.config['window_width'], self.config['incr'])

            self.redoFreqAxis(minFreq,maxFreq)

            self.statusLeft.setText("Ready")

    def showDenoiseDialog(self):
        """ Create the denoising dialog when the relevant button is pressed.
        """
        self.denoiseDialog = Dialogs.Denoise(DOC=self.DOC,minFreq=self.minFreq,maxFreq=self.maxFreq)
        self.denoiseDialog.show()
        self.denoiseDialog.activateWindow()
        self.denoiseDialog.activate.clicked.connect(self.denoise)
        self.denoiseDialog.undo.clicked.connect(self.denoise_undo)
        self.denoiseDialog.save.clicked.connect(self.denoise_save)

    def backup(self):
        """ Enables denoising to be undone. """
        if hasattr(self, 'audiodata_backup'):
            if self.audiodata_backup is not None:
                audiodata_backup_new = np.empty(
                    (np.shape(self.audiodata_backup)[0], np.shape(self.audiodata_backup)[1] + 1))
                audiodata_backup_new[:, :-1] = np.copy(self.audiodata_backup)
                audiodata_backup_new[:, -1] = np.copy(self.audiodata)
                self.audiodata_backup = audiodata_backup_new
            else:
                self.audiodata_backup = np.empty((np.shape(self.audiodata)[0], 1))
                self.audiodata_backup[:, 0] = np.copy(self.audiodata)
        else:
            self.audiodata_backup = np.empty((np.shape(self.audiodata)[0], 1))
            self.audiodata_backup[:, 0] = np.copy(self.audiodata)
        self.showFreq_backup = [self.minFreqShow, self.maxFreqShow]

    def decomposeWP(self, x=None):
        """ Listener for quickWP control button.
            Takes DATA and produces a WP decomposition.
        """
        print("Decomposing to WP...")
        ot = time.time()
        self.WFinst = WaveletFunctions.WaveletFunctions(data=self.audiodata, wavelet="dmey2", maxLevel=self.config['maxSearchDepth'], samplerate=self.sampleRate)
        self.WFinst.WaveletPacket(maxlevel=5, mode='symmetric', antialias=False)
        print("Done")
        print(time.time() - ot)

    def denoiseSeg(self):
        """ Listener for quickDenoise control button.
            Extracts a segment from DATA between START and STOP (in ms),
            denoises that segment, concats with rest of original DATA,
            and updates the original DATA.
        """

        if self.box1id > -1:
            start = self.listRectanglesa1[self.box1id].getRegion()[0] * 1000
            stop = self.listRectanglesa1[self.box1id].getRegion()[1] * 1000
        else:
            print("Can't play, no segment selected")
            return

        if self.media_obj.isPlaying():
            self.stopPlayback()

        # Since there is no dialog menu, settings are preset constants here:
        alg = "Wavelets"
        thrType = "soft"
        depth = 6   # can also use 0 to autoset
        wavelet = "dmey2"
        aaRec = True
        aaWP = True
        thr = 2.0  # this one is difficult to set universally...

        with pg.BusyCursor():
            opstartingtime = time.time()
            print("Denoising requested at " + time.strftime('%H:%M:%S', time.gmtime(opstartingtime)))
            self.statusLeft.setText("Denoising...")

            # extract the piece of audiodata under current segment
            denoised = self.audiodata[int(start * self.sampleRate//1000) : int(stop * self.sampleRate//1000)]
            WF = WaveletFunctions.WaveletFunctions(data=denoised, wavelet=wavelet, maxLevel=self.config['maxSearchDepth'], samplerate=self.sampleRate)

            if alg == "Wavelets":
                denoised = WF.waveletDenoise(thrType, thr, depth, aaRec=aaRec, aaWP=aaWP, thrfun="c")

            print("Denoising calculations completed in %.4f seconds" % (time.time() - opstartingtime))

            # update full audiodata
            self.audiodata[int(start * self.sampleRate//1000) : int(stop * self.sampleRate//1000)] = denoised

            # recalculate spectrogram
            sgRaw = self.sp.spectrogram(self.audiodata,mean_normalise=self.sgMeanNormalise,equal_loudness=self.sgEqualLoudness,onesided=self.sgOneSided,multitaper=self.sgMultitaper)
            maxsg = np.min(sgRaw)
            self.sg = np.abs(np.where(sgRaw==0,0.0,10.0 * np.log10(sgRaw/maxsg)))

            # Update the ampl image
            self.amplPlot.setData(np.linspace(0.0,self.datalength/self.sampleRate,num=self.datalength,endpoint=True),self.audiodata)

            # Update the spec & overview images.
            # Does not reset to start if the freqs aren't changed
            self.redoFreqAxis(self.minFreqShow,self.maxFreqShow, store=False)

            if hasattr(self,'spectrogramDialog'):
                self.spectrogramDialog.setValues(self.minFreq,self.maxFreq,self.minFreqShow,self.maxFreqShow)

            self.setColourLevels()

        print("Denoising completed in %s seconds" % round(time.time() - opstartingtime, 4))
        self.statusLeft.setText("Ready")

    def denoiseSegN(self):
        """ Listener for quickDenoise control button.
            Extracts a segment from DATA between START and STOP (in ms),
            denoises that segment, concats with rest of original DATA,
            and updates the original DATA.
        """

        if self.box1id > -1:
            start = self.listRectanglesa1[self.box1id].getRegion()[0] * 1000
            stop = self.listRectanglesa1[self.box1id].getRegion()[1] * 1000
        else:
            print("Can't play, no segment selected")
            return

        if self.media_obj.isPlaying():
            self.stopPlayback()

        # Since there is no dialog menu, settings are preset constants here:
        alg = "Wavelets"
        thrType = "soft"
        depth = 6   # can also use 0 to autoset
        wavelet = "dmey2"
        aaRec = True
        aaWP = True
        thr = 2.0  # this one is difficult to set universally...

        with pg.BusyCursor():
            opstartingtime = time.time()
            print("Denoising requested at " + time.strftime('%H:%M:%S', time.gmtime(opstartingtime)))
            self.statusLeft.setText("Denoising...")

            # extract the piece of audiodata under current segment
            denoised = self.audiodata[int(start * self.sampleRate//1000) : int(stop * self.sampleRate//1000)]
            WF = WaveletFunctions.WaveletFunctions(data=denoised, wavelet=wavelet, maxLevel=self.config['maxSearchDepth'], samplerate=self.sampleRate)

            if alg == "Wavelets":
                denoised = WF.waveletDenoise(thrType, thr, depth, aaRec=aaRec, aaWP=aaWP, thrfun="n")

            print("Denoising calculations completed in %.4f seconds" % (time.time() - opstartingtime))

            # update full audiodata
            self.audiodata[int(start * self.sampleRate//1000) : int(stop * self.sampleRate//1000)] = denoised

            # recalculate spectrogram
            sgRaw = self.sp.spectrogram(self.audiodata,mean_normalise=self.sgMeanNormalise,equal_loudness=self.sgEqualLoudness,onesided=self.sgOneSided,multitaper=self.sgMultitaper)
            maxsg = np.min(sgRaw)
            self.sg = np.abs(np.where(sgRaw==0,0.0,10.0 * np.log10(sgRaw/maxsg)))

            # Update the ampl image
            self.amplPlot.setData(np.linspace(0.0,self.datalength/self.sampleRate,num=self.datalength,endpoint=True),self.audiodata)

            # Update the spec & overview images.
            # Does not reset to start if the freqs aren't changed
            self.redoFreqAxis(self.minFreqShow,self.maxFreqShow, store=False)

            if hasattr(self,'spectrogramDialog'):
                self.spectrogramDialog.setValues(self.minFreq,self.maxFreq,self.minFreqShow,self.maxFreqShow)

            self.setColourLevels()

        print("Denoising completed in %s seconds" % round(time.time() - opstartingtime, 4))
        self.statusLeft.setText("Ready")


    def denoise(self):
        """ Listener for the denoising dialog.
        Calls the denoiser and then plots the updated data.
        """
        if self.CLI:
            # in CLI mode, default values will be retrieved from dialogs.
            self.denoiseDialog = Dialogs.Denoise(DOC=self.DOC,minFreq=self.minFreq,maxFreq=self.maxFreq)
            # values can be passed here explicitly, e.g.:
            # self.denoiseDialog.depth.setValue(10)
            # or could add an argument to pass custom defaults, e.g.:
            # self.denoiseDialog = Dialogs.Denoise(defaults=("wt", 1, 2, 'a')
        with pg.BusyCursor():
            opstartingtime = time.time()
            print("Denoising requested at " + time.strftime('%H:%M:%S', time.gmtime(opstartingtime)))
            self.statusLeft.setText("Denoising...")
            # Note: dialog returns all possible parameters
            if not self.DOC:
                [alg, depth, thrType, thr,wavelet,start,end,width,aaRec,aaWP] = self.denoiseDialog.getValues()
            else:
                wavelet = "dmey2"
                [alg, start, end, width] = self.denoiseDialog.getValues()
            self.backup()

            self.waveletDenoiser = WaveletFunctions.WaveletFunctions(data=self.audiodata, wavelet=wavelet, maxLevel=self.config['maxSearchDepth'], samplerate=self.sampleRate)

            if str(alg) == "Wavelets":
                if not self.DOC:
                    # pass dialog settings
                    self.audiodata = self.waveletDenoiser.waveletDenoise(thrType,float(str(thr)), depth, aaRec=aaRec, aaWP=aaWP)
                else:
                    # go with defaults
                    self.audiodata = self.waveletDenoiser.waveletDenoise(aaRec=True, aaWP=False)
                # here we override default 0-Fs/2 returns
                start = self.minFreqShow
                end = self.maxFreqShow

            elif str(alg) == "Bandpass --> Wavelets" and not self.DOC:
                if depth==0:
                    depth = None # why not let auto-fit?
                self.audiodata = self.sp.bandpassFilter(self.audiodata,start=start,end=end)
                self.audiodata = self.waveletDenoiser.waveletDenoise(thrType,float(str(thr)),depth, aaRec=aaRec, aaWP=aaWP)
            elif str(alg) == "Wavelets --> Bandpass" and not self.DOC:
                if depth==0:
                    depth = None # why not let auto-fit?
                self.audiodata = self.waveletDenoiser.waveletDenoise(thrType,float(str(thr)),depth, aaRec=aaRec, aaWP=aaWP)
                self.audiodata = self.sp.bandpassFilter(self.audiodata,self.sampleRate,start=start,end=end)

            elif str(alg) == "Bandpass":
                self.audiodata = self.sp.bandpassFilter(self.audiodata,self.sampleRate, start=start, end=end)
            elif str(alg) == "Butterworth Bandpass":
                self.audiodata = self.sp.ButterworthBandpass(self.audiodata, self.sampleRate, low=start, high=end)
            else:
                #"Median Filter"
                self.audiodata = self.sp.medianFilter(self.audiodata,int(str(width)))

            print("Denoising calculations completed in %.4f seconds" % (time.time() - opstartingtime))

            sgRaw = self.sp.spectrogram(self.audiodata,mean_normalise=self.sgMeanNormalise,equal_loudness=self.sgEqualLoudness,onesided=self.sgOneSided,multitaper=self.sgMultitaper)
            maxsg = np.min(sgRaw)
            self.sg = np.abs(np.where(sgRaw==0,0.0,10.0 * np.log10(sgRaw/maxsg)))

            self.amplPlot.setData(np.linspace(0.0,self.datalength/self.sampleRate,num=self.datalength,endpoint=True),self.audiodata)

            # Update the frequency axis
            self.redoFreqAxis(start, end, store=False)

            if hasattr(self,'spectrogramDialog'):
                self.spectrogramDialog.setValues(self.minFreq,self.maxFreq,self.minFreqShow,self.maxFreqShow)

            self.setColourLevels()

            print("Denoising completed in %s seconds" % round(time.time() - opstartingtime, 4))
            self.statusLeft.setText("Ready")

    def denoise_undo(self):
        """ Listener for undo button in denoising dialog.
        """
        print("Undoing",np.shape(self.audiodata_backup))
        if hasattr(self,'audiodata_backup'):
            if self.audiodata_backup is not None:
                if np.shape(self.audiodata_backup)[1]>0:
                    self.audiodata = np.copy(self.audiodata_backup[:,-1])
                    self.audiodata_backup = self.audiodata_backup[:,:-1]
                    self.sp.setNewData(self.audiodata,self.sampleRate)
                    sgRaw = self.sp.spectrogram(self.audiodata,mean_normalise=self.sgMeanNormalise,equal_loudness=self.sgEqualLoudness,onesided=self.sgOneSided,multitaper=self.sgMultitaper)
                    maxsg = np.min(sgRaw)
                    self.sg = np.abs(np.where(sgRaw == 0, 0.0, 10.0 * np.log10(sgRaw / maxsg)))
                    self.amplPlot.setData(
                        np.linspace(0.0, self.datalengthSec, num=self.datalength, endpoint=True),
                        self.audiodata)
                    if hasattr(self,'seg'):
                        self.seg.setNewData(self.audiodata,sgRaw,self.sampleRate,self.config['window_width'],self.config['incr'])

                    if hasattr(self, 'showFreq_backup'):
                        self.redoFreqAxis(self.showFreq_backup[0], self.showFreq_backup[1])
                    else:
                        self.redoFreqAxis(self.minFreq, self.maxFreq)
                    self.setColourLevels()

    def denoise_save(self):
        """ Listener for save button in denoising dialog.
        Adds _d to the filename and saves it as a new sound file.
        """
        filename = self.filename[:-4] + '_d' + self.filename[-4:]
        wavio.write(filename,self.audiodata.astype('int16'),self.sampleRate,scale='dtype-limits', sampwidth=2)
        self.statusLeft.setText("Saved")
        msg = SupportClasses.MessagePopup("d", "Saved", "Destination: " + '\n' + filename)
        msg.exec_()
        return

    def save_selected_sound(self, id=-1):
        """ Listener for 'Save selected sound' menu item.
        choose destination and give it a name
        """
        if self.box1id is None or self.box1id == -1:
            print("No box selected")
            msg = SupportClasses.MessagePopup("w", "No segment", "No sound selected to save")
            msg.exec_()
            return
        else:
            if type(self.listRectanglesa2[self.box1id]) == self.ROItype:
                x1 = self.listRectanglesa2[self.box1id].pos().x()
                x2 = x1 + self.listRectanglesa2[self.box1id].size().x()
            else:
                x1, x2 = self.listRectanglesa2[self.box1id].getRegion()
            x1 = math.floor(x1 * self.config['incr']) #/ self.sampleRate
            x2 = math.floor(x2 * self.config['incr']) #/ self.sampleRate
            filename, drop = QFileDialog.getSaveFileName(self, 'Save File as', self.SoundFileDir, '*.wav')
            if filename:
                wavio.write(str(filename), self.audiodata[int(x1):int(x2)].astype('int16'), self.sampleRate, scale='dtype-limits', sampwidth=2)
            # update the file list box
            self.fillFileList(os.path.basename(self.filename))

    def redoFreqAxis(self,start,end, store=True):
        """ This is the listener for the menu option to make the frequency axis tight (after bandpass filtering or just spectrogram changes)
            On the same go updates spectrogram and overview plots.
                store: boolean, indicates whether changes should be stored in the config
        """
        changedY = (start!=self.minFreqShow or end!=self.maxFreqShow)
        # Lots of updating can be avoided if the Y freqs aren't changing:
        if changedY:
            self.minFreqShow = max(start,self.minFreq)
            self.maxFreqShow = min(end,self.maxFreq)
            changedY = True

            if store:
                self.config['minFreq'] = start
                self.config['maxFreq'] = end

        # draw a spectrogram of proper height:
        height = self.sampleRate // 2 / np.shape(self.sg)[1]
        pixelstart = int(self.minFreqShow/height)
        pixelend = int(self.maxFreqShow/height)

        self.overviewImage.setImage(self.sg[:,pixelstart:pixelend])
        self.specPlot.setImage(self.sg[:,pixelstart:pixelend])

        # if Y freqs changed, some segments may appear/be dropped:
        if changedY:
            # Remove everything and redraw it
            self.removeSegments(delete=False)
            self.drawOverview()
            self.drawfigMain(remaking=True)

        QApplication.processEvents()

    def trainWaveletDialog(self):
        """ Create the wavelet training dialog for the relevant menu item
        """
        self.saveSegments()
        self.waveletTDialog = Dialogs.WaveletTrain(np.max(self.audiodata))
        self.waveletTDialog.show()
        self.waveletTDialog.activateWindow()
        self.dName=None
        self.waveletTDialog.browse.clicked.connect(self.browseTrainData)
        self.waveletTDialog.genGT.clicked.connect(self.prepareTrainData)
        self.waveletTDialog.train.clicked.connect(self.trainWavelet)
        self.waveletTDialog.browseTest.clicked.connect(self.browseTestData)
        self.waveletTDialog.test.clicked.connect(self.testWavelet)

    def testWavelet(self):
        if not hasattr(self, 'dNameTest'):
            msg = SupportClasses.MessagePopup("w", "Testing data", "Please specify testing data")
            msg.exec_()
            return
        if not hasattr(self, 'species'):
            msg = SupportClasses.MessagePopup("w", "Train first", "Please train the detector first!")
            msg.exec_()
            return

        with pg.BusyCursor():
            ind = self.species.find('>')
            if ind != -1:
                species = self.species.replace('>', '(')
                species = species + ')'
            else:
                species = self.species
            speciesData = json.load(open(os.path.join(self.filtersDir, species + '.txt')))
            ws = WaveletSegment.WaveletSegment(speciesData, 'dmey2')
            # Virginia: added window and increment as input. Window and inc are supposed to be in seconds
            window = 1
            inc = None
            Segments, TP, FP, TN, FN, = ws.waveletSegment_test(dirName=self.dNameTest, d=False, f=True, rf=True, withzeros=True, learnMode='recaa', savedetections=True, window=window, inc=inc)
            print('--Test summary--\n%d %d %d %d' %(TP, FP, TN, FN))
            if TP+FN != 0:
                recall = TP/(TP+FN)
            else:
                recall = 0
            if TP+FP != 0:
                precision = TP/(TP+FP)
            else:
                precision = 0
            if TN+FP != 0:
                specificity = TN/(TN+FP)
            else:
                specificity = 0
            if TP+FP+TN+FN != 0:
                accuracy = (TP+TN)/(TP+FP+TN+FN)
                self.waveletTDialog.note_step3.setText(' Detection summary:TPR:%.2f%% -- FPR:%.2f%%\n\t\t  Recall:%.2f%%\n\t\t  Precision:%.2f%%\n\t\t  Specificity:%.2f%%\n\t\t  Accuracy:%.2f%%' % (recall*100, 100-specificity*100, recall*100, precision*100, specificity*100, accuracy*100))

    def trainWavelet(self):
        """ Listener for the wavelet training dialog.
        """
        self.species = str(self.waveletTDialog.species.currentText()).title()
        minLen = float(self.waveletTDialog.minlen.text())
        maxLen = float(self.waveletTDialog.maxlen.text())
        minFrq = int(self.waveletTDialog.fLow.value())
        maxFrq = int(self.waveletTDialog.fHigh.value())
        fs = int(self.waveletTDialog.fs.value())
        if self.waveletTDialog.wind.checkState() == 0:
            wind = False
        else:
            wind = True
        if self.waveletTDialog.rain.checkState() == 0:
            rain = False
        else:
            rain = True
        if self.waveletTDialog.ff.checkState() == 0:
            ff = False
        else:
            ff = True
        # print("wind, rain, ff:", wind, rain, ff)
        speciesData = {'Name': self.species, 'SampleRate': fs, 'TimeRange': [minLen,maxLen], 'FreqRange': [minFrq, maxFrq]}
        ws = WaveletSegment.WaveletSegment(speciesData)
        # calculate f0_low and f0_high from GT
        if ff:
            f0_low = []     # int(self.waveletTDialog.f0Low.text())
            f0_high = []    # int(self.waveletTDialog.f0High.text())
            for root, dirs, files in os.walk(str(self.dName)):
                for file in files:
                    if file.endswith('.wav') and os.stat(root + '/' + file).st_size != 0 and file[:-4] + '-sec.txt' in files and file + '.data' in files:
                        wavFile = root + '/' + file[:-4]
                        datFile = root + '/' + file + '.data'
                        # calculate f0_low and f0_high from GT
                        if os.path.isfile(datFile):
                            with open(datFile) as f:
                                segments = json.load(f)
                            for seg in segments:
                                if seg[0] == -1:
                                    continue

                                if type(seg[4]) is not list:
                                    seg[4] = [seg[4]]

                                if seg[4][0].title() == self.species:
                                    secs = seg[1] - seg[0]
                                    wavobj = wavio.read(wavFile+'.wav', nseconds=secs, offset=seg[0])
                                    data = wavobj.data
                                    if np.shape(np.shape(data))[0] > 1:
                                        data = data[:, 0]
                                    sampleRate = wavobj.rate
                                    if data is not 'float':
                                        data = data.astype('float')
                                    if fs != sampleRate:
                                        data = librosa.core.audio.resample(data, sampleRate, fs)
                                    f0_l, f0_h = self.ff(data, speciesData)
                                    if f0_l != 0 and f0_h != 0:
                                        f0_low.append(f0_l)
                                        f0_high.append(f0_h)
            if len(f0_low) > 0 and len(f0_high) > 0:
                f0_low = np.min(f0_low)
                f0_high = np.max(f0_high)
            else:
                # user to enter?
                f0_low = minFrq
                f0_high = maxFrq

        # Get detection measures over all M,thr combinations
        with pg.BusyCursor():
            opstartingtime = time.time()
            speciesData = {'Name': self.species, 'SampleRate': fs, 'TimeRange': [minLen, maxLen],
                           'FreqRange': [minFrq, maxFrq]}
            # returns 2d lists of nodes over M x thr, or stats over M x thr
            #thrList = np.linspace(0.1, 1, num=self.waveletTDialog.setthr.value()) Virginia test to finde inconsistency
            thrList = np.linspace(0, 1, num=self.waveletTDialog.setthr.value())
            MList = np.linspace(0.25, 1.5, num=self.waveletTDialog.setM.value())
            # options for training are: recsep (old), recmulti (joint reconstruction), ethr (threshold energies), elearn (model from energies)
            # Virginia: added window and increment as input. Window and inc are supposed to be in seconds
            window = 1
            inc = None
            nodes, TP, FP, TN, FN = ws.waveletSegment_train(self.dName, thrList, MList, d=False,
                                                            f=True, rf=True, learnMode="recaa", window=window, inc=inc)
            print("Filtered nodes: ", nodes)
            print("TRAINING COMPLETED IN ", time.time() - opstartingtime)

            TPR = TP/(TP+FN)
            FPR = 1 - TN/(FP+TN)
            print("TP rate: ", TPR)
            print("FP rate: ", FPR)

        # Plot AUC and let the user to choose threshold and M
        self.thr = 0.5  # default, get updated when user double-clicks on ROC curve
        self.M = 0.25  # default, get updated when user double-clicks on ROC curve
        self.optimumNodesSel = []

        plt.style.use('ggplot')
        valid_markers = ([item[0] for item in mks.MarkerStyle.markers.items() if
                          item[1] is not 'nothing' and not item[1].startswith('tick') and not item[1].startswith(
                              'caret')])
        markers = np.random.choice(valid_markers, len(MList), replace=False)
        fig, ax = plt.subplots()
        for i in range(len(MList)):
            # each line - different M (rows of result arrays)
            ax.plot(FPR[i], TPR[i], marker=markers[i], label='M='+str(MList[i]))
        ax.set_title('Double click to choose TPR and FPR and set tolerance')
        ax.set_xlabel('False Positive Rate (FPR)')
        ax.set_ylabel('True Positive Rate (TPR)')
        fig.canvas.set_window_title('ROC Curve - %s' % (self.species))
        ax.set_ybound(0, 1)
        ax.set_xbound(0, 1)
        ax.yaxis.set_major_formatter(mtick.PercentFormatter(1, 0))
        ax.xaxis.set_major_formatter(mtick.PercentFormatter(1, 0))
        ax.legend()
        # plt.get_current_fig_manager().window.setWindowIcon(QtGui.QIcon('img/Avianz.ico'))
        def onclick(event):
            if event.dblclick:
                fpr_cl = event.xdata
                tpr_cl = event.ydata
                print("fpr_cl, tpr_cl: ",fpr_cl, tpr_cl)

                if tpr_cl is not None and fpr_cl is not None:
                    # TODO: Interpolate?, currently get the closest point
                    # get M and thr for closest point
                    distarr = (tpr_cl - TPR)**2 + (fpr_cl - FPR)**2
                    M_min_ind, thr_min_ind = np.unravel_index(np.argmin(distarr), distarr.shape)
                    tpr_cl = TPR[M_min_ind, thr_min_ind]
                    fpr_cl = FPR[M_min_ind, thr_min_ind]
                    msg = SupportClasses.MessagePopup("t", 'Set Tolerance - %s' % (self.species), 'Confirm %d%% Sensitivity with %d%% FPR?' % (tpr_cl*100, fpr_cl*100))
                    msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
                    reply = msg.exec_()
                    if reply == QMessageBox.Yes:
                        self.M = MList[M_min_ind]
                        self.thr = thrList[thr_min_ind]
                        # Get nodes for closest point
                        self.optimumNodesSel = nodes[M_min_ind][thr_min_ind]

                        # plt.close()   # If the user changed mind, let them choose another point without repeating heavy work
                        speciesData['Wind'] = wind
                        speciesData['Rain'] = rain
                        speciesData['F0'] = ff
                        if ff:
                            speciesData['F0Range'] = [f0_low, f0_high]
                        speciesData['WaveletParams'] = []
                        speciesData['WaveletParams'].append(self.thr)
                        speciesData['WaveletParams'].append(self.M)
                        speciesData['WaveletParams'].append(self.optimumNodesSel)

                        ind = self.species.find('>')
                        if ind != -1:
                            species = self.species.replace('>', '(')
                            species = species + ')'
                        else:
                            species = self.species
                        filename = os.path.join(self.filtersDir, species + '.txt')
                        if os.path.isfile(filename):
                            # Add it to the Filter list
                            msg = SupportClasses.MessagePopup("t", "Save Filter", 'Are you sure you want to Overwrite the existing filter\nfor %s?' %(species))
                            msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
                            reply = msg.exec_()
                            if reply == QMessageBox.Yes:
                                print("Saving new filter to ", filename)
                                f = open(filename, 'w')
                                f.write(json.dumps(speciesData))
                                f.close()
                                msg = SupportClasses.MessagePopup("d", "Training completed!", 'Training completed!\nFollow Step 3 and test on a separate dataset before actual use.')
                                msg.exec_()
                                self.FilterFiles.append(self.species)
                                self.waveletTDialog.test.setEnabled(True)
                            else:
                                # TODO: decide how to save the new filter when you don't want to overwrite
                                # previous filter (or forgot to rename previous filter before coming to this
                                # point and no need to loose all the heavy work done with grid search
                                filename = filename[:-4] + '_new.txt'
                                print("Saving new filter to ", filename[:-4] + '_new.txt', " (have to rename new filter)")
                                f = open(filename, 'w')
                                f.write(json.dumps(speciesData))
                                f.close()
                                # Add it to the Filter list
                                msg = SupportClasses.MessagePopup("d", "Training completed!", "Training completed! (but the filter saved with a new name)\nFollow Step 3 and test on a separate dataset before actual use.")
                                msg.exec_()
                                self.FilterFiles.append(self.species)
                                self.waveletTDialog.test.setEnabled(True)
                        else:
                            print("Saving new filter to ", filename)
                            f = open(filename, 'w')
                            f.write(json.dumps(speciesData))
                            f.close()
                            # Add it to the Filter list
                            msg = SupportClasses.MessagePopup("d", "Training completed!", 'Training completed!\nFollow Step 3 and test on a separate dataset before actual use.')
                            msg.exec_()
                            self.FilterFiles.append(self.species)
                            self.waveletTDialog.test.setEnabled(True)
        cid = fig.canvas.mpl_connect('button_press_event', onclick)
        plt.show()
        # plt.raise_()

    def ff(self, data, speciesData):
        # TODO: fs in speciesData could not be the actual fs of the audio, does it matter?
        sc = SupportClasses.preProcess(audioData=data, spInfo=speciesData, d=True, f=False)  # avoid bandpass filter
        data, sampleRate = sc.denoise_filter(level=8)
        sp = SignalProc.SignalProc([], 0, 256, 128) #SignalProc.SignalProc([], 0, 512, 256)
        sgRaw = sp.spectrogram(data, 256, 128, mean_normalise=True, onesided=True, multitaper=False)
        segment = Segment.Segment(data, sgRaw, sp, sampleRate, 256, 128)
        pitch, y, minfreq, W = segment.yin(minfreq=100)
        ind = np.squeeze(np.where(pitch > minfreq))
        pitch = pitch[ind]
        if pitch.size == 0:
            return 0, 0
        if ind.size < 2:
            f0 = pitch
            return f0, f0
        else:
            return round(np.min(pitch)), round(np.max(pitch))

    def prepareTrainData(self):
        """ Listener for the wavelet training dialog.
        """
        # Virginia: added window and increment to prepare annotation file
        # BE CAREFULL: they must be updated here
        window = 1
        inc= None
        species = self.waveletTDialog.species.currentText()
        #species = "Morepork" #Virginia: changed to test
        if species == 'Choose species...':
            msg = SupportClasses.MessagePopup("w", "Species Error", "Please specify the species!")
            msg.exec_()
            return
        if self.dName is None or self.dName == '':
            msg = SupportClasses.MessagePopup("w", "Data Error", "Please specify training data!")
            msg.exec_()
            return
        f_low = []
        f_high = []
        fs =[]
        len_min = []
        len_max = []
        for root, dirs, files in os.walk(str(self.dName)):
            for file in files:
                if file.endswith('.wav') and os.stat(root + '/' + file).st_size != 0 and file + '.data' in files:
                    wavFile = root + '/' + file
                    metaData = self.annotation2GT_OvWin(wavFile, species, window=window, inc=inc)
                    len_min.append(metaData[0])
                    len_max.append(metaData[1])
                    f_low.append(metaData[2])
                    f_high.append(metaData[3])
                    fs.append(metaData[4])
        self.waveletTDialog.minlen.setText(str(round(np.min(len_min),2)))
        self.waveletTDialog.maxlen.setText(str(round(np.max(len_max),2)))
        self.waveletTDialog.fLow.setRange(0, int(np.min(fs))/2)
        self.waveletTDialog.fLow.setValue(int(np.min(f_low)))
        self.waveletTDialog.fHigh.setRange(0, int(np.min(fs))/2)
        self.waveletTDialog.fHigh.setValue(int(np.max(f_high)))
        self.waveletTDialog.fs.setValue(int(np.min(fs)))
        self.waveletTDialog.fs.setRange(0, int(np.min(fs)))
        self.waveletTDialog.note_step2.setText('Above fields propagated using training data.\nAdjust if required.')
        self.waveletTDialog.train.setEnabled(True)        

        msg = SupportClasses.MessagePopup("d", "Preparation done!", "Follow Step 2 to complete training.")
        msg.exec_()
        return

    def annotation2GT(self, wavFile, species, duration=0):
        """
        This generates the ground truth for a given sound file
        Given the AviaNZ annotation, returns the ground truth as a txt file
        """
        # TODO: Allow empty files (no target calls) when testing but not when training - a flag
        #       Species combobox should compatible with labels in annotations, e.g. show Kiwi (Nth Is Brown) when
        #       variations of it exists Kiwi (Nth Is Brown)(M)1, Kiwi (Nth Is Brown)(M)2 etc.
        datFile = wavFile + '.data'
        eFile = datFile[:-9] + '-1sec.txt'
        if duration == 0:
            wavobj = wavio.read(wavFile)
            sampleRate = wavobj.rate
            data = wavobj.data
            duration = int(np.ceil(len(data) / sampleRate))  # number of secs
        GT = np.zeros((duration, 4))
        GT = GT.tolist()
        GT[:][1] = str(0)
        GT[:][2] = ''
        GT[:][3] = ''
        # fHigh and fLow for text boxes
        fLow = sampleRate/2
        fHigh = 0
        lenMin = duration
        lenMax =0
        if os.path.isfile(datFile):
            # print(datFile)
            with open(datFile) as f:
                segments = json.load(f)
            for seg in segments:
                if seg[0] == -1:
                    continue

                if type(seg[4]) is not list:
                    seg[4] = [seg[4]]

                if species in seg[4]:
                    # print("lenMin, seg[1]-seg[0]", lenMin, seg[1]-seg[0])
                    if lenMin > seg[1]-seg[0]:
                        lenMin = seg[1]-seg[0]
                    if lenMax < seg[1]-seg[0]:
                        lenMax = seg[1]-seg[0]
                    if fLow > seg[2]:
                        fLow = seg[2]
                    if fHigh < seg[3]:
                        fHigh = seg[3]
                    type = species
                    quality = ''
                    s = int(math.floor(seg[0]))
                    e = min(duration, int(math.ceil(seg[1])))
                    #print("start and end: ", s, e)
                    for i in range(s, e):
                        GT[i][1] = str(1)
                        GT[i][2] = type
                        GT[i][3] = quality

        # Empty files cannot be used now, and lead to problems
        if len(GT)==0:
            print("ERROR: no calls for this species in file", datFile)
            return

        for line in GT:
            if line[1] == 0.0:
                line[1] = '0'
            if line[2] == 0.0:
                line[2] = ''
            if line[3] == 0.0:
                line[3] = ''
        # now save GT as a .txt file
        for i in range(1, duration + 1):
            GT[i - 1][0] = str(i)  # add time as the first column to make GT readable
        # strings = (str(item) for item in GT)
        with open(eFile, "w") as f:
            for l, el in enumerate(GT):
                string = '\t'.join(map(str, el))
                for item in string:
                    f.write(item)
                f.write('\n')
            f.write('\n')
        # print(lenMin, lenMax, fLow, fHigh, sampleRate)
        return [lenMin, lenMax, fLow, fHigh, sampleRate]

    def annotation2GT_OvWin(self, wavFile, species, duration=0, window=1, inc=None):
        """
        This generates the ground truth for a given sound file
        Given the AviaNZ annotation, returns the ground truth as a txt file
        """
        # Virginia:now it generate a text file for overlapping window
        if inc == None:
            inc = window
            resol = window
        else:
            # Virginia: resolution is the "gcd" between window and inc. In this way I'm hoping to solve the case with
            # 75% overlap
            # maybe it is not useful
            resol = (math.gcd(int(100 * window), int(100 * inc))) / 100

        datFile = wavFile + '.data'
        eFile = datFile[:-9] + '-res' + str(float(resol)) + 'sec.txt'

        if duration == 0:
            wavobj = wavio.read(wavFile)
            sampleRate = wavobj.rate
            # Virginia: resolution length in samples
            res_sr = resol * sampleRate
            data = wavobj.data
            duration = int(np.ceil(len(data) / res_sr))
            # number of expected segments = number of center at inc distance

        GT = np.zeros((duration, 4))
        GT = GT.tolist()
        GT[:][1] = str(0)
        GT[:][2] = ''
        GT[:][3] = ''
        # fHigh and fLow for text boxes
        fLow = sampleRate / 2
        fHigh = 0
        lenMin = duration
        lenMax = 0
        if os.path.isfile(datFile):
            # print(datFile)
            with open(datFile) as f:
                segments = json.load(f)
            for seg in segments:
                if seg[0] == -1:
                    continue

                if type(seg[4]) is not list:
                    seg[4] = [seg[4]]

                if species in seg[4]:
                    # print("lenMin, seg[1]-seg[0]", lenMin, seg[1]-seg[0])
                    # Virginia: added this variable so the machine don't have to calculate it every rime
                    dur_segm = seg[1] - seg[0]
                    if lenMin > dur_segm:
                        lenMin = dur_segm
                    if lenMax < dur_segm:
                        lenMax = dur_segm
                    if fLow > seg[2]:
                        fLow = seg[2]
                    if fHigh < seg[3]:
                        fHigh = seg[3]
                    typeOfAnnot = species
                    quality = ''
                    # Virginia: start and end must be read in resol base
                    s = int(math.floor(seg[0] / resol))
                    e = int(math.ceil(seg[1] / resol))
                    # Virginia: start and end printed in seconds
                    print("start and end: ", s*resol, e*resol)
                    # Virginia: I'm adding the segments that have this interval in their window
                    # NB: if resol=window step2s=step2e=0
                    for i in range(s, e):
                        GT[i][1] = str(1)
                        GT[i][2] = typeOfAnnot
                        GT[i][3] = quality

                        # Empty files cannot be used now, and lead to problems

        for line in GT:
            if line[1] == 0.0:
                line[1] = '0'
            if line[2] == 0.0:
                line[2] = ''
            if line[3] == 0.0:
                line[3] = ''
                # now save GT as a .txt file
        for i in range(1, duration + 1):
            GT[i - 1][0] = str(i * resol)  # add time as the first column to make GT readable
        # strings = (str(item) for item in GT)
        with open(eFile, "w") as f:
            for l, el in enumerate(GT):
                string = '\t'.join(map(str, el))
                for item in string:
                    f.write(item)
                f.write('\n')
            f.write('\n')
        print(eFile)
        #print(lenMin, lenMax, fLow, fHigh, sampleRate)
        return [lenMin, lenMax, fLow, fHigh, sampleRate]

    def browseTrainData(self):
        """ Listener for the wavelet training dialog.
        """
        self.dName = QtGui.QFileDialog.getExistingDirectory(self, 'Choose Folder to Process')
        # get the species list from annotations
        self.waveletTDialog.fillFileList(self.dName)
        self.waveletTDialog.genGT.setEnabled(True)
        self.waveletTDialog.raise_()

    def browseTestData(self):
        """ Listener for the wavelet training dialog.
        """
        self.waveletTDialog.note_step3.clear()
        self.dNameTest = QtGui.QFileDialog.getExistingDirectory(self, 'Choose Folder to Test')
        self.waveletTDialog.fillFileList(self.dNameTest, False)
        self.waveletTDialog.test.setEnabled(True)
        self.waveletTDialog.raise_()

    def segmentationDialog(self):
        """ Create the segmentation dialog when the relevant button is pressed.
        """
        self.segmentDialog = Dialogs.Segmentation(np.max(self.audiodata),DOC=self.DOC, species=self.FilterFiles)
        self.segmentDialog.show()
        self.segmentDialog.activateWindow()
        self.segmentDialog.undo.clicked.connect(self.segment_undo)
        self.segmentDialog.activate.clicked.connect(self.segment)

    def segment(self):
        """ Listener for the segmentation dialog. Calls the relevant segmenter.
        """
        if self.CLI:
            self.segmentDialog = Dialogs.Segmentation(np.max(self.audiodata))

        opstartingtime = time.time()
        print('Segmenting requested at ' + time.strftime('%H:%M:%S', time.gmtime(opstartingtime)))
        # for undoing:
        self.prevSegments = copy.deepcopy(self.segments)

        self.segmentsToSave = True
        [alg, medThr,HarmaThr1,HarmaThr2,PowerThr,minfreq,minperiods,Yinthr,window,FIRThr1,CCThr1,species,resolution,species_cc] = self.segmentDialog.getValues()
        with pg.BusyCursor():
            species = str(species)
            self.statusLeft.setText('Segmenting...')
            # Delete old segments:
            # only this species, if using species-specific methods:
            if alg == 'Wavelets':
                if species == 'Choose species...':
                    msg = SupportClasses.MessagePopup("w", "Species Error", 'Please select your species!')
                    msg.exec_()
                    return

                # deleting from the end, because deleteSegments shifts IDs:
                for si in reversed(range(len(self.segments))):
                    if species in self.segments[si][4] or species+'?' in self.segments[si][4]:
                        self.deleteSegment(si)
            else:
                self.removeSegments()

            # NON-SPECIFIC methods here (produce "Don't Know"):
            if str(alg) == 'Default':
                newSegments = self.seg.bestSegments()
            elif str(alg) == 'Median Clipping':
                newSegments = self.seg.medianClip(float(str(medThr)), minSegment=self.config['minSegment'])
                newSegments = self.seg.checkSegmentOverlap(newSegments, minSegment=self.config['minSegment'])
            elif str(alg) == 'Harma':
                newSegments = self.seg.Harma(float(str(HarmaThr1)),float(str(HarmaThr2)),minSegment=self.config['minSegment'])
                newSegments = self.seg.checkSegmentOverlap(newSegments, minSegment=self.config['minSegment'])
            elif str(alg) == 'Power':
                newSegments = self.seg.segmentByPower(float(str(PowerThr)))
                newSegments = self.seg.checkSegmentOverlap(newSegments, minSegment=self.config['minSegment'])
            elif str(alg) == 'Onsets':
                newSegments = self.seg.onsets()
                newSegments = self.seg.checkSegmentOverlap(newSegments, minSegment=self.config['minSegment'])
            elif str(alg) == 'Fundamental Frequency':
                newSegments, pitch, times = self.seg.yin(int(str(minfreq)),int(str(minperiods)),float(str(Yinthr)),int(str(window)),returnSegs=True)
                newSegments = self.seg.checkSegmentOverlap(newSegments, minSegment=self.config['minSegment'])
            elif str(alg) == 'FIR':
                newSegments = self.seg.segmentByFIR(float(str(FIRThr1)))
                newSegments = self.seg.checkSegmentOverlap(newSegments, minSegment=self.config['minSegment'])
            # SPECIES-SPECIFIC methods from here:
            elif str(alg) == 'Wavelets':
                speciesData = json.load(open(os.path.join(self.filtersDir, species+'.txt')))
                ws = WaveletSegment.WaveletSegment(speciesData)
                newSegments = ws.waveletSegment(data=self.audiodata, sampleRate=self.sampleRate,
                                                d=False, f=True, wpmode="new")
            # TODO
            elif str(alg) == 'Cross-Correlation':
                if species_cc != 'Choose species...':
                    # need to load template/s
                    newSegments = self.findMatches(float(str(CCThr1)), species_cc)
                else:
                    newSegments = self.findMatches(float(str(CCThr1)))

            print('Segments: ', newSegments)

            # post process to remove short segments, wind, rain, and use F0 check.
            if species == 'All species' and species_cc == 'Choose species...' or str(alg) == 'Default' or str(alg) == 'Median Clipping' or str(alg) == 'Harma' or str(alg) == 'Power' or str(alg) == 'Onsets' or str(alg) == 'Fundamental Frequency' or str(alg) == 'FIR':
                post = SupportClasses.postProcess(audioData=self.audiodata, sampleRate=self.sampleRate, segments=newSegments, spInfo={})
                print(post.segments)
                post.wind(Tmean_wind = 20)
                print('After wind: ', post.segments)
                post.rainClick()
                print('After rain: ', post.segments)
            else:
                post = SupportClasses.postProcess(audioData=self.audiodata, sampleRate=self.sampleRate,
                                                  segments=newSegments, spInfo=speciesData)
                post.short()  #TODO: keep 'deleteShort' in filter file?
                if speciesData['Wind']:
                    # post.wind() - omitted in sppSpecific=T cases
                    print('After wind: ', post.segments)
                if speciesData['Rain']:
                    # post.rainClick() - omitted in sppSpecific=T cases
                    print('After rain: ', post.segments)
                if speciesData['F0']:
                    post.fundamentalFrq(self.filename, speciesData)
                    print('After ff: ', post.segments)

            newSegments = post.segments
            print("After post processing: ", newSegments)

            # Generate annotation friendly output.
            if str(alg)=='Wavelets':
                 if len(newSegments)>0:
                    y1 = self.convertFreqtoY(speciesData['FreqRange'][0]/2)
                    y2 = self.convertFreqtoY(speciesData['SampleRate']/2)
                    if speciesData['SampleRate']/2 > self.sampleRate:
                        y2 = self.convertFreqtoY(self.sampleRate/2-self.sampleRate*0.01)
                    for seg in newSegments:
                        self.addSegment(float(seg[0]), float(seg[1]), y1, y2,
                                        [species.title() + "?"],index=-1)
                        self.segmentsToSave = True
            elif str(alg)=='Cross-Correlation' and species_cc != 'Choose species...':
                if len(newSegments) > 0:
                    y1 = self.convertFreqtoY(speciesData['FreqRange'][0]/2)
                    y2 = self.convertFreqtoY(speciesData['SampleRate']/2)
                    if speciesData['SampleRate']/2 > self.sampleRate:
                        y2 = self.convertFreqtoY(self.sampleRate / 2 - self.sampleRate * 0.01)
                    for seg in newSegments:
                        self.addSegment(float(seg[0]), float(seg[1]), y1, y2,
                                        [species_cc.title() + "?"], index=-1)
                        self.segmentsToSave = True
            else:
                if len(newSegments)>0:
                    for seg in newSegments:
                        self.addSegment(seg[0],seg[1])
                        self.segmentsToSave = True

            self.lenNewSegments = len(newSegments)
            self.segmentDialog.undo.setEnabled(True)
            self.statusLeft.setText('Ready')
        print('Segmentation finished at %s' % (time.time() - opstartingtime))

    def segment_undo(self):
        """ Listener for undo button in segmentation dialog.
            Deletes everything, and re-adds segments from a backup.
        """
        # just in case:
        self.segmentDialog.undo.setEnabled(False)
        if not hasattr(self, 'prevSegments'):
            print("Nothing to undo!")
            return

        self.removeSegments()
        for seg in self.prevSegments:
            if seg[2] == 0 and seg[3] == 0:
                self.addSegment(seg[0], seg[1],0,0,seg[4], index=-1)
            else:
                self.addSegment(seg[0], seg[1],self.convertFreqtoY(seg[2]),self.convertFreqtoY(seg[3]),seg[4], index=-1)
            self.segmentsToSave = True

    def exportSeg(self, annotation=None):
        # find all the species
        species = set()
        species.add("All species")
        for seg in self.segments:
            for birdName in seg[4]:
                if birdName.endswith('?'):
                    species.add(birdName[:-1])
                else:
                    species.add(birdName)
        species = list(species)
        for root, dirs, files in os.walk(str(self.SoundFileDir)):
            for file in files:
                file = os.path.join(root, file)
                if fnmatch.fnmatch(file, self.filename[:-4] + "_*.xlsx"):
                    print("Removing file %s" % file)
                    os.remove(file)
        # excel should be split by page size, but for short files just give the file size
        datalen = self.config['maxFileShow'] if self.nFileSections>1 else self.datalengthSec
        out = SupportClasses.exportSegments(self.SoundFileDir, self.filename, datalen, self.segments, resolution=10, numpages=self.nFileSections, species=species)
        success = out.excel()
        # add user notification
        if success==0:
            print("Warning: xlsx output was not saved")
            return
        else:
            msg = SupportClasses.MessagePopup("d", "Segments Exported", "Check this directory for the excel output: " + '\n' + self.SoundFileDir)
            msg.exec_()
            return

    def findMatches(self,thr=0.4, species='Choose species...'):
        """ Calls the cross-correlation function to find matches like the currently highlighted box.
        It also check if you have selected a species, then allow to read those templates and match.
        """
        # print ("inside find Matches: ", species)
        segments = []
        if species != 'Choose species...' and os.path.exists('Sound Files/' + species):
            self.statusLeft.setText("Finding matches...")
            print("Reading template/s")
            # Todo: do more than one template and merge result?
            wavobj = wavio.read('Sound Files/'+species+'/train1_1.wav')

            # Parse wav format details based on file header:
            sampleRate = wavobj.rate
            audiodata = wavobj.data
            # minFreq = 0
            # maxFreq = self.sampleRate / 2.
            # fileLength = wavobj.nframes

            if audiodata.dtype is not 'float':
                audiodata = audiodata.astype('float')  # / 32768.0

            if np.shape(np.shape(audiodata))[0] > 1:
                audiodata = audiodata[:, 0]
            # downsample
            print("fs: ", sampleRate, self.sppInfo[str(species)][4])
            if sampleRate != self.sppInfo[str(species)][4]:
                audiodata = librosa.core.audio.resample(audiodata, sampleRate, self.sppInfo[str(species)][4])
                sampleRate = self.sppInfo[str(species)][4]
            datalength = np.shape(audiodata)[0]
            len_seg = datalength / sampleRate

            sp_temp = SignalProc.SignalProc([], 0, self.config['window_width'], self.config['incr'])
            sgRaw_temp = sp_temp.spectrogram(audiodata, self.config['window_width'],
                                        self.config['incr'], mean_normalise=self.sgMeanNormalise,
                                        equal_loudness=self.sgEqualLoudness, onesided=self.sgOneSided,
                                        multitaper=self.sgMultitaper)

            # Get the data for the spectrogram
            if self.sampleRate != self.sppInfo[str(species)][4]:
                data1 = librosa.core.audio.resample(self.audiodata, self.sampleRate, self.sppInfo[str(species)][4])
                sampleRate1 = self.sppInfo[str(species)][4]
            else:
                data1 = self.audiodata
                sampleRate1 = self.sampleRate
            sgRaw = self.sp.spectrogram(data1,mean_normalise=self.sgMeanNormalise,equal_loudness=self.sgEqualLoudness,onesided=self.sgOneSided,multitaper=self.sgMultitaper)
            indices = self.seg.findCCMatches(sgRaw_temp,sgRaw,thr)
            # scale indices to match with self.samplerate
            indices = [i*self.sampleRate/sampleRate1 for i in indices]
            # print('indices:', indices)
            # identifySegments(seg=indices, minlength=10)
            # indices are in spectrogram pixels, need to turn into times
            y1 = self.convertFreqtoY(self.sppInfo[str(species)][2]/2)
            if self.sppInfo[str(species)][4]/2 > self.sampleRate:
                y2 = self.convertFreqtoY(self.sampleRate / 2 - self.sampleRate * 0.01)
            else:
                y2 = self.convertFreqtoY(self.sppInfo[str(species)][4] / 2)
            for i in indices:
                if np.abs(i) > self.config['overlap_allowed']:
                    time = i*self.config['incr'] / self.sampleRate
                    # print(time, time + len_seg,self.segments)
                    # self.addSegment(time, time+len_seg,y1,y2,[species+'?'])
                    segments.append([time, time+len_seg])
        elif self.box1id is None or self.box1id == -1:
            print("No box selected")
            msg = SupportClasses.MessagePopup("w", "No segment", "No segment selected to match")
            msg.exec_()
            return []
        else:
            self.statusLeft.setText("Finding matches...")
            # Only want to draw new segments, so find out how many there are now
            seglen = len(self.segments)
            # Get the segment -- note that takes the full y range
            if type(self.listRectanglesa2[self.box1id]) == self.ROItype:
                x1 = self.listRectanglesa2[self.box1id].pos().x()
                x2 = x1 + self.listRectanglesa2[self.box1id].size().x()
            else:
                x1, x2 = self.listRectanglesa2[self.box1id].getRegion()
            # Get the data for the spectrogram
            sgRaw = self.sp.spectrogram(self.audiodata,mean_normalise=self.sgMeanNormalise,equal_loudness=self.sgEqualLoudness,onesided=self.sgOneSided,multitaper=self.sgMultitaper)
            segment = sgRaw[int(x1):int(x2),:]
            len_seg = (x2-x1) * self.config['incr'] / self.sampleRate
            indices = self.seg.findCCMatches(segment,sgRaw,thr)
            # indices are in spectrogram pixels, need to turn into times
            for i in indices:
                # Miss out the one selected: note the hack parameter
                if np.abs(i-x1) > self.config['overlap_allowed']:
                    time = i*self.config['incr'] / self.sampleRate
                    segments.append([time, time+len_seg])
                    # self.addSegment(time, time+len_seg,0,0,self.segments[self.box1id][4])
            self.statusLeft.setText("Ready")
        return segments

    def classifySegments(self):
        # TODO: Finish this
        # Note that this still works on 1 second -- species-specific parameter eventually (here twice: as 1 and in sec loop)
        print("Not implemented yet!")
        return

        if self.segments is None or len(self.segments) == 0:
            msg = SupportClasses.MessagePopup("w", "No segments", "No segments to recognise")
            msg.exec_()
            return
        else:
            with pg.BusyCursor():
                # TODO: Ask for species; brown kiwi for now
                # TODO: ***** TIDY UP WAVELET SEG, USE THIS!
                for i in range(len(self.segments)):
                    seglength = np.abs(self.segments[i][1] - self.segments[i][0])
                    if seglength <= 1:
                        # Recognise as is
                        label = WaveletSegment.computeWaveletEnergy(self.audiodata[self.segments[i][0]:self.segments[i][1]], self.sampleRate)
                        self.updateText(label,i)
                    else:
                        for sec in range(math.ceil(seglength)):
                            label = WaveletSegment.computeWaveletEnergy(self.audiodata[sec*self.sampleRate+self.segments[i][0]:(sec+1)*self.sampleRate+self.segments[i][0]], self.sampleRate)
                            # TODO: Check if the labels match, decide what to do if not
                        self.updateText(label,i)

    def recognise(self):
        # This will eventually call methods to do automatic recognition
        # Actually, will produce a dialog to ask which species, etc.
        # TODO
        pass

# ===============
# Code for playing sounds
    def playVisible(self):
        """ Listener for button to play the visible area.
        On PLAY, turns to PAUSE and two other buttons turn to STOPs.
        """
        if self.media_obj.isPlaying():
            self.pausePlayback()
        else:
            if self.media_obj.state() != QAudio.SuspendedState and not self.media_obj.keepSlider:
                # restart playback
                range = self.p_ampl.viewRange()[0]
                self.setPlaySliderLimits(range[0]*1000, range[1]*1000)
                # (else keep play slider range from before)
            self.playButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaPause))
            self.playSegButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaStop))
            self.playBandLimitedSegButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaStop))
            self.media_obj.pressedPlay(start=self.segmentStart, stop=self.segmentStop, audiodata=self.audiodata)

    def playSelectedSegment(self):
        """ Listener for PlaySegment button.
        Get selected segment start and end (or return if no segment selected).
        On PLAY, all three buttons turn to STOPs.
        """
        if self.media_obj.isPlaying():
            self.stopPlayback()
        else:
            if self.box1id > -1:
                self.stopPlayback()

                start = self.listRectanglesa1[self.box1id].getRegion()[0] * 1000
                stop = self.listRectanglesa1[self.box1id].getRegion()[1] * 1000

                self.setPlaySliderLimits(start, stop)
                self.playButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaPause))
                self.playSegButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaStop))
                self.playBandLimitedSegButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaStop))
                self.media_obj.filterSeg(start, stop, self.audiodata)
            else:
                print("Can't play, no segment selected")

    def playBandLimitedSegment(self):
        """ Listener for PlayBandlimitedSegment button.
        Gets the band limits of the segment, bandpass filters, then plays that.
        Currently uses FIR bandpass filter -- Butterworth is commented out.
        On PLAY, all three buttons turn to STOPs.
        """
        if self.media_obj.isPlaying():
            self.stopPlayback()
        else:
            if self.box1id > -1:
                self.stopPlayback()
                # check frequency limits, + small buffer bands
                bottom = max(0.1, self.minFreq, self.segments[self.box1id][2])
                top = min(self.segments[self.box1id][3], self.maxFreq-0.1)

                print("Extracting samples between %d-%d Hz" % (bottom, top))
                start = self.listRectanglesa1[self.box1id].getRegion()[0] * 1000
                stop = self.listRectanglesa1[self.box1id].getRegion()[1] * 1000
                self.setPlaySliderLimits(start, stop)
                self.playButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaPause))
                self.playSegButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaStop))
                self.playBandLimitedSegButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaStop))

                # filter the data into a temporary file or buffer
                self.media_obj.filterBand(self.segmentStart, self.segmentStop, bottom, top, self.audiodata, self.sp)
            else:
                print("Can't play, no segment selected")

    def pausePlayback(self):
        """ Restores the PLAY buttons, calls media_obj to pause playing."""
        self.media_obj.pressedPause()

        # Reset all button icons:
        self.playButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaPlay))
        self.playSegButton.setIcon(QtGui.QIcon('img/playsegment.png'))
        self.playBandLimitedSegButton.setIcon(QtGui.QIcon('img/playBandLimited.png'))

    def stopPlayback(self):
        """ Restores the PLAY buttons, slider, text, calls media_obj to stop playing."""
        self.media_obj.pressedStop()
        if not hasattr(self, 'segmentStart') or self.segmentStart is None:
            self.segmentStart = 0
        self.playSlider.setValue(-1000)
        self.bar.setValue(-1000)
        self.timePlayed.setText(self.convertMillisecs(self.segmentStart) + "/" + self.totalTime)

        # Reset all button icons:
        self.playButton.setIcon(self.style().standardIcon(QtGui.QStyle.SP_MediaPlay))
        self.playSegButton.setIcon(QtGui.QIcon('img/playsegment.png'))
        self.playBandLimitedSegButton.setIcon(QtGui.QIcon('img/playBandLimited.png'))

    def movePlaySlider(self):
        """ Listener called on sound notify (every 20 ms).
        Controls the slider, text timer, and listens for playback finish.
        """
        eltime = self.media_obj.processedUSecs() // 1000 + self.media_obj.timeoffset
        bufsize = 0.02

        # listener for playback finish. Note small buffer for catching up
        if eltime > (self.segmentStop-10):
            print("Stopped at %d ms" % eltime)
            self.stopPlayback()
        else:
            self.playSlider.setValue(eltime)
            self.timePlayed.setText(self.convertMillisecs(eltime) + "/" + self.totalTime)
            # playSlider.value() is in ms, need to convert this into spectrogram pixels
            self.bar.setValue(self.convertAmpltoSpec(eltime / 1000.0 - bufsize))

    def setPlaySliderLimits(self, start, end):
        """ Uses start/end in ms, does what it says, and also seeks file position marker.
        """
        offset = (self.startRead + self.startTime) * 1000 # in ms, absolute
        self.playSlider.setRange(start + offset, end + offset)
        self.segmentStart = self.playSlider.minimum() - offset # relative to file start
        self.segmentStop = self.playSlider.maximum() - offset # relative to file start

    def volSliderMoved(self, value):
        self.media_obj.applyVolSlider(value)

    def barMoved(self, evt):
        """ Listener for when the bar showing playback position moves.
        """
        self.playSlider.setValue(self.convertSpectoAmpl(evt.x()) * 1000)
        self.media_obj.seekToMs(self.convertSpectoAmpl(evt.x()) * 1000, self.segmentStart)

    def setOperatorReviewerDialog(self):
        """ Listener for Set Operator/Reviewer menu item.
        """
        if hasattr(self, 'operator') and hasattr(self, 'reviewer') :
            self.setOperatorReviewerDialog = Dialogs.OperatorReviewer(operator=self.operator,reviewer=self.reviewer)
        else:
            self.setOperatorReviewerDialog = Dialogs.OperatorReviewer(operator='', reviewer='')
        #self.setOperatorReviewerDialog.activateWindow()
        self.setOperatorReviewerDialog.activate.clicked.connect(self.changeOperator)
        self.setOperatorReviewerDialog.exec()

    def changeOperator(self):
        """ Listener for the operator/reviewer dialog.
        """
        name1, name2 = self.setOperatorReviewerDialog.getValues()
        self.operator = str(name1)
        self.reviewer = str(name2)
        self.statusRight.setText("Operator: " + self.operator + ", Reviewer: "+self.reviewer)
        self.setOperatorReviewerDialog.close()
        self.segmentsToSave = True

    def addNoiseData(self):
        """ Listener for the adding metadata about noise action """
        self.getNoiseDataDialog = Dialogs.addNoiseData(self.noiseLevel, self.noiseTypes)
        self.getNoiseDataDialog.activate.clicked.connect(self.getNoiseData)
        self.getNoiseDataDialog.exec()

    def getNoiseData(self):
        """ Collect data about the noise from the dialog """
        self.noiseLevel, self.noiseTypes = self.getNoiseDataDialog.getNoiseData()
        print(self.noiseLevel,self.noiseTypes)
        self.getNoiseDataDialog.close()
        self.segmentsToSave = True
        
    def saveImage(self, imageFile=''):
        if self.cheatsheet:
            self.showMaximized() # for nice spec images
        exporter = pge.ImageExporter(self.w_spec.scene())

        if imageFile=='':
            imageFile, drop = QFileDialog.getSaveFileName(self, "Save Image", "", "Images (*.png *.xpm *.jpg)")
            if not (imageFile.endswith('.png') or imageFile.endswith('.xpm') or imageFile.endswith('.jpg')):
                # exporter won't be able to deduce file type and will quit silently
                imageFile = imageFile + '.png'
        try:
            # works but requires devel (>=0.11) version of pyqtgraph:
            exporter.export(imageFile)
            print("Exporting spectrogram to file %s" % imageFile)
        except Exception as e:
            print("Warning: failed to save image")
            print(e)

    def changeSettings(self):
        """ Create the parameter tree when the Interface settings menu is pressed.
        """
        self.saveSegments()
        fn1 = self.config['BirdListShort']
        if '/' in fn1:
            ind = fn1[-1::-1].index('/')
            fn1 = fn1[-ind:]
        fn2 = self.config['BirdListLong']
        if fn2 is not None and '/' in fn2:
            ind = fn2[-1::-1].index('/')
            fn2 = fn2[-ind:]
        hasMultipleSegments = False
        for s in self.segments:
            if len(s[4])>1:
                hasMultipleSegments=True

        params = [
            {'name': 'Mouse settings', 'type' : 'group', 'children': [
                {'name': 'Use right button to make segments', 'type': 'bool', 'tip': 'If true, segments are drawn with right clicking.',
                 'value': self.config['drawingRightBtn']},
                {'name': 'Spectrogram mouse action', 'type': 'list', 'values':
                    {'Mark segments by clicking' : 1, 'Mark boxes by clicking' : 2, 'Mark boxes by dragging' : 3},
                 'value': self.config['specMouseAction']}
            ]},

            {'name': 'Paging', 'type': 'group', 'children': [
                {'name': 'Page size', 'type': 'float', 'value': self.config['maxFileShow'], 'limits': (5, 900),
                 'step': 5,
                 'suffix': ' sec'},
                {'name': 'Page overlap', 'type': 'float', 'value': self.config['fileOverlap'], 'limits': (0, 20),
                 'step': 2,
                 'suffix': ' sec'},
            ]},

            {'name': 'Annotation', 'type': 'group', 'children': [
                {'name': 'Annotation overview cell length', 'type': 'float',
                 'value': self.config['widthOverviewSegment'],
                 'limits': (5, 300), 'step': 5,
                 'suffix': ' sec'},
                {'name': 'Make boxes transparent', 'type': 'bool',
                     'value': self.config['transparentBoxes']},
                {'name': 'Segment colours', 'type': 'group', 'children': [
                    {'name': 'Confirmed segments', 'type': 'color', 'value': self.config['ColourNamed'],
                     'tip': "Correctly labeled segments"},
                    {'name': 'Possible', 'type': 'color', 'value': self.config['ColourPossible'],
                     'tip': "Segments that need further approval"},
                    {'name': "Don't know", 'type': 'color', 'value': self.config['ColourNone'],
                     'tip': "Segments that are not labelled"},
                    {'name': 'Currently selected', 'type': 'color', 'value': self.config['ColourSelected'],
                     'tip': "Currently delected segment"},
                ]},
                {'name': 'Check-ignore protocol', 'type': 'group', 'children': [
                    {'name': 'Show check-ignore marks', 'type': 'bool', 'value': self.config['protocolOn']},
                    {'name': 'Length of checking zone', 'type': 'float', 'value': self.config['protocolSize'],
                     'limits': (1, 300), 'step': 1, 'suffix': ' sec'},
                    {'name': 'Repeat zones every', 'type': 'float', 'value': self.config['protocolInterval'],
                     'limits': (1, 600), 'step': 1, 'suffix': ' sec'},
                ]}
            ]},

            {'name': 'Bird List', 'type': 'group', 'children': [
                {'name': 'Common Bird List', 'type': 'group', 'children': [
                    # {'name': 'Filename', 'type': 'text', 'value': self.config['BirdListShort']},
                    {'name': 'Filename', 'type': 'str', 'value': fn1, 'readonly': True},
                    {'name': 'Choose File', 'type': 'action'},
                ]},
                {'name': 'Full Bird List', 'type': 'group', 'children': [
                    # {'name': 'Filename', 'type': 'str', 'value': fn2,'readonly':True, 'tip': "Can be None"},
                    {'name': 'Filename', 'type': 'str', 'value': fn2, 'readonly': True},
                    {'name': 'No long list', 'type': 'bool',
                     'value': self.config['BirdListLong'] is None or self.config['BirdListLong'] == 'None',
                     'tip': "If you don't have a long list of birds"},
                    {'name': 'Choose File', 'type': 'action'}
                ]},
                {'name': 'Dynamically reorder bird list', 'type': 'bool', 'value': self.config['ReorderList']},
                {'name': 'Default to multiple species', 'type': 'bool', 'value': self.config['MultipleSpecies'],
                 'readonly': hasMultipleSegments},
            ]},
            {'name': 'Human classify', 'type': 'group', 'children': [
                {'name': 'Save corrections', 'type': 'bool', 'value': self.config['saveCorrections'],
                 'tip': "This helps the developers"},
            ]},

            {'name': 'Output parameters', 'type': 'group', 'children': [
                {'name': 'Show all pages', 'type': 'bool', 'value': self.config['showAllPages'],
                 'tip': "Where to show segments from when looking at outputs"},
                {'name': 'Auto save segments every', 'type': 'float', 'value': self.config['secsSave'],
                 'step': 5,
                 'limits': (5, 900),
                 'suffix': ' sec'},
            ]},

            {'name': 'User', 'type': 'group', 'children': [
                {'name': 'Operator', 'type': 'str', 'value': self.config['operator'],
                 'tip': "Person name"},

                {'name': 'Reviewer', 'type': 'str', 'value': self.config['reviewer'],
                 'tip': "Person name"},
            ]},
            {'name': 'Maximise window on startup', 'type': 'bool', 'value': self.config['StartMaximized']},
            {'name': 'Require noise data', 'type': 'bool', 'value': self.config['RequireNoiseData']},
        ]

        ## Create tree of Parameter objects
        self.p = Parameter.create(name='params', type='group', children=params)
        self.p.sigTreeStateChanged.connect(self.changeParams)
        ## Create ParameterTree widget
        self.t = ParameterTree()
        self.t.setParameters(self.p, showTop=False)
        self.t.show()
        self.t.setWindowTitle('AviaNZ - Interface Settings')
        self.t.setWindowIcon(QIcon('img/Avianz.ico'))
        self.t.setFixedSize(520, 900)

    def changeParams(self,param, changes):
        """ Update the config and the interface if anything changes in the tree
        """
        # first save the annotations
        self.saveSegments()

        for param, change, data in changes:
            path = self.p.childPath(param)
            if path is not None:
                childName = '.'.join(path)
            else:
                childName = param.name()

            if childName=='Output parameters.Auto save segments every':
                self.config['secsSave']=data
            elif childName=='Annotation.Annotation overview cell length':
                self.config['widthOverviewSegment']=data
            elif childName=='Annotation.Make boxes transparent':
                self.config['transparentBoxes']=data
                self.dragRectsTransparent()
            elif childName == 'Mouse settings.Use right button to make segments':
                self.config['drawingRightBtn'] = data
                if self.config['drawingRightBtn']:
                    self.MouseDrawingButton = QtCore.Qt.RightButton
                else:
                    self.MouseDrawingButton = QtCore.Qt.LeftButton
            elif childName == 'Mouse settings.Spectrogram mouse action':
                self.config['specMouseAction'] = data
                self.p_spec.enableDrag = data==3 and not self.readonly.isChecked()
            elif childName == 'Paging.Page size':
                self.config['maxFileShow'] = data
            elif childName=='Paging.Page overlap':
                self.config['fileOverlap'] = data
            elif childName == 'Maximise window on startup':
                self.config['StartMaximized'] = data
                if data:
                    self.showMaximized()
            elif childName == 'Bird List.Dynamically reorder bird list':
                self.config['ReorderList'] = data
            elif childName == 'Bird List.Default to multiple species':
                self.config['MultipleSpecies'] = data
            elif childName == 'Require noise data':
                self.config['RequireNoiseData'] = data
            elif childName=='Human classify.Save corrections':
                self.config['saveCorrections'] = data
            elif childName=='Bird List.Common Bird List.Filename':
                self.config['BirdListShort'] = data
            elif childName=='Bird List.Full Bird List.Filename':
                self.config['BirdListLong'] = data
            elif childName=='Annotation.Segment colours.Confirmed segments':
                rgbaNamed = list(data.getRgb())
                if rgbaNamed[3] > 100:
                    rgbaNamed[3] = 100
                self.config['ColourNamed'] = rgbaNamed
                self.ColourNamed = QtGui.QColor(self.config['ColourNamed'][0], self.config['ColourNamed'][1],
                                                self.config['ColourNamed'][2], self.config['ColourNamed'][3])
                self.ColourNamedDark = QtGui.QColor(self.config['ColourNamed'][0], self.config['ColourNamed'][1],
                                                    self.config['ColourNamed'][2], 255)
            elif childName=='Annotation.Segment colours.Possible':
                rgbaVal = list(data.getRgb())
                if rgbaVal[3] > 100:
                    rgbaVal[3] = 100
                self.config['ColourPossible'] = rgbaVal
                self.ColourPossible = QtGui.QColor(self.config['ColourPossible'][0], self.config['ColourPossible'][1],
                                                   self.config['ColourPossible'][2], self.config['ColourPossible'][3])
                self.ColourPossibleDark = QtGui.QColor(self.config['ColourPossible'][0],
                                                       self.config['ColourPossible'][1],
                                                       self.config['ColourPossible'][2], 255)
            elif childName=="Annotation.Segment colours.Don't know":
                rgbaVal = list(data.getRgb())
                if rgbaVal[3] > 100:
                    rgbaVal[3] = 100
                self.config['ColourNone'] = rgbaVal
                self.ColourNone = QtGui.QColor(self.config['ColourNone'][0], self.config['ColourNone'][1],
                                               self.config['ColourNone'][2], self.config['ColourNone'][3])
                self.ColourNoneDark = QtGui.QColor(self.config['ColourNone'][0], self.config['ColourNone'][1],
                                                   self.config['ColourNone'][2], 255)
            elif childName=='Annotation.Segment colours.Currently selected':
                rgbaVal = list(data.getRgb())
                if rgbaVal[3] > 100:
                    rgbaVal[3] = 100
                self.config['ColourSelected'] = rgbaVal
                # update the interface
                self.ColourSelected = QtGui.QColor(self.config['ColourSelected'][0], self.config['ColourSelected'][1],
                                                   self.config['ColourSelected'][2], self.config['ColourSelected'][3])
                self.ColourSelectedDark = QtGui.QColor(self.config['ColourSelected'][0], self.config['ColourSelected'][1],
                                                   self.config['ColourSelected'][2], 255)
            elif childName=='Annotation.Check-ignore protocol.Show check-ignore marks':
                self.config['protocolOn'] = data
                self.addRegularSegments()
            elif childName=='Annotation.Check-ignore protocol.Length of checking zone':
                self.config['protocolSize'] = data
                self.addRegularSegments()
            elif childName=='Annotation.Check-ignore protocol.Repeat zones every':
                self.config['protocolInterval'] = data
                self.addRegularSegments()
            elif childName=='Output parameters.Show all pages':
                self.config['showAllPages'] = data
            elif childName=='User.Operator':
                self.config['operator'] = data
                self.operator = data
                self.statusRight.setText("Operator: " + str(self.operator) + ", Reviewer: " + str(self.reviewer))
            elif childName=='User.Reviewer':
                self.config['reviewer'] = data
                self.reviewer = data
                self.statusRight.setText("Operator: " + str(self.operator) + ", Reviewer: " + str(self.reviewer))
            elif childName=='Bird List.Common Bird List.Choose File':
                filename, drop = QtGui.QFileDialog.getOpenFileName(self, 'Choose File', self.SoundFileDir, "Text files (*.txt)")
                if filename == '':
                    print("no list file selected")
                    return
                else:
                    self.shortBirdList = self.ConfigLoader.shortbl(filename, self.configdir)
                    if self.shortBirdList is not None:
                        self.config['BirdListShort'] = filename
                        self.p['Bird List','Common Bird List', 'Filename'] = filename
                    else:
                        self.shortBirdList = self.ConfigLoader.shortbl(self.config['BirdListShort'], self.configdir)
            elif childName=='Bird List.Full Bird List.Choose File':
                filename, drop = QtGui.QFileDialog.getOpenFileName(self, 'Choose File', self.SoundFileDir, "Text files (*.txt)")
                if filename == '':
                    print("no list file selected")
                    return
                else:
                    self.longBirdList = self.ConfigLoader.longbl(filename, self.configdir)
                    self.config['BirdListLong'] = filename
                    self.p['Bird List','Full Bird List','Filename'] = filename
                    self.p['Bird List','Full Bird List','No long list'] = False
            elif childName=='Bird List.Full Bird List.No long list':
                if param.value():
                    self.config['BirdListLong'] = 'None'
                    self.p['Bird List','Full Bird List','Filename'] = 'None'
                    self.longBirdList = None
                else:
                    if self.p['Bird List','Full Bird List','Filename'] is None or self.p['Bird List','Full Bird List','Filename'] == '' or self.p['Bird List','Full Bird List','Filename'] == 'None':
                        filename, drop = QtGui.QFileDialog.getOpenFileName(self, 'Choose File', self.SoundFileDir, "Text files (*.txt)")
                        if filename == '':
                            print("no list file selected")
                            return
                        else:
                            self.p['Bird List','Full Bird List','Filename'] = filename
                            self.config['BirdListLong'] = filename
                            self.longBirdList = self.ConfigLoader.longbl(self.config['BirdListLong'], self.configdir)

        self.saveConfig = True

        self.resetStorageArrays()
        self.loadFile()

# ============
# Various actions: deleting segments, saving, quitting
    def deleteSegment(self,id=-1,hr=False):
        """ Listener for delete segment button, or backspace key. Also called when segments are deleted by the
        human classify dialogs.
        Stops playback immediately in all cases.
        Deletes the segment that is selected, otherwise does nothing.
        Updates the overview segments as well.
        """
        print("deleting id:", id)
        if self.media_obj.isPlaying():
            # includes resetting playback buttons
            self.stopPlayback()

        if not hr and id<0:
            id = self.box1id

        if id>-1:
            startpoint = self.segments[id][0]-self.startRead
            endpoint = self.segments[id][1]-self.startRead
            species = self.segments[id][4]

            self.refreshOverviewWith(startpoint, endpoint, species, delete=True)

            if self.listRectanglesa1[id] is not None:
                try:
                    self.listRectanglesa1[id].sigRegionChangeFinished.disconnect()
                    self.listRectanglesa2[id].sigRegionChangeFinished.disconnect()
                except:
                    pass
                self.p_ampl.removeItem(self.listRectanglesa1[id])
                self.p_spec.removeItem(self.listRectanglesa2[id])
                self.p_spec.removeItem(self.listLabels[id])
            del self.listLabels[id]
            del self.segments[id]
            del self.listRectanglesa1[id]
            del self.listRectanglesa2[id]
            self.segmentsToSave = True

            self.box1id = -1
            # reset segment playback buttons
            self.playSegButton.setEnabled(False)
            self.playBandLimitedSegButton.setEnabled(False)
            self.quickDenButton.setEnabled(False)

    def deleteAll(self):
        """ Listener for delete all button.
        Checks if the user meant to do it, then calls removeSegments()
        """
        if len(self.segments) == 0:
            msg = SupportClasses.MessagePopup("w", "No segments", "No segments to delete")
            msg.exec_()
            return
        else:
            msg = SupportClasses.MessagePopup("t", "Delete All Segments?", "Are you sure you want to delete all segments?")
            msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            reply = msg.exec_()
            if reply == QMessageBox.Yes:
                self.removeSegments()
                self.segmentsToSave = True
                if self.Hartley:
                    self.segmentsToSave = False
                    if os.path.isfile(self.filename + '.data'):
                        os.remove(self.filename + '.data')
                    if os.path.isfile(self.filename[:-4] + '_output.xlsx'):
                        os.remove(self.filename[:-4] + '_output.xlsx')
                    self.listFiles.currentItem().setForeground(Qt.black)

            # reset segment playback buttons
            self.playSegButton.setEnabled(False)
            self.playBandLimitedSegButton.setEnabled(False)
            self.quickDenButton.setEnabled(False)

    def removeSegments(self,delete=True):
        """ Remove all the segments in response to the menu selection, or when a new file is loaded. """
        for r in self.listLabels:
            if r is not None:
                self.p_spec.removeItem(r)
        for r in self.listRectanglesa1:
            if r is not None:
                try:
                    r.sigRegionChangeFinished.disconnect()
                    self.p_ampl.removeItem(r)
                except:
                    pass
        for r in self.listRectanglesa2:
            if r is not None:
                try:
                    r.sigRegionChangeFinished.disconnect()
                    self.p_spec.removeItem(r)
                except:
                    pass

        # clear overview boxes and their count trackers
        for ovid in range(len(self.SegmentRects)):
            self.overviewSegments[ovid, :] = 0
            self.SegmentRects[ovid].setBrush(pg.mkBrush('w'))
            self.SegmentRects[ovid].update()

        if delete:
            self.segments=[]
            self.listRectanglesa1 = []
            self.listRectanglesa2 = []
            self.listLabels = []
            self.box1id = -1

    def saveSegments(self):
        """ Save the segmentation data as a json file.
        Name of the file is the name of the wave file + .data"""

        # def checkSave():
        #     msg = QMessageBox()
        #     msg.setIcon(QMessageBox.Information)
        #     msg.setText("Do you want to save?")
        #     msg.setInformativeText("You didn't identify any segments, are you sure you want to save this annotation?")
        #     msg.setWindowTitle("No segments")
        #     msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        #     msg.buttonClicked.connect(msgbtn)
        #     retval = msg.exec_()
        #     print "value of pressed message box button:", retval
        #     return retval

        if self.segmentsToSave:
            print("Saving segments to " + self.filename + '.data')
            if len(self.segments) > 0:
                if self.segments[0][0] > -1:
                    self.segments.insert(0, [-1, self.datalengthSec, self.operator, self.reviewer, [self.noiseLevel, self.noiseTypes]])
            else:
                self.segments.insert(0, [-1, self.datalengthSec, self.operator, self.reviewer, [self.noiseLevel, self.noiseTypes]])

            if isinstance(self.filename, str):
                file = open(self.filename + '.data', 'w')
            else:
                file = open(str(self.filename) + '.data', 'w')
            json.dump(self.segments,file)
            file.write("\n")
            #if self.Hartley:
                #if self.previousFile is not None:
                    #self.previousFile.setForeground(Qt.red)
            self.segmentsToSave = False
            del self.segments[0]
            if self.Hartley:
                self.exportToExcel_Hartley()
        else:
            print("Nothing to save")

    """def makeNewFilter(species,minLen,maxLen,minFrq,maxFrq,fs,f0_low,f0_high,thr,M,optimumNodes):
        # Write out a new dictionary for a filter for a particular species 

        dict = {'Name': species, 'SampleRate': fs, 'TimeRange': [minLen,maxLen], 'FreqRange': [minFrq, maxFrq], 'F0Range': [f0_low, f0_high], 'WaveletParams': [thr, M, optimumNodes]}

        # Check if file exists
        filename = self.config['FiltersDir']+species+'.txt'
        # TODO: More?
        if isfile(filename):
            print("File already exists, overwriting")
        json.dump(dict,filename)"""

    # TODO: Move this to SupportClasses, or just delete?
    def exportToExcel_Hartley(self):
        eFile = self.filename[:-4] + '_output.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Interval")
        for col in range(1,len(self.shortBirdList)+1):
            #print(self.config['BirdList'][col-1])
            ws.cell(row=1, column=1+col, value=self.shortBirdList[col-1])
        lastrow = 2

        props = np.zeros(len(self.shortBirdList))

        for seg in self.segments[lastrow-2:]:
            #print(lastrow)
            ws.cell(row=lastrow,column=1,value=lastrow-1)
            for col in range(1,len(self.shortBirdList)+1):
                if self.shortBirdList[col-1] in seg[4]:
                #if (self.shortBirdList[col-1]+',') in seg[4]:
                    #print(seg[4],"1")
                    ws.cell(row=lastrow, column=col+1, value=1)
                    props[col-1] += 1
                else:
                    #print(seg[4],"0")
                    ws.cell(row=lastrow, column=col+1, value=0)
            lastrow += 1
        
        wb.create_sheet(title='Summary', index=2)
        ws = wb['Summary']
        ws.cell(row=1, column=1, value="Species")
        ws.cell(row=1, column=2, value="Proportion")
        for row in range(1,len(self.shortBirdList)+1):
            ws.cell(row=row+1, column=1, value=self.shortBirdList[row-1])
            ws.cell(row=row+1, column=2, value=props[row-1]/(lastrow-2))

        wb.save(str(eFile))
        print("Saved to "+eFile)

    def restart(self):
        """ Listener for the restart option, which uses exit(1) to restart the program at the splash screen """
        print("Restarting")

        # Check if user requires noise data
        if self.config['RequireNoiseData'] and self.noiseLevel is None:
            self.addNoiseData()

        self.saveSegments()
        if self.saveConfig == True:
            try:
                print("Saving config file")
                json.dump(self.config, open(self.configfile, 'w'),indent=1)
            except Exception as e:
                print("ERROR while saving config file:")
                print(e)

        # Save the shortBirdList
        self.ConfigLoader.blwrite(self.shortBirdList, self.config['BirdListShort'], self.configdir)
        QApplication.exit(1)

    def closeEvent(self, event):
        """ Catch the user closing the window by clicking the Close button instead of quitting. """
        self.quit()

    def quit(self):
        """ Listener for the quit button, also called by closeEvent().
        Add in the operator and reviewer at the top, and then save the segments and the config file.
        """

        print("Quitting")

        # Check if user requires noise data
        if self.config['RequireNoiseData'] and self.noiseLevel is None:
            self.addNoiseData()

        self.saveSegments()
        if self.saveConfig:
            self.ConfigLoader.configwrite(self.config, self.configfile)

        # Save the shortBirdList
        self.ConfigLoader.blwrite(self.shortBirdList, self.config['BirdListShort'], self.configdir)
        QApplication.quit()

    def backupDatafiles(self):
        print("Backing up files in ", self.SoundFileDir)
        listOfDataFiles = QDir(self.SoundFileDir).entryList(['*.data'])
        for file in listOfDataFiles:
            source = self.SoundFileDir + '/' + file
            destination = source[:-5]+".backup"
            if os.path.isfile(destination):
                pass
                #print(destination," exists, not backing up")
            else:
                #print(source)
                #print(destination," doesn't exist")
                copyfile(source, destination)

    def eventFilter(self, obj, event):
        # This is an event filter for the context menu. It allows the user to select
        # multiple birds by stopping the menu being closed on first click
        if self.multipleBirds and event.type() in [QtCore.QEvent.MouseButtonRelease]:
            if isinstance(obj, QtGui.QMenu):
                if obj.activeAction():
                    if not obj.activeAction().menu(): 
                        #if the selected action does not have a submenu
                        #eat the event, but trigger the function
                        obj.activeAction().trigger()
                        return True
        return QMenu.eventFilter(self,obj, event)

# =============


@click.command()
@click.option('-c', '--cli', is_flag=True, help='Run in command-line mode')
@click.option('-s', '--cheatsheet', is_flag=True, help='Make the cheatsheet images')
@click.option('-z', '--zooniverse', is_flag=True, help='Make the Zooniverse images and sounds')
@click.option('-f', '--infile', type=click.Path(), help='Input wav file (mandatory in CLI mode)')
@click.option('-o', '--imagefile', type=click.Path(), help='If specified, a spectrogram will be saved to this file')
@click.argument('command', nargs=-1)
def mainlauncher(cli, cheatsheet, zooniverse, infile, imagefile, command):
    # determine location of config file and bird lists
    if platform.system() == 'Windows':
        # Win
        configdir = os.path.expandvars(os.path.join("%APPDATA%", "AviaNZ"))
    elif platform.system() == 'Linux' or platform.system() == 'Darwin':
        # Unix
        configdir = os.path.expanduser("~/.avianz/")
    else:
        print("ERROR: what OS is this? %s" % platform.system())
        sys.exit()

    # if config and bird files not found, copy from distributed backups.
    # so these files will always exist on load (although they could be corrupt)
    # (exceptions here not handled and should always result in crashes)
    necessaryFiles = ["AviaNZconfig.txt", "ListCommonBirds.txt", "ListDOCBirds.txt"]
    if not os.path.isdir(configdir):
        print("Creating config dir %s" % configdir)
        try:
            os.makedirs(configdir)
        except:
            print("ERROR: failed to make config dir")
            sys.exit()
    for f in necessaryFiles:
        if not os.path.isfile(os.path.join(configdir, f)):
            print("File %s not found in config dir, providing default" % f)
            try:
                shutil.copy2(os.path.join("Config", f), configdir)
            except:
                print("ERROR: failed to copy essential config files")
                sys.exit()

    # copy over filters to ~/.avianz/Filters/:
    filterdir = os.path.join(configdir, "Filters/")
    if not os.path.isdir(filterdir):
        print("Creating filter dir %s" % filterdir)
        os.makedirs(filterdir)
    for f in os.listdir("Filters"):
        ff = os.path.join("Filters", f) # Kiwi.txt
        if not os.path.isfile(os.path.join(filterdir, f)): # ~/.avianz/Filters/Kiwi.txt
            print("Filter %s not found, providing default" % f)
            try:
                shutil.copy2(ff, filterdir) # cp Filters/Kiwi.txt ~/.avianz/Filters/
            except Exception as e:
                print("Warning: failed to copy filter %s to %s" % (ff, filterdir))
                print(e)

    # run splash screen:
    if cli:
        print("Starting AviaNZ in CLI mode")
        if not cheatsheet and not zooniverse and not isinstance(infile, str):
            print("ERROR: valid input file (-f) is mandatory in CLI mode!")
            sys.exit()
        avianz = AviaNZ(configdir=configdir,CLI=True, cheatsheet=cheatsheet, zooniverse=zooniverse, firstFile=infile, imageFile=imagefile, command=command)
        print("Analysis complete, closing AviaNZ")
    else:
        print("Starting AviaNZ in GUI mode")
        # This screen asks what you want to do, then processes the response
        first = Dialogs.StartScreen()
        first.setWindowIcon(QtGui.QIcon('img/AviaNZ.ico'))
        first.show()
        app.exec_()

        task = first.getValues()

        if task == 1:
            avianz = AviaNZ(configdir=configdir)
            avianz.setWindowIcon(QtGui.QIcon('img/AviaNZ.ico'))
        elif task==2:
            avianz = AviaNZ_batch.AviaNZ_batchProcess(configdir=configdir)
            avianz.setWindowIcon(QtGui.QIcon('img/AviaNZ.ico'))
        elif task==4:
            avianz = AviaNZ_batch.AviaNZ_reviewAll(configdir=configdir)

        avianz.show()
        out = app.exec_()
        QApplication.closeAllWindows()
        if out == 1:
            mainlauncher()

# Start the application
app = QApplication(sys.argv)
mainlauncher()
