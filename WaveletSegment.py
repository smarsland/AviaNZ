# Version 0.2 10/7/17
# Author: Stephen Marsland

import pywt
import Wavelets
import wavio
import numpy as np
import os,json
import glob
import SignalProc

# Nirosha's approach of simultaneous segmentation and recognition using wavelets
    # (0) Bandpass filter with different parameters for each species
    # (1) 5 level wavelet packet decomposition
    # (2) Sort nodes of (1) into order by point-biserial correlation with training labels
        # This is based on energy
    # (3) Retain top nodes (up to 20)
    # (4) Re-sort to favour child nodes
    # (5) Reduce number using F_2 score
        # This is based on thresholded reconstruction
    # (6) Classify as call if OR of (5) is true

# TODO: This still needs tidying up
class WaveletSegment:
    # This class implements wavelet segmentation for the AviaNZ interface

    def __init__(self,data=[],sampleRate=0,spp='Kiwi',annotation=None):
        self.spp=spp
        self.annotation=annotation
        if data != []:
            self.data = data
            self.sampleRate = sampleRate
            if self.data.dtype is not 'float':
                self.data = self.data.astype('float') / 32768.0

        [lowd,highd,lowr,highr] = np.loadtxt('dmey.txt')
        self.wavelet = pywt.Wavelet(name="dmey2",filter_bank=[lowd,highd,lowr,highr])
        self.wavelet.orthogonal=True

        self.sp = SignalProc.SignalProc([],0,256,128)

        self.waveletDenoiser = Wavelets.Wavelets(data=data, wavelet=self.wavelet,maxLevel=20)

    def computeWaveletEnergy(self,fwData,sampleRate):
        """ Computes the energy of the nodes in the wavelet packet decomposition
        # There are 62 coefficients up to level 5 of the wavelet tree (without root), and 300 seconds in 5 mins
        # Hence coefs is a 62*300 matrix
        # The energy is the sum of the squares of the data in each node divided by the total in that level of the tree as a percentage.
        """
        n=len(fwData)/sampleRate
        coefs = np.zeros((62, n))
        for t in range(n):
            E = []
            for level in range(1,6):
                wp = pywt.WaveletPacket(data=fwData[t * sampleRate:(t + 1) * sampleRate], wavelet=self.wavelet, mode='symmetric', maxlevel=level)
                e = np.array([np.sum(n.data**2) for n in wp.get_level(level, "natural")])
                if np.sum(e)>0:
                    e = 100.0*e/np.sum(e)
                E = np.concatenate((E, e),axis=0)
            coefs[:, t] = E
        return coefs

    def fBetaScore(self,annotation, predicted,beta=2):
        """ Computes the beta scores given two sets of predictions """
        TP = np.sum(np.where((annotation==1)&(predicted==1),1,0))
        T = np.sum(annotation)
        P = np.sum(predicted)
        if T != 0:
            recall = float(TP)/T #TruePositive/#True
        else:
            recall = None
        if P!=0:
            precision = float(TP)/P #TruePositive/#Positive
        else:
            precision = None
        if recall != None and precision != None:
            fB=((1.+beta**2)*recall*precision)/(recall + beta**2*precision)
        else:
            fB=None
        if recall==None and precision==None:
            print "TP=%d \tFP=%d \tTN=%d \tFN=%d \tRecall=%s \tPrecision=%s \tfB=%s" %(TP,P-TP,len(annotation)-(P+T-TP),T-TP,recall,precision,fB)
        elif recall==None:
            print "TP=%d \tFP=%d \tTN=%d \tFN=%d \tRecall=%s \tPrecision=%0.2f \tfB=%s" %(TP,P-TP,len(annotation)-(P+T-TP),T-TP,recall,precision,fB)
        elif precision==None:
            print "TP=%d \tFP=%d \tTN=%d \tFN=%d \tRecall=%0.2f \tPrecision=%s \tfB=%s" %(TP,P-TP,len(annotation)-(P+T-TP),T-TP,recall,precision,fB)
        else:
            print "TP=%d \tFP=%d \tTN=%d \tFN=%d \tRecall=%0.2f \tPrecision=%0.2f \tfB=%0.2f" %(TP,P-TP,len(annotation)-(P+T-TP),T-TP,recall,precision,fB)
        #print TP, int(T), int(P), recall, precision, ((1.+beta**2)*recall*precision)/(recall + beta**2*precision)
        return fB

    def compute_r(self,annotation,waveletCoefs,nNodes=20):
        """ Computes the point-biserial correlations for a set of labels and a set of wavelet coefficients.
        r = (M_p - M_q) / S * sqrt(p*q), M_p = mean for those that are 0, S = std dev overall, p = proportion that are 0.
        """
        w0 = np.where(annotation==0)[0]
        w1 = np.where(annotation==1)[0]

        r = np.zeros(62)
        for count in range(62):
            r[count] = (np.mean(waveletCoefs[count,w1]) - np.mean(waveletCoefs[count,w0]))/np.std(waveletCoefs[count,:]) * np.sqrt(len(w0)*len(w1))/len(annotation)

        order = np.argsort(r)
        order = order[-1:-nNodes-1:-1]

        return order

    def sortListByChild(self,order):
        """ Inputs is a list sorted into order of correlation.
        This functions resort so that any children of the current node that are in the list go first.
        Assumes that there are five levels in the tree (easy to extend, though)
        """

        newlist = []
        currentIndex = 0
        # Need to keep track of where each level of the tree starts
        # Note that there is no root to the tree, hence the 0 then 2
        starts = [0, 2, 6, 14,30,62]
        while len(order)>0:
            if order[0]<30:
                # It could have children lower down the list
                # Build a list of the children of the first element of order
                level = int(np.log2(order[0]+2))
                nc = 2
                first = order[0]
                for l in range(level+1,6):
                    children = []
                    current = currentIndex
                    for i in range(nc):
                        children.append(starts[l] + 2*(first-starts[l-1])+i)
                    nc*=2
                    first = starts[l] + 2*(first-starts[l-1])
                    # Have to do it this annoying way since Python seems to ignore the next element if you delete one while iterating over the list
                    i=0
                    order_sub = []
                    while i < len(children):
                        if children[i] not in order:
                            del(children[i])
                        else:
                            order_sub.append(order.index(children[i]))
                            i+=1

                    # Sort into order
                    children = [x for (y, x) in sorted(zip(order_sub, children), key=lambda pair: pair[0])]

                    for a in children:
                        # If so, remove and insert at the current location in the new list
                        newlist.insert(current,a)
                        current += 1
                        order.remove(a)

            # Finally, add the first element
            newlist.append(order[0])
            currentIndex = newlist.index(order[0])+1
            del(order[0])

        return newlist

    def detectCalls(self,wp,node,sampleRate=0,n=300,species='Kiwi'):
        """ For a given node of the wavelet tree, make the tree, detect calls """
        # TODO: Pass in (or load) a species-specific threshold and use that

        if sampleRate==0:
            sampleRate=self.sampleRate
        import string
        # Add relevant nodes to the wavelet packet tree and then reconstruct the data
        new_wp = pywt.WaveletPacket(data=None, wavelet=self.wavelet, mode='symmetric')
        # First, turn the index into a leaf name.
        bin = self.waveletDenoiser.ConvertWaveletNodeName(node)
        # level = np.floor(np.log2(node))
        # first = int(2**level-1)
        # bin = np.binary_repr(node-first,width=int(level))
        # bin = string.replace(bin,'0','a',maxreplace=-1)
        # bin = string.replace(bin,'1','d',maxreplace=-1)
        # #print index+1, bin
        new_wp[bin] = wp[bin].data

        # Get the coefficients
        C = np.abs(new_wp.reconstruct(update=True))
        N = len(C)

        # wavio.write('E:\Rebecca SIPO\\train\\node'+str(node)+'.wav', C.astype('int16'), sampleRate, scale='dtype-limits', sampwidth=2)

        # Compute the number of samples in a window -- species specific
        M = int(0.8*sampleRate/2.0)     # TODO: set M in relation to the call length, M can make a big difference in res
        if species.title()=='Sipo':
            M=int(0.2*sampleRate/2.0)
        #print M

        # Compute the energy curve (a la Jinnai et al. 2012)
        E = np.zeros(N)
        E[M] = np.sum(C[:2 * M+1])
        for i in range(M + 1, N - M):
            E[i] = E[i - 1] - C[i - M - 1] + C[i + M]
        E = E / (2. * M)

        threshold = np.mean(C) + np.std(C)
        if species.title() == 'Sipo':
            threshold = np.mean(C) + np.std(C)/4
        # bittern
        # TODO: test
        #thresholds[np.where(waveletCoefs<=32)] = 0
        #thresholds[np.where(waveletCoefs>32)] = 0.3936 + 0.1829*np.log2(np.where(waveletCoefs>32))

        # If there is a call anywhere in the window, report it as a call
        E = np.where(E<threshold, 0, 1)
        detected = np.zeros(n)
        j = 0
        for i in range(0,N-sampleRate,sampleRate):
            detected[j] = np.max(E[i:i+sampleRate])
            j+=1

        return detected

    def detectCalls_test(self,wp,sampleRate, listnodes=[], species='Kiwi',trainTest=False): #default kiwi
        #for test recordings given the set of nodes
        import string
        # Add relevant nodes to the wavelet packet tree and then reconstruct the data
        # detected = np.zeros((300,len(listnodes)))
        detected = np.zeros((int(len(wp.data)/sampleRate),len(listnodes)))
        count = 0

        for index in listnodes:
            new_wp = pywt.WaveletPacket(data=None, wavelet=self.wavelet, mode='symmetric')

            # TODO: This primes the tree with zeros, which was necessary for denoising.
            # Is it needed here?
            #for level in range(wp.maxlevel + 1):
            #    for n in new_wp.get_level(level, 'natural'):
            #        n.data = np.zeros(len(wp.get_level(level, 'natural')[0].data))

            # First, turn the index into a leaf name.
            # level = np.floor(np.log2(index))
            # first = int(2**level-1)
            # bin = np.binary_repr(index-first,width=int(level))
            # bin = string.replace(bin,'0','a',maxreplace=-1)
            # bin = string.replace(bin,'1','d',maxreplace=-1)
            #print index+1, bin
            bin = self.waveletDenoiser.ConvertWaveletNodeName(index)
            new_wp[bin] = wp[bin].data

            # Get the coefficients
            C = np.abs(new_wp.reconstruct(update=True))
            N = len(C)

            # Compute the number of samples in a window -- species specific
            M = int(0.8*sampleRate/2.0)
            if species.title()=='Sipo':
                M = int(0.2 * sampleRate / 2.0)
            #print M

            # Compute the energy curve (a la Jinnai et al. 2012)
            E = np.zeros(N)
            E[M] = np.sum(C[:2 * M+1])
            for i in range(M + 1, N - M):
                E[i] = E[i - 1] - C[i - M - 1] + C[i + M]
            E = E / (2. * M)

            if species.title() == 'Sipo':
                threshold = np.mean(C) + np.std(C)/2
            else:
                threshold = np.mean(C) + np.std(C)/2

            # bittern
            # TODO: test
            #thresholds[np.where(waveletCoefs<=32)] = 0
            #thresholds[np.where(waveletCoefs>32)] = 0.3936 + 0.1829*np.log2(np.where(waveletCoefs>32))

            # If there is a call anywhere in the window, report it as a call
            E = np.where(E<threshold, 0, 1)
            j = 0
            for i in range(0,N-sampleRate,sampleRate):
                detected[j,count] = np.max(E[i:i+sampleRate])
                j+=1
            count += 1

        detected= np.max(detected,axis=1)
        if trainTest==True:
            return detected
        else:
            detected=np.where(detected>0)
            # print "det",detected
            if np.shape(detected)[1]>1:
                return self.identifySegments(np.squeeze(detected))
            elif np.shape(detected)[1]==1:
                return self.identifySegments(detected)
            else:
                return []

    def detectCalls1(self,wp,listnodes,sampleRate):
        # This way should be the best -- reconstruct from setting all of the relevant nodes. But it gives an error message
        # about different size coefficient arrays most of the time.
        import string
        # Add relevant nodes to the wavelet packet tree and then reconstruct the data
        new_wp = pywt.WaveletPacket(data=None, wavelet=self.wavelet, mode='symmetric')

        for index in listnodes:
            # First, turn the index into a leaf name.
            # TODO: This needs checking -- are they the right nodes?
            level = np.floor(np.log2(index))
            first = 2**level-1
            bin = np.binary_repr(index-first,width=int(level))
            bin = string.replace(bin,'0','a',maxreplace=-1)
            bin = string.replace(bin,'1','d',maxreplace=-1)
            #print index+1, bin
            new_wp[bin] = wp[bin].data

        # Get the coefficients
        C = np.abs(new_wp.reconstruct(update=True))
        N = len(C)

        # Compute the number of samples in a window -- species specific
        M = int(0.8*sampleRate/2.0)
        #print M

        # Compute the energy curve (a la Jinnai et al. 2012)
        E = np.zeros(N)
        E[M] = np.sum(C[:2 * M+1])
        for i in range(M + 1, N - M):
            E[i] = E[i - 1] - C[i - M - 1] + C[i + M]
        E = E / (2. * M)

        threshold = np.mean(C) + np.std(C)
        # bittern
        # TODO: test
        #thresholds[np.where(waveletCoefs<=32)] = 0
        #thresholds[np.where(waveletCoefs>32)] = 0.3936 + 0.1829*np.log2(np.where(waveletCoefs>32))

        # If there is a call anywhere in the window, report it as a call
        E = np.where(E<threshold, 0, 1)
        detected = np.zeros(300)
        j = 0
        for i in range(0,N-sampleRate,sampleRate):
            detected[j] = np.max(E[i:i+sampleRate])
            j+=1

        return detected

    def identifySegments(self, seg): #, maxgap=1, minlength=1):
        segments = []
        # print seg, type(seg)
        if len(seg)>0:
            for s in seg:
                segments.append([s, s+1])
        return segments

    def mergeSeg(self,segments):
        """ Combines segments from the wavelet segmenter."""
        indx = []
        for i in range(len(segments) - 1):
            if segments[i][1] == segments[i + 1][0]:
                indx.append(i)
        indx.reverse()
        for i in indx:
            segments[i][1] = segments[i + 1][1]
            del (segments[i + 1])
        return segments

    def loadData(self,fName,trainTest=True):
        # Load data
        filename = fName+'.wav' #'train/kiwi/train1.wav'
        filenameAnnotation = fName+'-sec.xlsx'#'train/kiwi/train1-sec.xlsx'
        # self.sampleRate, self.data = wavfile.read(filename)
        # if self.data.dtype is not 'float':
        #     self.data = self.data.astype('float') / 32768.0
        wavobj = wavio.read(filename)
        self.sampleRate = wavobj.rate
        self.data = wavobj.data
        if self.data.dtype is not 'float':
            self.data = self.data.astype('float') #/ 32768.0
        if np.shape(np.shape(self.data))[0]>1:
            self.data = np.squeeze(self.data[:,0])
        n=len(self.data)/self.sampleRate

        if trainTest==True:     #survey data don't have annotations
            # Get the segmentation from the excel file
            self.annotation = np.zeros(n)
            count = 0
            import xlrd
            wb=xlrd.open_workbook(filename = filenameAnnotation)
            ws=wb.sheet_by_index(0)
            col=ws.col(1)
            for row in range(1,n+1):
                self.annotation[count]=col[row].value
                count += 1
        #return self.data, self.sampleRate, self.annotation

def computeWaveletEnergy_1s(data,wavelet,choice='all',denoise=False):
    # Generate wavelet energy (all 62 nodes in a 5 level tree) given 1 sec data
    E=[]
    ws=WaveletSegment()

    for level in range(6):
        if wavelet == 'dmey2':
            [lowd, highd, lowr, highr] = np.loadtxt('dmey.txt')
            wavelet = pywt.Wavelet(filter_bank=[lowd, highd, lowr, highr])
            wavelet.orthogonal=True
        if denoise==True:
            data = ws.sp.waveletDenoise(data, thresholdType='soft', maxlevel=5)
        if choice=='bandpass':
            data=ws.sp.ButterworthBandpass(data,16000,low=500,high=7500)
        wp = pywt.WaveletPacket(data=data, wavelet=wavelet, mode='symmetric', maxlevel=level)
        e = np.array([np.sum(n.data**2) for n in wp.get_level(level, "natural")])
        if np.sum(e)>0:
            e = 100.0*e/np.sum(e)
        E = np.concatenate((E, e),axis=0)
    return E

def findCalls_train_learning(fName,species='Kiwi'):
    ws=WaveletSegment()
    f = np.genfromtxt("Sound Files\MLdata\wE.data",delimiter=',',dtype=None)
    ld = len(f[0])
    data = np.zeros((len(f),ld))

    names = []

    for i in range(len(f)):
        for j in range(ld-2):
            data[i,j] = f[i][j]
        data[i,ld-1] = f[i][ld-1]
        if not f[i][ld-2] in names:
            names.append(f[i][ld-2])
            data[i,ld-2] = len(names)
        else:
            data[i,ld-2] = names.index(f[i][ld-2])

    # Decide on a class to be the 1 to detect
    # It is choosing male kiwi as the positive class (10) here
    data[:,63] = 0
    ind = np.where(data[:,62] == 10)
    data[ind,63] = 1
    # ind = np.where(data[:,62] == 4)
    # data[ind,63] = 1

    # Compute point-biserial correlations and sort wrt it, return top nNodes
    nodes = ws.compute_r(data[:,62],data[:,:62].transpose())

    # Now for Nirosha's sorting
    # Basically, for each node, put any of its children (and their children, iteratively) that are in the list in front of it
    nodes = ws.sortListByChild(np.ndarray.tolist(nodes))

    # These nodes refer to the unrooted tree, so add 1 to get the real indices
    nodes = [n + 1 for n in nodes]

    # **** We actually need the real data :(
    # Generate a full 5 level wavelet packet decomposition
    # **** load newdata, spp
    f = np.genfromtxt("Sound Files\MLdata\data-1s.data",delimiter=',',dtype=None)
    f = np.squeeze(np.reshape(f,(np.shape(f)[0]*np.shape(f)[1],1)))
    #g = np.genfromtxt("Sound Files\MLdata\label-1s",delimiter=',',dtype=None)

    return f, data, nodes


def moretest():
    ws=WaveletSegment()
    wpFull = pywt.WaveletPacket(data=f, wavelet=ws.wavelet, mode='symmetric', maxlevel=5)
    # Now check the F2 values and add node if it improves F2
    listnodes = []
    bestBetaScore = 0
    detected = np.zeros(len(data))

    for node in nodes:
        testlist = listnodes[:]
        testlist.append(node)
        print testlist
        detected_c = ws.detectCalls(wpFull,node,16000,n=len(data))
        #update the detections
        det=np.maximum.reduce([detected,detected_c])
        fB = ws.fBetaScore(data[:,63], det)
        if fB > bestBetaScore:
            bestBetaScore = fB
            #now apend the detections of node c to detected
            detected=det
            listnodes.append(node)
        if bestBetaScore == 1:
            break
    # listnodes_all=listnodes_all.append(listnodes)
    return listnodes

def findCalls_train(fName,species='Kiwi'):
    # Load data and annotation
    ws=WaveletSegment()
    ws.loadData(fName)
    if species=='boom':
        fs=1000
    elif species.title()=='Sipo':
        fs=8000
    else:
        fs = 16000
    if ws.sampleRate != fs:
        ws.data = librosa.core.audio.resample(ws.data,ws.sampleRate,fs)
        ws.sampleRate=fs

    # Get the five level wavelet decomposition
    wData = ws.waveletDenoiser.waveletDenoise(ws.data, thresholdType='soft', wavelet='dmey',maxlevel=5)  # wavelet='dmey2' ??
    #print np.min(wData), np.max(wData)

    #librosa.output.write_wav('train/kiwi/D/', wData, sampleRate, norm=False)

    # Bandpass filter
    #fwData = bandpass(wData,sampleRate)
    # TODO: Params in here!
    # bittern
    #fwData = ButterworthBandpass(wData,sampleRate,low=100,high=400)
    # kiwi
    if species == 'Kiwi':
        fwData = ws.sp.ButterworthBandpass(wData,ws.sampleRate,low=1100,high=7000)
    elif species == 'Ruru':
        fwData = ws.sp.ButterworthBandpass(wData, ws.sampleRate, low=500, high=7000)
    elif species.title()=='Sipo':
        fwData = ws.sp.ButterworthBandpass(wData, ws.sampleRate, low=1200, high=3800)
    #print fwData

    #fwData = data
    waveletCoefs = ws.computeWaveletEnergy(fwData, ws.sampleRate)

    # Compute point-biserial correlations and sort wrt it, return top nNodes
    nodes = ws.compute_r(ws.annotation,waveletCoefs)
    print nodes

    # Now for Nirosha's sorting
    # Basically, for each node, put any of its children (and their children, iteratively) that are in the list in front of it
    nodes = ws.sortListByChild(np.ndarray.tolist(nodes))

    # These nodes refer to the unrooted tree, so add 1 to get the real indices
    nodes = [n + 1 for n in nodes]
    print nodes

    # Generate a full 5 level wavelet packet decomposition
    wpFull = pywt.WaveletPacket(data=fwData, wavelet=ws.wavelet, mode='symmetric', maxlevel=5)

    # Now check the F2 values and add node if it improves F2
    listnodes = []
    bestBetaScore = 0
    m=len(wData)/ws.sampleRate
    detected = np.zeros(m)

    for node in nodes:
        testlist = listnodes[:]
        testlist.append(node)
        print testlist
        detected_c = ws.detectCalls(wpFull,node,ws.sampleRate,n=m,species=species)
        #update the detections
        det=np.maximum.reduce([detected,detected_c])
        fB = ws.fBetaScore(ws.annotation, det)
        if fB > bestBetaScore:
            bestBetaScore = fB
            #now apend the detections of node c to detected
            detected=det
            listnodes.append(node)
        if bestBetaScore == 1:
            break
    return listnodes

# def findCalls_learn(fName=None,data=None, sampleRate=None, species='kiwi',trainTest=False):
#     # import xgboost as xgb
#     # from sklearn.externals import joblib
#
#     if species == 'Kiwi (M)':
#         clf = clf_maleKiwi
#     elif species == 'Kiwi (F)':
#         clf = clf_femaleKiwi
#     elif species == 'Ruru':
#         clf = clf_ruru
#
#     # Second by second, run through the data file and compute the wavelet energy, then classify them
#     segs = []
#     for i in range(0,len(data),sampleRate):
#         currentSec = data[i:(i+1)*sampleRate]
#         # Compute wavelet energy for this second
#         E = computeWaveletEnergy_1s(currentSec, 'dmey2')    # always calculate E on row data not denoised or bp
#         E = np.ones((1,len(E))) * E
#         #segs.append(int(clf.predict(E)[0]))
#         print clf.predict(E)[0]
#         # if int(clf.predict(E)[0]) == 1:
#         #     segs.append([float(i)/sampleRate,float(i+sampleRate)/sampleRate])
#         segs.append(int(clf.predict(E)[0]))
#     print segs
#     return segs

def findCalls_test(fName=None,data=None, sampleRate=None, species='Kiwi',trainTest=False):
    #data, sampleRate_o, annotation = loadData(fName)
    ws=WaveletSegment()
    # print species
    if species.title()=='Kiwi':
        nodes=[34,35,36,38,40,41,42,43,44,45,46,55]
    elif species.title()=='Ruru':
        nodes=[33,37,38]
    elif species.title()=='Sipo':
        nodes = [61,59,54,51,60,58,49,47]
        # print "SIPO nodes:", nodes
    # print nodes
    if fName!=None:
        ws.loadData(fName,trainTest)
    else:
        ws.data=data
        ws.sampleRate=sampleRate
    if species=='boom':
        fs=1000
    if species.title()=='Sipo':
        fs=8000
    else:
        fs = 16000
    if ws.sampleRate != fs:
        ws.data = librosa.core.audio.resample(ws.data,ws.sampleRate,fs)
        ws.sampleRate=fs
    wData = ws.waveletDenoiser.waveletDenoise(ws.data, thresholdType='soft', maxLevel=5)
    if species.title()=='Kiwi':
        fwData = ws.sp.ButterworthBandpass(wData,ws.sampleRate,low=1000,high=7000)
    elif species.title()=='Ruru':
        fwData = ws.sp.ButterworthBandpass(wData,ws.sampleRate,low=500,high=7000)
    elif species.title() == 'Sipo':
        fwData = ws.sp.ButterworthBandpass(wData, ws.sampleRate, low=1200, high=3800)
    wpFull = pywt.WaveletPacket(data=fwData, wavelet=ws.wavelet, mode='symmetric', maxlevel=5)
    # detect based on a previously defined nodeset, default for kiwi
    detected = ws.detectCalls_test(wpFull, ws.sampleRate,listnodes=nodes,species=species,trainTest=trainTest)
    if trainTest==True:
        print fName
        ws.fBetaScore(ws.annotation, detected)
    return ws.mergeSeg(detected)

def binary2seg(self,binary):
    segments=[]
    for i in range(len(binary)):
        if binary[i]==1:
            segments.append([i,i+1])
    return segments

def processFolder(folder_to_process = 'Sound Files/survey/5min', species='Kiwi'):
    #process survey recordings
    nfiles=len(glob.glob(os.path.join(folder_to_process,'*.wav')))
    detected=np.zeros((nfiles,300))
    i=0
    for filename in glob.glob(os.path.join(folder_to_process,'*.wav')):
        ws=WaveletSegment()
        ws.loadData(filename[:-4],trainTest=False)
        wData = ws.waveletDenoise(ws.data, thresholdType='soft', maxlevel=5)
        fwData = ws.sp.ButterworthBandpass(wData,ws.sampleRate,low=1000,high=7000)
        wpFull = pywt.WaveletPacket(data=fwData, wavelet=self.wavelet, mode='symmetric', maxlevel=5)
        detected[i,:] = ws.detectCalls_test(wpFull, ws.sampleRate, nodelist_kiwi) #detect based on a previously defined nodeset
    return detected

def processFolder_train(folder_to_process = 'E:/SONGSCAPE/MakeExecutable/AviaNZ_12thJune/Sound Files/MLPdata/train', species='Kiwi'):
    #Trainig on a set of files
    nfiles=len(glob.glob(os.path.join(folder_to_process,'*.wav')))
    for filename in glob.glob(os.path.join(folder_to_process,'*.wav')):
        nodes=findCalls_train(filename[:-4])
        print filename
        print "Node list:", nodes
        print "**********************************"

def genReport(folder_to_process,detected):
    #generate the report from the detections (yes-no)
    #ToDO: detailed report
    fnames=["" for x in range(np.shape(detected)[0])]
    presenceAbsence=["" for x in range(np.shape(detected)[0])]
    i=0
    for filename in glob.glob(os.path.join(folder_to_process,'*.wav')):
        fnames[i]=str(filename).split('\\')[-1:][0]
    for i in range(np.shape(detected)[0]):
        if sum(detected[i,:])>0:
            presenceAbsence[i]='Yes'
        else:
            presenceAbsence[i]='-'

    col_format = "{:<10}" + "," + "{:<3}" + "\n"
    with open(folder_to_process+"/presenceAbs.csv", 'w') as of:
        for x in zip(fnames,presenceAbsence):
            of.write(col_format.format(*x))

#Test
# nodelist_kiwi = [20, 31, 34, 35, 36, 38, 40, 41, 43, 44, 45, 46] # python
# nodelist_kiwi=[1,15,20,34,35,36,38,40,41,42,43,44,45,46,55] # python with new implimentation
#nodelist_kiwi=[34,35,36,38,40,41,42,43,44,45,46,55] # removed first three nodes from python with new implimentation
#nodelist_kiwi = [34, 35, 36, 38, 40, 41, 42, 43, 44, 45, 46, 55] # matlab
#nodelist_ruru=[33,37,38]




#### SIPO ############################################################
def annotation2GT(datFile):
    # Given the AviaNZ annotation returns the ground truth as an excel
    import math
    from openpyxl import load_workbook, Workbook
    wavFile=datFile[:-5]
    eFile = datFile[:-9]+'-sec.xlsx'
    wavobj = wavio.read(wavFile)
    sampleRate = wavobj.rate
    data = wavobj.data
    n=len(data)/sampleRate   # number of secs
    GT=np.zeros(n)
    with open(datFile) as f:
        segments = json.load(f)
    for seg in segments:
        s=int(math.floor(seg[0]))
        e=int(math.ceil(seg[1]))
        for i in range(s,e):
            GT[i]=1
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="time")
    ws.cell(row=1, column=2, value="call")
    ws.cell(row=1, column=3, value="call type")
    ws.cell(row=1, column=4, value="quality")
    r = 2
    for i in range(len(GT)):
        ws.cell(row=r, column=1, value=str(i + 1))
        ws.cell(row=r, column=2, value=str(int(GT[i])))
        r = r + 1
    wb.save(str(eFile))
    print GT

# generate GT for SIPO
# annotation2GT('E:\Rebecca SIPO\\train\Mt Cass coastal SIPO 07072012.wav.data')
# annotation2GT('E:\Rebecca SIPO\\train\\170617_163003_train.wav.data')

# Train
# nodes= findCalls_train('E:\Rebecca SIPO\\train\Mt Cass coastal SIPO 07072012',species='SIPO')
# print "Nodes for SIPO: ",nodes
# # # nodes_SIPO=[54,26,2]
# #           = [61L, 53L, 60L, 57L]
#
# nodes = findCalls_train('E:\Rebecca SIPO\\train\\170617_163003_train', species='SIPO')
# print "Nodes for SIPO: ",nodes
# [59L, 54L, 51L, 60L, 58L, 49L, 47L]

# d=findCalls_test(fName='E:\Rebecca SIPO\\train\Mt Cass coastal SIPO 07072012',data=None, sampleRate=None, species='SIPO',trainTest=False)
# print d

