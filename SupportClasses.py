
# SupportClasses.py
# Support classes for the AviaNZ program

# Version 3.0 14/09/20
# Authors: Stephen Marsland, Nirosha Priyadarshani, Julius Juodakis, Virginia Listanti

#    AviaNZ bioacoustic analysis program
#    Copyright (C) 2017--2020

#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.

#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.

#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

QtMM = True
try:
    import SupportClasses_GUI
except ImportError:
    print("No GUI")
    QtMM = False

import time
import math
import numpy as np
import os, json
import re
import sys
from tensorflow.keras.models import model_from_json
from tensorflow.keras.models import load_model

class Log(object):
    """ Used for logging info during batch processing.
        Stores most recent analysis for each species, to stay in sync w/ data files.
        Arguments:
        1. path to log file
        2. species
        3. list of other settings of the current analysis

        LOG FORMAT, for each analysis:
        #freetext line
        species
        settings line
        files, multiple lines
    """

    def __init__(self, path, species, settings):
        # in order to append, the previous log must:
        # 1. exist
        # 2. be writeable
        # 3. match current analysis
        # On init, we parse the existing log to see if appending is possible.
        # Actual append/create happens later.
        self.possibleAppend = False
        self.file = path
        self.species = species
        self.settings = ','.join(map(str, settings))
        self.oldAnalyses = []
        self.filesDone = []
        self.currentHeader = ""
        allans = []

        # now, check if the specified log can be resumed:
        if os.path.isfile(path):
            try:
                f = open(path, 'r+')
                print("Found log file at %s" % path)

                lines = [line.rstrip('\n') for line in f]
                f.close()
                lstart = 0
                lend = 1
                # parse to separate each analysis into
                # [freetext, species, settings, [files]]
                # (basically I'm parsing txt into json because I'm dumb)
                while lend<len(lines):
                    #print(lines[lend])
                    if len(lines[lend]) > 0:    # there are empty lines too
                        if lines[lend][0] == "#":
                            allans.append([lines[lstart], lines[lstart+1], lines[lstart+2],
                                            lines[lstart+3 : lend]])
                            lstart = lend
                    lend += 1
                allans.append([lines[lstart], lines[lstart+1], lines[lstart+2],
                                lines[lstart+3 : lend]])

                # parse the log thusly:
                # if current species analysis found, store parameters
                # and compare to check if it can be resumed.
                # store all other analyses for re-printing.
                for a in allans:
                    #print(a)
                    if a[1]==self.species:
                        print("Resumable analysis found")
                        # do not reprint this in log
                        if a[2]==self.settings:
                            self.currentHeader = a[0]
                            # (a1 and a2 match species & settings anyway)
                            self.filesDone = a[3]
                            self.possibleAppend = True
                    else:
                        # store this for re-printing to log
                        self.oldAnalyses.append(a)

            except IOError:
                # bad error: lacking permissions?
                print("ERROR: could not open log at %s" % path)

    def appendFile(self, filename):
        print('Appending %s to log' % filename)
        # attach file path to end of log
        self.file.write(filename)
        self.file.write("\n")
        self.file.flush()

    def appendHeader(self, header, species, settings):
        if header is None:
            header = "#Analysis started on " + time.strftime("%Y %m %d, %H:%M:%S") + ":"
        self.file.write(header)
        self.file.write("\n")
        self.file.write(species)
        self.file.write("\n")
        if type(settings) is list:
            settings = ','.join(settings)
        self.file.write(settings)
        self.file.write("\n")
        self.file.flush()

    def reprintOld(self):
        # push everything from oldAnalyses to log
        # To be called once starting a new log is confirmed
        for a in self.oldAnalyses:
            self.appendHeader(a[0], a[1], a[2])
            for f in a[3]:
                self.appendFile(f)


class ConfigLoader(object):
    """ This deals with reading main config files.
        Not much functionality, but lots of exception handling,
        so moved it out separately.

        Most of these functions return the contents of a corresponding JSON file.
    """

    def config(self, file):
        # At this point, the main config file should already be ensured to exist.
        # It will always be in user configdir, otherwise it would be impossible to find.
        print("Loading software settings from file %s" % file)
        try:
            f = open(file)
            config = json.load(f)
            f.close()
            return config
        except ValueError:
            # if JSON looks corrupt, quit:
            msg = SupportClasses_GUI.MessagePopup("w", "Bad config file", "ERROR: file " + file + " corrupt, delete it to restore default")
            msg.exec_()
            raise

    def filters(self, dir, bats=True):
        """ Returns a dict of filter JSONs,
            named after the corresponding file names.
            bats - include bat filters?
        """
        print("Loading call filters from folder %s" % dir)
        try:
            filters = [f for f in os.listdir(dir) if os.path.isfile(os.path.join(dir, f))]
        except Exception:
            print("Folder %s not found, no filters loaded" % dir)
            return None

        goodfilters = dict()
        for filtfile in filters:
            if not filtfile.endswith("txt"):
                continue
            # Very primitive way to recognize bat filters
            if not bats and filtfile.endswith("Bats.txt"):
                continue
            try:
                ff = open(os.path.join(dir, filtfile))
                filt = json.load(ff)
                ff.close()

                # skip this filter if it looks fishy:
                if not isinstance(filt, dict) or "species" not in filt or "SampleRate" not in filt or "Filters" not in filt or len(filt["Filters"])<1:
                    raise ValueError("Filter JSON format wrong, skipping")
                for subfilt in filt["Filters"]:
                    if not isinstance(subfilt, dict) or "calltype" not in subfilt or "WaveletParams" not in subfilt or "TimeRange" not in subfilt:
                        raise ValueError("Subfilter JSON format wrong, skipping")
                    if "thr" not in subfilt["WaveletParams"] or "nodes" not in subfilt["WaveletParams"] or len(subfilt["TimeRange"])<4:
                        raise ValueError("Subfilter JSON format wrong (details), skipping")

                # if filter passed checks, store it,
                # using filename (without extension) as the key
                goodfilters[filtfile[:-4]] = filt
            except Exception as e:
                print("Could not load filter:", filtfile, e)
        print("Loaded filters:", list(goodfilters.keys()))
        return goodfilters

    def CNNmodels(self, filters, dircnn, targetspecies):
        """ Returns a dict of target CNN models
            Filters - dict of loaded filter files
            Targetspecies - list of species names to load
            """
        print("Loading CNN models from folder %s" % dircnn)
        targetmodels = dict()
        for species in targetspecies:
            filt = filters[species]
            if "CNN" not in filt:
                continue
            elif filt["CNN"]:
                if species == "NZ Bats":
                    try:
                        model = load_model(os.path.join(dircnn, filt["CNN"]["CNN_name"]+'.h5'))
                        targetmodels[species] = [model, filt["CNN"]["win"], filt["CNN"]["inputdim"], filt["CNN"]["output"],
                                                 filt["CNN"]["windowInc"], filt["CNN"]["thr"]]
                        print('Loaded model:', os.path.join(dircnn, filt["CNN"]["CNN_name"]))
                    except Exception as e:
                        print("Could not load CNN model from file:", os.path.join(dircnn, filt["CNN"]["CNN_name"]), e)
                else:
                    try:
                        json_file = open(os.path.join(dircnn, filt["CNN"]["CNN_name"]) + '.json', 'r')
                        loaded_model_json = json_file.read()
                        json_file.close()
                        model = model_from_json(loaded_model_json)
                        model.load_weights(os.path.join(dircnn, filt["CNN"]["CNN_name"]) + '.h5')
                        print('Loaded model:', os.path.join(dircnn, filt["CNN"]["CNN_name"]))
                        model.compile(loss=filt["CNN"]["loss"], optimizer=filt["CNN"]["optimizer"], metrics=['accuracy'])
                        targetmodels[species] = [model, filt["CNN"]["win"], filt["CNN"]["inputdim"], filt["CNN"]["output"],
                                                 filt["CNN"]["windowInc"], filt["CNN"]["thr"]]
                    except Exception as e:
                        print("Could not load CNN model from file:", os.path.join(dircnn, filt["CNN"]["CNN_name"]))
                        print(e)
        print("Loaded CNN models:", list(targetmodels.keys()))
        return targetmodels

    def shortbl(self, file, configdir):
        # A fallback shortlist will be confirmed to exist in configdir.
        # This list is necessary
        print("Loading short species list from file %s" % file)
        try:
            if os.path.isabs(file):
                # user-picked files will have absolute paths
                shortblfile = file
            else:
                # initial file will have relative path,
                # to allow looking it up in various OSes.
                shortblfile = os.path.join(configdir, file)
            if not os.path.isfile(shortblfile):
                print("Warning: file %s not found, falling back to default" % shortblfile)
                shortblfile = os.path.join(configdir, "ListCommonBirds.txt")

            try:
                json_file = open(shortblfile)
                readlist = json.load(json_file)
                json_file.close()
                if len(readlist)>29:
                    print("Warning: short species list has %s entries, truncating to 30" % len(readlist))
                    readlist = readlist[:29]
                return readlist
            except ValueError as e:
                # if JSON looks corrupt, quit and suggest deleting:
                print(e)
                msg = SupportClasses_GUI.MessagePopup("w", "Bad species list", "ERROR: file " + shortblfile + " corrupt, delete it to restore default. Reverting to default.")
                msg.exec_()
                return None

        except Exception as e:
            # if file is not found at all, quit, user must recreate the file or change path
            print(e)
            msg = SupportClasses_GUI.MessagePopup("w", "Bad species list", "ERROR: Failed to load short species list from " + file + ". Reverting to default.")
            msg.exec_()
            return None

    def longbl(self, file, configdir):
        print("Loading long species list from file %s" % file)
        try:
            if os.path.isabs(file):
                # user-picked files will have absolute paths
                longblfile = file
            else:
                # initial file will have relative path,
                # to allow looking it up in various OSes.
                longblfile = os.path.join(configdir, file)
            if not os.path.isfile(longblfile):
                print("Warning: file %s not found, falling back to default" % longblfile)
                longblfile = os.path.join(configdir, "ListDOCBirds.txt")

            try:
                json_file = open(longblfile)
                readlist = json.load(json_file)
                json_file.close()
                return readlist
            except ValueError as e:
                print(e)
                msg = SupportClasses_GUI.MessagePopup("w", "Bad species list", "Warning: file " + longblfile + " corrupt, delete it to restore default. Reverting to default.")
                msg.exec_()
                return None

        except Exception as e:
            print(e)
            msg = SupportClasses_GUI.MessagePopup("w", "Bad species list", "Warning: Failed to load long species list from " + file + ". Reverting to default.")
            msg.exec_()
            return None

    def batl(self, file, configdir):
        print("Loading bat list from file %s" % file)
        try:
            if os.path.isabs(file):
                # user-picked files will have absolute paths
                blfile = file
            else:
                # initial file will have relative path,
                # to allow looking it up in various OSes.
                blfile = os.path.join(configdir, file)
            if not os.path.isfile(blfile):
                print("Warning: file %s not found, falling back to default" % blfile)
                blfile = os.path.join(configdir, "ListBats.txt")

            try:
                json_file = open(blfile)
                readlist = json.load(json_file)
                json_file.close()
                return readlist
            except ValueError as e:
                print(e)
                msg = SupportClasses_GUI.MessagePopup("w", "Bad species list", "Warning: file " + blfile + " corrupt, delete it to restore default. Reverting to default.")
                msg.exec_()
                return None

        except Exception as e:
            print(e)
            msg = SupportClasses_GUI.MessagePopup("w", "Bad species list", "Warning: Failed to load bat list from " + file + ". Reverting to default.")
            msg.exec_()
            return None

    def learningParams(self, file):
        print("Loading software settings from file %s" % file)
        try:
            configfile = open(file)
            config = json.load(configfile)
            configfile.close()
            return config
        except ValueError:
            # if JSON looks corrupt, quit:
            msg = SupportClasses_GUI.MessagePopup("w", "Bad config file", "ERROR: file " + file + " corrupt, delete it to restore default")
            msg.exec_()
            raise

    # Dumps the provided JSON array to the corresponding bird file.
    def blwrite(self, content, file, configdir):
        print("Updating species list in file %s" % file)
        try:
            if os.path.isabs(file):
                # user-picked files will have absolute paths
                file = file
            else:
                # initial file will have relative path,
                # to allow looking it up in various OSes.
                file = os.path.join(configdir, file)

            # no fallback in case file not found - don't want to write to random places.
            with open(file, 'w') as f:
                json.dump(content, f, indent=1)

        except Exception as e:
            print(e)
            msg = SupportClasses_GUI.MessagePopup("w", "Unwriteable species list", "Warning: Failed to write species list to " + file)
            msg.exec_()

    # Dumps the provided JSON array to the corresponding config file.
    def configwrite(self, content, file):
        print("Saving config to file %s" % file)
        try:
            # will always be an absolute path to the user configdir.
            with open(file, 'w') as f:
                json.dump(content, f, indent=1)
        except Exception as e:
            print("Warning: could not save config file:")
            print(e)


class ExcelIO():
    """ Exports the annotations to xlsx, with three sheets:
    time stamps, presence/absence, and per second presence/absence.
    Saves each species into a separate workbook,
    + an extra workbook for all species (to function as a readable segment printout).
    It makes the workbook if necessary.

    Inputs
        segments:   list of SegmentList objects, with additional filename attribute
        dirName:    xlsx will be stored here
        filename:   name of the wav file, to be recorded inside the xlsx
        action:     "append" or "overwrite" any found Excels
        pagelen:    page length, seconds (for filling out absence)
        numpages:   number of pages in this file (of size pagelen)
        speciesList:    list of species that are currently processed -- will force an xlsx output even if none were detected
        startTime:  timestamp for page start, or None to autodetect from file name
        precisionMS:  timestamp resolution for sheet 1: False=in s, True=in ms
        resolution: output resolution (sheet 3) in seconds
    """
    # functions for filling out the excel sheets:
    # First page lists all segments (of a species, if specified)
    # segsLL: list of SegmentList with filename attribute
    # startTime: offset from 0, when exporting a single page
    def writeToExcelp1(self, wb, segsLL, currsp, startTime, precisionMS):
        if precisionMS:
            timeStrFormat = "hh:mm:ss.zzz"
        else:
            timeStrFormat = "hh:mm:ss"
        from PyQt5.QtCore import QTime
        ws = wb['Time Stamps']
        r = ws.max_row + 1

        for segsl in segsLL:
            # extract segments for the current species
            # if species=="All", take ALL segments.
            if currsp=="Any sound":
                speciesSegs = segsl
            else:
                speciesSegs = [segsl[ix] for ix in segsl.getSpecies(currsp)]

            if len(speciesSegs)==0:
                continue

            if startTime is None:
                # if no startTime was provided, try to figure it out based on the filename
                DOCRecording = re.search('(\d{6})_(\d{6})', os.path.basename(segsl.filename)[:-8])

                if DOCRecording:
                    print("time stamp found", DOCRecording)
                    startTimeFile = DOCRecording.group(2)
                    startTimeFile = QTime(int(startTimeFile[:2]), int(startTimeFile[2:4]), int(startTimeFile[4:6]))
                else:
                    startTimeFile = QTime(0,0,0)
            else:
                startTimeFile = QTime(0,0,0).addSecs(startTime)

            # Loop over the segments
            for seg in speciesSegs:
                # Print the filename
                ws.cell(row=r, column=1, value=segsl.filename)

                # Time limits
                ws.cell(row=r, column=2, value=str(startTimeFile.addMSecs(seg[0]*1000).toString(timeStrFormat)))
                ws.cell(row=r, column=3, value=str(startTimeFile.addMSecs(seg[1]*1000).toString(timeStrFormat)))
                # Freq limits
                if seg[3]!=0:
                    ws.cell(row=r, column=4, value=int(seg[2]))
                    ws.cell(row=r, column=5, value=int(seg[3]))
                if currsp=="Any sound":
                    # print species and certainty and call type
                    text = [lab["species"] for lab in seg[4]]
                    ws.cell(row=r, column=6, value=", ".join(text))
                    text = [str(lab["certainty"]) for lab in seg[4]]
                    ws.cell(row=r, column=7, value=", ".join(text))
                    strct = []
                    for lab in seg[4]:
                        if "calltype" in lab:
                            strct.append(str(lab["calltype"]))
                        else:
                            strct.append("-")
                    ws.cell(row=r, column=8, value=", ".join(strct))
                else:
                    # only print certainty and call type
                    strcert = []
                    strct = []
                    for lab in seg[4]:
                        if lab["species"]==currsp:
                            strcert.append(str(lab["certainty"]))
                            if "calltype" in lab:
                                strct.append(str(lab["calltype"]))
                            else:
                                strct.append("-")
                    ws.cell(row=r, column=6, value=", ".join(strcert))
                    ws.cell(row=r, column=7, value=", ".join(strct))
                r += 1

    # This stores pres/abs and max certainty for the species in each file
    # segscert: a 2D list of segs x [start, end, certainty]
    def writeToExcelp2(self, wb, segscert, filename):
        ws = wb['Presence Absence']
        r = ws.max_row + 1

        ws.cell(row=r, column=1, value=filename)

        # segs: a 2D list of [start, end, certainty] for each seg
        if len(segscert)>0:
            pres = "Yes"
            certainty = [lab[2] for lab in segscert]
            certainty = max(certainty)
        else:
            pres = "No"
            certainty = 0
        ws.cell(row=r, column=2, value=pres)
        ws.cell(row=r, column=3, value=certainty)

    # This stores pres/abs (or max cert) for the species
    # in windows of size=resolution in each file
    # segscert: a 2D list of segs x [start, end, certainty]
    # pagenum: index of the current page, 0-base
    # totpages: total number of pages
    # pagelen: page length in s
    def writeToExcelp3(self, wb, segscert, filename, pagenum, pagelen, totpages, resolution):
        # writes binary output DETECTED (per s) from page PAGENUM of length PAGELEN
        starttime = pagenum * pagelen
        ws = wb['Per Time Period']
        r = ws.max_row + 1

        # print resolution "header"
        ws.cell(row=r, column=1, value=str(resolution) + ' secs resolution')
        ft = Font(color="808000")
        ws.cell(row=r, column=1).font=ft

        # print file name and page number
        ws.cell(row=r+1, column=1, value=filename)
        ws.cell(row=r+1, column=2, value=str(pagenum+1))

        detected = np.zeros(math.ceil(pagelen/resolution))
        # convert segs to max certainty at each second
        for seg in segscert:
            # segment start-end, relative to this page start:
            segStart = seg[0] - pagenum*pagelen
            segEnd = seg[1] - pagenum*pagelen
            # just in case of some old reversed segments:
            if segStart > segEnd:
                segStart, segEnd = segEnd, segStart

            # segment is completely outside the current page:
            if segEnd<0 or segStart>pagelen:
                continue

            # convert segment time in s to time in resol windows:
            # map [1..1.999 -> 1
            segStart = max(0, math.floor(segStart/resolution))
            # map 2.0001...3] -> 3
            segEnd = math.ceil(min(segEnd, pagelen)/resolution)
            # range 1:3 selects windows 1 & 2
            for t in range(segStart, segEnd):
                # store certainty if it's larger
                detected[t] = max(detected[t], seg[2])

        # fill the header and detection columns
        c = 3
        for t in range(len(detected)):
            # absolute (within-file) times:
            win_start = starttime + t*resolution
            win_end = min(win_start+resolution, int(pagelen * totpages))
            ws.cell(row=r, column=c, value="%d-%d" % (win_start, win_end))
            ws.cell(row=r, column=c).font = ft
            ws.cell(row=r+1, column=c, value=detected[t])
            c += 1

    def export(self, segments, dirName, action, pagelenarg=None, numpages=1, speciesList=[], startTime=None, precisionMS=False, resolution=10):
        # will export species present in self, + passed as arg, + "all species" excel
        speciesList = set(speciesList)
        for segl in segments:
            for seg in segl:
                speciesList.update([lab["species"] for lab in seg[4]])
        speciesList.add("Any sound")
        print("The following species were detected for export:", speciesList)

        # check source .wav file names -
        # ideally, we store relative paths, but that's not possible across drives:
        for segl in segments:
            try:
                segl.filename = str(os.path.relpath(segl.filename, dirName))
            except Exception as e:
                print("Falling back to absolute paths. Encountered exception:")
                print(e)
                segl.filename = str(os.path.abspath(segl.filename))

        # now, generate the actual files, SEPARATELY FOR EACH SPECIES:
        for species in speciesList:
            print("Exporting species %s" % species)
            # clean version for filename
            speciesClean = re.sub(r'\W', "_", species)

            # setup output files:
            # if an Excel exists, append (so multiple files go into one worksheet)
            # if not, create new
            eFile = os.path.join(dirName, 'DetectionSummary_' + speciesClean + '.xlsx')

            if action == "overwrite" or not os.path.isfile(eFile):
                # make a new workbook:
                wb = Workbook()

                # First sheet
                wb.create_sheet(title='Time Stamps', index=1)
                ws = wb['Time Stamps']
                ws.cell(row=1, column=1, value="File Name")
                if precisionMS:
                    ws.cell(row=1, column=2, value="start (hh:mm:ss.ms)")
                    ws.cell(row=1, column=3, value="end (hh:mm:ss.ms)")
                else:
                    ws.cell(row=1, column=2, value="start (hh:mm:ss)")
                    ws.cell(row=1, column=3, value="end (hh:mm:ss)")
                ws.cell(row=1, column=4, value="min freq. (Hz)")
                ws.cell(row=1, column=5, value="max freq. (Hz)")
                if species=="Any sound":
                    ws.cell(row=1, column=6, value="species")
                    ws.cell(row=1, column=7, value="certainty")
                    ws.cell(row=1, column=8, value="call type")
                else:
                    ws.cell(row=1, column=6, value="certainty")
                    ws.cell(row=1, column=7, value="call type")

                    # Second sheet
                    wb.create_sheet(title='Presence Absence', index=2)
                    ws = wb['Presence Absence']
                    ws.cell(row=1, column=1, value="File Name")
                    ws.cell(row=1, column=2, value="Present?")
                    ws.cell(row=1, column=3, value="Certainty, %")

                    # Third sheet
                    wb.create_sheet(title='Per Time Period', index=3)
                    ws = wb['Per Time Period']
                    ws.cell(row=1, column=1, value="File Name")
                    ws.cell(row=1, column=2, value="Page")
                    ws.cell(row=1, column=3, value="Maximum certainty of species presence (0 = absent)")

                # Hack to delete original sheet
                del wb['Sheet']
            elif action == "append":
                try:
                    wb = load_workbook(eFile)
                except Exception as e:
                    print("ERROR: cannot open file %s to append" % eFile)  # no read permissions or smth
                    print(e)
                    return 0
            else:
                print("ERROR: unrecognised action", action)
                return 0

            # export segments
            self.writeToExcelp1(wb, segments, species, startTime, precisionMS)

            if species!="Any sound":
                # loop over all SegmentLists, i.e. for each wav file:
                for segsl in segments:
                    # extract the certainty from each label for current species
                    # to a 2D list of segs x [start, end, certainty]
                    # (for this wav file)
                    speciesCerts = []
                    for seg in segsl:
                        for lab in seg[4]:
                            if lab["species"]==species:
                                speciesCerts.append([seg[0], seg[1], lab["certainty"]])

                    # export presence/absence and max certainty
                    self.writeToExcelp2(wb, speciesCerts, segsl.filename)

                    # either read duration from this SegList
                    # or need current page length if called from manual
                    # (assuming all pages are of same length as current data)
                    if pagelenarg is None:
                        pagelen = math.ceil(segsl.metadata["Duration"])
                    else:
                        pagelen = pagelenarg

                    # Generate pres/abs per custom resolution windows
                    for p in range(0, numpages):
                        self.writeToExcelp3(wb, speciesCerts, segsl.filename, p, pagelen, numpages, resolution)

            # Save the file
            try:
                wb.save(eFile)
            except Exception as e:
                print("ERROR: could not create new file %s" % eFile)  # no read permissions or smth
                print(e)
                return 0
        return 1

