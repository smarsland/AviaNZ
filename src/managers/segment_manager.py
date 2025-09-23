# SegmentManager.py
# Manages segment data operations for AviaNZ
# Extracted from AviaNZ_manual.py for better modularity

import copy
import numpy as np
from PyQt6.QtCore import QObject, pyqtSignal
from PyQt6.QtWidgets import QMessageBox
import pyqtgraph as pg
import pyqtgraph.functions as fn
from ..core import Segment
from ..ui import SupportClasses_GUI


class SegmentManager(QObject):
    """Manages segment data operations, selection, and overview display.
    
    This class handles all segment-related business logic including:
    - Creating, deleting, and modifying segments
    - Managing segment selection state
    - Updating overview display colors
    - Coordinating with UI for segment controls
    """
    
    # Signals for communicating with main window
    segment_added = pyqtSignal(object, int)  # Emitted when a segment is created (segment, index)
    segment_deleted = pyqtSignal(int)  # Emitted with segment index when deleted
    segment_updated = pyqtSignal(int)  # Emitted when segment bounds/labels change
    segment_selection_changed = pyqtSignal(int)  # Emitted with new selection (-1 for none)
    segment_labels_updated = pyqtSignal(int, list)  # Emitted when segment labels change (id, labels)
    overview_update_requested = pyqtSignal(object, bool)  # Emitted when overview needs update (segment, delete)
    segments_to_save_changed = pyqtSignal(bool)  # Emitted when save state changes
    
    def __init__(self, config_manager, audio_file_manager, display_manager):
        super().__init__()
        
        self.config_manager = config_manager
        self.audio_file_manager = audio_file_manager
        self.display_manager = display_manager
        
        # Segment data
        from ..core import Segment
        self.segments = Segment.SegmentList()
        self.segments.metadata = {}
        self.segments_to_save = False
        self.box1id = -1  # Currently selected segment (-1 = none)
        
        # Audio context
        self.startRead = 0
        self.datalengthSec = 0
        self.batmode = False
        
        # Overview display context (now accessed via display_manager)
        # self.overviewSegments, self.widthOverviewSegment, self.SegmentRects 
        # are no longer stored here to avoid stale references
        
    def set_audio_context(self, startRead, datalengthSec, batmode):
        """Set current audio and display context for segment operations."""
        self.startRead = startRead
        self.datalengthSec = datalengthSec
        self.batmode = batmode
        
    def set_overview_context(self, overviewSegments, widthOverviewSegment, SegmentRects):
        """Set overview tracking variables (deprecated - now uses display_manager directly)."""
        # This method is kept for compatibility but the values are no longer used
        # The update_overview_counters method now accesses display_manager arrays directly
        pass
        
    def refreshSegmentControls(self):
        """Toggles all the segment controls on/off when a segment
            is (de)selected. Call this after changing self.box1id.
            Remember to update this when segment controls change!
        """
        # Just return control state data, UI will handle button enabling
        segment_selected = self.box1id >= 0
        is_rectangle = False
        has_unconfirmed = False
        
        if segment_selected and self.box1id < len(self.segments):
            # Check if selected segment has frequency bounds (is rectangle)
            segment = self.segments[self.box1id]
            is_rectangle = segment[2] != 0 or segment[3] != 0
            
            # Check if has unconfirmed labels
            for sp in segment[4]:
                if sp["certainty"] < 100 and sp["species"] != "Don't Know":
                    has_unconfirmed = True
                    break
        
        return {
            'segment_selected': segment_selected,
            'is_rectangle': is_rectangle,
            'has_unconfirmed': has_unconfirmed,
            'batmode': getattr(self, 'batmode', False)
        }

    def addSegment(self, startpoint, endpoint, y1=0, y2=0, species=[], saveSeg=True, index=-1, remaking=False, coordsAbsolute=False):
        """Creates a new segment and adds it to the segments list.
        This is the data-handling portion of segment creation.
        
        Args:
            startpoint, endpoint: Time bounds (seconds from page start, unless coordsAbsolute)
            y1, y2: Frequency bounds (Hz), 0 means no frequency constraint
            species: List of species labels with certainties
            saveSeg: Whether to add to segments list (False for display-only segments)
            index: For remaking segments, which index to update (-1 = append)
            remaking: Whether this is recreating an existing segment
            coordsAbsolute: Whether times are relative to file start (True) or page start (False)
        
        Returns:
            Created Segment object, or None if creation failed
        """
        
        # Deselect current segment if any
        if self.box1id > -1:
            old_boxid = self.box1id
            self.box1id = -1
            self.segment_selection_changed.emit(-1)

        # Ensure proper order
        if startpoint > endpoint:
            startpoint, endpoint = endpoint, startpoint
        if y1 > y2:
            y1, y2 = y2, y1

        # Default species if empty
        if len(species) == 0:
            species = [{"species": "Don't Know", "certainty": 0, "filter": "M"}]
        else:
            species = copy.deepcopy(species)

        # Convert coordinates if needed
        if coordsAbsolute:
            # Convert from absolute times to relative-to-page times
            startpoint = startpoint - self.startRead
            endpoint = endpoint - self.startRead

        # Visibility check for display-only segments
        if not saveSeg:
            if endpoint < 0 or startpoint > self.datalengthSec:
                print("Warning: a segment was not shown")
                return None
            elif y1 != 0 and y2 != 0:
                # Would need freq show bounds from main window for this check
                pass

        # Create the segment object
        # Note: we convert time from _relative to page_ to _relative to file start_
        try:
            newSegment = Segment.Segment([startpoint + self.startRead, endpoint + self.startRead, y1, y2, species])
        except Exception as e:
            print(f"Error creating segment: {e}")
            return None

        # Add to data if requested
        if saveSeg:
            if remaking and index >= 0:
                self.segments[index] = newSegment
            else:
                self.segments.append(newSegment)
            self.segments_to_save = True
            self.segments_to_save_changed.emit(True)

        # Update overview if this segment affects the display
        if saveSeg or (startpoint >= 0 and endpoint <= self.datalengthSec):
            self.overview_update_requested.emit(newSegment, False)

        # Emit signal for UI to create graphics
        self.segment_added.emit(newSegment, len(self.segments) - 1 if saveSeg else -1)
        
        return newSegment

    def selectSegment(self, boxid):
        """Changes segment selection state and notifies UI."""
        self.box1id = boxid
        self.segment_selection_changed.emit(boxid)

    def deselectSegment(self, boxid):
        """Deselects segment and notifies UI."""
        self.box1id = -1
        self.segment_selection_changed.emit(-1)

    def updateRegion_spec(self, segment_index, new_start_spec, new_end_spec, new_y1_freq, new_y2_freq):
        """Handle spectrogram segment position updates.
        
        Args:
            segment_index: Index of the segment being updated
            new_start_spec, new_end_spec: New time bounds in spectrogram coordinates
            new_y1_freq, new_y2_freq: New frequency bounds in Hz
        """
        if segment_index < 0 or segment_index >= len(self.segments):
            return False
            
        # Update overview with old position
        self.overview_update_requested.emit(self.segments[segment_index], True)
        
        # Update segment data
        self.segments[segment_index][0] = new_start_spec + self.startRead
        self.segments[segment_index][1] = new_end_spec + self.startRead
        self.segments[segment_index][2] = new_y1_freq
        self.segments[segment_index][3] = new_y2_freq
        
        self.segments_to_save = True
        self.segments_to_save_changed.emit(True)
        
        # Update overview with new position
        self.overview_update_requested.emit(self.segments[segment_index], False)
        
        # Notify UI of the update
        self.segment_updated.emit(segment_index)
        return True

    def updateRegion_ampl(self, segment_index, new_start_ampl, new_end_ampl):
        """Handle amplitude plot segment position updates.
        
        Args:
            segment_index: Index of the segment being updated
            new_start_ampl, new_end_ampl: New time bounds in amplitude coordinates
        """
        if segment_index < 0 or segment_index >= len(self.segments):
            return False
            
        # Update overview with old position
        self.overview_update_requested.emit(self.segments[segment_index], True)
        
        # Update segment data
        self.segments[segment_index][0] = new_start_ampl + self.startRead
        self.segments[segment_index][1] = new_end_ampl + self.startRead
        
        self.segments_to_save = True
        self.segments_to_save_changed.emit(True)
        
        # Update overview with new position
        self.overview_update_requested.emit(self.segments[segment_index], False)
        
        # Notify UI of the update
        self.segment_updated.emit(segment_index)
        return True
        
    def addRegularSegments(self, protocolSize, protocolInterval, duration):
        """Add regular segments to the spectrogram based on protocol settings."""
        if self.box1id > -1:
            old_boxid = self.box1id
            self.box1id = -1
            self.segment_selection_changed.emit(-1)
            
        segtimes = [(seg[0], seg[1]) for seg in self.segments]
        i = 0
        print(f"Adding segments ({protocolSize} s every {protocolInterval} s)")
        
        while i < duration:
            # Check for segment presence to avoid overlaps
            segPresent = False
            for start, end in segtimes:
                if not (i + protocolSize <= start or i >= end):
                    segPresent = True
                    break
                    
            if not segPresent:
                species = [{"species": "Don't Know", "certainty": 0, "filter": "M"}]
                self.addSegment(i, i + protocolSize, y1=0, y2=0, species=species, coordsAbsolute=True)
                
            i += protocolInterval

    def update_overview_counters(self, segment, delete=False):
        """Update overview segment counters for data tracking.
        
        This method only updates the data counters - UI updates are handled
        by the main window/display manager in response to signals.
        
        Args:
            segment: The segment to update counters for
            delete: True if removing the segment, False if adding
        """
        # Always use display manager's arrays directly to avoid stale references
        if not hasattr(self.display_manager, 'overviewSegments') or self.display_manager.overviewSegments is None:
            return
            
        overviewSegments = self.display_manager.overviewSegments
        widthOverviewSegment = getattr(self.display_manager, 'widthOverviewSegment', 0)
        
        if widthOverviewSegment == 0:
            return
            
        # Calculate which overview segments this segment spans
        # Use audio processor from display manager to get proper conversion
        audio_processor = self.display_manager.audio_processor
        if audio_processor:
            inds = max(0, int(audio_processor.convertAmpltoSpec(segment[0] - self.startRead) / widthOverviewSegment))
            inde = min(int(audio_processor.convertAmpltoSpec(segment[1] - self.startRead) / widthOverviewSegment), 
                      len(overviewSegments) - 1)
        else:
            print("Warning: No audio processor available for overview calculation")
            return

        # Update counters based on label certainties
        for label in segment[4]:
            if label["certainty"] == 0:
                # "red" label counter
                if delete:
                    overviewSegments[inds:inde + 1, 0] -= 1
                else:
                    overviewSegments[inds:inde + 1, 0] += 1
            elif label["certainty"] == 100:
                # "green" label counter
                if delete:
                    overviewSegments[inds:inde + 1, 1] -= 1
                else:
                    overviewSegments[inds:inde + 1, 1] += 1
            else:
                # "yellow" label counter
                if delete:
                    overviewSegments[inds:inde + 1, 2] -= 1
                else:
                    overviewSegments[inds:inde + 1, 2] += 1

        # Prevent negative counters
        if np.any(overviewSegments < 0):
            print("Warning: something went wrong with overview colors!")
            overviewSegments[overviewSegments < 0] = 0

    def confirmSegment(self):
        """Confirm labels on the selected segment (set certainty to 100)."""
        segment_id = self.box1id
        
        if segment_id > -1:
            # Update overview counters with old labels
            self.update_overview_counters(self.segments[segment_id], delete=True)
            
            # Confirm all labels
            self.segments[segment_id].confirmLabels()
            
            # Update overview counters with new labels
            self.update_overview_counters(self.segments[segment_id])
            
            # Send signal for UI to update overview display
            self.overview_update_requested.emit(self.segments[segment_id], False)
            
            self.segments_to_save = True
            self.segments_to_save_changed.emit(True)
            
            # Notify UI of label changes
            self.segment_labels_updated.emit(segment_id, self.segments[segment_id][4])
            
            return True
        return False

    def deleteSegment(self, segment_id=-1):
        """Delete a segment by ID."""
        if segment_id < 0:
            segment_id = self.box1id

        if segment_id > -1 and segment_id < len(self.segments):
            # Update overview counters
            self.update_overview_counters(self.segments[segment_id], delete=True)
            
            # Send signal for UI to update overview display  
            self.overview_update_requested.emit(self.segments[segment_id], True)
            
            # Remove from data
            del self.segments[segment_id]
            self.segments_to_save = True
            self.segments_to_save_changed.emit(True)

            # Update selection
            self.box1id = -1
            self.segment_selection_changed.emit(-1)
            
            # Notify UI for graphics cleanup
            self.segment_deleted.emit(segment_id)
            
            return True
        return False
        
    def get_segment_count(self):
        """Get the number of segments."""
        return len(self.segments)
        
    def get_segment(self, index):
        """Get a segment by index."""
        if 0 <= index < len(self.segments):
            return self.segments[index]
        return None
        
    def get_selected_segment(self):
        """Get the currently selected segment."""
        if self.box1id >= 0 and self.box1id < len(self.segments):
            return self.segments[self.box1id]
        return None
        
    def get_selected_segment_id(self):
        """Get the ID of the currently selected segment."""
        return self.box1id
        
    def set_segments_data(self, segments):
        """Set the segments data (for loading from file)."""
        if segments:
            self.segments = segments
        else:
            self.segments = Segment.SegmentList()
        self.segments_to_save = False
        self.segments_to_save_changed.emit(False)
        
    def clear_segments(self):
        """Clear all segments."""
        self.segments = Segment.SegmentList()
        self.segments_to_save = False
        self.segments_to_save_changed.emit(False)
        self.box1id = -1
        self.segment_selection_changed.emit(-1)
        
    def get_segment_info_string(self, segment_id):
        """Get info string for a segment"""
        if segment_id < 0 or segment_id >= len(self.segments):
            return ""
        return self.segments[segment_id].infoString()
        
    def has_segments_to_save(self):
        """Check if there are unsaved changes to segments."""
        return self.segments_to_save
        
    def mark_segments_saved(self):
        """Mark segments as saved (no unsaved changes)."""
        self.segments_to_save = False
        self.segments_to_save_changed.emit(False)
        self.datalengthSec = 0  # Length of current page in seconds
        
        # Overview management
        self.overviewSegments = None
        self.widthOverviewSegment = 0
        
        # File metadata
        self.filename = None
        
    def set_dependencies(self, audio_processor, display_manager):
        """Set the audio processor and display manager dependencies"""
        self.audio_processor = audio_processor
        self.display_manager = display_manager
        
    def set_audio_context(self, startRead, datalengthSec, overviewSegments=None, widthOverviewSegment=0):
        """Set the current audio context for coordinate conversions"""
        self.startRead = startRead
        self.datalengthSec = datalengthSec
        if overviewSegments is not None:
            self.overviewSegments = overviewSegments
        if widthOverviewSegment > 0:
            self.widthOverviewSegment = widthOverviewSegment
    
    def set_segments(self, segments):
        """Set the segments list (when loading a file)"""
        if segments and hasattr(segments, 'metadata'):
            self.segments = segments
        else:
            # Create new SegmentList and copy data if needed
            self.segments = Segment.SegmentList()
            if segments:
                self.segments.extend(segments)
        self.segments_to_save = False
        self.box1id = -1
        
    def get_segments(self):
        """Get the current segments list"""
        return self.segments
        
    def get_segments_to_save(self):
        """Get whether segments need saving"""
        return self.segments_to_save
        
    def create_segment(self, startpoint, endpoint, y1=0, y2=0, species=[], saveSeg=True, index=-1, coordsAbsolute=False):
        """
        Create a new segment and handle the data operations.
        Returns the new segment object if successful, None if not visible.
        
        Args:
        startpoint, endpoint - in secs, either from page start, or absolute (then set coordsAbsolute=True)
        y1, y2 should be the frequencies (between 0 and Fs//2)
        species - list of labels (including certainties, .data format)
        saveSeg - store the created segment on self.segments. Set to False when drawing the saved ones.
        coordsAbsolute - set to True to accept start,end in absolute coords (from file start)
        """
        
        # Deselect any currently selected segment
        if self.box1id > -1:
            self.select_segment(-1)
            
        # Make sure startpoint and endpoint are in the right order
        if startpoint > endpoint:
            startpoint, endpoint = endpoint, startpoint
            
        # same for freqs
        if y1 > y2:
            y1, y2 = y2, y1
            
        # since we allow passing empty list here:
        if len(species) == 0:
            species = [{"species": "Don't Know", "certainty": 0, "filter": "M"}]
        else:
            species = copy.deepcopy(species)

        if coordsAbsolute:
            # convert from absolute times to relative-to-page times
            startpoint = startpoint - self.startRead
            endpoint = endpoint - self.startRead

        if not saveSeg:
            # check if this segment fits in the current spectrogram page
            if endpoint < 0 or startpoint > self.datalengthSec:
                print("Warning: a segment was not shown")
                return None
            elif y1!=0 and y2!=0 and self.audio_processor and (y1 > self.audio_processor.sp.maxFreqShow or y2 < self.audio_processor.sp.minFreqShow):
                print("Warning: a segment was not shown")
                return None
        else:
            self.segments_to_save = True
            self.segments_to_save_changed.emit(True)

        # Create a Segment. This will check for errors and standardize the labels
        # Note: we convert time from _relative to page_ to _relative to file start_
        newSegment = Segment.Segment([startpoint+self.startRead, endpoint+self.startRead, y1, y2, species])

        # Add the segment to the data
        if saveSeg:
            if index > -1:
                self.segments[index] = newSegment
            else:
                self.segments.append(newSegment)
                index = len(self.segments) - 1

        # Update overview data
        self.update_overview_counters(newSegment)
        
        # Emit signal for UI to create graphics
        self.segment_added.emit(newSegment, index)
        
        return newSegment
    
    def delete_segment(self, id=-1):
        """
        Delete a segment by ID.
        Returns True if deletion was successful, False otherwise.
        """
        if id < 0:
            id = self.box1id
            
        if id < 0 or id >= len(self.segments):
            return False
            
        # Update overview before deletion
        self.update_overview_counters(self.segments[id], delete=True)
        
        # Remove from segments list
        del self.segments[id]
        self.segments_to_save = True
        self.segments_to_save_changed.emit(True)
        
        # Deselect if this was the selected segment
        if self.box1id == id:
            self.box1id = -1
            self.segment_selection_changed.emit(-1)
        elif self.box1id > id:
            # Adjust selection index if needed
            self.box1id -= 1
            self.segment_selection_changed.emit(self.box1id)
            
        # Emit signal for UI cleanup
        self.segment_deleted.emit(id)
        
        return True
        
    def select_segment(self, boxid):
        """
        Select a segment by ID.
        """
        self.box1id = boxid
        self.segment_selection_changed.emit(boxid)
        
    def deselect_segment(self):
        """
        Deselect the currently selected segment.
        """
        self.box1id = -1
        self.segment_selection_changed.emit(-1)
        
    def confirm_segment_labels(self):
        """
        Confirm labels on the currently selected segment (set certainty to 100).
        Returns True if successful, False if no segment selected.
        """
        if self.box1id < 0:
            return False
            
        # Update overview with old segment first
        self.update_overview_counters(self.segments[self.box1id], delete=True)
        
        # Confirm the labels
        self.segments[self.box1id].confirmLabels()
        
        # Update overview with new segment
        self.update_overview_counters(self.segments[self.box1id])
        
        self.segments_to_save = True
        self.segments_to_save_changed.emit(True)
        
        # Emit signal for UI updates
        self.segment_labels_updated.emit(self.box1id, self.segments[self.box1id].getLabels())
        
        return True
        
    def update_segment_bounds(self, segment_id, start_sec, end_sec, y1=None, y2=None):
        """
        Update the time and/or frequency bounds of a segment.
        Times should be relative to current page start.
        """
        if segment_id < 0 or segment_id >= len(self.segments):
            return False
            
        # Update overview with old segment first
        self.update_overview_counters(self.segments[segment_id], delete=True)
        
        # Update segment bounds
        self.segments[segment_id][0] = start_sec + self.startRead  # absolute start time
        self.segments[segment_id][1] = end_sec + self.startRead    # absolute end time
        
        if y1 is not None:
            self.segments[segment_id][2] = y1
        if y2 is not None:
            self.segments[segment_id][3] = y2
            
        # Update overview with new segment
        self.update_overview_counters(self.segments[segment_id])
        
        # Send signal for UI overview update
        self.overview_update_requested.emit(self.segments[segment_id], False)
        
        self.segments_to_save = True
        self.segments_to_save_changed.emit(True)
        
        # Emit signal for UI updates
        self.segment_updated.emit(segment_id)
        
        return True
        
    def add_regular_segments(self, duration_total, protocol_size, protocol_interval):
        """
        Add regular segments at specified intervals.
        
        Args:
            duration_total: Total duration of the file in seconds
            protocol_size: Size of each segment in seconds  
            protocol_interval: Interval between segment starts in seconds
        """
        if self.box1id > -1:
            self.select_segment(-1)
            
        segtimes = [(seg[0], seg[1]) for seg in self.segments]
        i = 0
        print("Adding segments (%d s every %d s)" % (protocol_size, protocol_interval))
        
        while i < duration_total:
            # check for segment presence in case of double click or other issues
            if len(segtimes) > 0 and (i, i + protocol_size) in segtimes:
                print("segment already exists, skipping")
            else:
                self.create_segment(i, i + protocol_size, coordsAbsolute=True)
            i += protocol_interval
            
    def clear_all_segments(self):
        """Clear all segments"""
        from ..core import Segment
        self.segments = Segment.SegmentList()
        # Preserve metadata if it exists
        if hasattr(self, '_saved_metadata'):
            self.segments.metadata = self._saved_metadata
        else:
            self.segments.metadata = {}
        self.box1id = -1
        self.segments_to_save = True
        self.segments_to_save_changed.emit(True)
        self.segment_selection_changed.emit(-1)
        
    def load_segments_from_file(self, filename, file_length, operator, reviewer):
        """Load segments from data file"""
        import os
        
        # Initialize empty segment list
        self.segments = Segment.SegmentList()
        
        # Load any previous segments stored
        if os.path.isfile(filename) and os.stat(filename).st_size > 0:
            # Populate it, add the metadata attribute
            # (note: we're overwriting the JSON duration with actual full wav size)
            hasmetadata = self.segments.parseJSON(filename, file_length)
            if not hasmetadata:
                self.segments.metadata["Operator"] = operator
                self.segments.metadata["Reviewer"] = reviewer
                self.segments_to_save = True
        else:
            # Create empty list with metadata
            self.segments.metadata = {"Operator": operator, "Reviewer": reviewer}
            
        print(f"{len(self.segments)} segments read")
        
    def get_metadata(self):
        """Get segment metadata"""
        if hasattr(self.segments, 'metadata'):
            return self.segments.metadata
        return {}
        
    def restore_segments(self, prev_segments):
        """Restore segments from backup (for undo operations)"""
        import copy
        
        # Create new segment list and copy the previous segments
        self.segments = Segment.SegmentList()
        if hasattr(prev_segments, 'metadata'):
            self.segments.metadata = copy.deepcopy(prev_segments.metadata)
        
        # Copy all segment data
        for seg in prev_segments:
            self.segments.append(copy.deepcopy(seg))
            
        self.segments_to_save = True
        self.segments_to_save_changed.emit(True)
        
    def add_segment_object(self, segment_obj):
        """Add a pre-created segment object to the list"""
        self.segments.append(segment_obj)
        self.segments_to_save = True
        self.segments_to_save_changed.emit(True)
        
    def update_metadata(self, key, value):
        """Update a metadata field"""
        if hasattr(self.segments, 'metadata'):
            self.segments.metadata[key] = value
        else:
            self.segments.metadata = {key: value}
        self.segments_to_save = True
        self.segments_to_save_changed.emit(True)
        
    def export_segment_statistics(self, filename, audio_data, sample_rate, start_read, data_length_sec, 
                                  audio_processor, config_window_width, config_incr):
        """Calculate and export summary statistics for the currently marked segments.
        
        Args:
            filename: Base filename for the output CSV (without extension)
            audio_data: The audio data array
            sample_rate: Audio sample rate
            start_read: Start time of current page in seconds
            data_length_sec: Length of current page in seconds
            audio_processor: AudioProcessor instance for coordinate conversions
            config_window_width: Window width from config
            config_incr: Increment from config
            
        Returns:
            str: Path to the created CSV file
        """
        from ..core import Features
        
        output_file = filename.rsplit('.', 1)[0] + '_features.csv'
        
        with open(output_file, "w") as cs:
            cs.write("Start Time (sec),End Time (sec),Avg Power,Delta Power,Energy,Agg Entropy,Avg Entropy,Max Power,Max Freq\n")

            for seg in self.segments:
                # Skip segments that are not visible in current page
                if seg[1] <= start_read or seg[0] >= start_read + data_length_sec:
                    continue

                # Coordinates in seconds from current page start, bounded at page borders
                starttime = max(0, seg[0] - start_read)
                endtime = min(seg[1] - start_read, data_length_sec)

                # Extract audio data for this segment
                start_sample = int(starttime * sample_rate)
                end_sample = int(endtime * sample_rate)
                segment_data = audio_data[start_sample:end_sample]

                print(f"Calculating statistics for segment {starttime:.4f}-{endtime:.4f}s...")

                # Calculate features
                f = Features.Features(data=segment_data, sampleRate=sample_rate, 
                                    window_width=config_window_width, incr=config_incr)
                
                # Get frequency conversion bounds (500-8000 Hz)
                f1 = int(audio_processor.convertFreqtoY(500))
                f2 = int(audio_processor.convertFreqtoY(8000))
                
                avgPower, deltaPower, energy, aggEntropy, avgEntropy, maxPower, maxFreq = \
                    f.get_Raven_spectrogram_measurements(f1=f1, f2=f2)
                
                cs.write("%.4f,%.4f,%.2f,%.2f,%.2f,%.2f,%.2f,%.2f,%.2f\n" % 
                        (starttime, endtime, avgPower, deltaPower, energy, 
                         aggEntropy, avgEntropy, maxPower, maxFreq))
        
        return output_file
        
    def import_from_excel(self, excelfile, audiofile, species, calltype, 
                         colstart_ind, colend_ind, collow_ind, colhigh_ind):
        """Import segments from Excel file and create AviaNZ annotation file.
        
        Args:
            excelfile: Path to Excel file
            audiofile: Path to audio file (will create audiofile.data)
            species: Species name for all segments
            calltype: Call type (optional, can be "")
            colstart_ind: Column index for start times (1-based)
            colend_ind: Column index for end times (1-based)
            collow_ind: Column index for low frequencies (1-based)
            colhigh_ind: Column index for high frequencies (1-based)
            
        Returns:
            tuple: (success: bool, message: str, annotation_file: str)
        """
        import openpyxl
        import openpyxl.utils
        import soundfile as sf
        import json
        
        try:
            # Validation
            if calltype == "Add":
                return False, "calltype cannot be 'Add'", ""
            if species == "Other":
                return False, "species cannot be 'Other'", ""
                
            # Read Excel file
            book = openpyxl.load_workbook(excelfile)
            sheet = book.active
            
            # Convert column indices to letters
            colstart = openpyxl.utils.get_column_letter(colstart_ind)
            colend = openpyxl.utils.get_column_letter(colend_ind)
            collow = openpyxl.utils.get_column_letter(collow_ind)
            colhigh = openpyxl.utils.get_column_letter(colhigh_ind)
            
            # Read data from specified columns (skip header row)
            starttime = sheet[colstart+'2': colstart + str(sheet.max_row)]
            endtime = sheet[colend+'2': colend + str(sheet.max_row)]
            flow = sheet[collow+'2': collow + str(sheet.max_row)]
            fhigh = sheet[colhigh+'2': colhigh + str(sheet.max_row)]

            # Get audio file info
            info = sf.info(audiofile)
            duration = info.frames / info.samplerate

            # Create annotation segments
            annotation = []
            for i in range(len(starttime)):
                if calltype == "":
                    segment_data = [
                        float(starttime[i][0].value), 
                        float(endtime[i][0].value), 
                        float(flow[i][0].value),
                        float(fhigh[i][0].value),
                        [{"species": species, "certainty": 100.0, "filter": "M"}]
                    ]
                else:
                    segment_data = [
                        float(starttime[i][0].value), 
                        float(endtime[i][0].value), 
                        float(flow[i][0].value),
                        float(fhigh[i][0].value),
                        [{"species": species, "certainty": 100.0, "filter": "M", "calltype": calltype}]
                    ]
                annotation.append(segment_data)

            # Add metadata header
            annotation.insert(0, {"Operator": "", "Reviewer": "", "Duration": duration})
            
            # Save annotation file
            annotation_file = audiofile + '.data'
            with open(annotation_file, 'w') as f:
                json.dump(annotation, f)
            
            return True, f"Successfully saved annotation file: {annotation_file}", annotation_file
            
        except Exception as e:
            return False, f"ERROR: Generating annotation failed with error: {str(e)}", ""
            
    def import_from_freebird(self, sessiondir, freebird_list_file, configdir):
        """Import segments from Freebird XML files and create AviaNZ annotation files.
        
        Args:
            sessiondir: Path to Freebird session directory
            freebird_list_file: Path to Freebird species list file (.csv or .xlsx)
            configdir: Configuration directory path
            
        Returns:
            tuple: (success: bool, message: str, processed_files: list)
        """
        import xml.etree.ElementTree as ET
        import openpyxl
        import csv
        import os
        
        try:
            # Clean up session directory path
            if sessiondir.endswith(".session"):
                sessiondir = sessiondir[:-8]
                
            # Load species mapping from Freebird list
            spName = []
            spCode = []
            
            # Determine file path (absolute or relative to configdir)
            if not os.path.isabs(freebird_list_file):
                filename = os.path.join(configdir, freebird_list_file)
            else:
                filename = freebird_list_file

            # Read species mapping based on file extension
            if freebird_list_file.endswith('.csv'):
                try:
                    with open(filename, mode='r', encoding='utf-8') as f:
                        cs = csv.DictReader(f)
                        for l in cs:
                            if l['FreebirdCode'] != '':
                                spName.append(l['SpeciesName'])
                                spCode.append(int(l['FreebirdCode']))
                except Exception:
                    print("Warning: Could not read Freebird species list CSV")
                    
            elif freebird_list_file.endswith('.xlsx'):
                try:
                    book = openpyxl.load_workbook(filename)
                    sheet = book.active
                    name = sheet['A2': 'A' + str(sheet.max_row)]
                    code = sheet['C2': 'C' + str(sheet.max_row)]
        
                    for i in range(len(name)):
                        spName.append(str(name[i][0].value))
                    for i in range(len(code)):
                        if code[i][0].value is not None:
                            spCode.append(int(code[i][0].value))
                        else:
                            spCode.append(-1)
                except Exception:
                    print("Warning: Could not read Freebird species list Excel")

            spDict = dict(zip(spCode, spName))
            processed_files = []

            # Process each .tag file in the session directory
            for root, dirs, files in os.walk(sessiondir):
                for file in files:
                    if file.endswith('.tag'):
                        tagFile = os.path.join(root, file)
                        tagFileMinusExtension = tagFile.rsplit('.', 1)[0]
                        tagSegments = Segment.SegmentList()

                        # Read metadata from .setting file
                        operator = ""
                        reviewer = ""
                        duration = ""
                        
                        try:
                            stree = ET.parse(tagFileMinusExtension + '.setting')
                            stroot = stree.getroot()
                            for elem in stroot:
                                if elem.tag == 'Operator':
                                    operator = elem.text or ""
                                if elem.tag == 'Reviewer' and elem.text:
                                    reviewer = elem.text
                        except Exception:
                            print(f"Can't read {tagFileMinusExtension}.setting or missing data")
                            
                        try:
                            # Read duration from .p file
                            ptree = ET.parse(tagFileMinusExtension + '.p')
                            ptroot = ptree.getroot()
                            for elem in ptroot:
                                for elem2 in elem:
                                    if elem2.tag == 'DurationSecond':
                                        duration = elem2.text
                        except Exception:
                            print(f"Can't read {tagFileMinusExtension}.p or missing data")
                            # Fall back to loading the wav file
                            try:
                                sp = Spectrogram.Spectrogram(512, 256, 0, 0)
                                sp.readSoundFile(tagFileMinusExtension + '.wav', 0, 0)
                                duration = sp.fileLength / sp.audioFormat.sampleRate()
                            except Exception:
                                print(f"Could not determine duration for {tagFileMinusExtension}")
                                duration = 0
               
                        tagSegments.metadata = {"Operator": operator, "Reviewer": reviewer, "Duration": duration}
                                    
                        # Read segments from .tag file
                        try:
                            tree = ET.parse(tagFile)
                            troot = tree.getroot()
              
                            for elem in troot:
                                try:
                                    species_code = int(elem[0].text)
                                    if species_code in spDict:
                                        species = [{"species": spDict[species_code], "certainty": 100, "filter": "M"}]
                                        # Create segment: [start, end, low_freq, high_freq, species_list]
                                        start_time = float(elem[1].text)
                                        duration_seg = float(elem[2].text)
                                        end_time = start_time + duration_seg
                                        low_freq = float(elem[4].text)
                                        high_freq = float(elem[3].text)
                                        
                                        newSegment = Segment.Segment([start_time, end_time, low_freq, high_freq, species])
                                        tagSegments.append(newSegment)
                                    else:
                                        print(f"Species code {species_code} not in bird list for file {tagFile}")
                                except (KeyError, ValueError, IndexError) as e:
                                    print(f"Error processing segment in {tagFile}: {e}")
                        except Exception as e:
                            print(f"Can't read {tagFile} or missing data: {e}")
                    
                        # Save .data file (avoid saving in .session folder)
                        if root.endswith(".session"):
                           output_dir = root[:-8] 
                        else:
                            output_dir = root
                            
                        output_file = os.path.join(output_dir, tagFileMinusExtension + '.wav.data')
                        tagSegments.saveJSON(output_file)
                        processed_files.append(output_file)
             
            return True, f"Successfully processed {len(processed_files)} annotation files", processed_files
            
        except Exception as e:
            return False, f"ERROR: Freebird import failed with error: {str(e)}", []
