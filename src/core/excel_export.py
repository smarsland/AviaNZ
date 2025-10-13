
# Version 4.1 09/10/25
# Authors: Stephen Marsland, Nirosha Priyadarshani, Julius Juodakis, Virginia Listanti, Giotto Frean

#    AviaNZ bioacoustic analysis program
#    Copyright (C) 2017--2024

#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.

#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.

#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.

# Excel export functionality

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import re
import math
import numpy as np
import os


class ExcelIO():
    """ Exports the annotations to xlsx, with three sheets:
    time stamps, presence/absence, and per second presence/absence.
    Saves each species into a separate workbook,
    + an extra workbook for all species (to function as a readable segment printout).
    It makes the workbook if necessary.

    Inputs
        segments:   list of SegmentList objects, with additional filename attribute
        dirName:    xlsx will be stored here
        filename:   name of the wav file, to be recorded inside the xlsx
        action:     "append" or "overwrite" any found Excels
        pagelen:    page length, seconds (for filling out absence)
        numpages:   number of pages in this file (of size pagelen)
        speciesList:    list of species that are currently processed -- will force an xlsx output even if none were detected
        startTime:  timestamp for page start, or None to autodetect from file name
        precisionMS:  timestamp resolution for sheet 1: False=in s, True=in ms
        resolution: output resolution (sheet 3) in seconds
    """
    # functions for filling out the excel sheets:
    # First page lists all segments (of a species, if specified)
    # segsLL: list of SegmentList with filename attribute
    # startTime: offset from 0, when exporting a single page
    def writeToExcelp1(self, wb, segsLL, currsp, startTime, precisionMS):
        if precisionMS:
            timeStrFormat = "hh:mm:ss.zzz"
        else:
            timeStrFormat = "hh:mm:ss"
        # Avoid importing PyQt6 in core; use stdlib datetime for formatting
        from datetime import datetime, timedelta
        ws = wb['Time Stamps']
        r = ws.max_row + 1

        for segsl in segsLL:
            # extract segments for the current species
            # if species=="All", take ALL segments.
            if currsp=="Any sound":
                speciesSegs = segsl
            else:
                speciesSegs = [segsl[ix] for ix in segsl.getSpecies(currsp)]

            if len(speciesSegs)==0:
                continue

            if startTime is None:
                # if no startTime was provided, try to figure it out based on the filename
                DOCRecording = re.search(r'(\d{6})_(\d{6})', os.path.basename(segsl.filename)[:-8])

                if DOCRecording:
                    print("time stamp found", DOCRecording)
                    ts = DOCRecording.group(2)  # HHMMSS
                    startTimeFile = datetime.strptime(ts, "%H%M%S")
                else:
                    startTimeFile = datetime.strptime("000000", "%H%M%S")
            else:
                # startTime is in seconds offset
                startTimeFile = datetime.strptime("000000", "%H%M%S") + timedelta(seconds=startTime)

            # Loop over the segments
            #for seg in speciesSegs:
            for i in range(len(speciesSegs)):
                seg = speciesSegs[i]
                # Print the filename
                ws.cell(row=r, column=1, value=segsl.filename)

                # Time limits - use datetime formatting
                tstart = startTimeFile + timedelta(milliseconds=int(seg.start_time*1000))
                tend = startTimeFile + timedelta(milliseconds=int(seg.end_time*1000))
                if precisionMS:
                    # Format with milliseconds
                    ws.cell(row=r, column=2, value=tstart.strftime("%H:%M:%S.%f")[:-3])
                    ws.cell(row=r, column=3, value=tend.strftime("%H:%M:%S.%f")[:-3])
                else:
                    ws.cell(row=r, column=2, value=tstart.strftime("%H:%M:%S"))
                    ws.cell(row=r, column=3, value=tend.strftime("%H:%M:%S"))
                # Freq limits
                if seg.freq_high!=0:
                    ws.cell(row=r, column=4, value=int(seg.freq_low))
                    ws.cell(row=r, column=5, value=int(seg.freq_high))
                if currsp=="Any sound":
                    # print species and certainty and call type
                    text = [lab["species"] for lab in seg.labels]
                    ws.cell(row=r, column=6, value=", ".join(text))
                    text = [str(lab["certainty"]) for lab in seg.labels]
                    ws.cell(row=r, column=7, value=", ".join(text))
                    strct = []
                    for lab in seg.labels:
                        if "calltype" in lab:
                            strct.append(str(lab["calltype"]))
                        else:
                            strct.append("-")
                    ws.cell(row=r, column=8, value=", ".join(strct))
                else:
                    # only print certainty and call type
                    strcert = []
                    strct = []
                    for lab in seg.labels:
                        if lab["species"]==currsp:
                            strcert.append(str(lab["certainty"]))
                            if "calltype" in lab:
                                strct.append(str(lab["calltype"]))
                            else:
                                strct.append("-")
                    ws.cell(row=r, column=6, value=", ".join(strcert))
                    ws.cell(row=r, column=7, value=", ".join(strct))
                    # SRM: As desired for bittern (H. Caley)
                    # The segments are sorted into order of start time
                    Previous=1
                    while i-Previous>=0 and speciesSegs[i-Previous].end_time > seg.start_time:
                        Previous += 1
                    Next = 1
                    while i+Next<len(speciesSegs) and speciesSegs[i+Next].start_time < seg.end_time:
                        Next += 1
                    ws.cell(row=r, column=8, value=", ".join(str(Next+Previous-2)))
                r += 1

    # This stores pres/abs and max certainty for the species in each file
    # segscert: a 2D list of segs x [start, end, certainty]
    def writeToExcelp2(self, wb, segscert, filename):
        ws = wb['Presence Absence']
        r = ws.max_row + 1

        ws.cell(row=r, column=1, value=filename)

        # segs: a 2D list of [start, end, certainty] for each seg
        if len(segscert)>0:
            pres = "Yes"
            certainty = [lab[2] for lab in segscert]
            certainty = max(certainty)
        else:
            pres = "No"
            certainty = 0
        ws.cell(row=r, column=2, value=pres)
        ws.cell(row=r, column=3, value=certainty)

    # This stores pres/abs (or max cert) for the species
    # in windows of size=resolution in each file
    # segscert: a 2D list of segs x [start, end, certainty]
    # pagenum: index of the current page, 0-base
    # totpages: total number of pages
    # pagelen: page length in s
    def writeToExcelp3(self, wb, segscert, filename, pagenum, pagelen, totpages, resolution):
        # writes binary output DETECTED (per s) from page PAGENUM of length PAGELEN
        starttime = pagenum * pagelen
        ws = wb['Per Time Period']
        r = ws.max_row + 1

        # print resolution "header"
        ws.cell(row=r, column=1, value=str(resolution) + ' secs resolution')
        ft = Font(color="808000")
        ws.cell(row=r, column=1).font=ft

        # print file name and page number
        ws.cell(row=r+1, column=1, value=filename)
        ws.cell(row=r+1, column=2, value=str(pagenum+1))

        detected = np.zeros(math.ceil(pagelen/resolution))
        # convert segs to max certainty at each second
        for seg in segscert:
            # segment start-end, relative to this page start:
            segStart = seg[0] - pagenum*pagelen
            segEnd = seg[1] - pagenum*pagelen
            # just in case of some old reversed segments:
            if segStart > segEnd:
                segStart, segEnd = segEnd, segStart

            # segment is completely outside the current page:
            if segEnd<0 or segStart>pagelen:
                continue

            # convert segment time in s to time in resol windows:
            # map [1..1.999 -> 1
            segStart = max(0, math.floor(segStart/resolution))
            # map 2.0001...3] -> 3
            segEnd = math.ceil(min(segEnd, pagelen)/resolution)
            # range 1:3 selects windows 1 & 2
            for t in range(segStart, segEnd):
                # store certainty if it's larger
                detected[t] = max(detected[t], seg[2])

        # fill the header and detection columns
        c = 3
        for t in range(len(detected)):
            # absolute (within-file) times:
            win_start = starttime + t*resolution
            win_end = min(win_start+resolution, int(pagelen * totpages))
            ws.cell(row=r, column=c, value="%d-%d" % (win_start, win_end))
            ws.cell(row=r, column=c).font = ft
            ws.cell(row=r+1, column=c, value=detected[t])
            c += 1

    def export(self, segments, dirName, action, pagelenarg=None, numpages=1, speciesList=[], startTime=None, precisionMS=False, resolution=10):
        # will export species present in self, + passed as arg, + "all species" excel
        speciesList = set(speciesList)
        for segl in segments:
            for seg in segl:
                speciesList.update([lab["species"] for lab in seg.labels])
        speciesList.add("Any sound")
        print("The following species were detected for export:", speciesList)

        # check source .wav file names -
        # ideally, we store relative paths, but that's not possible across drives:
        for segl in segments:
            try:
                segl.filename = str(os.path.relpath(segl.filename, dirName))
            except Exception as e:
                print("Falling back to absolute paths. Encountered exception:")
                print(e)
                segl.filename = str(os.path.abspath(segl.filename))

        # now, generate the actual files, SEPARATELY FOR EACH SPECIES:
        for species in speciesList:
            print("Exporting species %s" % species)
            # clean version for filename
            speciesClean = re.sub(r'\W', "_", species)

            # setup output files:
            # if an Excel exists, append (so multiple files go into one worksheet)
            # if not, create new
            eFile = os.path.join(dirName, 'DetectionSummary_' + speciesClean + '.xlsx')

            if action == "overwrite" or not os.path.isfile(eFile):
                # make a new workbook:
                wb = Workbook()

                # First sheet
                wb.create_sheet(title='Time Stamps', index=1)
                ws = wb['Time Stamps']
                ws.cell(row=1, column=1, value="File Name")
                if precisionMS:
                    ws.cell(row=1, column=2, value="start (hh:mm:ss.ms)")
                    ws.cell(row=1, column=3, value="end (hh:mm:ss.ms)")
                else:
                    ws.cell(row=1, column=2, value="start (hh:mm:ss)")
                    ws.cell(row=1, column=3, value="end (hh:mm:ss)")
                ws.cell(row=1, column=4, value="min freq. (Hz)")
                ws.cell(row=1, column=5, value="max freq. (Hz)")
                if species=="Any sound":
                    ws.cell(row=1, column=6, value="species")
                    ws.cell(row=1, column=7, value="certainty")
                    ws.cell(row=1, column=8, value="call type")
                else:
                    ws.cell(row=1, column=6, value="certainty")
                    ws.cell(row=1, column=7, value="call type")
                    ws.cell(row=1, column=8, value="number of overlaps")

                    # Second sheet
                    wb.create_sheet(title='Presence Absence', index=2)
                    ws = wb['Presence Absence']
                    ws.cell(row=1, column=1, value="File Name")
                    ws.cell(row=1, column=2, value="Present?")
                    ws.cell(row=1, column=3, value="Certainty, %")

                    # Third sheet
                    wb.create_sheet(title='Per Time Period', index=3)
                    ws = wb['Per Time Period']
                    ws.cell(row=1, column=1, value="File Name")
                    ws.cell(row=1, column=2, value="Page")
                    ws.cell(row=1, column=3, value="Maximum certainty of species presence (0 = absent)")

                # Hack to delete original sheet
                del wb['Sheet']
            elif action == "append":
                try:
                    wb = load_workbook(eFile)
                except Exception as e:
                    print("ERROR: cannot open file %s to append" % eFile)
                    print(e)
                    return 0
            else:
                print("ERROR: unrecognised action", action)
                return 0

            # export segments
            self.writeToExcelp1(wb, segments, species, startTime, precisionMS)

            if species!="Any sound":
                # loop over all SegmentLists, i.e. for each wav file:
                for segsl in segments:
                    # extract the certainty from each label for current species
                    # to a 2D list of segs x [start, end, certainty]
                    # (for this wav file)
                    speciesCerts = []
                    for seg in segsl:
                        for lab in seg.labels:
                            if lab["species"]==species:
                                speciesCerts.append([seg.start_time, seg.end_time, lab["certainty"]])

                    # export presence/absence and max certainty
                    self.writeToExcelp2(wb, speciesCerts, segsl.filename)

                    # either read duration from this SegList
                    # or need current page length if called from manual
                    # (assuming all pages are of same length as current data)
                    if pagelenarg is None:
                        pagelen = math.ceil(segsl.metadata["Duration"])
                    else:
                        pagelen = pagelenarg

                    # Generate pres/abs per custom resolution windows
                    for p in range(0, numpages):
                        self.writeToExcelp3(wb, speciesCerts, segsl.filename, p, pagelen, numpages, resolution)

            # Save the file
            try:
                wb.save(eFile)
            except Exception as e:
                print("ERROR: could not create new file %s" % eFile)
                print(e)
                return 0
        return 1
