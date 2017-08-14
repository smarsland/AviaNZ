import os, json
from PyQt4.QtCore import *
from PyQt4.QtGui import *
import math

import wavio
import librosa
import numpy as np
from openpyxl import load_workbook, Workbook

from pyqtgraph.Qt import QtGui
from pyqtgraph.dockarea import *
import pyqtgraph as pg

import SignalProc
import Segment
import WaveletSegment

class AviaNZFindSpeciesInterface(QMainWindow):
    # Main class for the interface, which contains most of the user interface and plotting code

    def __init__(self,root=None,minSegment=50):
        # Basically allow the user to browse a folder and push a button to process that folder to find a target species
        # and sets up the window.
        super(AviaNZFindSpeciesInterface, self).__init__()
        self.root = root
        self.dirName=[]
        # self.minSegment=minSegment

        # Make the window and associated widgets
        QMainWindow.__init__(self, root)

        # add statusbar
        self.statusLeft = QLabel("Ready")
        self.statusLeft.setFrameStyle(QFrame.Panel)
        self.statusRight = QLabel("Processing file Current/Total")
        self.statusRight.setAlignment(Qt.AlignRight)
        self.statusRight.setFrameStyle(QFrame.Panel)
        statusStyle='QLabel {border:transparent}'
        self.statusLeft.setStyleSheet(statusStyle)
        self.statusRight.setStyleSheet(statusStyle)
        self.statusBar().addPermanentWidget(self.statusLeft,1)
        self.statusBar().addPermanentWidget(self.statusRight,1)

        # Set the message in the status bar
        self.statusLeft.setText("Ready")
        self.statusRight.setText("Processing file Current/Total")

        self.setWindowTitle('AviaNZ - Automatic Detection')
        self.createFrame()

    def createFrame(self):
        # Make the window and set its size
        self.area = DockArea()
        self.setCentralWidget(self.area)
        self.setFixedSize(500,300)

        # Make the docks
        # self.d_fileList = Dock("Folder",size=(300,150))
        self.d_detection = Dock("Automatic Detection",size=(350,100))
        self.d_detection.hideTitleBar()

        # self.area.addDock(self.d_fileList,'left')
        self.area.addDock(self.d_detection,'right')

        self.w_browse = QPushButton("  &Browse Folder")
        self.connect(self.w_browse, SIGNAL('clicked()'), self.browse)
        self.w_browse.setToolTip("Can select a folder with sub folders to process")
        self.w_dir = QLineEdit()
        self.w_dir.setText('')
        self.d_detection.addWidget(self.w_dir,row=0,col=1,colspan=2)
        self.d_detection.addWidget(self.w_browse,row=0,col=0)

        self.w_speLabel = QLabel("  Select Species")
        self.d_detection.addWidget(self.w_speLabel,row=1,col=0)
        self.w_spe = QComboBox()
        self.w_spe.addItems(["Kiwi", "Ruru","Any"])
        self.d_detection.addWidget(self.w_spe,row=1,col=1,colspan=2)

        self.w_processButton = QPushButton("&Process Folder")
        self.connect(self.w_processButton, SIGNAL('clicked()'), self.detect)
        self.d_detection.addWidget(self.w_processButton,row=10,col=2)
        self.w_processButton.setStyleSheet('QPushButton {background-color: #A3C1DA; font-weight: bold; font-size:14px}')

        self.statusLeft.setText("Ready")

        # Store the state of the docks
        self.state = self.area.saveState()

        # Plot everything
        self.show()

    def browse(self):
        # self.dirName = QtGui.QFileDialog.getExistingDirectory(self,'Choose Folder to Process',"Wav files (*.wav)")
        if self.dirName:
            self.dirName = QtGui.QFileDialog.getExistingDirectory(self,'Choose Folder to Process',str(self.dirName))
        else:
            self.dirName = QtGui.QFileDialog.getExistingDirectory(self,'Choose Folder to Process')
        print "Dir:", self.dirName
        # self.setWindowTitle('AviaNZ - '+ self.dirName)
        self.w_dir.setText(self.dirName)

        # files=[]
        # self.w_fileList.clear()
        # for f in os.listdir(self.dirName):
        #     # if f.endswith(".wav"): #os.path.isfile(f) and
        #     self.w_fileList.addItem(f)
        #     files.append(f)
        # print files

    def detect(self):
        with pg.BusyCursor():
            if self.dirName:
                self.statusLeft.setText("Processing...")
                # self.statusBar().showMessage("Processing...")
                i=self.w_spe.currentIndex()
                if i==0:
                    self.species="Kiwi"
                elif i==1:
                    self.species="Ruru"
                else: # All
                    self.species="Any"

                total=0
                for root, dirs, files in os.walk(str(self.dirName)):
                    for filename in files:
                        if filename.endswith('.wav'):
                            total=total+1
                cnt=0   # processed number of files
                self.statusRight.setText("Processing file " + "0/" + str(total))

                for root, dirs, files in os.walk(str(self.dirName)):
                    for filename in files:
                        if filename.endswith('.wav'):
                            # if not os.path.isfile(root+'/'+filename+'.data'): # if already processed then skip?
                            #     continue
                            cnt=cnt+1
                            self.statusRight.setText("Processing file " + str(cnt) + "/" + str(total))
                            self.filename=root+'/'+filename
                            self.loadFile()

                            self.seg = Segment.Segment(self.audiodata, self.sgRaw, self.sp, self.sampleRate)
                            # print self.algs.itemText(self.algs.currentIndex())
                            # if self.algs.currentText() == "Amplitude":
                            #     newSegments = self.seg.segmentByAmplitude(float(str(self.ampThr.text())))
                            # elif self.algs.currentText() == "Median Clipping":
                            #     newSegments = self.seg.medianClip(float(str(self.medThr.text())))
                            #     #print newSegments
                            # elif self.algs.currentText() == "Harma":
                            #     newSegments = self.seg.Harma(float(str(self.HarmaThr1.text())),float(str(self.HarmaThr2.text())))
                            # elif self.algs.currentText() == "Power":
                            #     newSegments = self.seg.segmentByPower(float(str(self.PowerThr.text())))
                            # elif self.algs.currentText() == "Onsets":
                            #     newSegments = self.seg.onsets()
                            #     #print newSegments
                            # elif self.algs.currentText() == "Fundamental Frequency":
                            #     newSegments, pitch, times = self.seg.yin(int(str(self.Fundminfreq.text())),int(str(self.Fundminperiods.text())),float(str(self.Fundthr.text())),int(str(self.Fundwindow.text())),returnSegs=True)
                            #     print newSegments
                            # elif self.algs.currentText() == "FIR":
                            #     print float(str(self.FIRThr1.text()))
                            #     # newSegments = self.seg.segmentByFIR(0.1)
                            #     newSegments = self.seg.segmentByFIR(float(str(self.FIRThr1.text())))
                            #     # print newSegments
                            # elif self.algs.currentText()=='Wavelets':
                            if self.species!='Any':
                                newSegments = WaveletSegment.findCalls_test(fName=None,data=self.audiodata, sampleRate=self.sampleRate, species=self.species,trainTest=False)
                            else:
                                newSegments=self.seg.bestSegments()

                            # Generate Binary output ('Binary)
                            n=math.ceil(float(self.datalength)/self.sampleRate)
                            detected=np.zeros(int(n))
                            for seg in newSegments:
                                for a in range(len(detected)):
                                    if math.floor(seg[0])<=a and a<math.ceil(seg[1]):
                                        detected[a]=1
                            self.saveSegments(detected, mode='Binary') # append

                            # Generate annotation friendly output ('Annotation')
                            # print "Generate annotation friendly output", newSegments
                            annotation=[]
                            if len(newSegments)>0:
                                if self.species!='Any': # alg="Wavelets"
                                    mergedSeg=self.mergeSeg(newSegments)
                                    if len(mergedSeg)>0:
                                        for seg in mergedSeg:
                                            annotation.append([float(seg[0]),float(seg[1]),0,0,self.species+'?'])
                                elif self.species=='Any':
                                    if len(newSegments)>0:
                                        for seg in newSegments:
                                            annotation.append([float(seg[0]),float(seg[1]),0,0,"Don't know"])
                                    # print annotation
                                self.saveSegments(annotation, mode='Annotation')

                            # Generate excel summary - time stamps [start(mm:ss) end(mm:ss)]
                            annotation=[]
                            for seg in newSegments:
                                annotation.append([self.convertMillisecs(seg[0]*1000),self.convertMillisecs(seg[1]*1000)])
                            # print annotation
                            self.saveSegments(annotation, mode='Excel') # append
                # self.statusBar().showMessage("Ready")
                self.statusLeft.setText("Ready")
            else:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setWindowIcon(QIcon('img/Avianz.ico'))
                msg.setText("Please select a folder to process!")
                msg.setWindowTitle("Select Folder")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()
            # return newSegments

    def mergeSeg(self,segments):
        indx=[]
        for i in range(len(segments)-1):
            if segments[i][1]==segments[i+1][0]:
                indx.append(i)
        indx.reverse()
        for i in indx:
            segments[i][1]=segments[i+1][1]
            del(segments[i+1])
        return segments

    def saveSegments(self,annotation,mode):
        # This saves the detections into three different formats: annotation, excel, and binary

        # method=self.algs.currentText()
        relfname=os.path.relpath(str(self.filename),str(self.dirName))
        eFile=self.dirName+'\DetectionSummary_'+self.species+'.xlsx'

        if mode=='Annotation':
            if isinstance(self.filename, str):
                file = open(self.filename + '.data', 'w')
            else:
                file = open(str(self.filename) + '.data', 'w')
            json.dump(annotation,file)

        elif mode=='Excel':
            if os.path.isfile(eFile):   #if the file is already there
                try:
                    wb = load_workbook(str(eFile))
                    ws=wb.get_sheet_by_name('TimeStamps')
                    c=1
                    r=ws.max_row+1 # TODO: get last row number from existing file
                    ws.cell(row=r,column=1,value=str(relfname))
                    for seg in annotation:
                        # ws.cell(row=r,column=c+1,value=seg[0]+'-'+seg[1])
                        # c=c+1
                        ws.cell(row=r, column=c + 1, value=seg[0])
                        ws.cell(row=r, column=c + 2, value=seg[1])
                        c = c + 2
                    wb.save(str(eFile))
                except:
                    print "Unable to open file"           #Does not exist OR no read permissions
            else:
                wb = Workbook()
                wb.create_sheet(title='TimeStamps',index=1)
                wb.create_sheet(title='PresenceAbsence',index=2)
                wb.create_sheet(title='PerSecond',index=3)

                ws = wb.get_sheet_by_name('TimeStamps')
                ws.cell(row=1,column=1, value="File Name")
                ws.cell(row=1, column=2, value="start(mm:ss)")
                c=3
                for i in range(100):
                    ws.cell(row=1, column=c , value="end")
                    ws.cell(row=1, column=c + 1, value="start")
                    c=c+2
                ws.cell(row=1, column=c , value="end")
                # ws.cell(row=1,column=2, value="Detections [start-end(mm:ss)]")
                c=1
                r=2
                ws.cell(row=r,column=c,value=str(relfname))
                for seg in annotation:
                    # ws.cell(row=r,column=c+1,value=seg[0]+'-'+seg[1])
                    # c=c+1
                    ws.cell(row=r, column=c + 1, value=seg[0])
                    ws.cell(row=r, column=c + 2, value=seg[1])
                    c = c + 2
                # Second sheet
                ws = wb.get_sheet_by_name('PresenceAbsence')
                ws.cell(row=1,column=1, value="File Name")
                ws.cell(row=1,column=2, value="Presence/Absence")

                # Third sheet
                ws = wb.get_sheet_by_name('PerSecond')
                ws.cell(row=1,column=1, value="File Name")
                ws.cell(row=1,column=2, value="Presence=1/Absence=0")
                c=2
                for i in range(900):
                    ws.cell(row=2,column=c, value="S "+str(i+1))
                    c=c+1
                first=wb.get_sheet_by_name('Sheet')
                wb.remove_sheet(first)
                wb.save(str(eFile))

            # Presence absence excel
            if os.path.isfile(eFile):   #if the file is already there
                try:
                    wb = load_workbook(str(eFile))
                    # ws=wb.create_sheet(title="PresenceAbsence",index=2)
                    ws=wb.get_sheet_by_name('PresenceAbsence')
                    r=ws.max_row+1 #
                    ws.cell(row=r,column=1,value=str(relfname))
                    if annotation:
                        ws.cell(row=r,column=2,value='Yes')
                    else:
                        ws.cell(row=r,column=2,value='_')
                    wb.save(str(eFile))
                except:
                    print "Unable to open file"           #Does not exist OR no read permissions

        else:   # mode=='Binary'
            # eFile=self.dirName+'\\3PerSecond_'+self.species+'_'+'.xlsx'
            if os.path.isfile(eFile):   #if the file is already there
                try:
                    wb = load_workbook(str(eFile))
                    ws=wb.get_sheet_by_name('PerSecond')
                    c=1
                    r=ws.max_row+1 # TODO: get last row number from existing file
                    ws.cell(row=r,column=1,value=str(relfname))
                    for seg in annotation:
                        ws.cell(row=r,column=c+1,value=seg)
                        c=c+1
                    wb.save(str(eFile))
                except:
                    print "Unable to open file"           #Does not exist OR no read permissions

    def loadFile(self):
        print self.filename
        wavobj = wavio.read(self.filename)
        self.sampleRate = wavobj.rate
        self.audiodata = wavobj.data
        print np.shape(self.audiodata)

        # None of the following should be necessary for librosa
        if self.audiodata.dtype is not 'float':
            self.audiodata = self.audiodata.astype('float') #/ 32768.0
        if np.shape(np.shape(self.audiodata))[0]>1:
            self.audiodata = self.audiodata[:,0]
        self.datalength = np.shape(self.audiodata)[0]
        print("Length of file is ",len(self.audiodata),float(self.datalength)/self.sampleRate,self.sampleRate)

        if (self.species=='Kiwi' or self.species=='Ruru') and self.sampleRate!=16000:
            self.audiodata = librosa.core.audio.resample(self.audiodata,self.sampleRate,16000)
            self.sampleRate=16000
            self.datalength = np.shape(self.audiodata)[0]
        print self.sampleRate
        self.minFreq = 0
        self.maxFreq = self.sampleRate/2.

        # Create an instance of the Signal Processing class
        if not hasattr(self,'sp'):
            # self.sp = SignalProc.SignalProc(self.audiodata, self.sampleRate)
            self.sp = SignalProc.SignalProc()

        # Get the data for the spectrogram
        # self.sgRaw = self.sp.spectrogram(self.audiodata,self.sampleRate,mean_normalise=True,onesided=True,multitaper=False)
        self.sgRaw = self.sp.spectrogram(self.audiodata, window_width=256, incr=128, window='Hann', mean_normalise=True, onesided=True,multitaper=False, need_even=False)
        maxsg = np.min(self.sgRaw)
        self.sg = np.abs(np.where(self.sgRaw==0,0.0,10.0 * np.log10(self.sgRaw/maxsg)))

        # Update the data that is seen by the other classes
        # TODO: keep an eye on this to add other classes as required
        if hasattr(self,'seg'):
            self.seg.setNewData(self.audiodata,self.sgRaw,self.sampleRate,256,128)
        else:
            self.seg = Segment.Segment(self.audiodata, self.sgRaw, self.sp, self.sampleRate)
        self.sp.setNewData(self.audiodata,self.sampleRate)

        # self.setWindowTitle('AviaNZ - ' + self.filename)

    def convertMillisecs(self,millisecs):
        seconds = (millisecs / 1000) % 60
        minutes = (millisecs / (1000 * 60)) % 60
        return "%02d" % minutes+":"+"%02d" % seconds
