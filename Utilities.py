
# Utilities.py
#
# A set of utility functions for AviaNZ

# Version 1.3 23/10/18
# Authors: Stephen Marsland, Nirosha Priyadarshani, Julius Juodakis

#    AviaNZ birdsong analysis program
#    Copyright (C) 2017--2018

#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.

#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.

#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.

import numpy as np
import wavio
import os, json
import math, re
import datetime
import WaveletSegment
import librosa
import matplotlib.markers as mks
import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
from scipy import signal
import shutil
from openpyxl import load_workbook, Workbook

def generateWavelet(name):
    """ This loads the PyWavelets package,
        generates a Wavelet object of the desired name,
        and exports it to Wavelets/name.txt as a 4 filter list.
    """
    import pywt
    import os
    wv = pywt.Wavelet(name)

    fname = 'Wavelets/' + name + '.txt'
    if(os.path.isfile(fname)):
        print("ERROR: file %s already exists!" % fname)
        return
    else:
        np.savetxt(fname, wv.filter_bank, delimiter='  ')
        print("File %s saved successfully" %fname)

def excel2data(dir, excelFile, sub=False):
    '''
    Generate .data from excel summary (GSK)
    '''
    # read excel file
    book = load_workbook(excelFile)
    sheet = book.active
    rid = sheet['A2': 'A919']
    endtime = sheet['C2': 'C919']
    date = sheet['D2': 'D919']
    type = sheet['E2': 'E919']
    qlty = sheet['F2': 'F919']
    rids = []
    endtimes=[]
    endtimes_sec = []
    dates = []
    types = []
    quality = []
    for i in range(len(rid)):
        rids.append(str(rid[i][0].value))
    for i in range(len(endtime)):
        endtimes.append(str(endtime[i][0].value))
        h, m, s = str(endtime[i][0].value).split(':')
        if len(h) > 2:
            x,h = h.split(' ')
        try:
            secs = int(h)*60*60+int(m)*60+int(s)
        except:
            print("here", i)
            print(str(endtime[i][0].value))
            print(h, m, s)
        endtimes_sec.append(str(secs))
    for i in range(len(date)):
        d, t = str(date[i][0].value).split(' ')
        y, m, d = d.split('-')
        if not sub:
            d = str(d)+str(m)+str(y[2:4])	# convert to DOC format
        else:
            d = str(y) + str(m) + str(d)
        dates.append(d)
    for i in range(len(type)):
        types.append(str(type[i][0].value))
    for i in range(len(qlty)):
        quality.append(str(qlty[i][0].value))

    # generate the .data files from excel
    for root, dirs, files in os.walk(str(dir)):
        for f1 in files:
            if f1.endswith('.wav'):
                if not sub:
                    DOCRecording = re.search('(\d{6})_(\d{6})', f1)
                else:
                    DOCRecording = re.search('(\d{8})_(\d{6})', f1)
                if DOCRecording:
                    startTime = DOCRecording.group(2)
                    recdate = DOCRecording.group(1)
                annotation=[]
                # find the recorder ID
                if not sub:
                    id = root.split('\\')[-1]
                else:
                    id = root.split('\\')[-2]
                # convert start time into seconds
                startSecs = int(startTime[0:2])*60*60+int(startTime[2:4])*60+int(startTime[4:6])
                # now find any matching segments from excel
                for i in range(len(rids)):
                    if id==rids[i] and recdate==dates[i] and startSecs<int(endtimes_sec[i]) and int(endtimes_sec[i])<startSecs+900:
                        if len(types[i])>1:
                            annotation.append([int(endtimes_sec[i])-startSecs-40, int(endtimes_sec[i])-startSecs, 500, 3900, types[i]+'_'+quality[i]])
                        else:
                            annotation.append([int(endtimes_sec[i])-startSecs-20, int(endtimes_sec[i])-startSecs, 500, 3900, types[i]+'_'+quality[i]])
                annotation.insert(0,[-1, startTime, "Doc_Operator", "Doc_Reviewer", -1])
                file = open(root + '\\' + f1 + '.data', 'w')
                json.dump(annotation, file)
                file.close()

def tag2data(dir,birdlist):
    '''
    Generate .data from Freebird tag files
    '''
    import os
    import openpyxl
    import xml.etree.ElementTree as ET

    # read freebird bird list
    book = openpyxl.load_workbook(birdlist)
    sheet = book.active
    name = sheet['A2': 'A353']
    code = sheet['B2': 'B353']
    spName = []
    spCode = []
    for i in range(len(name)):
        spName.append(str(name[i][0].value))
    for i in range(len(code)):
        spCode.append(int(code[i][0].value))

    spDict = dict(zip(spCode, spName))

    # generate the .data files from .tag
    for root, dirs, files in os.walk(str(dir)):
        for f1 in files:
            if f1.endswith('.tag'):
                DOCRecording = re.search('(\d{6})_(\d{6})', f1)
                if DOCRecording:
                    startTime = DOCRecording.group(2)
                annotation=[]
                tagFile = root + '/' + f1
                tree = ET.parse(tagFile)
                troot = tree.getroot()
                for elem in troot:
                    sp = spDict[int(elem[0].text)]
                    annotation.append([float(elem[1].text),float(elem[1].text)+float(elem[2].text),500,7500,sp])
                annotation.insert(0,[-1, startTime, "Doc_Operator", "Doc_Reviewer", -1])
                file = open(tagFile[:-4] + '.wav.data', 'w')
                json.dump(annotation, file)


def genDiffAnnotation(dir1, dir2):
    '''
    What was deleted by an additional filter?
    e.g. dir1 is Filter3 and dir2 is Filter4
    '''
    for root, dirs, files in os.walk(str(dir1)):
        for f1 in files:
            if f1.endswith('.data'):
                annotation1 = dir1 + '/' + f1
                try:
                    with open(annotation1) as ann1:
                        segments1 = json.load(ann1)
                    annotation2 = dir2 + '/' + f1
                    with open(annotation2) as ann2:
                        segments2 = json.load(ann2)
                except:
                    pass
                annotation = []
                for seg in segments1:
                    if seg[0] == -1:
                        annotation.append(seg)
                    if seg in segments2:
                        continue
                    else:
                        annotation.append(seg)

                file = open(str(annotation1), 'w')
                json.dump(annotation, file)

def countSegments(dir):
    """
    This counts the segments in the data files in a given folder
    """
    cnt = 0
    for root, dirs, files in os.walk(str(dir)):
        for f1 in files:
            if f1.endswith('.data'):
                annotation = dir + '/' + f1
                try:
                    with open(annotation) as ann:
                        segments = json.load(ann)
                except:
                    pass
                for seg in segments:
                    if seg[0] == -1:
                        pass
                    else:
                        cnt += 1
    print(cnt)

def delEmpAnn(dir):
    """
    This deletes empty data files in a given folder
    helps when manually reviewing auto detections, because reviwer can skip empty recordings without opening it (red and black color code from the list)
    """
    import os
    cnt = 0
    for root, dirs, files in os.walk(str(dir)):
        for f1 in files:
            if f1.endswith('.data'):
                annotation = root + '/' + f1
                try:
                    with open(annotation) as ann:
                        segments = json.load(ann)
                except:
                    break
                if len(segments)==0:
                    os.remove(annotation)
                    cnt += 1
                elif len(segments)==1 and segments[0][0]==-1:
                    os.remove(annotation)
                    cnt += 1
    print("#files deleted: ", cnt)

def extractSegments(wavFile, destination, copyName, species):
    """
    This extracts the sound segments given the annotation and the corresponding wav file. (Isabel's experiment data extraction)
    """
    datFile = wavFile+'.data'
    try:
        wavobj = wavio.read(wavFile)
        sampleRate = wavobj.rate
        data = wavobj.data
        if os.path.isfile(datFile):
            with open(datFile) as f:
                segments = json.load(f)
            cnt = 1
            for seg in segments:
                if seg[0] == -1:
                    continue
                if copyName:    # extract all - extracted sounds are saved with the same name as the corresponding segment in the annotation (e.g. Rawhiti exp.)
                    filename = destination + '\\' + seg[4] + '.wav'
                    s = int(seg[0] * sampleRate)
                    e = int(seg[1] * sampleRate)
                    temp = data[s:e]
                    wavio.write(filename, temp.astype('int16'), sampleRate, scale='dtype-limits', sampwidth=2)
                elif not species:   # extract all - extracted sounds are saved with the original file name followed by an index starting 1
                    ind = wavFile.rindex('/')
                    filename = destination + '\\' + str(wavFile[ind + 1:-4]) + '-' + str(cnt) + '.wav'
                    cnt += 1
                    s = int(seg[0] * sampleRate)
                    e = int(seg[1] * sampleRate)
                    temp = data[s:e]
                    wavio.write(filename, temp.astype('int16'), sampleRate, scale='dtype-limits', sampwidth=2)
                elif species == seg[4][0]:   # extract only specific calls - extracted sounds are saved with with the original file name followed by an index starting 1
                    ind = wavFile.rindex('/')
                    ind2 = wavFile.rindex('\\')
                    filename = destination + '\\' + str(wavFile[ind2+1:ind]) + '-' + str(wavFile[ind + 1:-4]) + '-' + str(seg[4][0]) + '-' + str(cnt) + '.wav'
                    cnt += 1
                    s = int((seg[0]-1) * sampleRate)
                    e = int((seg[1]+1) * sampleRate)
                    temp = data[s:e]
                    wavio.write(filename, temp.astype('int16'), sampleRate, scale='dtype-limits', sampwidth=2)
    except:
        print ("unsupported file: ", wavFile)
        # pass

# extractSegments('E:\ISABEL\Rawhiti Experiment Data Sorting NP\Station7-DOC16-Night-Expert\St7-Night-Expert-Trial1to7.wav', 'E:\ISABEL\Rawhiti Experiment Data Sorting NP\Station7-DOC16-Night-Expert\songs')

def extractSegments_batch(dirName, destination, copyName = True, species = None):
    """
    This extracts the sound segments in a directory (Isabel's Rawithi experiment data extraction)
    copyName is True when extracted sounds are saved with the same name as the corresponding segment in the annotation
    Specify 'species' when you want to extract specific calls, e.g. 'kiwi(M)1' to extract all v close brown kiwi male calls
    """
    for root, dirs, files in os.walk(str(dirName)):
        for filename in files:
            if filename.endswith('.wav') and filename+'.data' in files:
                filename = root + '/' + filename
                extractSegments(filename, destination, copyName=copyName, species = species)


def renameAnnotation(dirName, frm, to):
    '''
    Rename the annotations spp name
    '''
    for root, dirs, files in os.walk(str(dirName)):
        for filename in files:
            if filename.endswith('.data') and filename[:-5] in files:
                filename = root + '/' + filename
                with open(filename) as f:
                    segments = json.load(f)
                    chg = False
                    for seg in segments:
                        if seg[0] == -1:
                            continue
                        elif frm == seg[4]:
                            seg[4] = to
                            chg = True
                if chg:
                    file = open(str(filename), 'w')
                    json.dump(segments, file)
                    file.close()

def annotation2GT(wavFile, species, duration=0):
    """
    This generates the ground truth for a given sound file (currently for kiwi and bittern).
    Given the AviaNZ annotation, returns the ground truth as a txt file
    """
    datFile = wavFile+'.data'
    print(datFile)
    eFile = datFile[:-9]+'-1sec.txt'
    if duration == 0:
        wavobj = wavio.read(wavFile)
        sampleRate = wavobj.rate
        data = wavobj.data
        duration = len(data)/sampleRate   # number of secs
    GT = np.zeros((duration, 4))
    GT = GT.tolist()
    GT[:][1] = str(0)
    GT[:][2] = ''
    GT[:][3] = ''
    if os.path.isfile(datFile):
        with open(datFile) as f:
            segments = json.load(f)
        for seg in segments:
            if seg[0] == -1:
                continue
            if not 'Kiwi' in seg[4][0]:
                if seg[4][0][-1] == '?':
                    print("**", wavFile)
                continue
            elif species == 'Kiwi (Nth Is Brown)' or species == 'Kiwi' or species == 'Kiwi_Tokoeka_Haast' or species == 'Kiwi_Brown' or species == 'Kiwi_spp' or species == 'Kiwi_Okarito_Brown' or species == 'Kiwi_Tokoeka_Stewart_Is':
                # check M/F
                if '(M)' in str(seg[4][0]):        # if re.search('(M)', seg[4]):
                    type = 'M'
                elif '(F)' in str(seg[4][0]):      #if re.search('(F)', seg[4]):
                    type = 'F'
                elif '(D)' in str(seg[4][0]):
                    type = 'D'
                else:
                    type = 'K'
            elif species == 'Bittern':
                # check boom/inhalation
                if '(B)' in str(seg[4]):
                    type = 'B'
                elif '(I)' in str(seg[4]):
                    type = 'I'
                else:
                    type = ''

            # check quality
            if re.search('1', seg[4][0]):
                quality = '1'   # v close
            elif re.search('2', seg[4][0]):
                quality = '2'    # close
            elif re.search('3', seg[4][0]):
                quality = '3'   # fade
            elif re.search('4', seg[4][0]):
                quality = '4'  # v fade
            elif re.search('5', seg[4][0]):
                quality = '5'   # v v fade
            else:
                quality = ''

            s=int(math.floor(seg[0]))
            e=int(math.ceil(seg[1]))
            for i in range(s, e):
                GT[i][1] = str(1)
                GT[i][2] = type
                GT[i][3] = quality
        for line in GT:
            if line[1] == 0.0:
                line[1] = '0'
            if line[2] == 0.0:
                line[2] = ''
            if line[3] == 0.0:
                line[3] = ''

        # now save GT as a .txt file
        for i in range(1, duration + 1):
            GT[i-1][0] = str(i)   # add time as the first column to make GT readable
        out = open(eFile, "w")
        for line in GT:
            out.write("\t".join(line))
            out.write("\n")
        out.close()

def genGT(dirName,species='Kiwi',duration=0):
    """
    Given the directory where sound files and the annotations along with the species being considered,
    it generates the ground truth txt files using 'annotation2GT'
    If you know the duration of a recording pass it through 'duration'.
    """
    for root, dirs, files in os.walk(str(dirName)):
        for filename in files:
            if filename.endswith('.wav'):
                filename = root + '/' + filename
                annotation2GT(filename,species,duration=duration)
    print("Generated GT")

def genReport(dirName, species='Kiwi'):
    '''
    Convert the annotations collected from reviewers into a single excel summary using annotation2Report (for Fiordland)
    '''
    eFile = dirName + '/report_100.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="File name")
    ws.cell(row=1, column=2, value="start")
    ws.cell(row=1, column=3, value="end")
    ws.cell(row=1, column=4, value="call")
    r = 2
    for root, dirs, files in os.walk(str(dirName)):
        for filename in files:
            if filename.endswith('.data'):
                print(filename)
                fName=filename
                filename = root + '/' + filename
                with open(filename) as f:
                    segments = json.load(f)
                if len(segments) == 0:
                    continue
                if len(segments) == 1 and segments[0][0] == -1:
                    continue
                # Check if the filename is in standard DOC format
                # Which is xxxxxx_xxxxxx.wav or ccxx_cccc_xxxxxx_xxxxxx.wav (c=char, x=0-9), could have _ afterward
                # So this checks for the 6 ints _ 6 ints part anywhere in string
                DOCRecording = re.search('(\d{6})_(\d{6})', filename[-22:-9])

                if DOCRecording:
                    startTime = DOCRecording.group(2)

                    if int(startTime[:2]) > 8 or int(startTime[:2]) < 8:
                        print("Night time DOC recording")
                    else:
                        print("Day time DOC recording")
                        # TODO: And modify the order of the bird list
                    startTime = int(startTime[:2]) * 3600 + int(startTime[2:4]) * 60 + int(startTime[4:6])
                    print(startTime)
                    print(datetime.timedelta(seconds=startTime))

                c = 1
                ws.cell(row=r, column=c, value=str(filename.split(dirName)[-1][1:-5]))
                c = c + 1
                for seg in segments:
                    if seg[0] == -1:
                        continue
                    # x = re.search('Kiwi', seg[4][0])
                    listformat = isinstance(seg[4], list)
                    if listformat and not species in seg[4][0]:
                        continue
                    if not listformat and not species in seg[4]:
                        continue
                    s = int(math.floor(seg[0]))
                    s = datetime.timedelta(seconds=startTime+s)
                    e = int(math.ceil(seg[1]))
                    e = datetime.timedelta(seconds=startTime+e)

                    ws.cell(row=r, column=c, value=str(s))
                    c = c+1
                    ws.cell(row=r, column=c, value=str(e))
                    c = c + 1
                    if listformat:
                        ws.cell(row=r, column=c, value=seg[4][0])
                    else:
                        ws.cell(row=r, column=c, value=seg[4])
                    c = 2
                    r = r + 1
    wb.save(str(eFile))
    print("Generated Report")

def length(dirName):
    """
    Find min, max, total duration of recordings in a dir.
    """
    durations = []
    for root, dirs, files in os.walk(str(dirName)):
        for filename in files:
            if filename.endswith('.wav'):
                filename = root + '/' + filename
                wavobj = wavio.read(filename)
                sampleRate = wavobj.rate
                data = wavobj.data
                duration = len(data) / sampleRate  # number of secs
                durations.append(duration)
    print("min duration: ", min(durations), " secs")
    print("max duration: ", max(durations), " secs")
    print("mean duration: ", np.mean(durations), " secs")
    print("median duration: ", np.median(durations), " secs")
    print("total duration: ", sum(durations), " secs")

def mp3ToWav(dirName):
    """
    MP3 to WAV conversion
    """
    from os import path
    from pydub import AudioSegment
    for root, dirs, files in os.walk(str(dirName)):
        for filename in files:
            if filename.endswith('.mp3'):
                src = root + '/' + filename
                sound = AudioSegment.from_mp3(src)
                sound.export(src[:-4]+'.wav', format="wav")

def selectRandom(dirName, n, outDir):
    '''
    Randomly select n .wav and their .data from input dirName to outDir
    Careful the hack srcWav[42:] matters
    '''
    import random
    import os
    from shutil import copyfile
    # Get the complete file list
    filelist = []
    filelist2 = []
    for root, dirs, files in os.walk(str(dirName)):
        for filename in files:
            if filename.endswith('.wav') and filename + '.data' in files:
                filelist2.append(filename)
                filename = root + '/' + filename
                filelist.append(filename)
    # Randomly select n files and copy them to out folder
    ind = list(range(len(filelist)))
    random.shuffle(ind)
    inds = ind[0:n]
    selctedFiles = []
    selctedFiles2 = []
    for index in inds:
        selctedFiles.append(filelist[index])
        selctedFiles2.append(filelist2[index])
    for name in selctedFiles:
        print(name)     # save a .txt having the files chosen
    for i in range(n):
        srcWav = selctedFiles[i]
        srcData = srcWav + '.data'
        dstWav = outDir + '\\' + srcWav[23:]
        dstData = dstWav + '.data'
        # dstWav = outDir + '\\' + srcWav[23:].replace("\\", "_")
        # dstWav = dstWav.replace("/", "_")
        copyfile(srcWav, dstWav)
        copyfile(srcData, dstData)

# -------------------------- Fiordland kiwi-----------------------------------------------------------
#####################################################################################################
# (1) run the wavelet filter
# (2) post process to merge segments < 3 sec apart
# (3) then delete segments < 3 sec long
# (4) review in batch mode, choose single species

def mergeneighbours(dir='', maxGap=3):
    """
    This will merge neighbour segments in the annotation
    """
    for root, dirs, files in os.walk(str(dir)):
        for file in files:
            if file.endswith('.data') and file[:-5] in files:   # skip GT annotations
                file = root + '/' + file
                meta = None
                with open(file) as f:
                    segments = json.load(f)
                    if len(segments) > 0 and segments[0][0] == -1:
                        meta = segments[0]
                        del (segments[0])
                    indx = []
                    chg = False
                    for i in range(len(segments) - 1):
                        if segments[i + 1][0] - segments[i][1] < maxGap:
                            indx.append(i)
                            chg = True
                    indx.reverse()
                    for i in indx:
                        segments[i][1] = segments[i + 1][1]
                        del (segments[i + 1])
                if chg:
                    if meta:
                        segments.insert(0, meta)
                    file = open(file, 'w')
                    json.dump(segments, file)

def deleteShort(dir='', minLen=2):
    """
    This will delete short segments from the annotation
    """
    for root, dirs, files in os.walk(str(dir)):
        for file in files:
            if file.endswith('.data') and file[:-5] in files:   # skip GT annotations
                file = root + '/' + file
                with open(file) as f:
                    segments = json.load(f)
                    newSegments = []
                    chg = False
                    for seg in segments:
                        if seg[0] == -1:
                            newSegments.append(seg)
                        elif seg[1]-seg[0] > minLen:
                            newSegments.append(seg)
                        else:
                            chg = True
                            continue
                if chg:
                    file = open(file, 'w')
                    json.dump(newSegments, file)

def anotherFilter(dir1, dir2):
    '''
    Post prcessing - rule based. Post process auto detections in dir1 and compare with its GT (Reviewed after initial
    filter) in dir2.
    :param dir1: folder to process
    :param dir2: GT reviewed (after waveelt -> merge plus delete short)
    :return:
    '''
    # just need the node set, use this data structure to be compatible with code
    speciesData = {'Name': 'Kiwi', 'SampleRate': 16000, "TimeRange": [6, 32], "FreqRange": [800, 8000],
                   "WaveletParams": [0.5, 0.5, [35, 36, 39, 40, 43, 44, 45]]}  # kiwi filter

    for root, dirs, files in os.walk(dir1):
        for file in files:
            if file.endswith('.wav') and file + '.data' not in files:
                print('Done')
            elif file.endswith('.wav') and file + '.data' in files:
                with open(root + '/' + file + '.data') as f1:
                    print(root + '/' + file)
                    segments = json.load(f1)
                    # also load after review .data
                    revFile = root.replace(dir1, dir2) + '/' + file + '.data'
                    if os.path.isfile(revFile):
                        with open(revFile) as f2:
                            segments_after = json.load(f2)
                    else:
                        segments_after = []
                    i = 0
                    Flag1 = False
                    if segments[0][0] == -1:
                        Eng = [None] * (len(segments) - 1)
                    else:
                        Eng = [None] * len(segments)
                    for seg in segments:
                        if seg[0] == -1:
                            continue
                        else:
                            #  Read the segment
                            wavobj = wavio.read(root + '/' + file, seg[1] - seg[0], seg[0])
                            sampleRate = wavobj.rate
                            data = wavobj.data
                            if data.dtype != 'float':
                                data = data.astype('float')  # / 32768.0
                            if np.shape(np.shape(data))[0] > 1:
                                data = np.squeeze(data[:, 0])
                            if sampleRate != 16000:
                                data = librosa.core.audio.resample(data, sampleRate, 16000)
                                sampleRate = 16000
                            # Wavelet energy
                            WS = WaveletSegment.WaveletSegment(speciesData)
                            WS.sampleRate = sampleRate
                            WS.data = data
                            WE = WS.computeWaveletEnergy(data, sampleRate, 5, 'new', window=1, inc=1)
                            WE = np.mean(WE, axis=1)
                            Eng[i] = [WE[node - 1] for node in speciesData['WaveletParams'][2]]
                            total = np.sum(Eng[i])
                            Eng[i] /= total
                            if (((Eng[i][0] > 0.2) or (Eng[i][1] > 0.2)) and (
                                    (Eng[i][4] > 0.2) or (Eng[i][5] > 0.2))) or (
                                    Eng[i][0] > 0.4 or Eng[i][4] > 0.3):  # Eng[i][4] > 0.3) could save more males
                                                                          # node 8 - Eng[i][7] > 0.14 made every segment kiwi!
                                print("seg", i + 1, " KIWI")
                            else:
                                print("seg", i + 1, " nah")
                            # get GT
                            Flag2 = False
                            for s in segments_after:
                                if seg[0] in s:
                                    print("KIWI <----------------------------")
                                    Flag1 = True
                                    Flag2 = True
                            if not Flag2:
                                print("nah")
                            i += 1
                    if Flag1:
                        mx = np.max(Eng)
                        md = np.median(Eng)

                        x = [n+1 for n in range(i)]
                        y = speciesData['WaveletParams'][2]
                        plt.style.use('ggplot')
                        valid_markers = ([item[0] for item in mks.MarkerStyle.markers.items() if
                                          item[1] is not 'nothing' and not item[1].startswith('tick') and not item[1].startswith(
                                              'caret')])
                        markers = np.random.choice(valid_markers, i, replace=True)
                        plt.figure()
                        # fig, ax = pl.subplots()
                        for j in x:
                            plt.plot(y, Eng[j-1], marker=markers[j-1], label='Seg' + str(x[j-1]))
                        plt.plot(y, np.ones(np.shape(y))*mx, 'k--')
                        plt.plot(y, np.ones(np.shape(y))*md, 'c--')
                        plt.legend(x)
                        # plt.savefig(root + '/' + file +'.png')
                        plt.show()

def remDuplicates(dir):
    '''
    Remove duplicate segs accidently made by two filters for the same spp
    '''
    for root, dirs, files in os.walk(str(dir)):
        for file in files:
            if file.endswith('.data'):
                print(file)
                with open(root + '/' + file) as f:
                    segments = json.load(f)
                    newSegments = []
                    chg = False
                    for seg in segments:
                        if seg[4] == ['Kiwi(Tokoeka Fiordland)?']:
                            seg[4] = ['Kiwi (Tokoeka Fiordland)?']
                            chg = True
                        elif seg[4] == ['Kiwi(Tokoeka Fiordland)']:
                            seg[4] = ['Kiwi (Tokoeka Fiordland)']
                            chg = True
                        if seg not in newSegments:
                            newSegments.append(seg)
                if chg:
                    print(root+file)
                    file = open(root + '/' + file, 'w')
                    json.dump(newSegments, file)

def remWeka(dir, spp):
    '''
    Remove specific segs (one species) from .data files
    '''
    for root, dirs, files in os.walk(str(dir)):
        for file in files:
            if file.endswith('.data'):
                print(file)
                with open(root + '/' + file) as f:
                    segments = json.load(f)
                    newSegments = []
                    chg = False
                    for seg in segments:
                        if seg[4] == [spp] or seg[4] == [spp + '?']:
                            chg = True
                        elif seg not in newSegments:
                            newSegments.append(seg)
                if chg:
                    print(root+file)
                    file = open(root + '/' + file, 'w')
                    json.dump(newSegments, file)


def FiordlandSummary(dir):
   """
   This will count .wav and .data files in each dir and count segments
   """
   for dirName, subdirList, fileList in os.walk(str(dir), topdown=True):
       # print('Found directory: %s' % dirName)
       Recorder = re.search('(_\d{8})', dirName.split("\\")[-1])
       if Recorder:
           print("Recorder: ", dirName.split("\\")[-1])
           segs_rec = 0
           lsegs_rec = 0
           done_rec = 0
           torev_rec = 0
           for dirName2, subdirlist2, fileList2 in os.walk(str(dirName), topdown=True):
               Date = re.search('(\d{6})', dirName2.split("\\")[-1])
               if Date and len(dirName2.split("\\")[-1]) == 8 and dirName2.split("\\")[-2] != 'Bat':
                   print("Date: ", dirName2.split("\\")[-1])
                   segs_date = 0
                   lsegs_date = 0
                   done_date = 0
                   torev_date = 0
                   for file in fileList2:
                       print('\t%s' % dirName2 + '\\' + file)
                       if file.endswith('.wav') and os.path.getsize(dirName2 + '/' + file) < 100*1000:
                           continue
                       if file.endswith('.wav') and file + '.data' not in fileList2:  # no need to review
                           done_date += 1
                           done_rec += 1
                       elif file.endswith('.wav') and file + '.data' in fileList2:
                           torev_date += 1
                           torev_rec += 1
                           with open(dirName2 + '/' + file + '.data') as f:
                               segments = json.load(f)
                               if len(segments) > 1 and segments[0][0] == -1:
                                   segs_date += len(segments) - 1
                                   segs_rec += len(segments) - 1
                               elif len(segments) > 1 and segments[0][0] != -1:
                                   segs_date += len(segments)
                                   segs_rec += len(segments)
                   print("# files completed - date: ", done_date)
                   print("# files to review - date: ", torev_date)
           print("# files completed - rec: ", done_rec)
           print("# files to review - rec: ", torev_rec)

def post_analysis(dir0, dir1, dir2, dir3):
    '''
    dir0: wavelets
    dir1: postproc1
    dir2: postproc2
    dir3: revieved
    '''
    print('Recorder: ', dir0.split("\\")[-1])
    tp0 = 0
    fp0 = 0
    tp1 = 0
    fp1 = 0
    fn1 = 0
    tp2 = 0
    fp2 = 0
    fn2 = 0
    for root, dirs, files in os.walk(str(dir0)):
        for file in files:
            if file.endswith('.data'):
                with open(root + '/' + file) as f:
                    segments0 = json.load(f)
                postproc1datafile = dir1 + root.split("_")[-1] + '\\' + file
                if os.path.isfile(postproc1datafile):
                    with open(postproc1datafile) as f1:
                        segments1 = json.load(f1)
                else:
                    segments1 = []
                postproc2datafile = dir2 + root.split("_")[-1] + '\\' + file
                if os.path.isfile(postproc2datafile):
                    with open(postproc2datafile) as f2:
                        segments2 = json.load(f2)
                else:
                    segments2 = []
                GTdatafile = dir3 + root.split("_")[-1] + '\\' + file
                if os.path.isfile(GTdatafile):
                    with open(GTdatafile) as f3:
                        segmentsGT = json.load(f3)
                else:
                    segmentsGT = []
                for seg in segments0:
                    if seg[0] == -1:
                        continue
                    else:   # check if TP or FP
                        flag0 = False
                        for s in segmentsGT:
                            if seg[0] in s:
                                tp0 += 1
                                flag0 = True
                        if not flag0:
                            fp0 += 1
                for seg in segments1:
                    if seg[0] == -1:
                        continue
                    else:   # check if TP or FP
                        flag1 = False
                        for s in segmentsGT:
                            if seg[0] in s:
                                tp1 += 1
                                flag1 = True
                        if not flag1:
                            fp1 += 1
                for seg in segments2:
                    if seg[0] == -1:
                        continue
                    else:  # check if TP or FP
                        flag2 = False
                        for s in segmentsGT:
                            if seg[0] in s:
                                tp2 += 1
                                flag2 = True
                        if not flag2:
                            fp2 += 1
                # now check for FNs by postproc1 and 2
                for seg in segmentsGT:
                    if seg[0] == -1:
                        continue
                    else:
                        flag11 =False
                        for s1 in segments1:
                            if seg[0] == s1[0]:
                                flag11 = True
                                continue
                        if not flag11:
                            fn1 += 1
                            print('\tFN by postproc1 : ', root+'\\'+file, seg)
                for seg in segmentsGT:
                    if seg[0] == -1:
                        continue
                    else:
                        flag22 =False
                        for s2 in segments2:
                            if seg[0] == s2[0]:
                                flag22 = True
                                continue
                        if not flag22:
                            fn2 += 1
                            print('\tFN by postproc2 : ', root+'\\'+file, seg)

    print('\tTP, FP left after wavelet:\t\t', tp0, '\t', fp0)
    print('\tTP, FP left after postproc1:\t', tp1, '\t', fp1)
    print('\tTP, FP left after postproc2:\t', tp2, '\t', fp2)
    print('\tFN after postproc1:\t\t\t', fn1)
    print('\tFN after postproc2:\t\t\t', fn2, '\n\n')


# post_analysis('E:\Fiordland_kiwi_2018\Lake_Thompson__01052018\\1Wavelets\SOUTH301353_01052018',
#               'E:\Fiordland_kiwi_2018\Lake_Thompson__01052018\Postproc1\SOUTH301353_',
#               'E:\Fiordland_kiwi_2018\Lake_Thompson__01052018\Postproc2\SOUTH301353_',
#               'E:\Fiordland_kiwi_2018\Lake_Thompson__01052018\Reviewed\SOUTH301353_')

def generateMLData(dir1, dir2, dir3):
    '''
    For each segment, record energies in each node, hoping to use them in ML
    :param dir1 - directory with .wav and .data (not reviewed)
    :param dir2 - directory with corresponding .data (reviewed)
    :return: saves a .txt file with energy + GT for each segment
    '''
    speciesData = {'Name': 'Kiwi', 'SampleRate': 16000, "TimeRange": [6, 32], "FreqRange": [800, 8000],
                   "WaveletParams": [0.5, 0.5, [35, 36, 39, 40, 43, 44, 45]]}  # kiwi filter

    for root, dirs, files in os.walk(dir1):
        for file in files:
            if file.endswith('.data'):
                with open(root + '/' + file) as f1:
                    print(root + '/' + file)
                    segments = json.load(f1)
                    # also load after review .data
                    root2 = root
                    root2 = root2.replace(dir1, dir2)
                    revFile = root2 + '/' + file
                    if os.path.isfile(revFile):
                        with open(revFile) as f2:
                            segments_GT = json.load(f2)
                    else:
                        segments_GT = []
                    i = 0
                    for seg in segments:
                        if seg[0] == -1:
                            continue
                        else:
                            #  Read the segment
                            if os.path.isfile(root2 + '/' + file[:-5]):
                                wavobj = wavio.read(root2 + '/' + file[:-5], seg[1] - seg[0], seg[0])
                                sampleRate = wavobj.rate
                                data = wavobj.data
                            else:
                                break
                            if data.dtype != 'float':
                                data = data.astype('float')  # / 32768.0
                            if np.shape(np.shape(data))[0] > 1:
                                data = np.squeeze(data[:, 0])
                            if sampleRate != 16000:
                                data = librosa.core.audio.resample(data, sampleRate, 16000)
                                sampleRate = 16000
                            # Wavelet energy
                            WS = WaveletSegment.WaveletSegment(speciesData)
                            WS.sampleRate = sampleRate
                            WS.data = data
                            WE = WS.computeWaveletEnergy(data, sampleRate, 5, 'new', window=1, inc=1)
                            WE = np.mean(WE, axis=1)
                            Eng = [WE[node - 1] for node in speciesData['WaveletParams'][2]]
                            total = np.sum(Eng)
                            Eng /= total
                            # TP? FP? get GT
                            Flag2 = False
                            for s in segments_GT:
                                if seg[0] in s:
                                    Flag2 = True    # TP
                                    tgt = 1
                            if not Flag2:
                                tgt = 0      # FP
                            # Write seg to file
                            with open(dir3 + '\\test1.csv', 'a+') as f:
                                f.write("%s\t%f\t%f\t%f\t%f\t%f\t%f\t%f\t%d\n" %(root + '/' + file, Eng[0], Eng[1], Eng[2], Eng[3], Eng[4], Eng[5], Eng[6], tgt))
                            i += 1

def extract_1min(indir, outdir):
    '''
    Extract and saves first min of each recordings - used for noise analysis
    '''
    for root, dirs, files in os.walk(indir):
        for file in files:
            if file.endswith('.wav'):
                try:
                    wavobj = wavio.read(root + '/' + file, 60, 0)
                    sampleRate = wavobj.rate
                    data = wavobj.data
                    outfile = outdir + '\\' + root.split(indir)[-1].replace('\\', '_') + '_' + file
                    wavio.write(outfile, data.astype('int16'), sampleRate, scale='dtype-limits', sampwidth=2)
                except:
                    pass

def classify_wind(dir, len):
    '''
    Trying to automatically filter windy (gusty) recordings in to correct directory - 'no wind', 'wind', 'gust'
    '''
    for root, dirs, files in os.walk(dir):
        for file in files:
            if file.endswith('.wav'):
                try:
                    wavobj = wavio.read(root + '/' + file, len, 1)
                    sampleRate = wavobj.rate
                    data = wavobj.data
                except:
                    pass
                if data.dtype != 'float':
                    data = data.astype('float')  # / 32768.0
                if np.shape(np.shape(data))[0] > 1:
                    data = np.squeeze(data[:, 0])
                # now check for wind
                # pp = SupportClasses.postProcess(audioData=data, sampleRate=sampleRate, segments=[], spInfo={})
                we_mean = np.zeros(len)
                we_std = np.zeros(len)
                for w in range(len):
                    dataw = data[int(w * sampleRate):int((w + 1) * sampleRate)]
                    f, p = signal.welch(dataw, fs=sampleRate, window='hamming', nperseg=512, detrend=False)
                    p = np.log10(p)
                    limsup = int(p.__len__() * 2 * 500 / sampleRate)
                    liminf = int(p.__len__() * 2 * 50 / sampleRate)
                    a_wind = p[liminf:limsup]
                    we_mean[w] = np.mean(a_wind)
                    we_std[w] = np.std(a_wind)

                if np.mean(we_mean) > 2:
                    print(file, '\t', np.mean(we_mean), '\t', 'Too noisy')
                    shutil.move(root + '/' + file, dir+'\\TooNoisy\\'+file)
                # elif np.mean(we_mean) > 1.5:
                #     print(file, '\t', np.mean(we_mean), '\t', 'Noisy')
                #     shutil.move(root + '/' + file, dir + '\\Noisy\\' + file)
                # else:
                #     print(file, '\t', 'Good')
                #     shutil.move(root + '/' + file, dir + '\\Good\\' + file)
    #             power.append(np.mean(we_mean))
    #             filelist.append(file)
    # for i in range(np.shape(power)[0]):
    #     print(filelist[i], '\t', power[i])
    # plt.plot(power)
    # plt.show()

# classify_wind('D:\\NoiseDataset\Dagg-Breaksea__24042018', 59)

def filter1_filter2(dir1, dir2):
    '''
    How many calls were deleted by second filter - Case study on Lake thompson 11 recorders reviewed after first filter
    '''
    tp = 0
    fn = 0
    for root, dirs, files in os.walk(dir1):
        for file in files:
            if file.endswith('.data'):
                with open(root + '/' + file) as f1:
                    segments1 = json.load(f1)
                with open(root.replace(dir1, dir2) + '/' + file) as f2:
                    segments2 = json.load(f2)
                for seg in segments1:
                    if len(seg[4]) > 1 and seg[0] != -1:
                        print(file, seg)
                    if seg[0] == -1 or seg[4][0] == 'Faint':
                        continue
                    elif "Kiwi (Tokoeka Fiordland)" in seg[4] or "Kiwi (Tokoeka Fiordland)?" in seg[4]:  # check if TP or FN
                        flag = False
                        for s in segments2:
                            if seg[0] in s:
                                tp += 1
                                flag = True
                        if not flag:
                            fn += 1
    print('TP = ', tp)
    print('FN =', fn)
    print('Postproc2 detected %d out of %d kiwi segments detected with waveletfilter->human review', tp, tp+fn)

def prepareCARTdata(dirAudio, dirEng):
    '''
    Basically to prepare a single .csv that includes WE features and GT - used for LSK CART data JJ.
    :param dirAudio: dir including recordings and GT
    :param dirEng: dir including corresponding WE
    :return: saves the .csv in dirEng
    '''
    for root, dirs, files in os.walk(dirAudio):
        for file in files:
            if file.endswith('.wav') and file[:-4] + '-sec.txt' in files:
                GT = []
                f1 = open(root + '/' + file[:-4] + '-sec.txt', "r")
                for line in f1:
                    if line != '\n':
                        fields = line.split('\t')
                        GT.append(fields[1])
                # now read the WE
                WE = np.zeros((33, 300))
                f2 = open(root.replace(dirAudio, dirEng) + '/' + file + '.senergies', "r")
                i = 0
                for line in f2:
                    if line != '\n':
                        fields = line.split('\t')
                        for j in range(len(fields)):
                            WE[i, j] = fields[j]
                    i += 1
                WE = WE.transpose()
                WE[:, -1] = GT
            else:
                continue
            # write it to .csv
            with open(dirEng + '\\senergy.tsv', 'a+') as f:
                for k in range(np.shape(WE)[0]):
                    for m in range(33):
                        f.write(str(WE[k, m]))
                        if m != 32:
                            f.write('\t')
                    f.write('\n')